#!/bin/python
# The main script of SiPe.Sat Risat harvester

# ---------------------------- DOCUMENTATION ---------------------------- #

# The main file of SiPe.Sat
# This file is unified: all classes are declared in this file instead of
# splitted into separate directories/files

# Created on 2023-03-20

# [1] WEB ARTICLES USED AS REFERENCES:
# 
# (On 2023-03-20)
#  1. Tkinter Application to Switch Between Different Page Frames
#   -> https://www.geeksforgeeks.org/tkinter-application-to-switch-between-different-page-frames/
#  2. Python GUI examples (Tkinter Tutorial)
#   -> https://likegeeks.com/python-gui-examples-tkinter-tutorial/
#  3. Python Tkinter – Entry Widget
#   -> https://www.geeksforgeeks.org/python-tkinter-entry-widget/
#  4. How to justify text in label in Tkinter
#   -> https://stackoverflow.com/questions/37318060
#  5. RadioButton in Tkinter | Python
#   -> https://www.geeksforgeeks.org/radiobutton-in-tkinter-python/
#  6. Tkinter Grid
#   -> https://www.pythontutorial.net/tkinter/tkinter-grid/
# 
# (On 2023-03-21)
#  7. How do I close a tkinter window?
#   -> https://stackoverflow.com/questions/110923
#  8. How to set the text/value/content of an `Entry` widget using a button in tkinter
#   -> https://stackoverflow.com/questions/16373887
#  9. Python Switch Statement – Switch Case Example
#   -> https://www.freecodecamp.org/news/python-switch-statement-switch-case-example/
# 10. Disable / Enable Button in TKinter
#   -> https://stackoverflow.com/questions/53580507
#
# (On 2023-03-28)
# 11. How to change the Tkinter label text?
#   -> https://www.geeksforgeeks.org/how-to-change-the-tkinter-label-text/
# 12. How to get the text out of a scrolledtext widget?
#   -> https://stackoverflow.com/questions/53937400
# 13. Create temporary files and directories using Python-tempfile
#   -> https://www.geeksforgeeks.org/create-temporary-files-and-directories-using-python-tempfile/
# 14. Python path separator [duplicate]
#   -> https://stackoverflow.com/a/50738724
#
# (On 2023-04-04)
# 15. Xpath cheatsheet
#   -> https://devhints.io/xpath
# 16. Python Tkinter - How to set the default value of a radiobutton?
#   -> https://stackoverflow.com/a/51544948
# 17. Limiting python filedialog to a specific filetype
#   -> https://stackoverflow.com/a/46339932
# 18. How to place a default file name in the file dialog asksaveasfile python
#   -> https://stackoverflow.com/a/67092469
# 19. How to easily avoid Tkinter freezing?
#   -> https://stackoverflow.com/a/67489469
#
# (On 2023-04-11)
# 20. Openpyxl: How to merge cells using variable rows
#   -> https://stackoverflow.com/a/56278226
# 21. Python Program to Merge Excel Cells using openpyxl
#   -> https://www.codespeedy.com/python-program-to-merge-excel-cell/
# 22. XPATH: exclude elements which which has a specific child
#   -> https://stackoverflow.com/questions/44943638
# 23. XPath Select Nodes where all parent nodes do not contain specific attribute and value
#   -> https://stackoverflow.com/questions/17191971
# 24. Python: Using xpath locally / on a specific element
#   -> https://stackoverflow.com/a/4785929
#
# (On 2023-07-25)
# 25. Dropdown Menus – Tkinter
#   -> https://www.geeksforgeeks.org/dropdown-menus-tkinter/
# 26. Correct way to set scrollbar position in python tkinter
#   -> https://stackoverflow.com/questions/53157859

# [2] CODING CONVENTIONS:
# 1. Use single quote (') instead of double quotes (") when specifying strings
# 2. Equal signs used as function argument are not wrapped by empty space bars
# 3. 'SipesatScr...' class name prefix indicates a class of the superclass 'tk.Frame'

# [3] VARIABLE CONVENTIONS:
# 1. (harvest_category) 'Kategori'
#  -> ['r'] = The Risat Research menu category
#  -> ['c'] = The Risat ComService menu category
# 2. (harvest_datatype) 'Jenis Data' radio button in the harvester menu has the following possible IntVar values:
#  -> [0] = Data Ringkasan
#  -> [1] = Data Detil Lengkap
# 3. (harvest_output) 'Data Hasil Panenan' radio button in the harvester menu has the following possible StringVar values:
#  -> ['usulan'] = Risat Usulan Penelitian/AbdiMas
#  -> ['berdisetujui'] = Risat Berkas Penelitian/AbdiMas Disetujui
#  -> ['berdirevisi'] = Risat Berkas Penelitian/AbdiMas Direvisi
#  -> ['berditolak'] = Risat Berkas Penelitian/AbdiMas Ditolak
#  -> ['dana'] = Risat Dana Penelitian/AbdiMas
#  -> ['arsip'] = Risat Arsip Penelitian/AbdiMas
#  -> ['lapakhir'] = = Risat Laporan Akhir Penelitian/AbdiMas

# [4] APPLICATION MECHANISM:
# 1. Upon successful login, the username & password credentials are stored
#    as variables inside the 'controller' class instance.
#    These variables will then be flushed out upon loggin out.

# [5] "RISAT DATA_PROMPT" ARRAY CONVENTIONS:
# data_prompt = {
#   'http_response'     --> the http_response after posting
#   'html_content'      --> 'http_response', converted into an XML-compatible HTML string
#   'viewstate'         --> the view state hidden ASPX form value (e.g. "/wEPaA8FDzhkYjBkY2I5ODVlM2YzZmTXigQ6dU1GUsc1Dgno6Z11")
#   'viewstategen'      --> the viewstategen hidden ASPX form value (e.g. "28239525")
#   'eventvalidation'   --> the eventvalidation hidden ASPX form value (e.g. "/wEdABiKTaCYJNZ8hl3vzC5OJBTzquwmUN/b9MQs90SJn/")
#   'button_name'       --> the entry row's submit button 'name' attribute (e.g. "ctl00$ContentPlaceHolder1$danacair1$repusulan1$ctl01$btdetil1")
#                           this data array data is used both in 'dana detil' and 'arsip detil' pages
#   'kodetran_prop'     --> the entry row's hidden 'kodetran' tag attribute name (e.g. "ctl00$ContentPlaceHolder1$danacair1$repusulan1$ctl01$kodetran1")
#                           this data array data is used only in 'dana detil' pages
#   'kodetran_val'      --> the entry row's hidden 'kodetran' tag attribute value (e.g. "842BA8DC-F7FA-48CE-A8B8-3106876A3B1E")
#                           this data array data is used only in 'dana detil' pages
#   'stat'              --> backward compatibility of 'stat_prop'
#   'stat_prop'         --> the entry row's hidden 'stat' tag attribute name (e.g. "ctl00$ContentPlaceHolder1$danacair1$repusulan1$ctl01$stat1")
#                           this data array data is used only in 'dana detil' pages
#   'stat_val'          --> the entry row's hidden 'stat' tag attribute value (e.g. "M" for "belum terealisasi, or "C" for "terealisasi")
#                           this data array data is used only in 'dana detil' pages
#   'idat_prop'         --> the entry row's hidden 'data' tag attribute value (e.g. "ctl00$ContentPlaceHolder1$arsip1$repusul1$ctl01$idat1")
#                           this data array data is used only in 'arsip detil' pages
#   'idat_val'          --> the entry row's hidden 'data' tag attribute value (e.g. "12AAFE6D-7C03-4998-9CB6-1BF4ECA37831")
#                           this data array data is used only in 'arsip detil' pages
#   'itgl_prop'         --> the entry row's hidden 'tanggal' tag attribute value (e.g. "ctl00$ContentPlaceHolder1$arsip1$repusul1$ctl01$itgl1")
#                           this data array data is used only in 'arsip detil' pages
#   'itgl_val'          --> the entry row's hidden 'tanggal' tag attribute value (e.g. "2021/01/21")
#                           this data array data is used only in 'arsip detil' pages
# }

# [6] REFERENCES TO FILES INSIDE THE AUTHOR'S PERSONAL COMPUTER
#  1. The harvester script that gets past the Risat login page
#   -> /ssynthesia/ghostcity/ar/dumper-2/24__2023.02.13__requestsrisat.py

# [7] BACKENDHARVESTER RUN FUNCTIONS CONVETIONS
# Under the class BackEndHarvester, there are functions which look like
# the following:
# - run_harvest_r_0_dana()
# - run_harvest_r_1_data()
# - run_harvest_c_0_arsip()
# - run_harvest_c_0_lapakhir()
# - etc.
#
# These functions represent the radiobutton selection in the classes
# SipesatScrResearch and SipesatScrComService.
# The codes 'r', 'c', '0', 'dana', etc. are representatives of variable
# convetions as found in convention [3]
#
# Please refer to convention [3] for navigating in
# these BackEndHarvester functions

# [8] IMPORTANT TRICKS
#  1. Avoiding Tkinter freezing
#   -> Add the following code block
#   -> every time an element is changed:
#      self.update()  # --- avoids freezing
#  2. Manipulation of cell values
#   -> The two are equivalent:
#      sheet['A1'] = 'Boo!'
#      sheet.cell(row=1, column=1).value = 'Boo!'

# --------------------------- CODE PREAMBLE --------------------------- #

# Modules import
from datetime import datetime as dt
from lxml import html
from openpyxl.styles import Alignment
from os.path import sep
from tkinter import filedialog
from tkinter import IntVar
from tkinter import messagebox
from tkinter import OptionMenu
from tkinter import scrolledtext
from tkinter import StringVar
from tkinter import ttk
import openpyxl as xl
import requests as rq
import tempfile as tmp
import tkinter as tk

# ------------------------ CLASSES DECLARATION ------------------------ #

# Main class declaration
# This class is called first to manage frames and Tkinter instances
class MainGUI(tk.Tk):
    
    # The init function for the MainGUI class
    def __init__(self, *args, **kwargs):

        # Preamble logging
        print('[MainGUI] :: Starting the app ...')
        
        # Instantiating Tkinter instance
        tk.Tk.__init__(self, *args, **kwargs)
        
        # Configuring app's identity
        self.title(APP_NAME)
        
        # Establishing the GUI container
        container = tk.Frame(self)
        container.pack(side='top', fill='both', expand=True)
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)
        
        # Setting up the list of frames
        self.frames = {}
        
        # Initializing the credential variables and harvester arguments
        self.password = ''
        self.username = ''
        self.harvest_category = ''
        self.harvest_datatype = -1
        self.harvest_output = ''
        
        # Iterating through frame list
        for F in FRAME_CLASSES:
            
            # Initializing the frame
            frame = F(container, self)
            
            # Assigning the initialized frame into
            # the frame list
            self.frames[F] = frame
            
            # Setting the frame's layout
            # Sticky: 'nsew' stands for 'north-south-east-west'
            frame.grid(row=0, column=0, sticky='nsew')
        
        # By default displaying the first frame class specified in
        # the tuple 'FRAME_CLASSES'
        # This will be called upon at the beginning of the app
        self.raise_frame(FRAME_CLASSES[0])
        
    # The function that displays the frame
    def raise_frame(self, frame_choice):
        frame = self.frames[frame_choice]
        frame.tkraise()

# The login frame
# This screen asks Risat login credentials, i.e., staff ID and password
class SipesatScrLogin(tk.Frame):
    
    # The function that will be triggered when the 'ABOUT' button is pressed
    def on_about_button_click(self):
        about_title = 'TENTANG'
        about_content = 'SiPe.Sat -- Sistem Pemanen Risat Satya Wacana\n\nDibuat oleh Samarthya Lykamanuella\n\nDirektorat Riset dan Pengabdian Masyarakat (DRPM)\n\nUniversitas Kristen Satya Wacana\n\n@ 2023'
        # Showing the info box
        messagebox.showinfo(about_title, about_content)
    
    # The function that will be triggered when the 'LICENSE' button is pressed
    def on_license_button_click(self):
        license_title = 'LISENSI'
        license_content = 'Copyright (c) 2023 Samarthya Lykamanuella. All rights reserved.'
        messagebox.showinfo(license_title, license_content)
    
    # The function that will be triggered when the 'EXIT' button is pressed
    def on_exit_button_click(self):
        # Showing a message box that prompts the user whether to leave the app
        exit_title = 'KELUAR APLIKASI'
        exit_content = 'Apakah Anda yakin untuk keluar dari aplikasi?'
        res = messagebox.askquestion(exit_title, exit_content)
        if res == 'yes':
            # The user confirms to leave the app
            print('[SipesatScrLogin] :: See You :(')
            self.controller.destroy()
        else:
            # Just do nothing
            pass
    
    # The function that will be triggered when the 'LOGIN' form button is pressed
    def on_submit_button_click(self, username, password):

        # Avoiding Tkinter freezing
        self.update()  # --- avoids freezing
        
        # Validating the input credentials
        validation = BackEndLoginChecker()
        res = validation.validate(username, password)
        
        if res:
            # Login successful
            login_success_title = 'LOGIN BERHASIL'
            login_success_content = f'Anda telah berhasil log masuk.\nSelamat datang, {username}!'
            messagebox.showinfo(login_success_title, login_success_content)
            
            # Upon successful login, store the login credentials inside the 'controller'`s
            # class-wide variable
            self.controller.password = password
            self.controller.username = username
            print('[SipesatScrLogin] :: Credential variables set!')
            
            # Upon successful login, clear the password input
            self.pass_input.delete(0, tk.END)
            
            # Upon successful login, switch to main menu screen
            self.controller.raise_frame(SipesatScrMainMenu)
        else:
            # Login failed
            login_failed_title = 'LOGIN GAGAL'
            login_failed_content = 'Nama pengguna atau kata sandi salah!\n\nMungkin juga disebabkan karena koneksi internet yang bermasalah\n\nSilahkan coba lagi'
            messagebox.showerror(login_failed_title, login_failed_content)
    
    # The __init__ function
    def __init__(self, parent, controller):
        
        # Instantiating tkinter.Frame instance
        tk.Frame.__init__(self, parent)
        
        # Assigning the variables for use by other methods in the class
        self.parent = parent
        self.controller = controller
        
        # The header title displaying the screen's title 
        header_title = ttk.Label(self, text=STRING_HEADER_TITLE, font=FONT_HEADER_TITLE)
        header_title.grid(row=0, column=0, padx=10, pady=2)
        
        # The header description displaying additional information about the screen
        header_desc = ttk.Label(self, text=STRING_HEADER_DESC, font=FONT_HEADER_DESC)
        header_desc.grid(row=1, column=0, padx=10, pady=5)
        
        # The login display help
        help_text = 'Selamat datang di sistem pemanen Risat Satya Wacana!\nSilahkan masukkan kredensial berikut ini:'
        help_label = ttk.Label(self, text=help_text, font=FONT_REGULAR, anchor='w', justify='left')
        help_label.grid(row=2, column=0, padx=10, pady=10)
        
        # :::
        # The login form frame
        form_frame = tk.Frame(self)
        form_frame.grid(row=3, column=0, padx=10, pady=10)
        # Username display help
        user_text = 'ID Pegawai'
        user_label = ttk.Label(form_frame, text=user_text,)
        user_label.grid(row=0, column=0, padx=2, pady=2)
        # Username input prompt
        user_input = ttk.Entry(form_frame, width=50)
        user_input.grid(row=0, column=1, padx=2, pady=2)
        # Password display help
        pass_text = 'Password'
        pass_label = ttk.Label(form_frame, text=pass_text,)
        pass_label.grid(row=1, column=0, padx=2, pady=2)
        # Password input prompt
        pass_input = ttk.Entry(form_frame, width=50, show='*')
        pass_input.grid(row=1, column=1, padx=2, pady=2)
        self.pass_input = pass_input # --- for uses by other functions in the class
        # The submit button
        submit_text = 'LOGIN'
        submit_button = ttk.Button(form_frame, text=submit_text,
            command=lambda: self.on_submit_button_click(user_input.get(), pass_input.get()))
        submit_button.grid(row=2, column=0, padx=2, pady=2)
        
        # :::
        # The footer buttons frame container
        footer_frame = tk.Frame(self)
        footer_frame.grid(row=4, column=0, padx=10, pady=10)
        # The 'about' page button trigger
        about_text = 'TENTANG'
        about_button = ttk.Button(footer_frame, text=about_text,
            command=lambda: self.on_about_button_click())
        about_button.grid(row=0, column=0, padx=2, pady=2)
        # The 'license' page button trigger
        license_text = 'LISENSI'
        license_button = ttk.Button(footer_frame, text=license_text,
            command=lambda: self.on_license_button_click())
        license_button.grid(row=0, column=1, padx=2, pady=2)
        # The 'exit' page button trigger
        exit_text = 'KELUAR'
        exit_button = ttk.Button(footer_frame, text=exit_text,
            command=lambda: self.on_exit_button_click())
        exit_button.grid(row=0, column=2, padx=2, pady=2)

# The main menu
# This screen is displayed upon successful login
# Each button in this screen directs the user into the harvester menu
# of each category, e.g., 'Research' and 'Community Service'
class SipesatScrMainMenu(tk.Frame):
    
    # The function that will be triggered when the menu [1] button is selected/clicked
    def on_menu_1_button_click(self):
        self.controller.raise_frame(SipesatScrResearch)
    
    # The function that will be triggered when the menu [2] button is selected/clicked
    def on_menu_2_button_click(self):
        self.controller.raise_frame(SipesatScrComService)
    
    # The function that will be triggered when the logout button is selected/clicked
    def on_logout_button_click(self):
        # Showing a message box that prompts the user whehter to logout from the current session
        logout_title = 'LOG OUT'
        logout_content = 'Apakah Anda yakin untuk log keluar dari sesi ini?'
        res = messagebox.askquestion(logout_title, logout_content)
        if res == 'yes':
            # Upon logging out, flush out the credential variables
            self.controller.password = ''
            self.controller.username = ''
            print('[SipesatScrLogin] :: Credential variables flushed out!')
            
            # The user confirms to logout
            self.controller.raise_frame(SipesatScrLogin)
        else:
            # Just do nothing
            pass
    
    # The __init__ function
    def __init__(self, parent, controller):
        
        # Instantiating tkinter.Frame instance
        tk.Frame.__init__(self, parent)
        
        # Assigning the variables for use by other methods in the class
        self.parent = parent
        self.controller = controller
        
        # The header title displaying the screen's title 
        header_title = ttk.Label(self, text=STRING_HEADER_TITLE, font=FONT_HEADER_TITLE)
        header_title.grid(row=0, column=0, padx=10, pady=2)
        
        # The header description displaying additional information about the screen
        desc_text = 'Menu Utama'
        header_desc = ttk.Label(self, text=desc_text, font=FONT_HEADER_DESC)
        header_desc.grid(row=1, column=0, padx=10, pady=5)
        
        # The screen's display help
        help_text = 'Selamat datang di menu utama pemanen Risat Satya Wacana!\nPilih tombol di bawah ini untuk melanjutkan pemanenan'
        help_label = ttk.Label(self, text=help_text, font=FONT_REGULAR, anchor='w', justify='left')
        help_label.grid(row=2, column=0, padx=10, pady=10)
        
        # :::
        # The button layout frame
        layout_frame = tk.Frame(self)
        layout_frame.grid(row=3, column=0, padx=10, pady=10)
        # Menu [1]: Research
        menu_1_text = 'Risat Penelitian'
        menu_1_button = ttk.Button(layout_frame, text=menu_1_text, width=45,
            command=lambda: self.on_menu_1_button_click())
        menu_1_button.grid(row=0, column=0, padx=2, pady=2)
        # Menu [2]: Comm. Service
        menu_2_text = 'Risat Pengabdian Masyarakat'
        menu_2_button = ttk.Button(layout_frame, text=menu_2_text, width=45,
            command=lambda: self.on_menu_2_button_click())
        menu_2_button.grid(row=1, column=0, padx=2, pady=2)
        
        # Logout button
        logout_text = 'LOG KELUAR'
        logout_button = ttk.Button(self, text=logout_text, width=30,
            command=lambda: self.on_logout_button_click())
        logout_button.grid(row=4, column=0, padx=5, pady=5)

# The harvester menu for 'research'
# - Displayed before the actual harvesting process is executed, for
# preconfiguring what kind of data and which database to be harvested
# - Called upon following the 'SipesatScrMainMenu' screen
class SipesatScrResearch(tk.Frame):
    
    # The function that will be triggered when the 'kembali' button is hit
    def on_back_button_click(self):
        self.controller.raise_frame(SipesatScrMainMenu)
    
    # The function that will be triggered when the 'lanjut' button is hit
    def on_next_button_click(self):
        # Interpreting the value of the dropdown menu
        l = self.dropdown_harvest_type.get()  # --- temporary string
        match l:
            case 'Risat Usulan Penelitian':
                m = 'usulan'
            case 'Risat Berkas Disetujui Penelitian':
                m = 'berdisetujui'
            case 'Risat Berkas Direvisi Penelitian':
                m = 'berdirevisi'
            case 'Risat Telah Direview Penelitian':
                m = 'teldireview'
            case 'Risat Disetujui Penelitian':
                m = 'disetujui'
            case 'Risat Disetujui Dgn Revisi (DDR) Penelitian':
                m = 'ddr'
            case 'Risat Ditolak Penelitian':
                m = 'ditolak'
            case 'Risat Dana Penelitian':
                m = 'dana'
            case 'Risat Arsip Penelitian':
                m = 'arsip'
            case 'Risat Laporan Akhir Penelitian':
                m = 'lapakhir'
            case _:
                no_dropdown_selection_title = 'GALAT'
                no_dropdown_selection_content = 'Silahkan pilih menu untuk melanjutkan!'
                messagebox.showerror(no_dropdown_selection_title, no_dropdown_selection_content)
                return

        # Setting the harvest arguments using the controller's 'self' variables
        self.controller.harvest_category = 'r'
        self.controller.harvest_datatype = self.radio_data_type.get()
        self.controller.harvest_output = m
        
        # Calling the harvester screen, begin the harvesting process
        self.controller.raise_frame(SipesatScrHarvest)
    
    # The __init__ function
    def __init__(self, parent, controller):
        
        # The radiobutton variables, used to group different radiobutton together
        self.radio_data_type = IntVar()
        self.dropdown_harvest_type = StringVar()
        
        # Instantiating tkinter.Frame instance
        tk.Frame.__init__(self, parent)
        
        # Assigning the variables for use by other methods in the class
        self.parent = parent
        self.controller = controller
        
        # The header title displaying the screen's title 
        header_title = ttk.Label(self, text=STRING_HEADER_TITLE, font=FONT_HEADER_TITLE)
        header_title.grid(row=0, column=0, padx=10, pady=2)
        
        # The header description displaying additional information about the screen
        desc_text = 'Menu Utama » Risat Penelitian'
        header_desc = ttk.Label(self, text=desc_text, font=FONT_HEADER_DESC)
        header_desc.grid(row=1, column=0, padx=10, pady=5)
        
        # The screen's display help
        help_text = 'Silahkan tentukan basis data dan jenis data yang akan dipanen\nTekan "LANJUT" untuk memulai proses pemanenan data Risat'
        help_label = ttk.Label(self, text=help_text, font=FONT_REGULAR, anchor='w', justify='left')
        help_label.grid(row=2, column=0, padx=10, pady=10)
        
        # :::
        # Layout [1]: Data type
        layout_frame_1 = tk.Frame(self)
        layout_frame_1.grid(row=3, column=0, padx=10, pady=10)
        # The help label
        datatype_text = 'Jenis Data'
        datatype_label = ttk.Label(layout_frame_1, text=datatype_text, font=FONT_REGULAR)
        datatype_label.grid(row=0, column=0, padx=2, pady=5, columnspan=2)
        # Radio button 1 -- 'Summary'
        datatype_summary_text = 'Ringkasan Data'
        datatype_summary_radio = ttk.Radiobutton(layout_frame_1, text=datatype_summary_text,
            value=0, variable=self.radio_data_type)
        datatype_summary_radio.grid(row=1, column=0, padx=2, pady=2, sticky='w')
        # Radio button 2 -- 'Detailed data'
        datatype_detailed_text = 'Data Detil Lengkap'
        datatype_detailed_radio = ttk.Radiobutton(layout_frame_1, text=datatype_detailed_text,
            value=1, variable=self.radio_data_type)
        datatype_detailed_radio.grid(row=1, column=1, padx=2, pady=2, sticky='w')
        # Setting default radiobutton value
        self.radio_data_type.set(0)
        
        # :::
        # Layout [2]: Harvested data output
        layout_frame_2 = tk.Frame(self)
        layout_frame_2.grid(row=4, column=0, padx=10, pady=10)
        # The help label
        harvesttype_text = 'Data Hasil Panenan'
        harvesttype_label = ttk.Label(layout_frame_2, text=harvesttype_text, font=FONT_REGULAR)
        harvesttype_label.grid(row=0, column=0, padx=2, pady=5)
        # Dropdown that selects the Risat menu to harvest
        harvesttype_options = [
            'Risat Usulan Penelitian',
            'Risat Berkas Disetujui Penelitian',
            'Risat Berkas Direvisi Penelitian',
            'Risat Telah Direview Penelitian',
            'Risat Disetujui Penelitian',
            'Risat Disetujui Dgn Revisi (DDR) Penelitian',
            'Risat Ditolak Penelitian',
            'Risat Dana Penelitian',
            'Risat Laporan Akhir Penelitian',
            'Risat Arsip Penelitian'
        ]
        self.dropdown_harvest_type.set('--Pilih--')
        harvesttype_dropdown = OptionMenu(layout_frame_2, self.dropdown_harvest_type, *harvesttype_options)
        harvesttype_dropdown.grid(row=4, column=0, padx=2, pady=2, sticky='w')
        
        # :::
        # Layout [3]: The action buttons
        layout_frame_3 = tk.Frame(self)
        layout_frame_3.grid(row=5, column=0, padx=10, pady=10)
        # The 'back' button trigger
        # Clicking this button calls out 'SipesatScrMainMenu'
        back_text = 'KEMBALI'
        back_button = ttk.Button(layout_frame_3, text=back_text, width=40,
            command=lambda: self.on_back_button_click())
        back_button.grid(row=0, column=0, padx=2, pady=2)
        # The 'next' button trigger
        # Clicking this button proceeds the program to the harvester screen
        next_text = 'LANJUT'
        next_button = ttk.Button(layout_frame_3, text=next_text, width=40,
            command=lambda: self.on_next_button_click())
        next_button.grid(row=0, column=1, padx=2, pady=2)

# The harvester menu for 'community service'
# - Displayed before the actual harvesting process is executed, for
# preconfiguring what kind of data and which database to be harvested
# - Called upon following the 'SipesatScrMainMenu' screen
class SipesatScrComService(tk.Frame):
    
    # The function that will be triggered when the 'kembali' button is hit
    def on_back_button_click(self):
        self.controller.raise_frame(SipesatScrMainMenu)
    
    # The function that will be triggered when the 'lanjut' button is hit
    def on_next_button_click(self):
        # Interpreting the value of the dropdown menu
        l = self.dropdown_harvest_type.get()  # --- temporary string
        match l:
            case 'Risat Usulan Pengabdian Masyarakat':
                m = 'usulan'
            case 'Risat Berkas Disetujui Pengabdian Masyarakat':
                m = 'berdisetujui'
            case 'Risat Berkas Direvisi Pengabdian Masyarakat':
                m = 'berdirevisi'
            case 'Risat Ditolak Pengabdian Masyarakat':
                m = 'ditolak'
            case 'Risat Dana Pengabdian Masyarakat':
                m = 'dana'
            case 'Risat Arsip Pengabdian Masyarakat':
                m = 'arsip'
            case 'Risat Laporan Akhir Pengabdian Masyarakat':
                m = 'lapakhir'
            case _:
                no_dropdown_selection_title = 'GALAT'
                no_dropdown_selection_content = 'Silahkan pilih menu untuk melanjutkan!'
                messagebox.showerror(no_dropdown_selection_title, no_dropdown_selection_content)
                return

        # Setting the harvest arguments using the controller's 'self' variables
        self.controller.harvest_category = 'c'
        self.controller.harvest_datatype = self.radio_data_type.get()
        self.controller.harvest_output = m
        
        # Calling the harvester screen, begin the harvesting process
        self.controller.raise_frame(SipesatScrHarvest)
    
    # The __init__ function
    def __init__(self, parent, controller):
        
        # The radiobutton variables, used to group different radiobutton together
        self.radio_data_type = IntVar()
        self.dropdown_harvest_type = StringVar()
        
        # Instantiating tkinter.Frame instance
        tk.Frame.__init__(self, parent)
        
        # Assigning the variables for use by other methods in the class
        self.parent = parent
        self.controller = controller
        
        # The header title displaying the screen's title 
        header_title = ttk.Label(self, text=STRING_HEADER_TITLE, font=FONT_HEADER_TITLE)
        header_title.grid(row=0, column=0, padx=10, pady=2)
        
        # The header description displaying additional information about the screen
        desc_text = 'Menu Utama » Risat Pengabdian Masyarakat'
        header_desc = ttk.Label(self, text=desc_text, font=FONT_HEADER_DESC)
        header_desc.grid(row=1, column=0, padx=10, pady=5)
        
        # The screen's display help
        help_text = 'Silahkan tentukan basis data dan jenis data yang akan dipanen\nTekan "LANJUT" untuk memulai proses pemanenan data Risat'
        help_label = ttk.Label(self, text=help_text, font=FONT_REGULAR, anchor='w', justify='left')
        help_label.grid(row=2, column=0, padx=10, pady=10)
        
        # :::
        # Layout [1]: Data type
        layout_frame_1 = tk.Frame(self)
        layout_frame_1.grid(row=3, column=0, padx=10, pady=10)
        # The help label
        datatype_text = 'Jenis Data'
        datatype_label = ttk.Label(layout_frame_1, text=datatype_text, font=FONT_REGULAR)
        datatype_label.grid(row=0, column=0, padx=2, pady=5, columnspan=2)
        # Radio button 1 -- 'Summary'
        datatype_summary_text = 'Ringkasan Data'
        datatype_summary_radio = ttk.Radiobutton(layout_frame_1, text=datatype_summary_text,
            value=0, variable=self.radio_data_type)
        datatype_summary_radio.grid(row=1, column=0, padx=2, pady=2, sticky='w')
        # Radio button 2 -- 'Detailed data'
        datatype_detailed_text = 'Data Detil Lengkap'
        datatype_detailed_radio = ttk.Radiobutton(layout_frame_1, text=datatype_detailed_text,
            value=1, variable=self.radio_data_type)
        datatype_detailed_radio.grid(row=1, column=1, padx=2, pady=2, sticky='w')
        # Setting default radiobutton value
        self.radio_data_type.set(0)
        
        # :::
        # Layout [2]: Harvested data output
        layout_frame_2 = tk.Frame(self)
        layout_frame_2.grid(row=4, column=0, padx=10, pady=10)
        # The help label
        harvesttype_text = 'Data Hasil Panenan'
        harvesttype_label = ttk.Label(layout_frame_2, text=harvesttype_text, font=FONT_REGULAR)
        harvesttype_label.grid(row=0, column=0, padx=2, pady=5)
        # Dropdown that selects the Risat menu to harvest
        harvesttype_options = [
            'Risat Usulan Pengabdian Masyarakat',
            'Risat Berkas Disetujui Pengabdian Masyarakat',
            'Risat Berkas Direvisi Pengabdian Masyarakat',
            'Risat Ditolak Pengabdian Masyarakat',
            'Risat Dana Pengabdian Masyarakat',
            'Risat Laporan Akhir Pengabdian Masyarakat',
            'Risat Arsip Pengabdian Masyarakat'
        ]
        self.dropdown_harvest_type.set('--Pilih--')
        harvesttype_dropdown = OptionMenu(layout_frame_2, self.dropdown_harvest_type, *harvesttype_options)
        harvesttype_dropdown.grid(row=4, column=0, padx=2, pady=2, sticky='w')
        
        # :::
        # Layout [3]: The action buttons
        layout_frame_3 = tk.Frame(self)
        layout_frame_3.grid(row=5, column=0, padx=10, pady=10)
        # The 'back' button trigger
        # Clicking this button calls out 'SipesatScrMainMenu'
        back_text = 'KEMBALI'
        back_button = ttk.Button(layout_frame_3, text=back_text, width=40,
            command=lambda: self.on_back_button_click())
        back_button.grid(row=0, column=0, padx=2, pady=2)
        # The 'next' button trigger
        # Clicking this button proceeds the program to the harvester screen
        next_text = 'LANJUT'
        next_button = ttk.Button(layout_frame_3, text=next_text, width=40,
            command=lambda: self.on_next_button_click())
        next_button.grid(row=0, column=1, padx=2, pady=2)

# The harvester screen
# Displaying the current progress of the harvesting process
class SipesatScrHarvest(tk.Frame):
    
    # The function that will be triggered when the 'cancel' button is selected/clicked
    def on_cancel_button_click(self):
        print('[SipesatScrHarvest] :: Operation cancelled!')
        if self.controller.harvest_category == 'c':
            self.controller.raise_frame(SipesatScrComService)
        elif self.controller.harvest_category == 'r':
            self.controller.raise_frame(SipesatScrResearch)
    
    # The function that will be triggered when the 'start' button is selected/clicked
    def on_start_button_click(self):
        # Changes the helper display
        self.set_help_label('Data sedang dipanen. Silahkan menunggu.')
        # Disables the cancel button
        self.cancel_button.config(state='disabled')
        self.update()  # --- avoids freezing
        # Disables the start button
        self.start_button.config(state='disabled')
        self.update()  # --- avoids freezing
        # Calling the back-end harvester class
        harvester = BackEndHarvester()
        # Begin the harvesting process
        harvester.execute_harvester(
            self, # --- the control object, so that this screen's progress bar values,
                  # etc., can be manipulated by the harvest executor
            self.controller.username,
            self.controller.password,
            self.controller.harvest_category,
            self.controller.harvest_datatype,
            self.controller.harvest_output
        )

    # The function that will be triggered when the 'start' button,
    # which has been replaced with 'finish' button, is selected/clicked
    def on_finish_button_click(self):
        print('[SipesatScrHarvest] :: Finished the harvesting operation!')
        if self.controller.harvest_category == 'c':
            self.controller.raise_frame(SipesatScrComService)
        elif self.controller.harvest_category == 'r':
            self.controller.raise_frame(SipesatScrResearch)
        # Enables the cancel button
        self.cancel_button.config(state='normal')
        self.update()  # --- avoids freezing
        # Resets the helper text, message area, and the progress bar value
        self.set_progress_bar(0)
        self.clear_message_area()
        self.set_help_label('Tekan tombol "MULAI PANEN" untuk memulai proses pemanenan data Risat')
        self.update()  # --- avoids freezing
        # Reset the 'start' button
        start_text = 'MULAI PANEN'
        self.start_button.config(text=start_text, command=lambda: self.on_start_button_click())
        self.update()  # --- avoids freezing
        # Reset the header text
        desc_text = 'Jendela Pemanenan Data Risat'
        self.set_header_desc(desc_text)

    # The function that will be called by functions of
    # class member BackEndHarvester upon successful scraping
    def on_notify_successful_scraping(self):
        # Replace 'start' button with 'finish' button
        finish_text = 'SELESAI'
        self.start_button.config(text=finish_text, command=lambda: self.on_finish_button_click())
        self.update()  # --- avoids freezing

        # Change the helper status display
        self.set_help_label('Data selesai dipanen!')

        # Re-enable the 'start' button
        self.start_button.config(state='normal')
        self.update()  # --- avoids freezing

        # Showing info window
        harvest_success_title = 'PEMANENAN DATA SELESAI'
        harvest_success_content = f'Data telah selesai dipanen dan disimpan!'
        messagebox.showinfo(harvest_success_title, harvest_success_content)

    # The function that changes the header description of this screen
    def set_header_desc(self, string):
        self.header_desc.config(text=string)

    # The function that changes the help label of this screen
    def set_help_label(self, string):
        self.help_label.config(text=string)

    # The function that changes the progress label of this screen
    # i.e., changes the percentage display of the progress bar
    #
    # This function should not be run directly
    # Only access through 'set_progress_bar()' function
    def set_progress_label(self, string):
        self.progress_label.config(text=string)

    # The function that changes the progress bar progression of this screen
    def set_progress_bar(self, value):
        # The value of the variable 'value' must be
        # in the range of 0 - 100
        if value > 100:
            value = 100
        elif value < 0:
            value = 0

        # Begin setting the progress bar's progression value
        self.progress_bar['value'] = value

        # Setting the progress label value
        value_string = f'Status: {str(value)}%'
        self.set_progress_label(value_string)
        self.update()  # --- avoids freezing

    # The function that clears this screen's message area's harvest logging output
    def clear_message_area(self):
        self.message_area.delete(1.0, tk.END)

    # The function that sets (i.e., clear and write) this screen's message area's
    # harvest logging output
    def set_message_area(self, long_string):
        self.clear_message_area() # --- clearing the content first
        self.message_area.insert(tk.INSERT, long_string)

    # This function sets the y-position of the message area's scrollbar
    def set_message_area_yview(self, pos):
        '''
        :param pos: a float number between 0 and 1
        '''
        self.message_area.yview_moveto(pos)

    # The function that appends long string to this screen's message area's
    # harvest logging output
    # New line character is concatenated in between the existing message area's
    # content and the string to be appended, by default
    def append_message_area(self, long_string, new_line=True):
        # Getting the current content of the message area
        str_ = self.message_area.get('1.0', tk.END).strip()
        # Appending 'long_string' to the existing content,
        # Concat with a new line character, if specified by the argument
        if new_line:
            str_ += '\n' + long_string
        else:
            str_ += long_string
        # Applying to the message area GUI element
        # then set the scrollbar position
        self.set_message_area(str_)
        self.set_message_area_yview(1.0)
    
    # The __init__ function
    def __init__(self, parent, controller):
        
        # -------------------- GUI LAYOUT -------------------- #

        # List of GUI elements that can be accessed by
        # other (non-'__init__') functions of this class
        #
        # self.header_desc      --> the header description
        # self.help_label       --> the label that indicates the status of the harvest process
        # self.progress_label   --> to display the percentage of the progress bar
        # self.progress_bar     --> the progress bar that indicates the progression of the harvest
        # self.message_area     --> the displayer of the harvest log
        
        # Instantiating tkinter.Frame instance
        tk.Frame.__init__(self, parent)
        
        # Assigning the variables for use by other methods in the class
        self.parent = parent
        self.controller = controller
        
        # The header title displaying the screen's title 
        header_title = ttk.Label(self, text=STRING_HEADER_TITLE, font=FONT_HEADER_TITLE)
        header_title.grid(row=0, column=0, padx=10, pady=2)
        
        # The header description displaying additional information about the screen
        desc_text = 'Jendela Pemanenan Data Risat'
        self.header_desc = ttk.Label(self, text=desc_text, font=FONT_HEADER_DESC)
        self.header_desc.grid(row=1, column=0, padx=10, pady=5)
        
        # The screen's display help
        help_text = 'Tekan tombol "MULAI PANEN" untuk memulai proses pemanenan data Risat'
        self.help_label = ttk.Label(self, text=help_text, font=FONT_REGULAR, anchor='w', justify='left')
        self.help_label.grid(row=2, column=0, padx=10, pady=10)
        
        # :::
        # Layout harvester progress bar display
        progress_frame = tk.Frame(self)
        progress_frame.grid(row=3, column=0, padx=10, pady=10)
        # The progress status/value display
        progress_text = 'Status: 0%'
        self.progress_label = ttk.Label(progress_frame, text=progress_text, font=FONT_PROGRESS_VALUE, anchor='center', justify='center')
        self.progress_label.grid(row=0, column=0, padx=2, pady=5)
        # The progress bar
        self.progress_bar = ttk.Progressbar(progress_frame, length=600)
        self.progress_bar['value'] = 0
        self.progress_bar.grid(row=1, column=0, padx=2, pady=2, sticky='we')
        
        # The log message displayer
        self.message_area = scrolledtext.ScrolledText(self, width=100, height=10)
        self.message_area.grid(row=4, column=0, padx=5, pady=5)
        
        # :::
        # Layout: The action buttons
        layout_actions = tk.Frame(self)
        layout_actions.grid(row=5, column=0, padx=10, pady=10)
        
        # Cancel button
        # - For debug purposes only: to allow testing out switching between screens
        #   in an efficient manner
        # - Upon release, this button should be disabled
        cancel_text = 'BATALKAN'
        self.cancel_button = ttk.Button(layout_actions, text=cancel_text, width=30,
            command=lambda: self.on_cancel_button_click())
        self.cancel_button.grid(row=0, column=0, padx=2, pady=2)
        # Setting the 'disabled' state of the button
        # Possible values: 'normal' and 'disabled'
        self.cancel_button['state'] = 'normal'
        
        # The 'start' button trigger
        # Clicking this button proceeds the program to begin the harvesting process
        start_text = 'MULAI PANEN'
        self.start_button = ttk.Button(layout_actions, text=start_text, width=40,
            command=lambda: self.on_start_button_click())
        self.start_button.grid(row=0, column=1, padx=2, pady=2)

# - The class that checks whether the input username/password credentials in
#   'SipesatScrLogin' screen are correct
class BackEndLoginChecker():
    
    # The __init__ function
    def __init__(self, *args, **kwargs):
        
        # Instantiating 'requests.Session'
        self.session = rq.Session()
                
        # Initializing variables
        self.password = ''
        self.username = ''
    
    # The variable that checks the validity of the input Risat credentials
    def validate(self, username, password):
        
        # Setting the variables according to the passed arguments
        self.password = password
        self.username = username

        # Preamble logging
        print('[BackEndLoginChecker] :: Validating Risat credential ...')
        
        # Opening the Risat homepage
        risat_homepage = 'https://risat.uksw.edu/login.aspx?ReturnUrl=%2f'
        r = self.session.get(risat_homepage)
        content = html.fromstring(r.content)
        
        # Obtaining the computer-generated hidden values of ASPX (before login)
        viewstate = content.xpath('//*[@id="__VIEWSTATE"]/@value')[0]
        viewstategen = content.xpath('//*[@id="__VIEWSTATEGENERATOR"]/@value')[0]
        eventvalidation = content.xpath('//*[@id="__EVENTVALIDATION"]/@value')[0]
        
        # Preparing the http handler URL and payload
        HANDLER_URL = 'https://risat.uksw.edu/login.aspx?ReturnUrl=%2f'
        PAYLOAD = {
            # The values below are computer-generated
            '__VIEWSTATE' : viewstate,
            '__VIEWSTATEGENERATOR' : viewstategen,
            '__EVENTVALIDATION' : eventvalidation,
            # The values below are the login information provided by the prompt in the previous code block
            'txnip1': username,
            'txpwd1': password,
            'btlogin1': True
        }
        
        # Logging in to Risat administrator page
        post = self.session.post(HANDLER_URL, data=PAYLOAD)
        response = post.text # --- Obtaining the response text
        content = html.fromstring(response) # --- Scraping the HTML code
        
        # Finding the existence of HTML elements that exist when the input
        # login credentials are wrong
        try:
            prooftest = content.xpath('//div[@class="card-body"]/div[@class="form-group f12"]/text()')[0]
            if prooftest.__contains__('User atau password anda salah'):
                # The input credentials are wrong
                return False
        except IndexError:
            # The element is not present in the HTML page,
            # meaning the login credentials must be correct
            return True
        
        # The default state
        return False

# The essence of this application: the harvester class
# This class sorts harvest input arguments (such as data category,
# data type, and data output) and then execute harvesting the Risat data
class BackEndHarvester():
    
    # The __init__ function
    # This function instantiates 'requests.Session'
    # It does nothing else
    def __init__(self, *args, **kwargs):
        
        # Instantiating 'requests.Session'
        self.session = rq.Session()

        # Preparing the temporary folder that will be used
        # in the recursive scraper of the detail pages
        self.tmpdir = tmp.mkdtemp(prefix=SCRAPE_TEMP_DIR_PREFIX)
    
    # The harvest executor
    # Specify the username and password credentials to mitigate session timeout
    # Please refer to convention [3] for the possible values of 'category', 'datatype', and 'output'
    #
    # In the function 'execute_harvester()' below,
    # the argument 'control' refers to the SipesatScrHarvest
    # harvesting screen/GUI, which progress bar, labels, and
    # message area will be manipulated by this harvest executor
    # function as the harvesting process progresses
    def execute_harvester(self, control, username, password, category, datatype, output):
        
        # :::
        # Determining the cases of the harvesting arguments passed
        # This switching-cases require Python version >= v3.10

        # Preamble logging
        print(f'[BackEndHarvester] :: SELECTED_CASE: category={category}, datatype={datatype}, output={output}')
        
        # --------------------- RISAT PENELITIAN --------------------- #
        if category == 'r': # --- category selected: 'Risat Penelitian'
            
            # Determining the cases of the datatype
            if datatype == 0: # --- 'data ringkasan'
            
                # Determining the cases of the data output
                match output:
                    case 'usulan':
                        self.run_harvest_r_0_usulan(control, username, password)
                    case 'berdisetujui':
                        self.run_harvest_r_0_berdisetujui(control, username, password)
                    case 'berdirevisi':
                        self.run_harvest_r_0_berdirevisi(control, username, password)
                    case 'teldireview':
                        self.run_harvest_r_0_teldireview(control, username, password)
                    case 'disetujui':
                        self.run_harvest_r_0_disetujui(control, username, password)
                    case 'ddr':
                        self.run_harvest_r_0_ddr(control, username, password)
                    case 'ditolak':
                        self.run_harvest_r_0_ditolak(control, username, password)
                    case 'arsip':
                        self.run_harvest_r_0_arsip(control, username, password)
                    case 'dana':
                        self.run_harvest_r_0_dana(control, username, password)
                    case 'lapakhir':
                        self.run_harvest_r_0_lapakhir(control, username, password)
                        
            elif datatype == 1: # --- 'data detil lengkap'
                
                # Determining the cases of the data output
                match output:
                    case 'usulan':
                        self.run_harvest_r_1_usulan(control, username, password)
                    case 'berdisetujui':
                        self.run_harvest_r_1_berdisetujui(control, username, password)
                    case 'berdirevisi':
                        self.run_harvest_r_1_berdirevisi(control, username, password)
                    case 'teldireview':
                        self.run_harvest_r_1_teldireview(control, username, password)
                    case 'disetujui':
                        self.run_harvest_r_1_disetujui(control, username, password)
                    case 'ddr':
                        self.run_harvest_r_1_ddr(control, username, password)
                    case 'ditolak':
                        self.run_harvest_r_1_ditolak(control, username, password)
                    case 'arsip':
                        self.run_harvest_r_1_arsip(control, username, password)
                    case 'dana':
                        self.run_harvest_r_1_dana(control, username, password)
                    case 'lapakhir':
                        self.run_harvest_r_1_lapakhir(control, username, password)
        
        # ---------------- RISAT PENGABDIAN MASYARAKAT ---------------- #
        elif category == 'c': # --- category selected: 'Pengabdian Masyarakat'
            
            # Determining the cases of the datatype
            if datatype == 0: # --- 'data ringkasan'

                # Determining the cases of the data output
                match output:
                    case 'usulan':
                        self.run_harvest_c_0_usulan(control, username, password)
                    case 'berdisetujui':
                        self.run_harvest_c_0_berdisetujui(control, username, password)
                    case 'berdirevisi':
                        self.run_harvest_c_0_berdirevisi(control, username, password)
                    case 'ditolak':
                        self.run_harvest_c_0_ditolak(control, username, password)
                    case 'arsip':
                        self.run_harvest_c_0_arsip(control, username, password)
                    case 'dana':
                        self.run_harvest_c_0_dana(control, username, password)
                    case 'lapakhir':
                        self.run_harvest_c_0_lapakhir(control, username, password)
                        
            elif datatype == 1: # --- 'data detil lengkap'
                
                # Determining the cases of the data output
                match output:
                    case 'usulan':
                        self.run_harvest_c_1_usulan(control, username, password)
                    case 'berdisetujui':
                        self.run_harvest_c_1_berdisetujui(control, username, password)
                    case 'berdirevisi':
                        self.run_harvest_c_1_berdirevisi(control, username, password)
                    case 'ditolak':
                        pass
                    case 'arsip':
                        self.run_harvest_c_1_arsip(control, username, password)
                    case 'dana':
                        self.run_harvest_c_1_dana(control, username, password)
                    case 'lapakhir':
                        self.run_harvest_c_1_lapakhir(control, username, password)
        
        # ======================== END ======================== #

    # This function recursively obtains the detail pages of Risat
    # For example, the 'Dana Penelitian Detil' and 'Abdimas Arsip Detil'
    #
    # Requires no 'data_prompt' passed as argument,
    # but 'mode' argument is mandated to be passed
    # Also requires 'username' and 'password' credentials to be
    # passed as arguments
    #
    # Returns an array which elements are the paths to the
    # temporary files where each detail pages are stored into
    #
    # The possible values of 'mode' argument are as follows:
    # mode='dana_penelitian'        --> obtains the detail pages of 'Risat Dana Penelitian'
    # mode='arsip_penelitian'        --> obtains the detail pages of 'Risat Arsip Penelitian'
    # mode='dana_pengabdian'        --> obtains the detail pages of 'Risat Dana Pengabdian'
    # mode='arsip_pengabdian'        --> obtains the detail pages of 'Risat Arsip Pengabdian'
    #
    # !!! THIS FUNCTION IS DEPRECATED !!!
    # PyFoldDefault
    def get_auto_risat_detil(self, mode, username, password):
        # PyFoldDefault

        '''
        This function is deprecated
        '''

        # :::
        # Determining the mode of the risat detil pages to be scraped recursively

        # --------------------- DANA PENELITIAN --------------------- #
        if mode == 'dana_penelitian':  # --- category selected: 'Dana Penelitian'

            # Preparing the 'data_prompt' arrays
            data_prompt = self.get_risat_login(username, password)
            data_prompt = self.get_risat_penelitian(data_prompt)
            data_prompt = self.get_risat_penelitian_dana_penelitian(data_prompt)
            
            # Parsing XML tree content
            content = data_prompt['html_content']
            
            # The base XPath location, pointing to each entry row
            base = '//div[@id="ContentPlaceHolder1_upd1"]/div[@class="mw-100"]//table[@width="100%"]//tr[@valign="top"]'
            
            # Reading the HTML entry row hidden ASPX values
            # These variables below are *arrays*, so they have indices
            all_kodetran_prop = content.xpath(base + '//input[1][@type="hidden"]/@name')
            all_kodetran_val = content.xpath(base + '//input[1][@type="hidden"]/@value')
            all_stat_prop = content.xpath(base + '//input[2][@type="hidden"]/@name')
            all_stat_val = content.xpath(base + '//input[2][@type="hidden"]/@value')
            all_submitbtn = content.xpath(base + '//input[@type="submit"]/@name')

            # The temporary 'data_prompt' to be used to revert back to the table page
            # which lists the detil pages
            temporary_prompt = data_prompt

            # The array which stores the list of temporary files
            # that stores the HTTP response of each detail page scraped
            scrape_array = []

            # Iterating through each entry row element
            # Assumes all the arrays in the previous code block
            # are of the same length/size
            for i in range(len(all_kodetran_prop)-1):

                # Preparing the AJAX payload
                detail_prompt = {
                    'viewstate' : temporary_prompt['viewstate'],
                    'viewstategen' : temporary_prompt['viewstategen'],
                    'eventvalidation' : temporary_prompt['eventvalidation'],
                    'button_name' : all_submitbtn[i],
                    'kodetran_prop' : all_kodetran_prop[i],
                    'kodetran_val' : all_kodetran_val[i],      
                    'stat_prop' : all_stat_prop[i],
                    'stat_val' : all_stat_val[i]
                }
                
                # Obtaining the response data of each individual entry row detail page
                data = self.get_risat_penelitian_dana_penelitian_detil(detail_prompt)
                content = data['html_content']
                response = data['http_response']

                # Writing to the temporary file in the temporary directory
                tmppath = self.tmpdir + sep + 'dana_penelitian-detil-' + str(i) + '.html'
                print(f'+ Saving to temporary path: {tmppath}.')
                fo = open(tmppath, 'w')
                fo.write(response)
                fo.close()

                # Appending to the array
                scrape_array.insert(i, tmppath)
                
                # Logging: printing the basic information
                judul = content.xpath('//span[@id="ContentPlaceHolder1_danacair1_kiri1_txjudul1"]/text()')[0].replace(
                    '\r', '').replace('\n', '').strip()
                print(f'ENTRY_NO: {i} --> {judul}')
                
                # Reopening the "Dana Penelitian" list page,
                # then assign the AJAX response to the temporary array 'temporary_prompt'
                # The 'data' array is obtained from opening individual entry row detail page
                temporary_prompt = self.get_risat_penelitian_dana_penelitian(data)
                continue

            # Returning the list of array
            return scrape_array

        # --------------------- ARSIP PENELITIAN --------------------- #
        elif mode == 'arsip_penelitian':  # --- category selected: 'Arsip Penelitian'
            pass

        # ---------------------- DANA PENGABDIAN ---------------------- #
        elif mode == 'dana_pengabdian':  # --- category selected: 'Dana Pengabdian'
            pass

        # ---------------------- ARSIP PENGABDIAN ---------------------- #
        elif mode == 'arsip_pengabdian':  # --- category selected: 'Arsip Pengabdian'
            pass

        # ======================== END ======================== #

    # This function gets past the barrier of Risat login page
    # - Requires two arguments: the username and password credentials
    # - Returns 'data_prompt' array
    def get_risat_login(self, username, password):
        # Preamble logging
        print(f'[BackEndHarvester] :: Logging in to Risat as {username} to start a new ASPX session ...')

        # Opening the Risat homepage
        print('+ Opening Risat homepage...')
        risat_homepage = 'https://risat.uksw.edu/login.aspx?ReturnUrl=%2f'
        r = self.session.get(risat_homepage)
        content = html.fromstring(r.content)
        
        # Obtaining the computer-generated hidden values of ASPX (before login)
        viewstate = content.xpath('//*[@id="__VIEWSTATE"]/@value')[0]
        viewstategen = content.xpath('//*[@id="__VIEWSTATEGENERATOR"]/@value')[0]
        eventvalidation = content.xpath('//*[@id="__EVENTVALIDATION"]/@value')[0]
        
        # Preparing the http handler URL and payload
        HANDLER_URL = 'https://risat.uksw.edu/login.aspx?ReturnUrl=%2f'
        PAYLOAD = {
            # The values below are computer-generated
            '__VIEWSTATE' : viewstate,
            '__VIEWSTATEGENERATOR' : viewstategen,
            '__EVENTVALIDATION' : eventvalidation,
            # The values below are the login information provided by the prompt in the previous code block
            'txnip1': username,
            'txpwd1': password,
            'btlogin1': True
        }
        
        # Logging in to Risat administrator page
        print('+ Logging in...')
        post = self.session.post(HANDLER_URL, data=PAYLOAD)
        response = post.text # --- Obtaining the response text
        content = html.fromstring(response) # --- Scraping the HTML code
        
        # Obtaining the computer-generated hidden values of ASPX (after login)
        viewstate = content.xpath('//*[@id="__VIEWSTATE"]/@value')[0]
        viewstategen = content.xpath('//*[@id="__VIEWSTATEGENERATOR"]/@value')[0]
        eventvalidation = content.xpath('//*[@id="__EVENTVALIDATION"]/@value')[0]
        
        # Building the 'data_prompt' array
        data_prompt = {
            'http_response' : response,
            'html_content' : content,
            'viewstate' : viewstate,
            'viewstategen' : viewstategen,
            'eventvalidation' : eventvalidation
        }
        
        # Returning the http response string
        return data_prompt

    # This function opens "Penelitian" menu tab
    # Requires 'data_prompt' array obtained from the login function
    # Returns another 'data_prompt' array
    def get_risat_penelitian(self, data_prompt):
        # Logging the calling of the function
        print('+ Opening "Penelitian" menu...')
        
        # Preparing the http handler URL and payload
        HANDLER_URL = 'https://risat.uksw.edu/bp3mpage.aspx'
        PAYLOAD = {
            # The values below are computer-generated
            '__VIEWSTATE' : data_prompt['viewstate'],
            '__VIEWSTATEGENERATOR' : data_prompt['viewstategen'],
            '__EVENTVALIDATION' : data_prompt['eventvalidation'],
            'ctl00$menu2': True
        }
        
        # Posting the http payloads
        print('+ Posting http payloads...')
        post = self.session.post(HANDLER_URL, data=PAYLOAD)
        response = post.text # --- Obtaining the response text
        content = html.fromstring(response) # --- Scraping the HTML code
        
        # Obtaining the computer-generated hidden values of ASPX (after login)
        viewstate = content.xpath('//*[@id="__VIEWSTATE"]/@value')[0]
        viewstategen = content.xpath('//*[@id="__VIEWSTATEGENERATOR"]/@value')[0]
        eventvalidation = content.xpath('//*[@id="__EVENTVALIDATION"]/@value')[0]
        
        # Building the 'data_prompt' array
        data_prompt = {
            'http_response' : response,
            'html_content' : content,
            'viewstate' : viewstate,
            'viewstategen' : viewstategen,
            'eventvalidation' : eventvalidation
        }
        
        # Returning the http response string
        return data_prompt

    # This function opens "Usulan Penelitian" menu after opening the tab "Penelitian"
    # - Requires 'data_prompt' array as an unary argument obtained from get_risat_penelitian() function
    # - Returns also another 'data_prompt' array
    def get_risat_penelitian_usulan_penelitian(self, data_prompt):
        # Logging the calling of the function
        print('+ Opening "Penelitian --> Usulan Penelitian" menu...')

        # Preparing the http handler URL and payload
        HANDLER_URL = 'https://risat.uksw.edu/bp3mpage.aspx'
        PAYLOAD = {
            # The values below are computer-generated
            '__VIEWSTATE': data_prompt['viewstate'],
            '__VIEWSTATEGENERATOR': data_prompt['viewstategen'],
            '__EVENTVALIDATION': data_prompt['eventvalidation'],
            '__EVENTTARGET': 'ctl00$ContentPlaceHolder1$menu1'
        }

        # Posting the http payloads
        print('+ Posting http payloads...')
        post = self.session.post(HANDLER_URL, data=PAYLOAD)
        response = post.text  # --- Obtaining the response text
        content = html.fromstring(response)  # --- Scraping the HTML code

        # Obtaining the computer-generated hidden values of ASPX (after login)
        viewstate = content.xpath('//*[@id="__VIEWSTATE"]/@value')[0]
        viewstategen = content.xpath('//*[@id="__VIEWSTATEGENERATOR"]/@value')[0]
        eventvalidation = content.xpath('//*[@id="__EVENTVALIDATION"]/@value')[0]

        # Building the 'data_prompt' array
        data_prompt = {
            'http_response': response,
            'html_content': content,
            'viewstate': viewstate,
            'viewstategen': viewstategen,
            'eventvalidation': eventvalidation
        }

        # Returning 'data_prompt' array
        return data_prompt

    # This function opens the detail page of "Usulan Penelitian" entry row
    # Requires one argument:
    # - the data prompt value array in compliance with convention [5] of this file
    # The 'data_prompt' argument passed into this function must be
    # rooted from the function 'get_risat_penelitian_dana_penelitian()'
    def get_risat_penelitian_usulan_penelitian_detil(self, data_prompt):
        # Logging the calling of the function
        print('+ Opening "Usulan Penelitian" entry row detail page...')

        # Preparing the http handler URL and payload
        LOGIN_HANDLER_URL = 'https://risat.uksw.edu/bp3mpage.aspx'
        LOGIN_PAYLOAD = {
            # The values below are computer-generated
            '__VIEWSTATE': data_prompt['viewstate'],
            '__VIEWSTATEGENERATOR': data_prompt['viewstategen'],
            '__EVENTVALIDATION': data_prompt['eventvalidation'],
            # The entry row's submit button to 'hit'
            data_prompt['button_name']: 'Detil',
            data_prompt['kodetran_prop']: data_prompt['kodetran_val'],
            data_prompt['stat_prop']: data_prompt['stat_val']
        }

        # Posting the http payloads
        print('+ Posting http payloads...')
        post = self.session.post(LOGIN_HANDLER_URL, data=LOGIN_PAYLOAD)
        response = post.text  # --- Obtaining the response text
        content = html.fromstring(response)  # --- Scraping the HTML code

        # Obtaining the computer-generated hidden values of ASPX (after login)
        viewstate = content.xpath('//*[@id="__VIEWSTATE"]/@value')[0]
        viewstategen = content.xpath('//*[@id="__VIEWSTATEGENERATOR"]/@value')[0]
        eventvalidation = content.xpath('//*[@id="__EVENTVALIDATION"]/@value')[0]

        # Building the 'data_prompt' array
        data_prompt = {
            'http_response': response,
            'html_content': content,
            'viewstate': viewstate,
            'viewstategen': viewstategen,
            'eventvalidation': eventvalidation
        }

        # Returning the http response string
        return data_prompt

    # This function opens "Berkas Disetujui Penelitian" menu after opening the tab "Penelitian"
    # - Requires 'data_prompt' array as an unary argument obtained from get_risat_penelitian() function
    # - Returns also another 'data_prompt' array
    def get_risat_penelitian_berdisetujui_penelitian(self, data_prompt):
        # Logging the calling of the function
        print('+ Opening "Penelitian --> Berkas Disetujui Penelitian" menu...')

        # Preparing the http handler URL and payload
        HANDLER_URL = 'https://risat.uksw.edu/bp3mpage.aspx'
        PAYLOAD = {
            # The values below are computer-generated
            '__VIEWSTATE' : data_prompt['viewstate'],
            '__VIEWSTATEGENERATOR' : data_prompt['viewstategen'],
            '__EVENTVALIDATION' : data_prompt['eventvalidation'],
            '__EVENTTARGET' : 'ctl00$ContentPlaceHolder1$menu3'
        }

        # Posting the http payloads
        print('+ Posting http payloads...')
        post = self.session.post(HANDLER_URL, data=PAYLOAD)
        response = post.text # --- Obtaining the response text
        content = html.fromstring(response) # --- Scraping the HTML code

        # Obtaining the computer-generated hidden values of ASPX (after login)
        viewstate = content.xpath('//*[@id="__VIEWSTATE"]/@value')[0]
        viewstategen = content.xpath('//*[@id="__VIEWSTATEGENERATOR"]/@value')[0]
        eventvalidation = content.xpath('//*[@id="__EVENTVALIDATION"]/@value')[0]

        # Building the 'data_prompt' array
        data_prompt = {
            'http_response' : response,
            'html_content' : content,
            'viewstate' : viewstate,
            'viewstategen' : viewstategen,
            'eventvalidation' : eventvalidation
        }

        # Returning 'data_prompt' array
        return data_prompt

    # This function opens the detail page of "Berkas Disetujui Penelitian" entry row
    # Requires one argument:
    # - the data prompt value array in compliance with convention [5] of this file
    # The 'data_prompt' argument passed into this function must be
    # rooted from the function 'get_risat_penelitian_berdisetujui_penelitian()'
    def get_risat_penelitian_berdisetujui_penelitian_detil(self, data_prompt):
        # Logging the calling of the function
        print('+ Opening "Berdisetujui Penelitian" entry row detail page...')

        # Preparing the http handler URL and payload
        LOGIN_HANDLER_URL = 'https://risat.uksw.edu/bp3mpage.aspx'
        LOGIN_PAYLOAD = {
            # The values below are computer-generated
            '__VIEWSTATE': data_prompt['viewstate'],
            '__VIEWSTATEGENERATOR': data_prompt['viewstategen'],
            '__EVENTVALIDATION': data_prompt['eventvalidation'],
            # The entry row's submit button to 'hit'
            data_prompt['button_name']: 'Detil',
            data_prompt['kodetran_prop']: data_prompt['kodetran_val'],
            data_prompt['stat_prop']: data_prompt['stat_val']
        }

        # Posting the http payloads
        print('+ Posting http payloads...')
        post = self.session.post(LOGIN_HANDLER_URL, data=LOGIN_PAYLOAD)
        response = post.text  # --- Obtaining the response text
        content = html.fromstring(response)  # --- Scraping the HTML code

        # Obtaining the computer-generated hidden values of ASPX (after login)
        viewstate = content.xpath('//*[@id="__VIEWSTATE"]/@value')[0]
        viewstategen = content.xpath('//*[@id="__VIEWSTATEGENERATOR"]/@value')[0]
        eventvalidation = content.xpath('//*[@id="__EVENTVALIDATION"]/@value')[0]

        # Building the 'data_prompt' array
        data_prompt = {
            'http_response': response,
            'html_content': content,
            'viewstate': viewstate,
            'viewstategen': viewstategen,
            'eventvalidation': eventvalidation
        }

        # Returning the http response string
        return data_prompt

    # This function opens "Berkas Direvisi Penelitian" menu after opening the tab "Penelitian"
    # - Requires 'data_prompt' array as an unary argument obtained from get_risat_penelitian() function
    # - Returns also another 'data_prompt' array
    def get_risat_penelitian_berdirevisi_penelitian(self, data_prompt):
        # Logging the calling of the function
        print('+ Opening "Penelitian --> Berkas Direvisi Penelitian" menu...')

        # Preparing the http handler URL and payload
        HANDLER_URL = 'https://risat.uksw.edu/bp3mpage.aspx'
        PAYLOAD = {
            # The values below are computer-generated
            '__VIEWSTATE' : data_prompt['viewstate'],
            '__VIEWSTATEGENERATOR' : data_prompt['viewstategen'],
            '__EVENTVALIDATION' : data_prompt['eventvalidation'],
            '__EVENTTARGET' : 'ctl00$ContentPlaceHolder1$menu5'
        }

        # Posting the http payloads
        print('+ Posting http payloads...')
        post = self.session.post(HANDLER_URL, data=PAYLOAD)
        response = post.text # --- Obtaining the response text
        content = html.fromstring(response) # --- Scraping the HTML code

        # Obtaining the computer-generated hidden values of ASPX (after login)
        viewstate = content.xpath('//*[@id="__VIEWSTATE"]/@value')[0]
        viewstategen = content.xpath('//*[@id="__VIEWSTATEGENERATOR"]/@value')[0]
        eventvalidation = content.xpath('//*[@id="__EVENTVALIDATION"]/@value')[0]

        # Building the 'data_prompt' array
        data_prompt = {
            'http_response' : response,
            'html_content' : content,
            'viewstate' : viewstate,
            'viewstategen' : viewstategen,
            'eventvalidation' : eventvalidation
        }

        # Returning 'data_prompt' array
        return data_prompt

    # This function opens the detail page of "Berkas Direvisi Penelitian" entry row
    # Requires one argument:
    # - the data prompt value array in compliance with convention [5] of this file
    # The 'data_prompt' argument passed into this function must be
    # rooted from the function 'get_risat_penelitian_berdirevisi_penelitian()'
    def get_risat_penelitian_berdirevisi_penelitian_detil(self, data_prompt):
        # Logging the calling of the function
        print('+ Opening "Berdirevisi Penelitian" entry row detail page...')

        # Preparing the http handler URL and payload
        LOGIN_HANDLER_URL = 'https://risat.uksw.edu/bp3mpage.aspx'
        LOGIN_PAYLOAD = {
            # The values below are computer-generated
            '__VIEWSTATE': data_prompt['viewstate'],
            '__VIEWSTATEGENERATOR': data_prompt['viewstategen'],
            '__EVENTVALIDATION': data_prompt['eventvalidation'],
            # The entry row's submit button to 'hit'
            data_prompt['button_name']: 'Detil',
            data_prompt['kodetran_prop']: data_prompt['kodetran_val'],
            data_prompt['stat_prop']: data_prompt['stat_val']
        }

        # Posting the http payloads
        print('+ Posting http payloads...')
        post = self.session.post(LOGIN_HANDLER_URL, data=LOGIN_PAYLOAD)
        response = post.text  # --- Obtaining the response text
        content = html.fromstring(response)  # --- Scraping the HTML code

        # Obtaining the computer-generated hidden values of ASPX (after login)
        viewstate = content.xpath('//*[@id="__VIEWSTATE"]/@value')[0]
        viewstategen = content.xpath('//*[@id="__VIEWSTATEGENERATOR"]/@value')[0]
        eventvalidation = content.xpath('//*[@id="__EVENTVALIDATION"]/@value')[0]

        # Building the 'data_prompt' array
        data_prompt = {
            'http_response': response,
            'html_content': content,
            'viewstate': viewstate,
            'viewstategen': viewstategen,
            'eventvalidation': eventvalidation
        }

        # Returning the http response string
        return data_prompt

    # This function opens "Telah Direview Penelitian" menu after opening the tab "Penelitian"
    # - Requires 'data_prompt' array as an unary argument obtained from get_risat_penelitian() function
    # - Returns also another 'data_prompt' array
    def get_risat_penelitian_teldireview_penelitian(self, data_prompt):
        # Logging the calling of the function
        print('+ Opening "Penelitian --> Telah Direview Penelitian" menu...')

        # Preparing the http handler URL and payload
        HANDLER_URL = 'https://risat.uksw.edu/bp3mpage.aspx'
        PAYLOAD = {
            # The values below are computer-generated
            '__VIEWSTATE' : data_prompt['viewstate'],
            '__VIEWSTATEGENERATOR' : data_prompt['viewstategen'],
            '__EVENTVALIDATION' : data_prompt['eventvalidation'],
            '__EVENTTARGET' : 'ctl00$ContentPlaceHolder1$menu6'
        }

        # Posting the http payloads
        print('+ Posting http payloads...')
        post = self.session.post(HANDLER_URL, data=PAYLOAD)
        response = post.text # --- Obtaining the response text
        content = html.fromstring(response) # --- Scraping the HTML code

        # Obtaining the computer-generated hidden values of ASPX (after login)
        viewstate = content.xpath('//*[@id="__VIEWSTATE"]/@value')[0]
        viewstategen = content.xpath('//*[@id="__VIEWSTATEGENERATOR"]/@value')[0]
        eventvalidation = content.xpath('//*[@id="__EVENTVALIDATION"]/@value')[0]

        # Building the 'data_prompt' array
        data_prompt = {
            'http_response' : response,
            'html_content' : content,
            'viewstate' : viewstate,
            'viewstategen' : viewstategen,
            'eventvalidation' : eventvalidation
        }

        # Returning 'data_prompt' array
        return data_prompt

    # This function opens the detail page of "Telah Direview Penelitian" entry row
    # Requires one argument:
    # - the data prompt value array in compliance with convention [5] of this file
    # The 'data_prompt' argument passed into this function must be
    # rooted from the function 'get_risat_penelitian_teldireview_penelitian()'
    def get_risat_penelitian_teldireview_penelitian_detil(self, data_prompt):
        # Logging the calling of the function
        print('+ Opening "Telah Direview Penelitian" entry row detail page...')

        # Preparing the http handler URL and payload
        LOGIN_HANDLER_URL = 'https://risat.uksw.edu/bp3mpage.aspx'
        LOGIN_PAYLOAD = {
            # The values below are computer-generated
            '__VIEWSTATE': data_prompt['viewstate'],
            '__VIEWSTATEGENERATOR': data_prompt['viewstategen'],
            '__EVENTVALIDATION': data_prompt['eventvalidation'],
            # The entry row's submit button to 'hit'
            data_prompt['button_name']: 'Detil',
            data_prompt['kodetran_prop']: data_prompt['kodetran_val'],
            data_prompt['stat_prop']: data_prompt['stat_val']
        }

        # Posting the http payloads
        print('+ Posting http payloads...')
        post = self.session.post(LOGIN_HANDLER_URL, data=LOGIN_PAYLOAD)
        response = post.text  # --- Obtaining the response text
        content = html.fromstring(response)  # --- Scraping the HTML code

        # Obtaining the computer-generated hidden values of ASPX (after login)
        viewstate = content.xpath('//*[@id="__VIEWSTATE"]/@value')[0]
        viewstategen = content.xpath('//*[@id="__VIEWSTATEGENERATOR"]/@value')[0]
        eventvalidation = content.xpath('//*[@id="__EVENTVALIDATION"]/@value')[0]

        # Building the 'data_prompt' array
        data_prompt = {
            'http_response': response,
            'html_content': content,
            'viewstate': viewstate,
            'viewstategen': viewstategen,
            'eventvalidation': eventvalidation
        }

        # Returning the http response string
        return data_prompt

    # This function opens "Disetujui Penelitian" menu after opening the tab "Penelitian"
    # - Requires 'data_prompt' array as an unary argument obtained from get_risat_penelitian() function
    # - Returns also another 'data_prompt' array
    def get_risat_penelitian_disetujui_penelitian(self, data_prompt):
        # Logging the calling of the function
        print('+ Opening "Penelitian --> Disetujui Penelitian" menu...')

        # Preparing the http handler URL and payload
        HANDLER_URL = 'https://risat.uksw.edu/bp3mpage.aspx'
        PAYLOAD = {
            # The values below are computer-generated
            '__VIEWSTATE' : data_prompt['viewstate'],
            '__VIEWSTATEGENERATOR' : data_prompt['viewstategen'],
            '__EVENTVALIDATION' : data_prompt['eventvalidation'],
            '__EVENTTARGET' : 'ctl00$ContentPlaceHolder1$menu7'
        }

        # Posting the http payloads
        print('+ Posting http payloads...')
        post = self.session.post(HANDLER_URL, data=PAYLOAD)
        response = post.text # --- Obtaining the response text
        content = html.fromstring(response) # --- Scraping the HTML code

        # Obtaining the computer-generated hidden values of ASPX (after login)
        viewstate = content.xpath('//*[@id="__VIEWSTATE"]/@value')[0]
        viewstategen = content.xpath('//*[@id="__VIEWSTATEGENERATOR"]/@value')[0]
        eventvalidation = content.xpath('//*[@id="__EVENTVALIDATION"]/@value')[0]

        # Building the 'data_prompt' array
        data_prompt = {
            'http_response' : response,
            'html_content' : content,
            'viewstate' : viewstate,
            'viewstategen' : viewstategen,
            'eventvalidation' : eventvalidation
        }

        # Returning 'data_prompt' array
        return data_prompt

    # This function opens the detail page of "Disetujui Penelitian" entry row
    # Requires one argument:
    # - the data prompt value array in compliance with convention [5] of this file
    # The 'data_prompt' argument passed into this function must be
    # rooted from the function 'get_risat_penelitian_disetujui_penelitian()'
    def get_risat_penelitian_disetujui_penelitian_detil(self, data_prompt):
        # Logging the calling of the function
        print('+ Opening "Disetujui Penelitian" entry row detail page...')

        # Preparing the http handler URL and payload
        LOGIN_HANDLER_URL = 'https://risat.uksw.edu/bp3mpage.aspx'
        LOGIN_PAYLOAD = {
            # The values below are computer-generated
            '__VIEWSTATE': data_prompt['viewstate'],
            '__VIEWSTATEGENERATOR': data_prompt['viewstategen'],
            '__EVENTVALIDATION': data_prompt['eventvalidation'],
            # The entry row's submit button to 'hit'
            data_prompt['button_name']: 'Detil',
            data_prompt['kodetran_prop']: data_prompt['kodetran_val'],
            data_prompt['stat_prop']: data_prompt['stat_val']
        }

        # Posting the http payloads
        print('+ Posting http payloads...')
        post = self.session.post(LOGIN_HANDLER_URL, data=LOGIN_PAYLOAD)
        response = post.text  # --- Obtaining the response text
        content = html.fromstring(response)  # --- Scraping the HTML code

        # Obtaining the computer-generated hidden values of ASPX (after login)
        viewstate = content.xpath('//*[@id="__VIEWSTATE"]/@value')[0]
        viewstategen = content.xpath('//*[@id="__VIEWSTATEGENERATOR"]/@value')[0]
        eventvalidation = content.xpath('//*[@id="__EVENTVALIDATION"]/@value')[0]

        # Building the 'data_prompt' array
        data_prompt = {
            'http_response': response,
            'html_content': content,
            'viewstate': viewstate,
            'viewstategen': viewstategen,
            'eventvalidation': eventvalidation
        }

        # Returning the http response string
        return data_prompt

    # This function opens "Disetujui Dgn Revisi (DDR) Penelitian" menu after opening the tab "Penelitian"
    # - Requires 'data_prompt' array as an unary argument obtained from get_risat_penelitian() function
    # - Returns also another 'data_prompt' array
    def get_risat_penelitian_ddr_penelitian(self, data_prompt):
        # Logging the calling of the function
        print('+ Opening "Penelitian --> Disetujui Dgn Revisi (DDR) Penelitian" menu...')

        # Preparing the http handler URL and payload
        HANDLER_URL = 'https://risat.uksw.edu/bp3mpage.aspx'
        PAYLOAD = {
            # The values below are computer-generated
            '__VIEWSTATE' : data_prompt['viewstate'],
            '__VIEWSTATEGENERATOR' : data_prompt['viewstategen'],
            '__EVENTVALIDATION' : data_prompt['eventvalidation'],
            '__EVENTTARGET' : 'ctl00$ContentPlaceHolder1$menu4'
        }

        # Posting the http payloads
        print('+ Posting http payloads...')
        post = self.session.post(HANDLER_URL, data=PAYLOAD)
        response = post.text # --- Obtaining the response text
        content = html.fromstring(response) # --- Scraping the HTML code

        # Obtaining the computer-generated hidden values of ASPX (after login)
        viewstate = content.xpath('//*[@id="__VIEWSTATE"]/@value')[0]
        viewstategen = content.xpath('//*[@id="__VIEWSTATEGENERATOR"]/@value')[0]
        eventvalidation = content.xpath('//*[@id="__EVENTVALIDATION"]/@value')[0]

        # Building the 'data_prompt' array
        data_prompt = {
            'http_response' : response,
            'html_content' : content,
            'viewstate' : viewstate,
            'viewstategen' : viewstategen,
            'eventvalidation' : eventvalidation
        }

        # Returning 'data_prompt' array
        return data_prompt

    # This function opens the detail page of "Disetujui Dgn Revisi (DDR) Penelitian" entry row
    # Requires one argument:
    # - the data prompt value array in compliance with convention [5] of this file
    # The 'data_prompt' argument passed into this function must be
    # rooted from the function 'get_risat_penelitian_ddr_penelitian()'
    def get_risat_penelitian_ddr_penelitian_detil(self, data_prompt):
        # Logging the calling of the function
        print('+ Opening "Disetujui Dgn Revisi (DDR) Penelitian" entry row detail page...')

        # Preparing the http handler URL and payload
        LOGIN_HANDLER_URL = 'https://risat.uksw.edu/bp3mpage.aspx'
        LOGIN_PAYLOAD = {
            # The values below are computer-generated
            '__VIEWSTATE': data_prompt['viewstate'],
            '__VIEWSTATEGENERATOR': data_prompt['viewstategen'],
            '__EVENTVALIDATION': data_prompt['eventvalidation'],
            # The entry row's submit button to 'hit'
            data_prompt['button_name']: 'Detil',
            data_prompt['kodetran_prop']: data_prompt['kodetran_val'],
            data_prompt['stat_prop']: data_prompt['stat_val']
        }

        # Posting the http payloads
        print('+ Posting http payloads...')
        post = self.session.post(LOGIN_HANDLER_URL, data=LOGIN_PAYLOAD)
        response = post.text  # --- Obtaining the response text
        content = html.fromstring(response)  # --- Scraping the HTML code

        # Obtaining the computer-generated hidden values of ASPX (after login)
        viewstate = content.xpath('//*[@id="__VIEWSTATE"]/@value')[0]
        viewstategen = content.xpath('//*[@id="__VIEWSTATEGENERATOR"]/@value')[0]
        eventvalidation = content.xpath('//*[@id="__EVENTVALIDATION"]/@value')[0]

        # Building the 'data_prompt' array
        data_prompt = {
            'http_response': response,
            'html_content': content,
            'viewstate': viewstate,
            'viewstategen': viewstategen,
            'eventvalidation': eventvalidation
        }

        # Returning the http response string
        return data_prompt

    # This function opens "Ditolak Penelitian" menu after opening the tab "Penelitian"
    # - Requires 'data_prompt' array as an unary argument obtained from get_risat_penelitian() function
    # - Returns also another 'data_prompt' array
    def get_risat_penelitian_ditolak_penelitian(self, data_prompt):
        # Logging the calling of the function
        print('+ Opening "Penelitian --> Ditolak Penelitian" menu...')

        # Preparing the http handler URL and payload
        HANDLER_URL = 'https://risat.uksw.edu/bp3mpage.aspx'
        PAYLOAD = {
            # The values below are computer-generated
            '__VIEWSTATE' : data_prompt['viewstate'],
            '__VIEWSTATEGENERATOR' : data_prompt['viewstategen'],
            '__EVENTVALIDATION' : data_prompt['eventvalidation'],
            '__EVENTTARGET' : 'ctl00$ContentPlaceHolder1$menu8'
        }

        # Posting the http payloads
        print('+ Posting http payloads...')
        post = self.session.post(HANDLER_URL, data=PAYLOAD)
        response = post.text # --- Obtaining the response text
        content = html.fromstring(response) # --- Scraping the HTML code

        # Obtaining the computer-generated hidden values of ASPX (after login)
        viewstate = content.xpath('//*[@id="__VIEWSTATE"]/@value')[0]
        viewstategen = content.xpath('//*[@id="__VIEWSTATEGENERATOR"]/@value')[0]
        eventvalidation = content.xpath('//*[@id="__EVENTVALIDATION"]/@value')[0]

        # Building the 'data_prompt' array
        data_prompt = {
            'http_response' : response,
            'html_content' : content,
            'viewstate' : viewstate,
            'viewstategen' : viewstategen,
            'eventvalidation' : eventvalidation
        }

        # Returning 'data_prompt' array
        return data_prompt

    # This function opens the detail page of "Ditolak Penelitian" entry row
    # Requires one argument:
    # - the data prompt value array in compliance with convention [5] of this file
    # The 'data_prompt' argument passed into this function must be
    # rooted from the function 'get_risat_penelitian_ditolak_penelitian()'
    def get_risat_penelitian_ditolak_penelitian_detil(self, data_prompt):
        # Logging the calling of the function
        print('+ Opening "Ditolak Penelitian" entry row detail page...')

        # Preparing the http handler URL and payload
        LOGIN_HANDLER_URL = 'https://risat.uksw.edu/bp3mpage.aspx'
        LOGIN_PAYLOAD = {
            # The values below are computer-generated
            '__VIEWSTATE': data_prompt['viewstate'],
            '__VIEWSTATEGENERATOR': data_prompt['viewstategen'],
            '__EVENTVALIDATION': data_prompt['eventvalidation'],
            # The entry row's submit button to 'hit'
            data_prompt['button_name']: 'Detil',
            data_prompt['kodetran_prop']: data_prompt['kodetran_val'],
            data_prompt['stat_prop']: data_prompt['stat_val']
        }

        # Posting the http payloads
        print('+ Posting http payloads...')
        post = self.session.post(LOGIN_HANDLER_URL, data=LOGIN_PAYLOAD)
        response = post.text  # --- Obtaining the response text
        content = html.fromstring(response)  # --- Scraping the HTML code

        # Obtaining the computer-generated hidden values of ASPX (after login)
        viewstate = content.xpath('//*[@id="__VIEWSTATE"]/@value')[0]
        viewstategen = content.xpath('//*[@id="__VIEWSTATEGENERATOR"]/@value')[0]
        eventvalidation = content.xpath('//*[@id="__EVENTVALIDATION"]/@value')[0]

        # Building the 'data_prompt' array
        data_prompt = {
            'http_response': response,
            'html_content': content,
            'viewstate': viewstate,
            'viewstategen': viewstategen,
            'eventvalidation': eventvalidation
        }

        # Returning the http response string
        return data_prompt

    # This function opens "Dana Penelitian" menu after opening the tab "Penelitian"
    # - Requires 'data_prompt' array as an unary argument obtained from get_risat_penelitian() function
    # - Returns also another 'data_prompt' array
    def get_risat_penelitian_dana_penelitian(self, data_prompt):
        # Logging the calling of the function
        print('+ Opening "Penelitian --> Dana Penelitian" menu...')
        
        # Preparing the http handler URL and payload
        HANDLER_URL = 'https://risat.uksw.edu/bp3mpage.aspx'
        PAYLOAD = {
            # The values below are computer-generated
            '__VIEWSTATE' : data_prompt['viewstate'],
            '__VIEWSTATEGENERATOR' : data_prompt['viewstategen'],
            '__EVENTVALIDATION' : data_prompt['eventvalidation'],
            '__EVENTTARGET' : 'ctl00$ContentPlaceHolder1$menu10'
        }
        
        # Posting the http payloads
        print('+ Posting http payloads...')
        post = self.session.post(HANDLER_URL, data=PAYLOAD)
        response = post.text # --- Obtaining the response text
        content = html.fromstring(response) # --- Scraping the HTML code
        
        # Obtaining the computer-generated hidden values of ASPX (after login)
        viewstate = content.xpath('//*[@id="__VIEWSTATE"]/@value')[0]
        viewstategen = content.xpath('//*[@id="__VIEWSTATEGENERATOR"]/@value')[0]
        eventvalidation = content.xpath('//*[@id="__EVENTVALIDATION"]/@value')[0]
        
        # Building the 'data_prompt' array
        data_prompt = {
            'http_response' : response,
            'html_content' : content,
            'viewstate' : viewstate,
            'viewstategen' : viewstategen,
            'eventvalidation' : eventvalidation
        }
        
        # Returning 'data_prompt' array
        return data_prompt

    # This function opens the detail page of "Dana Penelitian" entry row
    # Requires one argument:
    # - the data prompt value array in compliance with convention [5] of this file
    # The 'data_prompt' argument passed into this function must be
    # rooted from the function 'get_risat_penelitian_dana_penelitian()'
    def get_risat_penelitian_dana_penelitian_detil(self, data_prompt):
        # Logging the calling of the function
        print('+ Opening "Dana Penelitian" entry row detail page...')
        
        # Preparing the http handler URL and payload
        LOGIN_HANDLER_URL = 'https://risat.uksw.edu/bp3mpage.aspx'
        LOGIN_PAYLOAD = {
            # The values below are computer-generated
            '__VIEWSTATE' : data_prompt['viewstate'],
            '__VIEWSTATEGENERATOR' : data_prompt['viewstategen'],
            '__EVENTVALIDATION' : data_prompt['eventvalidation'],
            # The entry row's submit button to 'hit'
            data_prompt['button_name'] : 'Detil',
            data_prompt['kodetran_prop'] : data_prompt['kodetran_val'],
            data_prompt['stat_prop'] : data_prompt['stat_val']
        }
        
        # Posting the http payloads
        print('+ Posting http payloads...')
        post = self.session.post(LOGIN_HANDLER_URL, data=LOGIN_PAYLOAD)
        response = post.text # --- Obtaining the response text
        content = html.fromstring(response) # --- Scraping the HTML code
        
        # Obtaining the computer-generated hidden values of ASPX (after login)
        viewstate = content.xpath('//*[@id="__VIEWSTATE"]/@value')[0]
        viewstategen = content.xpath('//*[@id="__VIEWSTATEGENERATOR"]/@value')[0]
        eventvalidation = content.xpath('//*[@id="__EVENTVALIDATION"]/@value')[0]
        
        # Building the 'data_prompt' array
        data_prompt = {
            'http_response' : response,
            'html_content' : content,
            'viewstate' : viewstate,
            'viewstategen' : viewstategen,
            'eventvalidation' : eventvalidation
        }
        
        # Returning the http response string
        return data_prompt

    # This function opens "Pelaksanaan Kegiatan" menu after opening the tab "Penelitian"
    # - Requires 'data_prompt' array as an unary argument obtained from get_risat_penelitian() function
    # - Returns also another 'data_prompt' array
    def get_risat_penelitian_pelak_kegi(self, data_prompt):
        # Logging the calling of the function
        print('+ Opening "Penelitian --> Pelaksanaan Kegiatan" menu...')

        # Preparing the http handler URL and payload
        HANDLER_URL = 'https://risat.uksw.edu/bp3mpage.aspx'
        PAYLOAD = {
            # The values below are computer-generated
            '__VIEWSTATE' : data_prompt['viewstate'],
            '__VIEWSTATEGENERATOR' : data_prompt['viewstategen'],
            '__EVENTVALIDATION' : data_prompt['eventvalidation'],
            '__EVENTTARGET' : 'ctl00$ContentPlaceHolder1$menu9'
        }

        # Posting the http payloads
        print('+ Posting http payloads...')
        post = self.session.post(HANDLER_URL, data=PAYLOAD)
        response = post.text # --- Obtaining the response text
        content = html.fromstring(response) # --- Scraping the HTML code

        # Obtaining the computer-generated hidden values of ASPX (after login)
        viewstate = content.xpath('//*[@id="__VIEWSTATE"]/@value')[0]
        viewstategen = content.xpath('//*[@id="__VIEWSTATEGENERATOR"]/@value')[0]
        eventvalidation = content.xpath('//*[@id="__EVENTVALIDATION"]/@value')[0]

        # Building the 'data_prompt' array
        data_prompt = {
            'http_response' : response,
            'html_content' : content,
            'viewstate' : viewstate,
            'viewstategen' : viewstategen,
            'eventvalidation' : eventvalidation
        }

        # Returning 'data_prompt' array
        return data_prompt

    # This function opens "Laporan Akhir" menu after opening the tab "Penelitian -> Pelaksanaan Kegiatan"
    # - Requires 'data_prompt' array as an unary argument obtained from get_risat_penelitian_pelak_kegi() function
    # - Returns also another 'data_prompt' array
    def get_risat_penelitian_pelak_kegi_lapakhir(self, data_prompt):
        # Logging the calling of the function
        print('+ Opening "Penelitian --> Pelaksanaan Kegiatan --> Laporan Akhir" menu...')

        # Preparing the http handler URL and payload
        HANDLER_URL = 'https://risat.uksw.edu/bp3mpage.aspx'
        PAYLOAD = {
            # The values below are computer-generated
            '__VIEWSTATE' : data_prompt['viewstate'],
            '__VIEWSTATEGENERATOR' : data_prompt['viewstategen'],
            '__EVENTVALIDATION' : data_prompt['eventvalidation'],
            '__EVENTTARGET' : 'ctl00$ContentPlaceHolder1$submenu3'
        }

        # Posting the http payloads
        print('+ Posting http payloads...')
        post = self.session.post(HANDLER_URL, data=PAYLOAD)
        response = post.text # --- Obtaining the response text
        content = html.fromstring(response) # --- Scraping the HTML code

        # Obtaining the computer-generated hidden values of ASPX (after login)
        viewstate = content.xpath('//*[@id="__VIEWSTATE"]/@value')[0]
        viewstategen = content.xpath('//*[@id="__VIEWSTATEGENERATOR"]/@value')[0]
        eventvalidation = content.xpath('//*[@id="__EVENTVALIDATION"]/@value')[0]

        # Building the 'data_prompt' array
        data_prompt = {
            'http_response' : response,
            'html_content' : content,
            'viewstate' : viewstate,
            'viewstategen' : viewstategen,
            'eventvalidation' : eventvalidation
        }

        # Returning 'data_prompt' array
        return data_prompt

    # This function opens "Arsip" menu after opening the tab "Penelitian"
    # - Requires 'data_prompt' array as an unary argument obtained from get_risat_penelitian() function
    # - Returns also another 'data_prompt' array
    def get_risat_penelitian_arsip(self, data_prompt):
        # Logging the calling of the function
        print('+ Opening "Penelitian --> Arsip" menu...')

        # Preparing the http handler URL and payload
        HANDLER_URL = 'https://risat.uksw.edu/bp3mpage.aspx'
        PAYLOAD = {
            # The values below are computer-generated
            '__VIEWSTATE' : data_prompt['viewstate'],
            '__VIEWSTATEGENERATOR' : data_prompt['viewstategen'],
            '__EVENTVALIDATION' : data_prompt['eventvalidation'],
            '__EVENTTARGET' : 'ctl00$ContentPlaceHolder1$menu11'
        }

        # Posting the http payloads
        print('+ Posting http payloads...')
        post = self.session.post(HANDLER_URL, data=PAYLOAD)
        response = post.text # --- Obtaining the response text
        content = html.fromstring(response) # --- Scraping the HTML code

        # Obtaining the computer-generated hidden values of ASPX (after login)
        viewstate = content.xpath('//*[@id="__VIEWSTATE"]/@value')[0]
        viewstategen = content.xpath('//*[@id="__VIEWSTATEGENERATOR"]/@value')[0]
        eventvalidation = content.xpath('//*[@id="__EVENTVALIDATION"]/@value')[0]

        # Building the 'data_prompt' array
        data_prompt = {
            'http_response' : response,
            'html_content' : content,
            'viewstate' : viewstate,
            'viewstategen' : viewstategen,
            'eventvalidation' : eventvalidation
        }

        # Returning 'data_prompt' array
        return data_prompt

    # This function opens the detail page of "Arsip" entry row
    # Requires one argument:
    # - the data prompt value array in compliance with convention [5] of this file
    # The 'data_prompt' argument passed into this function must be
    # rooted from the function 'get_risat_penelitian_arsip()'
    def get_risat_penelitian_arsip_detil(self, data_prompt):
        # Logging the calling of the function
        print('+ Opening "Arsip" entry row detail page...')

        # Preparing the http handler URL and payload
        LOGIN_HANDLER_URL = 'https://risat.uksw.edu/bp3mpage.aspx'
        LOGIN_PAYLOAD = {
            # The values below are computer-generated
            '__VIEWSTATE' : data_prompt['viewstate'],
            '__VIEWSTATEGENERATOR' : data_prompt['viewstategen'],
            '__EVENTVALIDATION' : data_prompt['eventvalidation'],
            # The entry row's submit button to 'hit'
            data_prompt['button_name'] : 'Detil',
            data_prompt['idat_prop'] : data_prompt['idat_val'],
            data_prompt['itgl_prop'] : data_prompt['itgl_val']
        }

        # Posting the http payloads
        print('+ Posting http payloads...')
        post = self.session.post(LOGIN_HANDLER_URL, data=LOGIN_PAYLOAD)
        response = post.text # --- Obtaining the response text
        content = html.fromstring(response) # --- Scraping the HTML code

        # Obtaining the computer-generated hidden values of ASPX (after login)
        viewstate = content.xpath('//*[@id="__VIEWSTATE"]/@value')[0]
        viewstategen = content.xpath('//*[@id="__VIEWSTATEGENERATOR"]/@value')[0]
        eventvalidation = content.xpath('//*[@id="__EVENTVALIDATION"]/@value')[0]

        # Building the 'data_prompt' array
        data_prompt = {
            'http_response' : response,
            'html_content' : content,
            'viewstate' : viewstate,
            'viewstategen' : viewstategen,
            'eventvalidation' : eventvalidation
        }

        # Returning the http response string
        return data_prompt

    # This function opens "Pengabdian" menu tab
    # Requires 'data_prompt' array obtained from the login function
    # Returns another 'data_prompt' array
    def get_risat_pengabdian(self, data_prompt):
        # Logging the calling of the function
        print('+ Opening "Pengabdian" menu...')

        # Preparing the http handler URL and payload
        HANDLER_URL = 'https://risat.uksw.edu/bp3mpageabdimas.aspx'
        PAYLOAD = {
            # The values below are computer-generated
            '__VIEWSTATE' : data_prompt['viewstate'],
            '__VIEWSTATEGENERATOR' : data_prompt['viewstategen'],
            '__EVENTVALIDATION' : data_prompt['eventvalidation'],
            'ctl00$menu3': True
        }

        # Posting the http payloads
        print('+ Posting http payloads...')
        post = self.session.post(HANDLER_URL, data=PAYLOAD)
        response = post.text # --- Obtaining the response text
        content = html.fromstring(response) # --- Scraping the HTML code

        # Obtaining the computer-generated hidden values of ASPX (after login)
        viewstate = content.xpath('//*[@id="__VIEWSTATE"]/@value')[0]
        viewstategen = content.xpath('//*[@id="__VIEWSTATEGENERATOR"]/@value')[0]
        eventvalidation = content.xpath('//*[@id="__EVENTVALIDATION"]/@value')[0]

        # Building the 'data_prompt' array
        data_prompt = {
            'http_response' : response,
            'html_content' : content,
            'viewstate' : viewstate,
            'viewstategen' : viewstategen,
            'eventvalidation' : eventvalidation
        }

        # Returning the http response string
        return data_prompt

    # This function opens "Usulan Pengabdian" menu after opening the tab "Pengabdian"
    # - Requires 'data_prompt' array as an unary argument obtained from get_risat_pengabdian() function
    # - Returns also another 'data_prompt' array
    def get_risat_pengabdian_usulan_pengabdian(self, data_prompt):
        # Logging the calling of the function
        print('+ Opening "Pengabdian --> Usulan Pengabdian" menu...')

        # Preparing the http handler URL and payload
        HANDLER_URL = 'https://risat.uksw.edu/bp3mpageabdimas.aspx'
        PAYLOAD = {
            # The values below are computer-generated
            '__VIEWSTATE': data_prompt['viewstate'],
            '__VIEWSTATEGENERATOR': data_prompt['viewstategen'],
            '__EVENTVALIDATION': data_prompt['eventvalidation'],
            '__EVENTTARGET': 'ctl00$ContentPlaceHolder1$menu1'
        }

        # Posting the http payloads
        print('+ Posting http payloads...')
        post = self.session.post(HANDLER_URL, data=PAYLOAD)
        response = post.text  # --- Obtaining the response text
        content = html.fromstring(response)  # --- Scraping the HTML code

        # Obtaining the computer-generated hidden values of ASPX (after login)
        viewstate = content.xpath('//*[@id="__VIEWSTATE"]/@value')[0]
        viewstategen = content.xpath('//*[@id="__VIEWSTATEGENERATOR"]/@value')[0]
        eventvalidation = content.xpath('//*[@id="__EVENTVALIDATION"]/@value')[0]

        # Building the 'data_prompt' array
        data_prompt = {
            'http_response': response,
            'html_content': content,
            'viewstate': viewstate,
            'viewstategen': viewstategen,
            'eventvalidation': eventvalidation
        }

        # Returning 'data_prompt' array
        return data_prompt

    # This function opens the detail page of Usulan Pengabdian" entry row
    # Requires one argument:
    # - the data prompt value array in compliance with convention [5] of this file
    # The 'data_prompt' argument passed into this function must be
    # rooted from the function 'get_risat_pengabdian_dana_pengabdian()'
    def get_risat_pengabdian_usulan_pengabdian_detil(self, data_prompt):
        # Logging the calling of the function
        print('+ Opening "Usulan Pengabdian" entry row detail page...')

        # Preparing the http handler URL and payload
        LOGIN_HANDLER_URL = 'https://risat.uksw.edu/bp3mpageabdimas.aspx'
        LOGIN_PAYLOAD = {
            # The values below are computer-generated
            '__VIEWSTATE' : data_prompt['viewstate'],
            '__VIEWSTATEGENERATOR' : data_prompt['viewstategen'],
            '__EVENTVALIDATION' : data_prompt['eventvalidation'],
            # The entry row's submit button to 'hit'
            data_prompt['button_name'] : 'Detil',
            data_prompt['kodetran_prop'] : data_prompt['kodetran_val'],
            data_prompt['stat_prop'] : data_prompt['stat_val']
        }

        # Posting the http payloads
        print('+ Posting http payloads...')
        post = self.session.post(LOGIN_HANDLER_URL, data=LOGIN_PAYLOAD)
        response = post.text # --- Obtaining the response text
        content = html.fromstring(response) # --- Scraping the HTML code

        # Obtaining the computer-generated hidden values of ASPX (after login)
        viewstate = content.xpath('//*[@id="__VIEWSTATE"]/@value')[0]
        viewstategen = content.xpath('//*[@id="__VIEWSTATEGENERATOR"]/@value')[0]
        eventvalidation = content.xpath('//*[@id="__EVENTVALIDATION"]/@value')[0]

        # Building the 'data_prompt' array
        data_prompt = {
            'http_response' : response,
            'html_content' : content,
            'viewstate' : viewstate,
            'viewstategen' : viewstategen,
            'eventvalidation' : eventvalidation
        }

        # Returning the http response string
        return data_prompt

    # This function opens "Berkas Disetujui Pengabdian" menu after opening the tab "Pengabdian"
    # - Requires 'data_prompt' array as an unary argument obtained from get_risat_pengabdian() function
    # - Returns also another 'data_prompt' array
    def get_risat_pengabdian_berdisetujui_pengabdian(self, data_prompt):
        # Logging the calling of the function
        print('+ Opening "Pengabdian --> Berkas Disetujui Pengabdian" menu...')

        # Preparing the http handler URL and payload
        HANDLER_URL = 'https://risat.uksw.edu/bp3mpageabdimas.aspx'
        PAYLOAD = {
            # The values below are computer-generated
            '__VIEWSTATE': data_prompt['viewstate'],
            '__VIEWSTATEGENERATOR': data_prompt['viewstategen'],
            '__EVENTVALIDATION': data_prompt['eventvalidation'],
            '__EVENTTARGET': 'ctl00$ContentPlaceHolder1$menu3'
        }

        # Posting the http payloads
        print('+ Posting http payloads...')
        post = self.session.post(HANDLER_URL, data=PAYLOAD)
        response = post.text  # --- Obtaining the response text
        content = html.fromstring(response)  # --- Scraping the HTML code

        # Obtaining the computer-generated hidden values of ASPX (after login)
        viewstate = content.xpath('//*[@id="__VIEWSTATE"]/@value')[0]
        viewstategen = content.xpath('//*[@id="__VIEWSTATEGENERATOR"]/@value')[0]
        eventvalidation = content.xpath('//*[@id="__EVENTVALIDATION"]/@value')[0]

        # Building the 'data_prompt' array
        data_prompt = {
            'http_response': response,
            'html_content': content,
            'viewstate': viewstate,
            'viewstategen': viewstategen,
            'eventvalidation': eventvalidation
        }

        # Returning 'data_prompt' array
        return data_prompt

    # This function opens "Berkas Disetujui Pengabdian" menu after opening the tab "Pengabdian"
    # - Requires 'data_prompt' array as an unary argument obtained from get_risat_pengabdian() function
    # - Returns also another 'data_prompt' array
    def get_risat_pengabdian_berdisetujui_pengabdian(self, data_prompt):
        # Logging the calling of the function
        print('+ Opening "Pengabdian --> Berkas Disetujui Pengabdian" menu...')

        # Preparing the http handler URL and payload
        HANDLER_URL = 'https://risat.uksw.edu/bp3mpageabdimas.aspx'
        PAYLOAD = {
            # The values below are computer-generated
            '__VIEWSTATE': data_prompt['viewstate'],
            '__VIEWSTATEGENERATOR': data_prompt['viewstategen'],
            '__EVENTVALIDATION': data_prompt['eventvalidation'],
            '__EVENTTARGET': 'ctl00$ContentPlaceHolder1$menu3'
        }

        # Posting the http payloads
        print('+ Posting http payloads...')
        post = self.session.post(HANDLER_URL, data=PAYLOAD)
        response = post.text  # --- Obtaining the response text
        content = html.fromstring(response)  # --- Scraping the HTML code

        # Obtaining the computer-generated hidden values of ASPX (after login)
        viewstate = content.xpath('//*[@id="__VIEWSTATE"]/@value')[0]
        viewstategen = content.xpath('//*[@id="__VIEWSTATEGENERATOR"]/@value')[0]
        eventvalidation = content.xpath('//*[@id="__EVENTVALIDATION"]/@value')[0]

        # Building the 'data_prompt' array
        data_prompt = {
            'http_response': response,
            'html_content': content,
            'viewstate': viewstate,
            'viewstategen': viewstategen,
            'eventvalidation': eventvalidation
        }

        # Returning 'data_prompt' array
        return data_prompt

    # This function opens the detail page of "Berkas Disetujui Pengabdian" entry row
    # Requires one argument:
    # - the data prompt value array in compliance with convention [5] of this file
    # The 'data_prompt' argument passed into this function must be
    # rooted from the function 'get_risat_pengabdian_dana_pengabdian()'
    def get_risat_pengabdian_berdisetujui_pengabdian_detil(self, data_prompt):
        # Logging the calling of the function
        print('+ Opening "Berkas Disetujui Pengabdian" entry row detail page...')

        # Preparing the http handler URL and payload
        LOGIN_HANDLER_URL = 'https://risat.uksw.edu/bp3mpageabdimas.aspx'
        LOGIN_PAYLOAD = {
            # The values below are computer-generated
            '__VIEWSTATE' : data_prompt['viewstate'],
            '__VIEWSTATEGENERATOR' : data_prompt['viewstategen'],
            '__EVENTVALIDATION' : data_prompt['eventvalidation'],
            # The entry row's submit button to 'hit'
            data_prompt['button_name'] : 'Detil',
            data_prompt['kodetran_prop'] : data_prompt['kodetran_val'],
            data_prompt['stat_prop'] : data_prompt['stat_val']
        }

        # Posting the http payloads
        print('+ Posting http payloads...')
        post = self.session.post(LOGIN_HANDLER_URL, data=LOGIN_PAYLOAD)
        response = post.text # --- Obtaining the response text
        content = html.fromstring(response) # --- Scraping the HTML code

        # Obtaining the computer-generated hidden values of ASPX (after login)
        viewstate = content.xpath('//*[@id="__VIEWSTATE"]/@value')[0]
        viewstategen = content.xpath('//*[@id="__VIEWSTATEGENERATOR"]/@value')[0]
        eventvalidation = content.xpath('//*[@id="__EVENTVALIDATION"]/@value')[0]

        # Building the 'data_prompt' array
        data_prompt = {
            'http_response' : response,
            'html_content' : content,
            'viewstate' : viewstate,
            'viewstategen' : viewstategen,
            'eventvalidation' : eventvalidation
        }

        # Returning the http response string
        return data_prompt

    # This function opens "Berkas Direvisi Pengabdian" menu after opening the tab "Pengabdian"
    # - Requires 'data_prompt' array as an unary argument obtained from get_risat_pengabdian() function
    # - Returns also another 'data_prompt' array
    def get_risat_pengabdian_berdirevisi_pengabdian(self, data_prompt):
        # Logging the calling of the function
        print('+ Opening "Pengabdian --> Berkas Direvisi Pengabdian" menu...')

        # Preparing the http handler URL and payload
        HANDLER_URL = 'https://risat.uksw.edu/bp3mpageabdimas.aspx'
        PAYLOAD = {
            # The values below are computer-generated
            '__VIEWSTATE': data_prompt['viewstate'],
            '__VIEWSTATEGENERATOR': data_prompt['viewstategen'],
            '__EVENTVALIDATION': data_prompt['eventvalidation'],
            '__EVENTTARGET': 'ctl00$ContentPlaceHolder1$menu5'
        }

        # Posting the http payloads
        print('+ Posting http payloads...')
        post = self.session.post(HANDLER_URL, data=PAYLOAD)
        response = post.text  # --- Obtaining the response text
        content = html.fromstring(response)  # --- Scraping the HTML code

        # Obtaining the computer-generated hidden values of ASPX (after login)
        viewstate = content.xpath('//*[@id="__VIEWSTATE"]/@value')[0]
        viewstategen = content.xpath('//*[@id="__VIEWSTATEGENERATOR"]/@value')[0]
        eventvalidation = content.xpath('//*[@id="__EVENTVALIDATION"]/@value')[0]

        # Building the 'data_prompt' array
        data_prompt = {
            'http_response': response,
            'html_content': content,
            'viewstate': viewstate,
            'viewstategen': viewstategen,
            'eventvalidation': eventvalidation
        }

        # Returning 'data_prompt' array
        return data_prompt

    # This function opens the detail page of "Berkas Direvisi Pengabdian" entry row
    # Requires one argument:
    # - the data prompt value array in compliance with convention [5] of this file
    # The 'data_prompt' argument passed into this function must be
    # rooted from the function 'get_risat_pengabdian_dana_pengabdian()'
    def get_risat_pengabdian_berdirevisi_pengabdian_detil(self, data_prompt):
        # Logging the calling of the function
        print('+ Opening "Berkas Direvisi Pengabdian" entry row detail page...')

        # Preparing the http handler URL and payload
        LOGIN_HANDLER_URL = 'https://risat.uksw.edu/bp3mpageabdimas.aspx'
        LOGIN_PAYLOAD = {
            # The values below are computer-generated
            '__VIEWSTATE' : data_prompt['viewstate'],
            '__VIEWSTATEGENERATOR' : data_prompt['viewstategen'],
            '__EVENTVALIDATION' : data_prompt['eventvalidation'],
            # The entry row's submit button to 'hit'
            data_prompt['button_name'] : 'Detil',
            data_prompt['kodetran_prop'] : data_prompt['kodetran_val'],
            data_prompt['stat_prop'] : data_prompt['stat_val']
        }

        # Posting the http payloads
        print('+ Posting http payloads...')
        post = self.session.post(LOGIN_HANDLER_URL, data=LOGIN_PAYLOAD)
        response = post.text # --- Obtaining the response text
        content = html.fromstring(response) # --- Scraping the HTML code

        # Obtaining the computer-generated hidden values of ASPX (after login)
        viewstate = content.xpath('//*[@id="__VIEWSTATE"]/@value')[0]
        viewstategen = content.xpath('//*[@id="__VIEWSTATEGENERATOR"]/@value')[0]
        eventvalidation = content.xpath('//*[@id="__EVENTVALIDATION"]/@value')[0]

        # Building the 'data_prompt' array
        data_prompt = {
            'http_response' : response,
            'html_content' : content,
            'viewstate' : viewstate,
            'viewstategen' : viewstategen,
            'eventvalidation' : eventvalidation
        }

        # Returning the http response string
        return data_prompt

    # This function opens "Ditolak Pengabdian" menu after opening the tab "Pengabdian"
    # - Requires 'data_prompt' array as an unary argument obtained from get_risat_pengabdian() function
    # - Returns also another 'data_prompt' array
    def get_risat_pengabdian_ditolak_pengabdian(self, data_prompt):
        # Logging the calling of the function
        print('+ Opening "Pengabdian --> Ditolak Pengabdian" menu...')

        # Preparing the http handler URL and payload
        HANDLER_URL = 'https://risat.uksw.edu/bp3mpageabdimas.aspx'
        PAYLOAD = {
            # The values below are computer-generated
            '__VIEWSTATE' : data_prompt['viewstate'],
            '__VIEWSTATEGENERATOR' : data_prompt['viewstategen'],
            '__EVENTVALIDATION' : data_prompt['eventvalidation'],
            '__EVENTTARGET' : 'ctl00$ContentPlaceHolder1$menu8'
        }

        # Posting the http payloads
        print('+ Posting http payloads...')
        post = self.session.post(HANDLER_URL, data=PAYLOAD)
        response = post.text # --- Obtaining the response text
        content = html.fromstring(response) # --- Scraping the HTML code

        # Obtaining the computer-generated hidden values of ASPX (after login)
        viewstate = content.xpath('//*[@id="__VIEWSTATE"]/@value')[0]
        viewstategen = content.xpath('//*[@id="__VIEWSTATEGENERATOR"]/@value')[0]
        eventvalidation = content.xpath('//*[@id="__EVENTVALIDATION"]/@value')[0]

        # Building the 'data_prompt' array
        data_prompt = {
            'http_response' : response,
            'html_content' : content,
            'viewstate' : viewstate,
            'viewstategen' : viewstategen,
            'eventvalidation' : eventvalidation
        }

        # Returning 'data_prompt' array
        return data_prompt

    # This function opens "Dana Pengabdian" menu after opening the tab "Pengabdian"
    # - Requires 'data_prompt' array as an unary argument obtained from get_risat_pengabdian() function
    # - Returns also another 'data_prompt' array
    def get_risat_pengabdian_dana_pengabdian(self, data_prompt):
        # Logging the calling of the function
        print('+ Opening "Pengabdian --> Dana Pengabdian" menu...')

        # Preparing the http handler URL and payload
        HANDLER_URL = 'https://risat.uksw.edu/bp3mpageabdimas.aspx'
        PAYLOAD = {
            # The values below are computer-generated
            '__VIEWSTATE' : data_prompt['viewstate'],
            '__VIEWSTATEGENERATOR' : data_prompt['viewstategen'],
            '__EVENTVALIDATION' : data_prompt['eventvalidation'],
            '__EVENTTARGET' : 'ctl00$ContentPlaceHolder1$menu10'
        }

        # Posting the http payloads
        print('+ Posting http payloads...')
        post = self.session.post(HANDLER_URL, data=PAYLOAD)
        response = post.text # --- Obtaining the response text
        content = html.fromstring(response) # --- Scraping the HTML code

        # Obtaining the computer-generated hidden values of ASPX (after login)
        viewstate = content.xpath('//*[@id="__VIEWSTATE"]/@value')[0]
        viewstategen = content.xpath('//*[@id="__VIEWSTATEGENERATOR"]/@value')[0]
        eventvalidation = content.xpath('//*[@id="__EVENTVALIDATION"]/@value')[0]

        # Building the 'data_prompt' array
        data_prompt = {
            'http_response' : response,
            'html_content' : content,
            'viewstate' : viewstate,
            'viewstategen' : viewstategen,
            'eventvalidation' : eventvalidation
        }

        # Returning 'data_prompt' array
        return data_prompt

    # This function opens the detail page of "Dana Pengabdian" entry row
    # Requires one argument:
    # - the data prompt value array in compliance with convention [5] of this file
    # The 'data_prompt' argument passed into this function must be
    # rooted from the function 'get_risat_pengabdian_dana_pengabdian()'
    def get_risat_pengabdian_dana_pengabdian_detil(self, data_prompt):
        # Logging the calling of the function
        print('+ Opening "Dana Pengabdian" entry row detail page...')

        # Preparing the http handler URL and payload
        LOGIN_HANDLER_URL = 'https://risat.uksw.edu/bp3mpageabdimas.aspx'
        LOGIN_PAYLOAD = {
            # The values below are computer-generated
            '__VIEWSTATE' : data_prompt['viewstate'],
            '__VIEWSTATEGENERATOR' : data_prompt['viewstategen'],
            '__EVENTVALIDATION' : data_prompt['eventvalidation'],
            # The entry row's submit button to 'hit'
            data_prompt['button_name'] : 'Detil',
            data_prompt['kodetran_prop'] : data_prompt['kodetran_val'],
            data_prompt['stat_prop'] : data_prompt['stat_val']
        }

        # Posting the http payloads
        print('+ Posting http payloads...')
        post = self.session.post(LOGIN_HANDLER_URL, data=LOGIN_PAYLOAD)
        response = post.text # --- Obtaining the response text
        content = html.fromstring(response) # --- Scraping the HTML code

        # Obtaining the computer-generated hidden values of ASPX (after login)
        viewstate = content.xpath('//*[@id="__VIEWSTATE"]/@value')[0]
        viewstategen = content.xpath('//*[@id="__VIEWSTATEGENERATOR"]/@value')[0]
        eventvalidation = content.xpath('//*[@id="__EVENTVALIDATION"]/@value')[0]

        # Building the 'data_prompt' array
        data_prompt = {
            'http_response' : response,
            'html_content' : content,
            'viewstate' : viewstate,
            'viewstategen' : viewstategen,
            'eventvalidation' : eventvalidation
        }

        # Returning the http response string
        return data_prompt

    # This function opens "Pelaksanaan Kegiatan" menu after opening the tab "Pengabdian"
    # - Requires 'data_prompt' array as an unary argument obtained from get_risat_pengabdian() function
    # - Returns also another 'data_prompt' array
    def get_risat_pengabdian_pelak_kegi(self, data_prompt):
        # Logging the calling of the function
        print('+ Opening "Pengabdian --> Pelaksanaan Kegiatan" menu...')

        # Preparing the http handler URL and payload
        HANDLER_URL = 'https://risat.uksw.edu/bp3mpageabdimas.aspx'
        PAYLOAD = {
            # The values below are computer-generated
            '__VIEWSTATE' : data_prompt['viewstate'],
            '__VIEWSTATEGENERATOR' : data_prompt['viewstategen'],
            '__EVENTVALIDATION' : data_prompt['eventvalidation'],
            '__EVENTTARGET' : 'ctl00$ContentPlaceHolder1$menu9'
        }

        # Posting the http payloads
        print('+ Posting http payloads...')
        post = self.session.post(HANDLER_URL, data=PAYLOAD)
        response = post.text # --- Obtaining the response text
        content = html.fromstring(response) # --- Scraping the HTML code

        # Obtaining the computer-generated hidden values of ASPX (after login)
        viewstate = content.xpath('//*[@id="__VIEWSTATE"]/@value')[0]
        viewstategen = content.xpath('//*[@id="__VIEWSTATEGENERATOR"]/@value')[0]
        eventvalidation = content.xpath('//*[@id="__EVENTVALIDATION"]/@value')[0]

        # Building the 'data_prompt' array
        data_prompt = {
            'http_response' : response,
            'html_content' : content,
            'viewstate' : viewstate,
            'viewstategen' : viewstategen,
            'eventvalidation' : eventvalidation
        }

        # Returning 'data_prompt' array
        return data_prompt

    # This function opens "Laporan Akhir" menu after opening the tab "Pengabdian -> Pelaksanaan Kegiatan"
    # - Requires 'data_prompt' array as an unary argument obtained from get_risat_pengabdian_pelak_kegi() function
    # - Returns also another 'data_prompt' array
    def get_risat_pengabdian_pelak_kegi_lapakhir(self, data_prompt):
        # Logging the calling of the function
        print('+ Opening "Pengabdian --> Pelaksanaan Kegiatan --> Laporan Akhir" menu...')

        # Preparing the http handler URL and payload
        HANDLER_URL = 'https://risat.uksw.edu/bp3mpageabdimas.aspx'
        PAYLOAD = {
            # The values below are computer-generated
            '__VIEWSTATE' : data_prompt['viewstate'],
            '__VIEWSTATEGENERATOR' : data_prompt['viewstategen'],
            '__EVENTVALIDATION' : data_prompt['eventvalidation'],
            '__EVENTTARGET' : 'ctl00$ContentPlaceHolder1$submenu3'
        }

        # Posting the http payloads
        print('+ Posting http payloads...')
        post = self.session.post(HANDLER_URL, data=PAYLOAD)
        response = post.text # --- Obtaining the response text
        content = html.fromstring(response) # --- Scraping the HTML code

        # Obtaining the computer-generated hidden values of ASPX (after login)
        viewstate = content.xpath('//*[@id="__VIEWSTATE"]/@value')[0]
        viewstategen = content.xpath('//*[@id="__VIEWSTATEGENERATOR"]/@value')[0]
        eventvalidation = content.xpath('//*[@id="__EVENTVALIDATION"]/@value')[0]

        # Building the 'data_prompt' array
        data_prompt = {
            'http_response' : response,
            'html_content' : content,
            'viewstate' : viewstate,
            'viewstategen' : viewstategen,
            'eventvalidation' : eventvalidation
        }

        # Returning 'data_prompt' array
        return data_prompt

    # This function opens "Arsip" menu after opening the tab "Pengabdian"
    # - Requires 'data_prompt' array as an unary argument obtained from get_risat_pengabdian() function
    # - Returns also another 'data_prompt' array
    def get_risat_pengabdian_arsip(self, data_prompt):
        # Logging the calling of the function
        print('+ Opening "Pengabdian --> Arsip" menu...')

        # Preparing the http handler URL and payload
        HANDLER_URL = 'https://risat.uksw.edu/bp3mpageabdimas.aspx'
        PAYLOAD = {
            # The values below are computer-generated
            '__VIEWSTATE' : data_prompt['viewstate'],
            '__VIEWSTATEGENERATOR' : data_prompt['viewstategen'],
            '__EVENTVALIDATION' : data_prompt['eventvalidation'],
            '__EVENTTARGET' : 'ctl00$ContentPlaceHolder1$menu11'
        }

        # Posting the http payloads
        print('+ Posting http payloads...')
        post = self.session.post(HANDLER_URL, data=PAYLOAD)
        response = post.text # --- Obtaining the response text
        content = html.fromstring(response) # --- Scraping the HTML code

        # Obtaining the computer-generated hidden values of ASPX (after login)
        viewstate = content.xpath('//*[@id="__VIEWSTATE"]/@value')[0]
        viewstategen = content.xpath('//*[@id="__VIEWSTATEGENERATOR"]/@value')[0]
        eventvalidation = content.xpath('//*[@id="__EVENTVALIDATION"]/@value')[0]

        # Building the 'data_prompt' array
        data_prompt = {
            'http_response' : response,
            'html_content' : content,
            'viewstate' : viewstate,
            'viewstategen' : viewstategen,
            'eventvalidation' : eventvalidation
        }

        # Returning 'data_prompt' array
        return data_prompt

    # This function opens the detail page of "Arsip" entry row
    # Requires one argument:
    # - the data prompt value array in compliance with convention [5] of this file
    # The 'data_prompt' argument passed into this function must be
    # rooted from the function 'get_risat_pengabdian_arsip()'
    def get_risat_pengabdian_arsip_detil(self, data_prompt):
        # Logging the calling of the function
        print('+ Opening "Arsip" entry row detail page...')

        # Preparing the http handler URL and payload
        LOGIN_HANDLER_URL = 'https://risat.uksw.edu/bp3mpageabdimas.aspx'
        LOGIN_PAYLOAD = {
            # The values below are computer-generated
            '__VIEWSTATE' : data_prompt['viewstate'],
            '__VIEWSTATEGENERATOR' : data_prompt['viewstategen'],
            '__EVENTVALIDATION' : data_prompt['eventvalidation'],
            # The entry row's submit button to 'hit'
            data_prompt['button_name'] : 'Detil',
            data_prompt['idat_prop'] : data_prompt['idat_val'],
            data_prompt['itgl_prop'] : data_prompt['itgl_val']
        }

        # Posting the http payloads
        print('+ Posting http payloads...')
        post = self.session.post(LOGIN_HANDLER_URL, data=LOGIN_PAYLOAD)
        response = post.text # --- Obtaining the response text
        content = html.fromstring(response) # --- Scraping the HTML code

        # Obtaining the computer-generated hidden values of ASPX (after login)
        viewstate = content.xpath('//*[@id="__VIEWSTATE"]/@value')[0]
        viewstategen = content.xpath('//*[@id="__VIEWSTATEGENERATOR"]/@value')[0]
        eventvalidation = content.xpath('//*[@id="__EVENTVALIDATION"]/@value')[0]

        # Building the 'data_prompt' array
        data_prompt = {
            'http_response' : response,
            'html_content' : content,
            'viewstate' : viewstate,
            'viewstategen' : viewstategen,
            'eventvalidation' : eventvalidation
        }

        # Returning the http response string
        return data_prompt

    # This function harvests "Risat Usulan Penelitian > Ringkasan Data" data
    # and then store the harvested data as an excel file
    #
    # Required arguments:
    # - control             --> for updating the progress bar and
    #                           message area of the screen SipesatScrHarvest
    # - username, password  --> the Risat administrator username and password
    def run_harvest_r_0_usulan(self, control, username, password):
        # SipesatScrHarvest messenger
        control.set_header_desc('Panen Data "Risat Usulan Penelitian > Ringkasan Data"')
        control.set_help_label('Data sedang dipanen. Silahkan menunggu.')
        control.set_progress_bar(0)
        control.clear_message_area()

        # Preamble logging
        control.append_message_area(f'+ Memulai pemanenan data ...')
        control.append_message_area(f'+ Pemanenan dimulai pada: {str(dt.now())}')
        control.set_progress_bar(5)

        # Preparing the 'data_prompt' arrays
        control.append_message_area(f'+ Log masuk Risat sebagai [{username}] ...')
        control.set_progress_bar(10)
        data_prompt = self.get_risat_login(username, password)
        data_prompt = self.get_risat_penelitian(data_prompt)
        data_prompt = self.get_risat_penelitian_usulan_penelitian(data_prompt)

        # Parsing XML tree content
        control.append_message_area(f'+ Membaca halaman web ...')
        control.set_progress_bar(15)
        content = data_prompt['html_content']

        # Establishing the export spreadsheet file
        control.append_message_area(f'+ Mempersiapkan file spreadsheet luaran ...')
        control.set_progress_bar(20)
        workbook = xl.Workbook()
        sheet = workbook.active
        sheet.title = 'Usulan Penelitian Ringkasan'

        # Preparing the sheet header
        control.append_message_area(f'+ Mempersiapkan kepala lembar spreadsheet ...')
        control.set_progress_bar(25)
        sheet['A1'].value = 'No.'
        sheet['B1'].value = 'Judul'
        sheet['C1'].value = 'Ketua'
        sheet['D1'].value = 'Jml. Anggota'
        sheet['E1'].value = 'Tgl. Usulan'
        sheet['F1'].value = 'Bidang Fokus'
        sheet['G1'].value = 'Rencana Biaya'
        sheet['H1'].value = 'Lama Kegiatan'

        # The base XPath location, pointing to each entry row
        base = '//div[@class="mw-100"]//div[@class="form-group f12"]/table[@width="100%"]//tr[@valign="top"]'

        # ---
        # Obtaining the data row values
        control.append_message_area(f'+ Mendapatkan data pada baris tabel ...')
        control.set_progress_bar(30)

        # HYPOTHESIS:
        # Xpath cannot detect 'tbody' element.
        # So instead of using 'table/tbody/tr', use 'table//tr' instead
        #
        # RESULT:
        # The hypothesis is correct.
        # Therefore, don't mention 'tbody' in any of the following Xpath paths

        a1 = [str(i)
              for i in range(1, len(content.xpath(base))+1)]

        b1 = [l.strip()
              for l in content.xpath(base + '//span[@class="hijau"]/text()')]

        c1 = [l.replace('Ketua:', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[2]/td/table//tr/td[1]/text()')]

        d1 = [l.replace('Jumlah Anggota:', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[2]/td/table//tr/td[3]/text()')]

        e1 = [l.replace('Tgl Usulan:', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[3]/td[1]/text()')]

        f1 = [l.replace('Bidang Fokus:', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[4]/td[1]/text()')]

        g1 = [l.replace('Rencana Biaya:', '').replace('Rp.', '').replace(',', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[5]/td[1]/text()')]

        h1 = [l.replace('Lama Kegiatan:', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[6]/td[1]/text()[1]')]

        # The starting row coordinate of the active sheet
        row_start = 2

        # DEBUG
        # Please comment out after use
        # ---
        # print(a1, b1, c1, d1, e1, f1, g1, h1, i1, j1, k1, l1)

        # Iterating through each table row and write to the table
        # Assumes the lists a1, b1, c1, ... have the same array size
        control.append_message_area(f'+ Melakukan iterasi terhadap baris tabel dan menulis spreadsheet luaran ...')
        control.set_progress_bar(35)
        for i in range(len(a1)):

            # Noisy preamble logging
            # Please don't use this -_-
            # ---
            # control.append_message_area(f'ITERASI [{i}]')

            # Updating the progress bar status
            control.set_progress_bar(35 + round(45*(i+1)/(len(a1))))

            # Painting the scraped data to the output spreadsheet row
            sheet[f'A{row_start}'] = a1[i]
            sheet[f'B{row_start}'] = b1[i]
            sheet[f'C{row_start}'] = c1[i]
            sheet[f'D{row_start}'] = d1[i]
            sheet[f'E{row_start}'] = e1[i]
            sheet[f'F{row_start}'] = f1[i]
            sheet[f'G{row_start}'] = g1[i]
            sheet[f'H{row_start}'] = h1[i]

            # Incrementing the 'row_start' iterator
            # Then continue the loop
            row_start += 1
            continue
        
        # Post-loop logging: successfully painted the output spreadsheet file
        control.append_message_area(f'+ Tabel sukses dipanen!')
        control.set_progress_bar(85)

        # Asking for the spreadsheet name to save as
        # ---
        # Logging and setting the progress bar
        control.append_message_area(f'+ Menyimpan spreadsheet luaran ...')
        control.set_progress_bar(90)
        # Dealing with file name prompt and saving
        # Using loop to mitigate the user clicking 'cancel'
        # in the file name dialog prompt
        while True:
            # Opening the dialog prompt
            output_spreadsheet = filedialog.asksaveasfilename(
                filetypes=[('Excel files', '*.xlsx')],
                initialfile='Sipesat - Usulan Penelitian Ringkasan Risat.xlsx',
                title='Simpan sebagai ...'
            )
            
            # 'cancel' button in the dialog prompt is clicked
            if len(output_spreadsheet) == 0:
                # Showing confirmation
                x = messagebox.askyesno(
                    'Nama File Kosong',
                    'Apakah Anda yakin ingin melanjutkan tanpa menyimpan file spreadsheet hasil pemanenan?'
                )
                # Determining whether to break or to continue the loop
                # based on the inversed value of 'x'
                if x:
                    control.append_message_area(f'+ Finalisasi pemanenan data tanpa menyimpan file spreadsheet luaran ...')
                    workbook.close()  # --- closing the workbook without saving
                    break
                else:
                    continue  # --- continuing the loop
            # File name does not end in spreadsheet extension
            elif output_spreadsheet[-5:] != '.xlsx':
                output_spreadsheet = output_spreadsheet + '.xlsx'
            
            # Saving the spreadsheet
            control.append_message_area(f'LOKASI_SPREADSHEET_LUARAN: {output_spreadsheet}')
            control.set_progress_bar(95)
            workbook.save(output_spreadsheet)
    
            # Closing the openpyxl workbook
            control.append_message_area(f'+ Menutup file spreadsheet ...')
            control.set_progress_bar(98)
            workbook.close()

            # Breaking the loop
            break

        # Notify for a successful scraping
        control.append_message_area(f'+ Pemanenan selesai pada: {str(dt.now())}')
        control.set_progress_bar(100)
        control.on_notify_successful_scraping()

    # This function harvests "Risat Berkas Disetujui Penelitian > Ringkasan Data" data
    # and then store the harvested data as an excel file
    #
    # Required arguments:
    # - control             --> for updating the progress bar and
    #                           message area of the screen SipesatScrHarvest
    # - username, password  --> the Risat administrator username and password
    def run_harvest_r_0_berdisetujui(self, control, username, password):
        # SipesatScrHarvest messenger
        control.set_header_desc('Panen Data "Risat Berkas Disetujui Penelitian > Ringkasan Data"')
        control.set_help_label('Data sedang dipanen. Silahkan menunggu.')
        control.set_progress_bar(0)
        control.clear_message_area()

        # Preamble logging
        control.append_message_area(f'+ Memulai pemanenan data ...')
        control.append_message_area(f'+ Pemanenan dimulai pada: {str(dt.now())}')
        control.set_progress_bar(5)

        # Preparing the 'data_prompt' arrays
        control.append_message_area(f'+ Log masuk Risat sebagai [{username}] ...')
        control.set_progress_bar(10)
        data_prompt = self.get_risat_login(username, password)
        data_prompt = self.get_risat_penelitian(data_prompt)
        data_prompt = self.get_risat_penelitian_berdisetujui_penelitian(data_prompt)

        # Parsing XML tree content
        control.append_message_area(f'+ Membaca halaman web ...')
        control.set_progress_bar(15)
        content = data_prompt['html_content']

        # Establishing the export spreadsheet file
        control.append_message_area(f'+ Mempersiapkan file spreadsheet luaran ...')
        control.set_progress_bar(20)
        workbook = xl.Workbook()
        sheet = workbook.active
        sheet.title = 'Berkas Disetujui Penelitian Ringkasan'

        # Preparing the sheet header
        control.append_message_area(f'+ Mempersiapkan kepala lembar spreadsheet ...')
        control.set_progress_bar(25)
        sheet['A1'].value = 'No.'
        sheet['B1'].value = 'Judul'
        sheet['C1'].value = 'Ketua'
        sheet['D1'].value = 'Jml. Anggota'
        sheet['E1'].value = 'Tgl. Usulan'
        sheet['F1'].value = 'Bidang Fokus'
        sheet['G1'].value = 'Rencana Biaya'
        sheet['H1'].value = 'Lama Kegiatan'

        # The base XPath location, pointing to each entry row
        base = '//div[@class="mw-100"]//div[@class="form-group f12"]/table[@width="100%"]//tr[@valign="top"]'

        # ---
        # Obtaining the data row values
        control.append_message_area(f'+ Mendapatkan data pada baris tabel ...')
        control.set_progress_bar(30)

        # HYPOTHESIS:
        # Xpath cannot detect 'tbody' element.
        # So instead of using 'table/tbody/tr', use 'table//tr' instead
        #
        # RESULT:
        # The hypothesis is correct.
        # Therefore, don't mention 'tbody' in any of the following Xpath paths

        a1 = [str(i)
              for i in range(1, len(content.xpath(base))+1)]

        b1 = [l.strip()
              for l in content.xpath(base + '//span[@class="hijau"]/text()')]

        c1 = [l.replace('Ketua:', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[2]/td/table//tr/td[1]/text()')]

        d1 = [l.replace('Jumlah Anggota:', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[2]/td/table//tr/td[3]/text()')]

        e1 = [l.replace('Tgl Usulan:', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[3]/td[1]/text()')]

        f1 = [l.replace('Bidang Fokus:', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[4]/td[1]/text()')]

        g1 = [l.replace('Rencana Biaya:', '').replace('Rp.', '').replace(',', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[5]/td[1]/text()')]

        h1 = [l.replace('Lama Kegiatan:', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[6]/td[1]/text()[1]')]

        # The starting row coordinate of the active sheet
        row_start = 2

        # DEBUG
        # Please comment out after use
        # ---
        # print(a1, b1, c1, d1, e1, f1, g1, h1, i1, j1, k1, l1)

        # Iterating through each table row and write to the table
        # Assumes the lists a1, b1, c1, ... have the same array size
        control.append_message_area(f'+ Melakukan iterasi terhadap baris tabel dan menulis spreadsheet luaran ...')
        control.set_progress_bar(35)
        for i in range(len(a1)):

            # Noisy preamble logging
            # Please don't use this -_-
            # ---
            # control.append_message_area(f'ITERASI [{i}]')

            # Updating the progress bar status
            control.set_progress_bar(35 + round(45*(i+1)/(len(a1))))

            # Painting the scraped data to the output spreadsheet row
            sheet[f'A{row_start}'] = a1[i]
            sheet[f'B{row_start}'] = b1[i]
            sheet[f'C{row_start}'] = c1[i]
            sheet[f'D{row_start}'] = d1[i]
            sheet[f'E{row_start}'] = e1[i]
            sheet[f'F{row_start}'] = f1[i]
            sheet[f'G{row_start}'] = g1[i]
            sheet[f'H{row_start}'] = h1[i]

            # Incrementing the 'row_start' iterator
            # Then continue the loop
            row_start += 1
            continue

        # Post-loop logging: successfully painted the output spreadsheet file
        control.append_message_area(f'+ Tabel sukses dipanen!')
        control.set_progress_bar(85)

        # Asking for the spreadsheet name to save as
        # ---
        # Logging and setting the progress bar
        control.append_message_area(f'+ Menyimpan spreadsheet luaran ...')
        control.set_progress_bar(90)
        # Dealing with file name prompt and saving
        # Using loop to mitigate the user clicking 'cancel'
        # in the file name dialog prompt
        while True:
            # Opening the dialog prompt
            output_spreadsheet = filedialog.asksaveasfilename(
                filetypes=[('Excel files', '*.xlsx')],
                initialfile='Sipesat - Berkas Disetujui Penelitian Ringkasan Risat.xlsx',
                title='Simpan sebagai ...'
            )

            # 'cancel' button in the dialog prompt is clicked
            if len(output_spreadsheet) == 0:
                # Showing confirmation
                x = messagebox.askyesno(
                    'Nama File Kosong',
                    'Apakah Anda yakin ingin melanjutkan tanpa menyimpan file spreadsheet hasil pemanenan?'
                )
                # Determining whether to break or to continue the loop
                # based on the inversed value of 'x'
                if x:
                    control.append_message_area(f'+ Finalisasi pemanenan data tanpa menyimpan file spreadsheet luaran ...')
                    workbook.close()  # --- closing the workbook without saving
                    break
                else:
                    continue  # --- continuing the loop
            # File name does not end in spreadsheet extension
            elif output_spreadsheet[-5:] != '.xlsx':
                output_spreadsheet = output_spreadsheet + '.xlsx'

            # Saving the spreadsheet
            control.append_message_area(f'LOKASI_SPREADSHEET_LUARAN: {output_spreadsheet}')
            control.set_progress_bar(95)
            workbook.save(output_spreadsheet)

            # Closing the openpyxl workbook
            control.append_message_area(f'+ Menutup file spreadsheet ...')
            control.set_progress_bar(98)
            workbook.close()

            # Breaking the loop
            break

        # Notify for a successful scraping
        control.append_message_area(f'+ Pemanenan selesai pada: {str(dt.now())}')
        control.set_progress_bar(100)
        control.on_notify_successful_scraping()

    # This function harvests "Risat Berkas Direvisi Penelitian > Ringkasan Data" data
    # and then store the harvested data as an excel file
    #
    # Required arguments:
    # - control             --> for updating the progress bar and
    #                           message area of the screen SipesatScrHarvest
    # - username, password  --> the Risat administrator username and password
    def run_harvest_r_0_berdirevisi(self, control, username, password):
        # SipesatScrHarvest messenger
        control.set_header_desc('Panen Data "Risat Berkas Direvisi Penelitian > Ringkasan Data"')
        control.set_help_label('Data sedang dipanen. Silahkan menunggu.')
        control.set_progress_bar(0)
        control.clear_message_area()

        # Preamble logging
        control.append_message_area(f'+ Memulai pemanenan data ...')
        control.append_message_area(f'+ Pemanenan dimulai pada: {str(dt.now())}')
        control.set_progress_bar(5)

        # Preparing the 'data_prompt' arrays
        control.append_message_area(f'+ Log masuk Risat sebagai [{username}] ...')
        control.set_progress_bar(10)
        data_prompt = self.get_risat_login(username, password)
        data_prompt = self.get_risat_penelitian(data_prompt)
        data_prompt = self.get_risat_penelitian_berdirevisi_penelitian(data_prompt)

        # Parsing XML tree content
        control.append_message_area(f'+ Membaca halaman web ...')
        control.set_progress_bar(15)
        content = data_prompt['html_content']

        # Establishing the export spreadsheet file
        control.append_message_area(f'+ Mempersiapkan file spreadsheet luaran ...')
        control.set_progress_bar(20)
        workbook = xl.Workbook()
        sheet = workbook.active
        sheet.title = 'Berkas Direvisi Penelitian Ringkasan'

        # Preparing the sheet header
        control.append_message_area(f'+ Mempersiapkan kepala lembar spreadsheet ...')
        control.set_progress_bar(25)
        sheet['A1'].value = 'No.'
        sheet['B1'].value = 'Judul'
        sheet['C1'].value = 'Ketua'
        sheet['D1'].value = 'Jml. Anggota'
        sheet['E1'].value = 'Tgl. Usulan'
        sheet['F1'].value = 'Bidang Fokus'
        sheet['G1'].value = 'Rencana Biaya'
        sheet['H1'].value = 'Lama Kegiatan'

        # The base XPath location, pointing to each entry row
        base = '//div[@class="mw-100"]//div[@class="form-group f12"]/table[@width="100%"]//tr[@valign="top"]'

        # ---
        # Obtaining the data row values
        control.append_message_area(f'+ Mendapatkan data pada baris tabel ...')
        control.set_progress_bar(30)

        # HYPOTHESIS:
        # Xpath cannot detect 'tbody' element.
        # So instead of using 'table/tbody/tr', use 'table//tr' instead
        #
        # RESULT:
        # The hypothesis is correct.
        # Therefore, don't mention 'tbody' in any of the following Xpath paths

        a1 = [str(i)
              for i in range(1, len(content.xpath(base))+1)]

        b1 = [l.strip()
              for l in content.xpath(base + '//span[@class="hijau"]/text()')]

        c1 = [l.replace('Ketua:', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[2]/td/table//tr/td[1]/text()')]

        d1 = [l.replace('Jumlah Anggota:', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[2]/td/table//tr/td[3]/text()')]

        e1 = [l.replace('Tgl Usulan:', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[3]/td[1]/text()')]

        f1 = [l.replace('Bidang Fokus:', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[4]/td[1]/text()')]

        g1 = [l.replace('Rencana Biaya:', '').replace('Rp.', '').replace(',', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[5]/td[1]/text()')]

        h1 = [l.replace('Lama Kegiatan:', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[6]/td[1]/text()[1]')]

        # The starting row coordinate of the active sheet
        row_start = 2

        # DEBUG
        # Please comment out after use
        # ---
        # print(a1, b1, c1, d1, e1, f1, g1, h1, i1, j1, k1, l1)

        # Iterating through each table row and write to the table
        # Assumes the lists a1, b1, c1, ... have the same array size
        control.append_message_area(f'+ Melakukan iterasi terhadap baris tabel dan menulis spreadsheet luaran ...')
        control.set_progress_bar(35)
        for i in range(len(a1)):

            # Noisy preamble logging
            # Please don't use this -_-
            # ---
            # control.append_message_area(f'ITERASI [{i}]')

            # Updating the progress bar status
            control.set_progress_bar(35 + round(45*(i+1)/(len(a1))))

            # Painting the scraped data to the output spreadsheet row
            sheet[f'A{row_start}'] = a1[i]
            sheet[f'B{row_start}'] = b1[i]
            sheet[f'C{row_start}'] = c1[i]
            sheet[f'D{row_start}'] = d1[i]
            sheet[f'E{row_start}'] = e1[i]
            sheet[f'F{row_start}'] = f1[i]
            sheet[f'G{row_start}'] = g1[i]
            sheet[f'H{row_start}'] = h1[i]

            # Incrementing the 'row_start' iterator
            # Then continue the loop
            row_start += 1
            continue

        # Post-loop logging: successfully painted the output spreadsheet file
        control.append_message_area(f'+ Tabel sukses dipanen!')
        control.set_progress_bar(85)

        # Asking for the spreadsheet name to save as
        # ---
        # Logging and setting the progress bar
        control.append_message_area(f'+ Menyimpan spreadsheet luaran ...')
        control.set_progress_bar(90)
        # Dealing with file name prompt and saving
        # Using loop to mitigate the user clicking 'cancel'
        # in the file name dialog prompt
        while True:
            # Opening the dialog prompt
            output_spreadsheet = filedialog.asksaveasfilename(
                filetypes=[('Excel files', '*.xlsx')],
                initialfile='Sipesat - Berkas Direvisi Penelitian Ringkasan Risat.xlsx',
                title='Simpan sebagai ...'
            )

            # 'cancel' button in the dialog prompt is clicked
            if len(output_spreadsheet) == 0:
                # Showing confirmation
                x = messagebox.askyesno(
                    'Nama File Kosong',
                    'Apakah Anda yakin ingin melanjutkan tanpa menyimpan file spreadsheet hasil pemanenan?'
                )
                # Determining whether to break or to continue the loop
                # based on the inversed value of 'x'
                if x:
                    control.append_message_area(f'+ Finalisasi pemanenan data tanpa menyimpan file spreadsheet luaran ...')
                    workbook.close()  # --- closing the workbook without saving
                    break
                else:
                    continue  # --- continuing the loop
            # File name does not end in spreadsheet extension
            elif output_spreadsheet[-5:] != '.xlsx':
                output_spreadsheet = output_spreadsheet + '.xlsx'

            # Saving the spreadsheet
            control.append_message_area(f'LOKASI_SPREADSHEET_LUARAN: {output_spreadsheet}')
            control.set_progress_bar(95)
            workbook.save(output_spreadsheet)

            # Closing the openpyxl workbook
            control.append_message_area(f'+ Menutup file spreadsheet ...')
            control.set_progress_bar(98)
            workbook.close()

            # Breaking the loop
            break

        # Notify for a successful scraping
        control.append_message_area(f'+ Pemanenan selesai pada: {str(dt.now())}')
        control.set_progress_bar(100)
        control.on_notify_successful_scraping()

    # This function harvests "Risat Telah Direview Penelitian > Ringkasan Data" data
    # and then store the harvested data as an excel file
    #
    # Required arguments:
    # - control             --> for updating the progress bar and
    #                           message area of the screen SipesatScrHarvest
    # - username, password  --> the Risat administrator username and password
    def run_harvest_r_0_teldireview(self, control, username, password):
        # SipesatScrHarvest messenger
        control.set_header_desc('Panen Data "Risat Telah Direview Penelitian > Ringkasan Data"')
        control.set_help_label('Data sedang dipanen. Silahkan menunggu.')
        control.set_progress_bar(0)
        control.clear_message_area()

        # Preamble logging
        control.append_message_area(f'+ Memulai pemanenan data ...')
        control.append_message_area(f'+ Pemanenan dimulai pada: {str(dt.now())}')
        control.set_progress_bar(5)

        # Preparing the 'data_prompt' arrays
        control.append_message_area(f'+ Log masuk Risat sebagai [{username}] ...')
        control.set_progress_bar(10)
        data_prompt = self.get_risat_login(username, password)
        data_prompt = self.get_risat_penelitian(data_prompt)
        data_prompt = self.get_risat_penelitian_teldireview_penelitian(data_prompt)

        # Parsing XML tree content
        control.append_message_area(f'+ Membaca halaman web ...')
        control.set_progress_bar(15)
        content = data_prompt['html_content']

        # Establishing the export spreadsheet file
        control.append_message_area(f'+ Mempersiapkan file spreadsheet luaran ...')
        control.set_progress_bar(20)
        workbook = xl.Workbook()
        sheet = workbook.active
        sheet.title = 'Telah Direview Penelitian Ringkasan'

        # Preparing the sheet header
        control.append_message_area(f'+ Mempersiapkan kepala lembar spreadsheet ...')
        control.set_progress_bar(25)
        # ---
        # Preparing the "NOMOR" header
        sheet.merge_cells(
            start_row=1, start_column=1, end_row=2, end_column=1
        )
        sheet['A1'].value = 'No.'
        sheet['A1'].alignment = Alignment(horizontal='center')
        # ---
        # Preparing the "JUDUL" header
        sheet.merge_cells(
            start_row=1, start_column=2, end_row=1, end_column=11
        )
        sheet['B1'].value = 'IDENTITAS'
        sheet['B1'].alignment = Alignment(horizontal='center')
        # Preparing the "JUDUL" sub-headers
        sheet['B2'].value = 'Judul'
        sheet['C2'].value = 'Ketua'
        sheet['D2'].value = 'Jml. Anggota'
        sheet['E2'].value = 'Tgl. Usulan'
        sheet['F2'].value = 'Bidang Fokus'
        sheet['G2'].value = 'Rencana Biaya'
        sheet['H2'].value = 'Lama Kegiatan'
        sheet['I2'].value = 'Biaya Setelah Revisi'
        sheet['J2'].value = 'Catatan Revisi'
        sheet['K2'].value = 'File Revisi'
        # ---
        # Preparing the "REVIEWER 1" header
        sheet.merge_cells(
            start_row=1, start_column=12, end_row=1, end_column=15
        )
        sheet['L1'].value = 'REVIEWER 1'
        sheet['L1'].alignment = Alignment(horizontal='center')
        # Preparing the "REVIEWER 1" sub-headers
        sheet['L2'].value = 'Nama Reviewer'
        sheet['M2'].value = 'Nilai'
        sheet['N2'].value = 'Rekomendasi Dana'
        sheet['O2'].value = 'Komentar'
        # ---
        # Preparing the "REVIEWER 2" header
        sheet.merge_cells(
            start_row=1, start_column=16, end_row=1, end_column=19
        )
        sheet['P1'].value = 'REVIEWER 2'
        sheet['P1'].alignment = Alignment(horizontal='center')
        # Preparing the "REVIEWER 2" sub-headers
        sheet['P2'].value = 'Nama Reviewer'
        sheet['Q2'].value = 'Nilai'
        sheet['R2'].value = 'Rekomendasi Dana'
        sheet['S2'].value = 'Komentar'

        # The base XPath location, pointing to each entry row
        base = '//div[@class="mw-100"]//div[@class="form-group f12"]/table[@width="100%"]//tr[@valign="top"]'

        # ---
        # Obtaining the data row values
        control.append_message_area(f'+ Mendapatkan data pada baris tabel ...')
        control.set_progress_bar(30)

        # HYPOTHESIS:
        # Xpath cannot detect 'tbody' element.
        # So instead of using 'table/tbody/tr', use 'table//tr' instead
        #
        # RESULT:
        # The hypothesis is correct.
        # Therefore, don't mention 'tbody' in any of the following Xpath paths

        a1 = [str(i)
              for i in range(1, len(content.xpath(base))+1)]

        b2 = [l.strip()
              for l in content.xpath(base + '//span[@class="hijau"]/text()')]

        c2 = [l.replace('Ketua:', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[2]/td/table//tr/td[1]/text()')]

        d2 = [l.replace('Jumlah Anggota:', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[2]/td/table//tr/td[3]/text()')]

        e2 = [l.replace('Tgl Usulan:', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[3]/td[1]/text()')]

        f2 = [l.replace('Bidang Fokus:', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[4]/td[1]/text()')]

        g2 = [l.replace('Rencana Biaya:', '').replace('Rp.', '').replace(',', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[5]/td[1]/text()')]

        h2 = [l.replace('Lama Kegiatan:', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[6]/td[1]/text()[1]')]

        i2 = [l.replace('Rp.', '').replace(',', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[10]/td/table//tr[1]/td[3]/text()')]

        j2 = [l.strip()
              for l in content.xpath(base + '/td[2]/table//tr[10]/td/table//tr[2]/td[3]/text()')]

        k2 = [l.strip()
              for l in content.xpath(base + '/td[2]/table//tr[10]/td/table//tr[3]/td[3]/a/@href')]

        # The 'Reviewer 1' all-content stripper
        rev_1 = [l.strip() for l in content.xpath(base + '/td[2]/table//tr[8]/td/table[@width="100%"]//tr[1]/td[1]/text()')]

        # Calculation for data pattern search
        # This is equal to the number of entries
        # 'rev_1_n' is always a multiple of 5
        rev_1_n = int(len(rev_1) / 5)

        l2 = []
        for i in range(rev_1_n):
            loc = 1 + (i * 5)  # --- the location of the data in the array
            l2.append(rev_1[loc].strip())

        m2 = []
        for i in range(rev_1_n):
            loc = 2 + (i * 5)  # --- the location of the data in the array
            m2.append(rev_1[loc].strip())

        n2 = []
        for i in range(rev_1_n):
            loc = 3 + (i * 5)  # --- the location of the data in the array
            n2.append(rev_1[loc].replace('Rp.', '').replace(',', '').strip())

        o2 = []
        for i in range(rev_1_n):
            loc = 4 + (i * 5)  # --- the location of the data in the array
            o2.append(rev_1[loc].strip())

        # The 'Reviewer 2' all-content stripper
        rev_2 = [l.strip() for l in content.xpath(base + '/td[2]/table//tr[8]/td/table[@width="100%"]//tr[1]/td[3]/text()')]

        # Calculation for data pattern search
        # This is equal to the number of entries
        # 'rev_2_n' is always a multiple of 5
        rev_2_n = int(len(rev_2) / 5)

        p2 = []
        for i in range(rev_2_n):
            loc = 1 + (i * 5)  # --- the location of the data in the array
            p2.append(rev_2[loc].strip())

        q2 = []
        for i in range(rev_2_n):
            loc = 2 + (i * 5)  # --- the location of the data in the array
            q2.append(rev_2[loc].strip())

        r2 = []
        for i in range(rev_2_n):
            loc = 3 + (i * 5)  # --- the location of the data in the array
            r2.append(rev_2[loc].replace('Rp.', '').replace(',', '').strip())

        s2 = []
        for i in range(rev_2_n):
            loc = 4 + (i * 5)  # --- the location of the data in the array
            s2.append(rev_2[loc].strip())

        # The starting row coordinate of the active sheet
        row_start = 3

        # DEBUG
        # Please comment out after use
        # ---
        # print(a1, b2, c2, d2, e2, f2, g2, h2, i2, j2, k2, l2, m2, n2, o2, p2, q2, r2, s2)

        # Iterating through each table row and write to the table
        # Assumes the lists a1, b2, c2, ... have the same array size
        control.append_message_area(f'+ Melakukan iterasi terhadap baris tabel dan menulis spreadsheet luaran ...')
        control.set_progress_bar(35)
        for i in range(len(a1)):

            # Noisy preamble logging
            # Please don't use this -_-
            # ---
            # control.append_message_area(f'ITERASI [{i}]')

            # Updating the progress bar status
            control.set_progress_bar(35 + round(45*(i+1)/(len(a1))))

            # Painting the scraped data to the output spreadsheet row
            sheet[f'A{row_start}'] = a1[i]
            sheet[f'B{row_start}'] = b2[i]
            sheet[f'C{row_start}'] = c2[i]
            sheet[f'D{row_start}'] = d2[i]
            sheet[f'E{row_start}'] = e2[i]
            sheet[f'F{row_start}'] = f2[i]
            sheet[f'G{row_start}'] = g2[i]
            sheet[f'H{row_start}'] = h2[i]
            sheet[f'I{row_start}'] = i2[i]
            sheet[f'J{row_start}'] = j2[i]
            sheet[f'K{row_start}'] = k2[i]
            sheet[f'L{row_start}'] = l2[i]
            sheet[f'M{row_start}'] = m2[i]
            sheet[f'N{row_start}'] = n2[i]
            sheet[f'O{row_start}'] = o2[i]
            sheet[f'P{row_start}'] = p2[i]
            sheet[f'Q{row_start}'] = q2[i]
            sheet[f'R{row_start}'] = r2[i]
            sheet[f'S{row_start}'] = s2[i]

            # Incrementing the 'row_start' iterator
            # Then continue the loop
            row_start += 1
            continue

        # Post-loop logging: successfully painted the output spreadsheet file
        control.append_message_area(f'+ Tabel sukses dipanen!')
        control.set_progress_bar(85)

        # Asking for the spreadsheet name to save as
        # ---
        # Logging and setting the progress bar
        control.append_message_area(f'+ Menyimpan spreadsheet luaran ...')
        control.set_progress_bar(90)
        # Dealing with file name prompt and saving
        # Using loop to mitigate the user clicking 'cancel'
        # in the file name dialog prompt
        while True:
            # Opening the dialog prompt
            output_spreadsheet = filedialog.asksaveasfilename(
                filetypes=[('Excel files', '*.xlsx')],
                initialfile='Sipesat - Telah Direview Penelitian Ringkasan Risat.xlsx',
                title='Simpan sebagai ...'
            )

            # 'cancel' button in the dialog prompt is clicked
            if len(output_spreadsheet) == 0:
                # Showing confirmation
                x = messagebox.askyesno(
                    'Nama File Kosong',
                    'Apakah Anda yakin ingin melanjutkan tanpa menyimpan file spreadsheet hasil pemanenan?'
                )
                # Determining whether to break or to continue the loop
                # based on the inversed value of 'x'
                if x:
                    control.append_message_area(f'+ Finalisasi pemanenan data tanpa menyimpan file spreadsheet luaran ...')
                    workbook.close()  # --- closing the workbook without saving
                    break
                else:
                    continue  # --- continuing the loop
            # File name does not end in spreadsheet extension
            elif output_spreadsheet[-5:] != '.xlsx':
                output_spreadsheet = output_spreadsheet + '.xlsx'

            # Saving the spreadsheet
            control.append_message_area(f'LOKASI_SPREADSHEET_LUARAN: {output_spreadsheet}')
            control.set_progress_bar(95)
            workbook.save(output_spreadsheet)

            # Closing the openpyxl workbook
            control.append_message_area(f'+ Menutup file spreadsheet ...')
            control.set_progress_bar(98)
            workbook.close()

            # Breaking the loop
            break

        # Notify for a successful scraping
        control.append_message_area(f'+ Pemanenan selesai pada: {str(dt.now())}')
        control.set_progress_bar(100)
        control.on_notify_successful_scraping()

    # This function harvests "Risat Disetujui Penelitian > Ringkasan Data" data
    # and then store the harvested data as an excel file
    #
    # Required arguments:
    # - control             --> for updating the progress bar and
    #                           message area of the screen SipesatScrHarvest
    # - username, password  --> the Risat administrator username and password
    def run_harvest_r_0_disetujui(self, control, username, password):
        # SipesatScrHarvest messenger
        control.set_header_desc('Panen Data "Risat Disetujui Penelitian > Ringkasan Data"')
        control.set_help_label('Data sedang dipanen. Silahkan menunggu.')
        control.set_progress_bar(0)
        control.clear_message_area()

        # Preamble logging
        control.append_message_area(f'+ Memulai pemanenan data ...')
        control.append_message_area(f'+ Pemanenan dimulai pada: {str(dt.now())}')
        control.set_progress_bar(5)

        # Preparing the 'data_prompt' arrays
        control.append_message_area(f'+ Log masuk Risat sebagai [{username}] ...')
        control.set_progress_bar(10)
        data_prompt = self.get_risat_login(username, password)
        data_prompt = self.get_risat_penelitian(data_prompt)
        data_prompt = self.get_risat_penelitian_disetujui_penelitian(data_prompt)

        # Parsing XML tree content
        control.append_message_area(f'+ Membaca halaman web ...')
        control.set_progress_bar(15)
        content = data_prompt['html_content']

        # Establishing the export spreadsheet file
        control.append_message_area(f'+ Mempersiapkan file spreadsheet luaran ...')
        control.set_progress_bar(20)
        workbook = xl.Workbook()
        sheet = workbook.active
        sheet.title = 'Disetujui Penelitian Ringkasan'

        # Preparing the sheet header
        control.append_message_area(f'+ Mempersiapkan kepala lembar spreadsheet ...')
        control.set_progress_bar(25)
        # ---
        # Preparing the "NOMOR" header
        sheet.merge_cells(
            start_row=1, start_column=1, end_row=2, end_column=1
        )
        sheet['A1'].value = 'No.'
        sheet['A1'].alignment = Alignment(horizontal='center')
        # ---
        # Preparing the "JUDUL" header
        sheet.merge_cells(
            start_row=1, start_column=2, end_row=1, end_column=11
        )
        sheet['B1'].value = 'IDENTITAS'
        sheet['B1'].alignment = Alignment(horizontal='center')
        # Preparing the "JUDUL" sub-headers
        sheet['B2'].value = 'Judul'
        sheet['C2'].value = 'Ketua'
        sheet['D2'].value = 'Jml. Anggota'
        sheet['E2'].value = 'Tgl. Usulan'
        sheet['F2'].value = 'Bidang Fokus'
        sheet['G2'].value = 'Rencana Biaya'
        sheet['H2'].value = 'Lama Kegiatan'
        sheet['I2'].value = 'Biaya Setelah Revisi'
        sheet['J2'].value = 'Catatan Revisi'
        sheet['K2'].value = 'File Revisi'
        # ---
        # Preparing the "REVIEWER 1" header
        sheet.merge_cells(
            start_row=1, start_column=12, end_row=1, end_column=15
        )
        sheet['L1'].value = 'REVIEWER 1'
        sheet['L1'].alignment = Alignment(horizontal='center')
        # Preparing the "REVIEWER 1" sub-headers
        sheet['L2'].value = 'Nama Reviewer'
        sheet['M2'].value = 'Nilai'
        sheet['N2'].value = 'Rekomendasi Dana'
        sheet['O2'].value = 'Komentar'
        # ---
        # Preparing the "REVIEWER 2" header
        sheet.merge_cells(
            start_row=1, start_column=16, end_row=1, end_column=19
        )
        sheet['P1'].value = 'REVIEWER 2'
        sheet['P1'].alignment = Alignment(horizontal='center')
        # Preparing the "REVIEWER 2" sub-headers
        sheet['P2'].value = 'Nama Reviewer'
        sheet['Q2'].value = 'Nilai'
        sheet['R2'].value = 'Rekomendasi Dana'
        sheet['S2'].value = 'Komentar'

        # The base XPath location, pointing to each entry row
        base = '//div[@class="mw-100"]//div[@class="form-group f12"]/table[@width="100%"]//tr[@valign="top"]'

        # ---
        # Obtaining the data row values
        control.append_message_area(f'+ Mendapatkan data pada baris tabel ...')
        control.set_progress_bar(30)

        # HYPOTHESIS:
        # Xpath cannot detect 'tbody' element.
        # So instead of using 'table/tbody/tr', use 'table//tr' instead
        #
        # RESULT:
        # The hypothesis is correct.
        # Therefore, don't mention 'tbody' in any of the following Xpath paths

        a1 = [str(i)
              for i in range(1, len(content.xpath(base))+1)]

        b2 = [l.strip()
              for l in content.xpath(base + '//span[@class="hijau"]/text()')]

        c2 = [l.replace('Ketua:', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[2]/td/table//tr/td[1]/text()')]

        d2 = [l.replace('Jumlah Anggota:', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[2]/td/table//tr/td[3]/text()')]

        e2 = [l.replace('Tgl Usulan:', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[3]/td[1]/text()')]

        f2 = [l.replace('Bidang Fokus:', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[4]/td[1]/text()')]

        g2 = [l.replace('Rencana Biaya:', '').replace('Rp.', '').replace(',', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[5]/td[1]/text()')]

        h2 = [l.replace('Lama Kegiatan:', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[6]/td[1]/text()[1]')]

        i2 = [l.replace('Rp.', '').replace(',', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[10]/td/table//tr[1]/td[3]/text()')]

        j2 = [l.strip()
              for l in content.xpath(base + '/td[2]/table//tr[10]/td/table//tr[2]/td[3]/text()')]

        k2 = [l.strip()
              for l in content.xpath(base + '/td[2]/table//tr[10]/td/table//tr[3]/td[3]/a/@href')]

        # The 'Reviewer 1' all-content stripper
        rev_1 = [l.strip() for l in content.xpath(base + '/td[2]/table//tr[8]/td/table[@width="100%"]//tr[1]/td[1]/text()')]

        # Calculation for data pattern search
        # This is equal to the number of entries
        # 'rev_1_n' is always a multiple of 5
        rev_1_n = int(len(rev_1) / 5)

        l2 = []
        for i in range(rev_1_n):
            loc = 1 + (i * 5)  # --- the location of the data in the array
            l2.append(rev_1[loc].strip())

        m2 = []
        for i in range(rev_1_n):
            loc = 2 + (i * 5)  # --- the location of the data in the array
            m2.append(rev_1[loc].strip())

        n2 = []
        for i in range(rev_1_n):
            loc = 3 + (i * 5)  # --- the location of the data in the array
            n2.append(rev_1[loc].replace('Rp.', '').replace(',', '').strip())

        o2 = []
        for i in range(rev_1_n):
            loc = 4 + (i * 5)  # --- the location of the data in the array
            o2.append(rev_1[loc].strip())

        # The 'Reviewer 2' all-content stripper
        rev_2 = [l.strip() for l in content.xpath(base + '/td[2]/table//tr[8]/td/table[@width="100%"]//tr[1]/td[3]/text()')]

        # Calculation for data pattern search
        # This is equal to the number of entries
        # 'rev_2_n' is always a multiple of 5
        rev_2_n = int(len(rev_2) / 5)

        p2 = []
        for i in range(rev_2_n):
            loc = 1 + (i * 5)  # --- the location of the data in the array
            p2.append(rev_2[loc].strip())

        q2 = []
        for i in range(rev_2_n):
            loc = 2 + (i * 5)  # --- the location of the data in the array
            q2.append(rev_2[loc].strip())

        r2 = []
        for i in range(rev_2_n):
            loc = 3 + (i * 5)  # --- the location of the data in the array
            r2.append(rev_2[loc].replace('Rp.', '').replace(',', '').strip())

        s2 = []
        for i in range(rev_2_n):
            loc = 4 + (i * 5)  # --- the location of the data in the array
            s2.append(rev_2[loc].strip())

        # The starting row coordinate of the active sheet
        row_start = 3

        # DEBUG
        # Please comment out after use
        # ---
        # print(a1, b2, c2, d2, e2, f2, g2, h2, i2, j2, k2, l2, m2, n2, o2, p2, q2, r2, s2)

        # Iterating through each table row and write to the table
        # Assumes the lists a1, b2, c2, ... have the same array size
        control.append_message_area(f'+ Melakukan iterasi terhadap baris tabel dan menulis spreadsheet luaran ...')
        control.set_progress_bar(35)
        for i in range(len(a1)):

            # Noisy preamble logging
            # Please don't use this -_-
            # ---
            # control.append_message_area(f'ITERASI [{i}]')

            # Updating the progress bar status
            control.set_progress_bar(35 + round(45*(i+1)/(len(a1))))

            # Painting the scraped data to the output spreadsheet row
            sheet[f'A{row_start}'] = a1[i]
            sheet[f'B{row_start}'] = b2[i]
            sheet[f'C{row_start}'] = c2[i]
            sheet[f'D{row_start}'] = d2[i]
            sheet[f'E{row_start}'] = e2[i]
            sheet[f'F{row_start}'] = f2[i]
            sheet[f'G{row_start}'] = g2[i]
            sheet[f'H{row_start}'] = h2[i]
            sheet[f'I{row_start}'] = i2[i]
            sheet[f'J{row_start}'] = j2[i]
            sheet[f'K{row_start}'] = k2[i]
            sheet[f'L{row_start}'] = l2[i]
            sheet[f'M{row_start}'] = m2[i]
            sheet[f'N{row_start}'] = n2[i]
            sheet[f'O{row_start}'] = o2[i]
            sheet[f'P{row_start}'] = p2[i]
            sheet[f'Q{row_start}'] = q2[i]
            sheet[f'R{row_start}'] = r2[i]
            sheet[f'S{row_start}'] = s2[i]

            # Incrementing the 'row_start' iterator
            # Then continue the loop
            row_start += 1
            continue

        # Post-loop logging: successfully painted the output spreadsheet file
        control.append_message_area(f'+ Tabel sukses dipanen!')
        control.set_progress_bar(85)

        # Asking for the spreadsheet name to save as
        # ---
        # Logging and setting the progress bar
        control.append_message_area(f'+ Menyimpan spreadsheet luaran ...')
        control.set_progress_bar(90)
        # Dealing with file name prompt and saving
        # Using loop to mitigate the user clicking 'cancel'
        # in the file name dialog prompt
        while True:
            # Opening the dialog prompt
            output_spreadsheet = filedialog.asksaveasfilename(
                filetypes=[('Excel files', '*.xlsx')],
                initialfile='Sipesat - Disetujui Penelitian Ringkasan Risat.xlsx',
                title='Simpan sebagai ...'
            )

            # 'cancel' button in the dialog prompt is clicked
            if len(output_spreadsheet) == 0:
                # Showing confirmation
                x = messagebox.askyesno(
                    'Nama File Kosong',
                    'Apakah Anda yakin ingin melanjutkan tanpa menyimpan file spreadsheet hasil pemanenan?'
                )
                # Determining whether to break or to continue the loop
                # based on the inversed value of 'x'
                if x:
                    control.append_message_area(f'+ Finalisasi pemanenan data tanpa menyimpan file spreadsheet luaran ...')
                    workbook.close()  # --- closing the workbook without saving
                    break
                else:
                    continue  # --- continuing the loop
            # File name does not end in spreadsheet extension
            elif output_spreadsheet[-5:] != '.xlsx':
                output_spreadsheet = output_spreadsheet + '.xlsx'

            # Saving the spreadsheet
            control.append_message_area(f'LOKASI_SPREADSHEET_LUARAN: {output_spreadsheet}')
            control.set_progress_bar(95)
            workbook.save(output_spreadsheet)

            # Closing the openpyxl workbook
            control.append_message_area(f'+ Menutup file spreadsheet ...')
            control.set_progress_bar(98)
            workbook.close()

            # Breaking the loop
            break

        # Notify for a successful scraping
        control.append_message_area(f'+ Pemanenan selesai pada: {str(dt.now())}')
        control.set_progress_bar(100)
        control.on_notify_successful_scraping()

    # This function harvests "Risat Disetujui Dng Revisi (DDR) Penelitian > Ringkasan Data" data
    # and then store the harvested data as an excel file
    #
    # Required arguments:
    # - control             --> for updating the progress bar and
    #                           message area of the screen SipesatScrHarvest
    # - username, password  --> the Risat administrator username and password
    def run_harvest_r_0_ddr(self, control, username, password):
        # SipesatScrHarvest messenger
        control.set_header_desc('Panen Data "Risat Disetujui Dgn Revisi (DDR) Penelitian > Ringkasan Data"')
        control.set_help_label('Data sedang dipanen. Silahkan menunggu.')
        control.set_progress_bar(0)
        control.clear_message_area()

        # Preamble logging
        control.append_message_area(f'+ Memulai pemanenan data ...')
        control.append_message_area(f'+ Pemanenan dimulai pada: {str(dt.now())}')
        control.set_progress_bar(5)

        # Preparing the 'data_prompt' arrays
        control.append_message_area(f'+ Log masuk Risat sebagai [{username}] ...')
        control.set_progress_bar(10)
        data_prompt = self.get_risat_login(username, password)
        data_prompt = self.get_risat_penelitian(data_prompt)
        data_prompt = self.get_risat_penelitian_ddr_penelitian(data_prompt)

        # Parsing XML tree content
        control.append_message_area(f'+ Membaca halaman web ...')
        control.set_progress_bar(15)
        content = data_prompt['html_content']

        # Establishing the export spreadsheet file
        control.append_message_area(f'+ Mempersiapkan file spreadsheet luaran ...')
        control.set_progress_bar(20)
        workbook = xl.Workbook()
        sheet = workbook.active
        sheet.title = 'DDR Penelitian Ringkasan'

        # Preparing the sheet header
        control.append_message_area(f'+ Mempersiapkan kepala lembar spreadsheet ...')
        control.set_progress_bar(25)
        # ---
        # Preparing the "NOMOR" header
        sheet.merge_cells(
            start_row=1, start_column=1, end_row=2, end_column=1
        )
        sheet['A1'].value = 'No.'
        sheet['A1'].alignment = Alignment(horizontal='center')
        # ---
        # Preparing the "JUDUL" header
        sheet.merge_cells(
            start_row=1, start_column=2, end_row=1, end_column=11
        )
        sheet['B1'].value = 'IDENTITAS'
        sheet['B1'].alignment = Alignment(horizontal='center')
        # Preparing the "JUDUL" sub-headers
        sheet['B2'].value = 'Judul'
        sheet['C2'].value = 'Ketua'
        sheet['D2'].value = 'Jml. Anggota'
        sheet['E2'].value = 'Tgl. Usulan'
        sheet['F2'].value = 'Bidang Fokus'
        sheet['G2'].value = 'Rencana Biaya'
        sheet['H2'].value = 'Lama Kegiatan'
        sheet['I2'].value = 'Biaya Setelah Revisi'
        sheet['J2'].value = 'Catatan Revisi'
        sheet['K2'].value = 'File Revisi'
        # ---
        # Preparing the "REVIEWER 1" header
        sheet.merge_cells(
            start_row=1, start_column=12, end_row=1, end_column=15
        )
        sheet['L1'].value = 'REVIEWER 1'
        sheet['L1'].alignment = Alignment(horizontal='center')
        # Preparing the "REVIEWER 1" sub-headers
        sheet['L2'].value = 'Nama Reviewer'
        sheet['M2'].value = 'Nilai'
        sheet['N2'].value = 'Rekomendasi Dana'
        sheet['O2'].value = 'Komentar'
        # ---
        # Preparing the "REVIEWER 2" header
        sheet.merge_cells(
            start_row=1, start_column=16, end_row=1, end_column=19
        )
        sheet['P1'].value = 'REVIEWER 2'
        sheet['P1'].alignment = Alignment(horizontal='center')
        # Preparing the "REVIEWER 2" sub-headers
        sheet['P2'].value = 'Nama Reviewer'
        sheet['Q2'].value = 'Nilai'
        sheet['R2'].value = 'Rekomendasi Dana'
        sheet['S2'].value = 'Komentar'

        # The base XPath location, pointing to each entry row
        base = '//div[@class="mw-100"]//div[@class="form-group f12"]/table[@width="100%"]//tr[@valign="top"]'

        # ---
        # Obtaining the data row values
        control.append_message_area(f'+ Mendapatkan data pada baris tabel ...')
        control.set_progress_bar(30)

        # HYPOTHESIS:
        # Xpath cannot detect 'tbody' element.
        # So instead of using 'table/tbody/tr', use 'table//tr' instead
        #
        # RESULT:
        # The hypothesis is correct.
        # Therefore, don't mention 'tbody' in any of the following Xpath paths

        a1 = [str(i)
              for i in range(1, len(content.xpath(base))+1)]

        b2 = [l.strip()
              for l in content.xpath(base + '//span[@class="hijau"]/text()')]

        c2 = [l.replace('Ketua:', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[2]/td/table//tr/td[1]/text()')]

        d2 = [l.replace('Jumlah Anggota:', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[2]/td/table//tr/td[3]/text()')]

        e2 = [l.replace('Tgl Usulan:', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[3]/td[1]/text()')]

        f2 = [l.replace('Bidang Fokus:', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[4]/td[1]/text()')]

        g2 = [l.replace('Rencana Biaya:', '').replace('Rp.', '').replace(',', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[5]/td[1]/text()')]

        h2 = [l.replace('Lama Kegiatan:', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[6]/td[1]/text()[1]')]

        i2 = [l.replace('Rp.', '').replace(',', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[10]/td/table//tr[1]/td[3]/text()')]

        j2 = [l.strip()
              for l in content.xpath(base + '/td[2]/table//tr[10]/td/table//tr[2]/td[3]/text()')]

        k2 = [l.strip()
              for l in content.xpath(base + '/td[2]/table//tr[10]/td/table//tr[3]/td[3]/a/@href')]

        # The 'Reviewer 1' all-content stripper
        rev_1 = [l.strip() for l in content.xpath(base + '/td[2]/table//tr[8]/td/table[@width="100%"]//tr[1]/td[1]/text()')]

        # Calculation for data pattern search
        # This is equal to the number of entries
        # 'rev_1_n' is always a multiple of 5
        rev_1_n = int(len(rev_1) / 5)

        l2 = []
        for i in range(rev_1_n):
            loc = 1 + (i * 5)  # --- the location of the data in the array
            l2.append(rev_1[loc].strip())

        m2 = []
        for i in range(rev_1_n):
            loc = 2 + (i * 5)  # --- the location of the data in the array
            m2.append(rev_1[loc].strip())

        n2 = []
        for i in range(rev_1_n):
            loc = 3 + (i * 5)  # --- the location of the data in the array
            n2.append(rev_1[loc].replace('Rp.', '').replace(',', '').strip())

        o2 = []
        for i in range(rev_1_n):
            loc = 4 + (i * 5)  # --- the location of the data in the array
            o2.append(rev_1[loc].strip())

        # The 'Reviewer 2' all-content stripper
        rev_2 = [l.strip() for l in content.xpath(base + '/td[2]/table//tr[8]/td/table[@width="100%"]//tr[1]/td[3]/text()')]

        # Calculation for data pattern search
        # This is equal to the number of entries
        # 'rev_2_n' is always a multiple of 5
        rev_2_n = int(len(rev_2) / 5)

        p2 = []
        for i in range(rev_2_n):
            loc = 1 + (i * 5)  # --- the location of the data in the array
            p2.append(rev_2[loc].strip())

        q2 = []
        for i in range(rev_2_n):
            loc = 2 + (i * 5)  # --- the location of the data in the array
            q2.append(rev_2[loc].strip())

        r2 = []
        for i in range(rev_2_n):
            loc = 3 + (i * 5)  # --- the location of the data in the array
            r2.append(rev_2[loc].replace('Rp.', '').replace(',', '').strip())

        s2 = []
        for i in range(rev_2_n):
            loc = 4 + (i * 5)  # --- the location of the data in the array
            s2.append(rev_2[loc].strip())

        # The starting row coordinate of the active sheet
        row_start = 3

        # DEBUG
        # Please comment out after use
        # ---
        # print(a1, b2, c2, d2, e2, f2, g2, h2, i2, j2, k2, l2, m2, n2, o2, p2, q2, r2, s2)

        # Iterating through each table row and write to the table
        # Assumes the lists a1, b2, c2, ... have the same array size
        control.append_message_area(f'+ Melakukan iterasi terhadap baris tabel dan menulis spreadsheet luaran ...')
        control.set_progress_bar(35)
        for i in range(len(a1)):

            # Noisy preamble logging
            # Please don't use this -_-
            # ---
            # control.append_message_area(f'ITERASI [{i}]')

            # Updating the progress bar status
            control.set_progress_bar(35 + round(45*(i+1)/(len(a1))))

            # Painting the scraped data to the output spreadsheet row
            sheet[f'A{row_start}'] = a1[i]
            sheet[f'B{row_start}'] = b2[i]
            sheet[f'C{row_start}'] = c2[i]
            sheet[f'D{row_start}'] = d2[i]
            sheet[f'E{row_start}'] = e2[i]
            sheet[f'F{row_start}'] = f2[i]
            sheet[f'G{row_start}'] = g2[i]
            sheet[f'H{row_start}'] = h2[i]
            sheet[f'I{row_start}'] = i2[i]
            sheet[f'J{row_start}'] = j2[i]
            sheet[f'K{row_start}'] = k2[i]
            sheet[f'L{row_start}'] = l2[i]
            sheet[f'M{row_start}'] = m2[i]
            sheet[f'N{row_start}'] = n2[i]
            sheet[f'O{row_start}'] = o2[i]
            sheet[f'P{row_start}'] = p2[i]
            sheet[f'Q{row_start}'] = q2[i]
            sheet[f'R{row_start}'] = r2[i]
            sheet[f'S{row_start}'] = s2[i]

            # Incrementing the 'row_start' iterator
            # Then continue the loop
            row_start += 1
            continue

        # Post-loop logging: successfully painted the output spreadsheet file
        control.append_message_area(f'+ Tabel sukses dipanen!')
        control.set_progress_bar(85)

        # Asking for the spreadsheet name to save as
        # ---
        # Logging and setting the progress bar
        control.append_message_area(f'+ Menyimpan spreadsheet luaran ...')
        control.set_progress_bar(90)
        # Dealing with file name prompt and saving
        # Using loop to mitigate the user clicking 'cancel'
        # in the file name dialog prompt
        while True:
            # Opening the dialog prompt
            output_spreadsheet = filedialog.asksaveasfilename(
                filetypes=[('Excel files', '*.xlsx')],
                initialfile='Sipesat - Disetujui Dgn Revisi (DDR) Penelitian Ringkasan Risat.xlsx',
                title='Simpan sebagai ...'
            )

            # 'cancel' button in the dialog prompt is clicked
            if len(output_spreadsheet) == 0:
                # Showing confirmation
                x = messagebox.askyesno(
                    'Nama File Kosong',
                    'Apakah Anda yakin ingin melanjutkan tanpa menyimpan file spreadsheet hasil pemanenan?'
                )
                # Determining whether to break or to continue the loop
                # based on the inversed value of 'x'
                if x:
                    control.append_message_area(f'+ Finalisasi pemanenan data tanpa menyimpan file spreadsheet luaran ...')
                    workbook.close()  # --- closing the workbook without saving
                    break
                else:
                    continue  # --- continuing the loop
            # File name does not end in spreadsheet extension
            elif output_spreadsheet[-5:] != '.xlsx':
                output_spreadsheet = output_spreadsheet + '.xlsx'

            # Saving the spreadsheet
            control.append_message_area(f'LOKASI_SPREADSHEET_LUARAN: {output_spreadsheet}')
            control.set_progress_bar(95)
            workbook.save(output_spreadsheet)

            # Closing the openpyxl workbook
            control.append_message_area(f'+ Menutup file spreadsheet ...')
            control.set_progress_bar(98)
            workbook.close()

            # Breaking the loop
            break

        # Notify for a successful scraping
        control.append_message_area(f'+ Pemanenan selesai pada: {str(dt.now())}')
        control.set_progress_bar(100)
        control.on_notify_successful_scraping()

    # This function harvests "Risat Ditolak Penelitian > Ringkasan Data" data
    # and then store the harvested data as an excel file
    #
    # Required arguments:
    # - control             --> for updating the progress bar and
    #                           message area of the screen SipesatScrHarvest
    # - username, password  --> the Risat administrator username and password
    def run_harvest_r_0_ditolak(self, control, username, password):
        # SipesatScrHarvest messenger
        control.set_header_desc('Panen Data "Risat Ditolak Penelitian > Ringkasan Data"')
        control.set_help_label('Data sedang dipanen. Silahkan menunggu.')
        control.set_progress_bar(0)
        control.clear_message_area()

        # Preamble logging
        control.append_message_area(f'+ Memulai pemanenan data ...')
        control.append_message_area(f'+ Pemanenan dimulai pada: {str(dt.now())}')
        control.set_progress_bar(5)

        # Preparing the 'data_prompt' arrays
        control.append_message_area(f'+ Log masuk Risat sebagai [{username}] ...')
        control.set_progress_bar(10)
        data_prompt = self.get_risat_login(username, password)
        data_prompt = self.get_risat_penelitian(data_prompt)
        data_prompt = self.get_risat_penelitian_ditolak_penelitian(data_prompt)

        # Parsing XML tree content
        control.append_message_area(f'+ Membaca halaman web ...')
        control.set_progress_bar(15)
        content = data_prompt['html_content']

        # Establishing the export spreadsheet file
        control.append_message_area(f'+ Mempersiapkan file spreadsheet luaran ...')
        control.set_progress_bar(20)
        workbook = xl.Workbook()
        sheet = workbook.active
        sheet.title = 'Ditolak Penelitian Ringkasan'

        # Preparing the sheet header
        control.append_message_area(f'+ Mempersiapkan kepala lembar spreadsheet ...')
        control.set_progress_bar(25)
        # ---
        # Preparing the "NOMOR" header
        sheet.merge_cells(
            start_row=1, start_column=1, end_row=2, end_column=1
        )
        sheet['A1'].value = 'No.'
        sheet['A1'].alignment = Alignment(horizontal='center')
        # ---
        # Preparing the "JUDUL" header
        sheet.merge_cells(
            start_row=1, start_column=2, end_row=1, end_column=11
        )
        sheet['B1'].value = 'IDENTITAS'
        sheet['B1'].alignment = Alignment(horizontal='center')
        # Preparing the "JUDUL" sub-headers
        sheet['B2'].value = 'Judul'
        sheet['C2'].value = 'Ketua'
        sheet['D2'].value = 'Jml. Anggota'
        sheet['E2'].value = 'Tgl. Usulan'
        sheet['F2'].value = 'Bidang Fokus'
        sheet['G2'].value = 'Rencana Biaya'
        sheet['H2'].value = 'Lama Kegiatan'
        sheet['I2'].value = 'Biaya Setelah Revisi'
        sheet['J2'].value = 'Catatan Revisi'
        sheet['K2'].value = 'File Revisi'
        # ---
        # Preparing the "REVIEWER 1" header
        sheet.merge_cells(
            start_row=1, start_column=12, end_row=1, end_column=15
        )
        sheet['L1'].value = 'REVIEWER 1'
        sheet['L1'].alignment = Alignment(horizontal='center')
        # Preparing the "REVIEWER 1" sub-headers
        sheet['L2'].value = 'Nama Reviewer'
        sheet['M2'].value = 'Nilai'
        sheet['N2'].value = 'Rekomendasi Dana'
        sheet['O2'].value = 'Komentar'
        # ---
        # Preparing the "REVIEWER 2" header
        sheet.merge_cells(
            start_row=1, start_column=16, end_row=1, end_column=19
        )
        sheet['P1'].value = 'REVIEWER 2'
        sheet['P1'].alignment = Alignment(horizontal='center')
        # Preparing the "REVIEWER 2" sub-headers
        sheet['P2'].value = 'Nama Reviewer'
        sheet['Q2'].value = 'Nilai'
        sheet['R2'].value = 'Rekomendasi Dana'
        sheet['S2'].value = 'Komentar'

        # The base XPath location, pointing to each entry row
        base = '//div[@class="mw-100"]//div[@class="form-group f12"]/table[@width="100%"]//tr[@valign="top"]'

        # ---
        # Obtaining the data row values
        control.append_message_area(f'+ Mendapatkan data pada baris tabel ...')
        control.set_progress_bar(30)

        # HYPOTHESIS:
        # Xpath cannot detect 'tbody' element.
        # So instead of using 'table/tbody/tr', use 'table//tr' instead
        #
        # RESULT:
        # The hypothesis is correct.
        # Therefore, don't mention 'tbody' in any of the following Xpath paths

        a1 = [str(i)
              for i in range(1, len(content.xpath(base))+1)]

        b2 = [l.strip()
              for l in content.xpath(base + '//span[@class="hijau"]/text()')]

        c2 = [l.replace('Ketua:', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[2]/td/table//tr/td[1]/text()')]

        d2 = [l.replace('Jumlah Anggota:', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[2]/td/table//tr/td[3]/text()')]

        e2 = [l.replace('Tgl Usulan:', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[3]/td[1]/text()')]

        f2 = [l.replace('Bidang Fokus:', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[4]/td[1]/text()')]

        g2 = [l.replace('Rencana Biaya:', '').replace('Rp.', '').replace(',', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[5]/td[1]/text()')]

        h2 = [l.replace('Lama Kegiatan:', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[6]/td[1]/text()[1]')]

        i2 = [l.replace('Rp.', '').replace(',', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[10]/td/table//tr[1]/td[3]/text()')]

        j2 = [l.strip()
              for l in content.xpath(base + '/td[2]/table//tr[10]/td/table//tr[2]/td[3]/text()')]

        k2 = [l.strip()
              for l in content.xpath(base + '/td[2]/table//tr[10]/td/table//tr[3]/td[3]/a/@href')]

        # The 'Reviewer 1' all-content stripper
        rev_1 = [l.strip() for l in content.xpath(base + '/td[2]/table//tr[8]/td/table[@width="100%"]//tr[1]/td[1]/text()')]

        # Calculation for data pattern search
        # This is equal to the number of entries
        # 'rev_1_n' is always a multiple of 5
        rev_1_n = int(len(rev_1) / 5)

        l2 = []
        for i in range(rev_1_n):
            loc = 1 + (i * 5)  # --- the location of the data in the array
            l2.append(rev_1[loc].strip())

        m2 = []
        for i in range(rev_1_n):
            loc = 2 + (i * 5)  # --- the location of the data in the array
            m2.append(rev_1[loc].strip())

        n2 = []
        for i in range(rev_1_n):
            loc = 3 + (i * 5)  # --- the location of the data in the array
            n2.append(rev_1[loc].replace('Rp.', '').replace(',', '').strip())

        o2 = []
        for i in range(rev_1_n):
            loc = 4 + (i * 5)  # --- the location of the data in the array
            o2.append(rev_1[loc].strip())

        # The 'Reviewer 2' all-content stripper
        rev_2 = [l.strip() for l in content.xpath(base + '/td[2]/table//tr[8]/td/table[@width="100%"]//tr[1]/td[3]/text()')]

        # Calculation for data pattern search
        # This is equal to the number of entries
        # 'rev_2_n' is always a multiple of 5
        rev_2_n = int(len(rev_2) / 5)

        p2 = []
        for i in range(rev_2_n):
            loc = 1 + (i * 5)  # --- the location of the data in the array
            p2.append(rev_2[loc].strip())

        q2 = []
        for i in range(rev_2_n):
            loc = 2 + (i * 5)  # --- the location of the data in the array
            q2.append(rev_2[loc].strip())

        r2 = []
        for i in range(rev_2_n):
            loc = 3 + (i * 5)  # --- the location of the data in the array
            r2.append(rev_2[loc].replace('Rp.', '').replace(',', '').strip())

        s2 = []
        for i in range(rev_2_n):
            loc = 4 + (i * 5)  # --- the location of the data in the array
            s2.append(rev_2[loc].strip())

        # The starting row coordinate of the active sheet
        row_start = 3

        # DEBUG
        # Please comment out after use
        # ---
        # print(a1, b2, c2, d2, e2, f2, g2, h2, i2, j2, k2, l2, m2, n2, o2, p2, q2, r2, s2)

        # Iterating through each table row and write to the table
        # Assumes the lists a1, b2, c2, ... have the same array size
        control.append_message_area(f'+ Melakukan iterasi terhadap baris tabel dan menulis spreadsheet luaran ...')
        control.set_progress_bar(35)
        for i in range(len(a1)):

            # Noisy preamble logging
            # Please don't use this -_-
            # ---
            # control.append_message_area(f'ITERASI [{i}]')

            # Updating the progress bar status
            control.set_progress_bar(35 + round(45*(i+1)/(len(a1))))

            # Painting the scraped data to the output spreadsheet row
            sheet[f'A{row_start}'] = a1[i]
            sheet[f'B{row_start}'] = b2[i]
            sheet[f'C{row_start}'] = c2[i]
            sheet[f'D{row_start}'] = d2[i]
            sheet[f'E{row_start}'] = e2[i]
            sheet[f'F{row_start}'] = f2[i]
            sheet[f'G{row_start}'] = g2[i]
            sheet[f'H{row_start}'] = h2[i]
            sheet[f'I{row_start}'] = i2[i]
            sheet[f'J{row_start}'] = j2[i]
            sheet[f'K{row_start}'] = k2[i]
            sheet[f'L{row_start}'] = l2[i]
            sheet[f'M{row_start}'] = m2[i]
            sheet[f'N{row_start}'] = n2[i]
            sheet[f'O{row_start}'] = o2[i]
            sheet[f'P{row_start}'] = p2[i]
            sheet[f'Q{row_start}'] = q2[i]
            sheet[f'R{row_start}'] = r2[i]
            sheet[f'S{row_start}'] = s2[i]

            # Incrementing the 'row_start' iterator
            # Then continue the loop
            row_start += 1
            continue

        # Post-loop logging: successfully painted the output spreadsheet file
        control.append_message_area(f'+ Tabel sukses dipanen!')
        control.set_progress_bar(85)

        # Asking for the spreadsheet name to save as
        # ---
        # Logging and setting the progress bar
        control.append_message_area(f'+ Menyimpan spreadsheet luaran ...')
        control.set_progress_bar(90)
        # Dealing with file name prompt and saving
        # Using loop to mitigate the user clicking 'cancel'
        # in the file name dialog prompt
        while True:
            # Opening the dialog prompt
            output_spreadsheet = filedialog.asksaveasfilename(
                filetypes=[('Excel files', '*.xlsx')],
                initialfile='Sipesat - Ditolak Penelitian Ringkasan Risat.xlsx',
                title='Simpan sebagai ...'
            )

            # 'cancel' button in the dialog prompt is clicked
            if len(output_spreadsheet) == 0:
                # Showing confirmation
                x = messagebox.askyesno(
                    'Nama File Kosong',
                    'Apakah Anda yakin ingin melanjutkan tanpa menyimpan file spreadsheet hasil pemanenan?'
                )
                # Determining whether to break or to continue the loop
                # based on the inversed value of 'x'
                if x:
                    control.append_message_area(f'+ Finalisasi pemanenan data tanpa menyimpan file spreadsheet luaran ...')
                    workbook.close()  # --- closing the workbook without saving
                    break
                else:
                    continue  # --- continuing the loop
            # File name does not end in spreadsheet extension
            elif output_spreadsheet[-5:] != '.xlsx':
                output_spreadsheet = output_spreadsheet + '.xlsx'

            # Saving the spreadsheet
            control.append_message_area(f'LOKASI_SPREADSHEET_LUARAN: {output_spreadsheet}')
            control.set_progress_bar(95)
            workbook.save(output_spreadsheet)

            # Closing the openpyxl workbook
            control.append_message_area(f'+ Menutup file spreadsheet ...')
            control.set_progress_bar(98)
            workbook.close()

            # Breaking the loop
            break

        # Notify for a successful scraping
        control.append_message_area(f'+ Pemanenan selesai pada: {str(dt.now())}')
        control.set_progress_bar(100)
        control.on_notify_successful_scraping()

    # This function harvests "Risat Dana Penelitian > Ringkasan Data" data
    # and then store the harvested data as an excel file
    #
    # Required arguments:
    # - control             --> for updating the progress bar and
    #                           message area of the screen SipesatScrHarvest
    # - username, password  --> the Risat administrator username and password
    def run_harvest_r_0_dana(self, control, username, password):

        # SipesatScrHarvest messenger
        control.set_header_desc('Panen Data "Risat Dana Penelitian > Ringkasan Data"')
        control.set_help_label('Data sedang dipanen. Silahkan menunggu.')
        control.set_progress_bar(0)
        control.clear_message_area()

        # Preamble logging
        control.append_message_area(f'+ Memulai pemanenan data ...')
        control.append_message_area(f'+ Pemanenan dimulai pada: {str(dt.now())}')
        control.set_progress_bar(5)

        # Preparing the 'data_prompt' arrays
        control.append_message_area(f'+ Log masuk Risat sebagai [{username}] ...')
        control.set_progress_bar(10)
        data_prompt = self.get_risat_login(username, password)
        data_prompt = self.get_risat_penelitian(data_prompt)
        data_prompt = self.get_risat_penelitian_dana_penelitian(data_prompt)

        # Parsing XML tree content
        control.append_message_area(f'+ Membaca halaman web ...')
        control.set_progress_bar(15)
        content = data_prompt['html_content']

        # Establishing the export spreadsheet file
        control.append_message_area(f'+ Mempersiapkan file spreadsheet luaran ...')
        control.set_progress_bar(20)
        workbook = xl.Workbook()
        sheet = workbook.active
        sheet.title = 'Dana Penelitian Ringkasan'

        # Preparing the sheet header
        control.append_message_area(f'+ Mempersiapkan kepala lembar spreadsheet ...')
        control.set_progress_bar(25)
        sheet['A1'].value = 'No.'
        sheet['B1'].value = 'Judul'
        sheet['C1'].value = 'Ketua'
        sheet['D1'].value = 'Jml. Anggota'
        sheet['E1'].value = 'Tgl. Usulan'
        sheet['F1'].value = 'Bidang Fokus'
        sheet['G1'].value = 'Rencana Biaya'
        sheet['H1'].value = 'Lama Kegiatan'
        sheet['I1'].value = 'Dana Disetujui'
        sheet['J1'].value = 'Tgl. Persetujuan'
        sheet['K1'].value = 'Dana Ditransfer'
        sheet['L1'].value = 'Tgl. Ditransfer'

        # The base XPath location, pointing to each entry row
        base = '//div[@class="mw-100"]//div[@class="form-group f12"]/table[@width="100%"]//tr[@valign="top"]'

        # ---
        # Obtaining the data row values
        control.append_message_area(f'+ Mendapatkan data pada baris tabel ...')
        control.set_progress_bar(30)

        # HYPOTHESIS:
        # Xpath cannot detect 'tbody' element.
        # So instead of using 'table/tbody/tr', use 'table//tr' instead
        #
        # RESULT:
        # The hypothesis is correct.
        # Therefore, don't mention 'tbody' in any of the following Xpath paths

        a1 = [str(i)
              for i in range(1, len(content.xpath(base))+1)]

        b1 = [l.strip()
              for l in content.xpath(base + '//span[@class="hijau"]/text()')]

        c1 = [l.replace('Ketua:', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[2]/td/table//tr/td[1]/text()')]

        d1 = [l.replace('Jumlah Anggota:', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[2]/td/table//tr/td[3]/text()')]

        e1 = [l.replace('Tgl Usulan:', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[3]/td[1]/text()')]

        f1 = [l.replace('Bidang Fokus:', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[4]/td[1]/text()')]

        g1 = [l.replace('Rencana Biaya:', '').replace('Rp.', '').replace(',', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[5]/td[1]/text()')]

        h1 = [l.replace('Lama Kegiatan:', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[6]/td[1]/text()[1]')]

        i1 = [l.replace('Rp.', '').replace(',', '').strip()
              for l in content.xpath(base + '/td[3]/text()')]

        j1 = [l.strip()
              for l in content.xpath(base + '/td[4]/text()')]

        k1 = [l.replace('Rp.', '').replace(',', '').strip()
              for l in content.xpath(base + '/td[5]/text()')]

        l1 = [l.replace('Terealisasi', '').strip()
              for l in content.xpath(base + '/td[6]/text()')]

        # The starting row coordinate of the active sheet
        row_start = 2

        # DEBUG
        # Please comment out after use
        # ---
        # print(a1, b1, c1, d1, e1, f1, g1, h1, i1, j1, k1, l1)

        # Iterating through each table row and write to the table
        # Assumes the lists a1, b1, c1, ... have the same array size
        control.append_message_area(f'+ Melakukan iterasi terhadap baris tabel dan menulis spreadsheet luaran ...')
        control.set_progress_bar(35)
        for i in range(len(a1)):

            # Noisy preamble logging
            # Please don't use this -_-
            # ---
            # control.append_message_area(f'ITERASI [{i}]')

            # Updating the progress bar status
            control.set_progress_bar(35 + round(45*(i+1)/(len(a1))))

            # Painting the scraped data to the output spreadsheet row
            sheet[f'A{row_start}'] = a1[i]
            sheet[f'B{row_start}'] = b1[i]
            sheet[f'C{row_start}'] = c1[i]
            sheet[f'D{row_start}'] = d1[i]
            sheet[f'E{row_start}'] = e1[i]
            sheet[f'F{row_start}'] = f1[i]
            sheet[f'G{row_start}'] = g1[i]
            sheet[f'H{row_start}'] = h1[i]
            sheet[f'I{row_start}'] = i1[i]
            sheet[f'J{row_start}'] = j1[i]
            sheet[f'K{row_start}'] = k1[i]
            sheet[f'L{row_start}'] = l1[i]

            # Incrementing the 'row_start' iterator
            # Then continue the loop
            row_start += 1
            continue
        
        # Post-loop logging: successfully painted the output spreadsheet file
        control.append_message_area(f'+ Tabel sukses dipanen!')
        control.set_progress_bar(85)

        # Asking for the spreadsheet name to save as
        # ---
        # Logging and setting the progress bar
        control.append_message_area(f'+ Menyimpan spreadsheet luaran ...')
        control.set_progress_bar(90)
        # Dealing with file name prompt and saving
        # Using loop to mitigate the user clicking 'cancel'
        # in the file name dialog prompt
        while True:
            # Opening the dialog prompt
            output_spreadsheet = filedialog.asksaveasfilename(
                filetypes=[('Excel files', '*.xlsx')],
                initialfile='Sipesat - Dana Penelitian Ringkasan Risat.xlsx',
                title='Simpan sebagai ...'
            )
            
            # 'cancel' button in the dialog prompt is clicked
            if len(output_spreadsheet) == 0:
                # Showing confirmation
                x = messagebox.askyesno(
                    'Nama File Kosong',
                    'Apakah Anda yakin ingin melanjutkan tanpa menyimpan file spreadsheet hasil pemanenan?'
                )
                # Determining whether to break or to continue the loop
                # based on the inversed value of 'x'
                if x:
                    control.append_message_area(f'+ Finalisasi pemanenan data tanpa menyimpan file spreadsheet luaran ...')
                    workbook.close()  # --- closing the workbook without saving
                    break
                else:
                    continue  # --- continuing the loop
            # File name does not end in spreadsheet extension
            elif output_spreadsheet[-5:] != '.xlsx':
                output_spreadsheet = output_spreadsheet + '.xlsx'
            
            # Saving the spreadsheet
            control.append_message_area(f'LOKASI_SPREADSHEET_LUARAN: {output_spreadsheet}')
            control.set_progress_bar(95)
            workbook.save(output_spreadsheet)
    
            # Closing the openpyxl workbook
            control.append_message_area(f'+ Menutup file spreadsheet ...')
            control.set_progress_bar(98)
            workbook.close()

            # Breaking the loop
            break

        # Notify for a successful scraping
        control.append_message_area(f'+ Pemanenan selesai pada: {str(dt.now())}')
        control.set_progress_bar(100)
        control.on_notify_successful_scraping()

    # This function harvests "Risat Arsip Penelitian > Ringkasan Data" data
    # and then store the harvested data as an excel file
    #
    # Required arguments:
    # - control             --> for updating the progress bar and
    #                           message area of the screen SipesatScrHarvest
    # - username, password  --> the Risat administrator username and password
    def run_harvest_r_0_arsip(self, control, username, password):
        # SipesatScrHarvest messenger
        control.set_header_desc('Panen Data "Risat Arsip Penelitian > Ringkasan Data"')
        control.set_help_label('Data sedang dipanen. Silahkan menunggu.')
        control.set_progress_bar(0)
        control.clear_message_area()

        # Preamble logging
        control.append_message_area(f'+ Memulai pemanenan data ...')
        control.append_message_area(f'+ Pemanenan dimulai pada: {str(dt.now())}')
        control.set_progress_bar(5)

        # Preparing the 'data_prompt' arrays
        control.append_message_area(f'+ Log masuk Risat sebagai [{username}] ...')
        control.set_progress_bar(10)
        data_prompt = self.get_risat_login(username, password)
        data_prompt = self.get_risat_penelitian(data_prompt)
        data_prompt = self.get_risat_penelitian_arsip(data_prompt)

        # Parsing XML tree content
        control.append_message_area(f'+ Membaca halaman web ...')
        control.set_progress_bar(15)
        content = data_prompt['html_content']

        # Establishing the export spreadsheet file
        control.append_message_area(f'+ Mempersiapkan file spreadsheet luaran ...')
        control.set_progress_bar(20)
        workbook = xl.Workbook()
        sheet = workbook.active
        sheet.title = 'Arsip Penelitian Ringkasan'

        # Preparing the sheet header
        control.append_message_area(f'+ Mempersiapkan kepala lembar spreadsheet ...')
        control.set_progress_bar(25)
        sheet['A1'].value = 'No.'
        sheet['B1'].value = 'Judul'
        sheet['C1'].value = 'Ketua'
        sheet['D1'].value = 'Bidang Fokus'
        sheet['E1'].value = 'Tgl. Usulan'
        sheet['F1'].value = 'Lama Kegiatan'

        # The base XPath location, pointing to each entry row
        base = '//div[@id="ContentPlaceHolder1_arsip1_updg1"]//div[@class="panel-body f12"]/table[@width="100%"]'

        # ---
        # Obtaining the data row values
        control.append_message_area(f'+ Mendapatkan data pada baris tabel ...')
        control.set_progress_bar(30)

        # HYPOTHESIS:
        # Xpath cannot detect 'tbody' element.
        # So instead of using 'table/tbody/tr', use 'table//tr' instead
        #
        # RESULT:
        # The hypothesis is correct.
        # Therefore, don't mention 'tbody' in any of the following Xpath paths

        a1 = [str(i)
              for i in range(1, len(content.xpath(base)) + 1)]

        b1 = [l.strip()
              for l in content.xpath(base + '//tr[1][@valign="top"]/td[@colspan="4"]/b/text()')]

        c1 = [l.replace('Ketua:', '').strip()
              for l in content.xpath(base + '//tr[3][@valign="top"]/td[2]/table//tr[1]/td/text()')]

        d1 = [l.replace('Bidang Fokus:', '').strip()
              for l in content.xpath(base + '//tr[3][@valign="top"]/td[2]/table//tr[2]/td/text()')]

        e1 = [l.replace('Tgl Usulan:', '').strip()
              for l in content.xpath(base + '//tr[3][@valign="top"]/td[2]/table//tr[3]/td/text()')]

        f1 = [l.replace('Lama Kegiatan:', '').strip()
              for l in content.xpath(base + '//tr[3][@valign="top"]/td[2]/table//tr[4]/td/text()')]

        # The starting row coordinate of the active sheet
        row_start = 2

        # DEBUG
        # Please comment out after use
        # ---
        '''
        print()
        print('-------------------------------------------------------------------------------------------------------')
        for a in (a1, b1, c1, d1, e1, f1):
            print('LEN', len(a))
            print('CONTENT', a)
            print()
        print('-------------------------------------------------------------------------------------------------------', end='')
        print()
        '''

        # Iterating through each table row and write to the table
        # Assumes the lists a1, b1, c1, ... have the same array size
        control.append_message_area(f'+ Melakukan iterasi terhadap baris tabel dan menulis spreadsheet luaran ...')
        control.set_progress_bar(35)
        for i in range(len(a1)):
            # Noisy preamble logging
            # Please don't use this -_-
            # ---
            # control.append_message_area(f'ITERASI [{i}]')

            # Updating the progress bar status
            control.set_progress_bar(35 + round(45 * (i + 1) / (len(a1))))

            # Painting the scraped data to the output spreadsheet row
            sheet[f'A{row_start}'] = a1[i]
            sheet[f'B{row_start}'] = b1[i]
            sheet[f'C{row_start}'] = c1[i]
            sheet[f'D{row_start}'] = d1[i]
            sheet[f'E{row_start}'] = e1[i]
            sheet[f'F{row_start}'] = f1[i]

            # Incrementing the 'row_start' iterator
            # Then continue the loop
            row_start += 1
            continue

        # Post-loop logging: successfully painted the output spreadsheet file
        control.append_message_area(f'+ Tabel sukses dipanen!')
        control.set_progress_bar(85)

        # Asking for the spreadsheet name to save as
        # ---
        # Logging and setting the progress bar
        control.append_message_area(f'+ Menyimpan spreadsheet luaran ...')
        control.set_progress_bar(90)
        # Dealing with file name prompt and saving
        # Using loop to mitigate the user clicking 'cancel'
        # in the file name dialog prompt
        while True:
            # Opening the dialog prompt
            output_spreadsheet = filedialog.asksaveasfilename(
                filetypes=[('Excel files', '*.xlsx')],
                initialfile='Sipesat - Arsip Penelitian Ringkasan Risat.xlsx',
                title='Simpan sebagai ...'
            )

            # 'cancel' button in the dialog prompt is clicked
            if len(output_spreadsheet) == 0:
                # Showing confirmation
                x = messagebox.askyesno(
                    'Nama File Kosong',
                    'Apakah Anda yakin ingin melanjutkan tanpa menyimpan file spreadsheet hasil pemanenan?'
                )
                # Determining whether to break or to continue the loop
                # based on the inversed value of 'x'
                if x:
                    control.append_message_area(
                        f'+ Finalisasi pemanenan data tanpa menyimpan file spreadsheet luaran ...')
                    workbook.close()  # --- closing the workbook without saving
                    break
                else:
                    continue  # --- continuing the loop
            # File name does not end in spreadsheet extension
            elif output_spreadsheet[-5:] != '.xlsx':
                output_spreadsheet = output_spreadsheet + '.xlsx'

            # Saving the spreadsheet
            control.append_message_area(f'LOKASI_SPREADSHEET_LUARAN: {output_spreadsheet}')
            control.set_progress_bar(95)
            workbook.save(output_spreadsheet)

            # Closing the openpyxl workbook
            control.append_message_area(f'+ Menutup file spreadsheet ...')
            control.set_progress_bar(98)
            workbook.close()

            # Breaking the loop
            break

        # Notify for a successful scraping
        control.append_message_area(f'+ Pemanenan selesai pada: {str(dt.now())}')
        control.set_progress_bar(100)
        control.on_notify_successful_scraping()

    # This function harvests "Risat Laporan Akhir Penelitian > Data Ringkasan" data
    # and then store the harvested data as an excel file
    #
    # Required arguments:
    # - control             --> for updating the progress bar and
    #                           message area of the screen SipesatScrHarvest
    # - username, password  --> the Risat administrator username and password
    def run_harvest_r_0_lapakhir(self, control, username, password):

        # SipesatScrHarvest messenger
        control.set_header_desc('Panen Data "Risat Laporan Akhir Penelitian > Data Ringkasan"')
        control.set_help_label('Data sedang dipanen. Silahkan menunggu.')
        control.set_progress_bar(0)
        control.clear_message_area()

        # Preamble logging
        control.append_message_area(f'+ Memulai pemanenan data ...')
        control.append_message_area(f'+ Pemanenan dimulai pada: {str(dt.now())}')
        control.set_progress_bar(5)

        # Preparing the 'data_prompt' arrays
        control.append_message_area(f'+ Log masuk Risat sebagai [{username}] ...')
        control.set_progress_bar(10)
        data_prompt = self.get_risat_login(username, password)
        data_prompt = self.get_risat_penelitian(data_prompt)
        data_prompt = self.get_risat_penelitian_pelak_kegi(data_prompt)
        data_prompt = self.get_risat_penelitian_pelak_kegi_lapakhir(data_prompt)

        # Parsing XML tree content
        control.append_message_area(f'+ Membaca halaman web ...')
        control.set_progress_bar(15)
        content = data_prompt['html_content']

        # Establishing the export spreadsheet file
        control.append_message_area(f'+ Mempersiapkan file spreadsheet luaran ...')
        control.set_progress_bar(20)
        workbook = xl.Workbook()
        sheet = workbook.active
        sheet.title = 'Lap. Akhir Ringkasan Penelitian'

        # Preparing the sheet header
        control.append_message_area(f'+ Mempersiapkan kepala lembar spreadsheet ...')
        control.set_progress_bar(25)
        # ---
        # Preparing the headers
        sheet['A1'].value = 'No.'
        sheet['B1'].value = 'Judul'
        sheet['C1'].value = 'Ketua'
        sheet['D1'].value = 'Bidang Fokus'
        sheet['E1'].value = 'Tgl. Usulan'
        sheet['F1'].value = 'Jumlah Anggota'
        sheet['G1'].value = 'Biaya'
        sheet['H1'].value = 'Dana Disetujui'
        sheet['I1'].value = 'Lama Kegiatan'

        # The base XPath location, pointing to each entry row
        base = '//div[@id="ContentPlaceHolder1_lapakhirbp3m1_updg1"]//div[@class="panel-body f12"]/table[@width="100%"]'

        # ---
        # Obtaining the data row values
        control.append_message_area(f'+ Mendapatkan data pada baris tabel ...')
        control.set_progress_bar(30)

        # HYPOTHESIS:
        # Xpath cannot detect 'tbody' element.
        # So instead of using 'table/tbody/tr', use 'table//tr' instead
        #
        # RESULT:
        # The hypothesis is correct.
        # Therefore, don't mention 'tbody' in any of the following Xpath paths

        a2 = [str(i)
              for i in range(1, len(content.xpath(base)) + 1)]

        b2 = [l.strip()
              for l in content.xpath(base + '//tr[@valign="top"]/td[@colspan="3"]/b/text()')]

        c2 = [l.replace('Ketua:', '').strip()
              for l in content.xpath(base + '//tr[2][not(descendant::table) and not(ancestor::table[@class="table"])]/td[2]/text()')]

        d2 = [l.replace('Bidang Fokus:', '').strip()
              for l in content.xpath(base + '//tr[3][not(descendant::table) and not(ancestor::table[@class="table"])]/td[@colspan="3"]/text()')]

        e2 = [l.replace('Tgl Usulan:', '').strip()
              for l in content.xpath(base + '//tr[4][not(descendant::table) and not(ancestor::table[@class="table"])]/td[@colspan="3"]/text()')]

        f2 = [l.replace('Jml Anggota:', '').strip()
              for l in content.xpath(base + '//tr[5][not(descendant::table) and not(ancestor::table[@class="table"])]/td[@colspan="3"]/text()')]

        g2 = [l.replace('Biaya:', '').replace('Rp.', '').replace(',', '').split('disetujui')[0].strip()
              for l in content.xpath(base + '//tr[6][not(descendant::table) and not(ancestor::table[@class="table"])]/td[@colspan="3"]/text()')]

        h2 = [l.replace('Biaya:', '').replace('Rp.', '').replace(',', '').split('disetujui')[1].strip()
              for l in content.xpath(base + '//tr[6][not(descendant::table) and not(ancestor::table[@class="table"])]/td[@colspan="3"]/text()')]

        i2 = [l.replace('Lama Kegiatan:', '').strip()
              for l in content.xpath(base + '//tr[7][not(descendant::table) and not(ancestor::table[@class="table"])]/td[@colspan="3"]/text()')]

        # The starting row coordinate of the active sheet
        row_start = 2

        # DEBUG
        # Please comment out after use
        # ---
        '''
        print()
        print('-------------------------------------------------------------------------------------------------------')
        for a in (a2, b2, c2, d2, e2, f2, g2, h2, i2, j_):
            print('LEN', len(a))
            print('CONTENT', a)
            print()
        print('-------------------------------------------------------------------------------------------------------', end='')
        print()
        '''

        # Iterating through each table row and write to the table
        # Assumes the lists a1, b1, c1, ... have the same array size
        control.append_message_area(f'+ Melakukan iterasi terhadap baris tabel dan menulis spreadsheet luaran ...')
        control.set_progress_bar(35)
        for i in range(len(a2)):
            # Noisy preamble logging
            # Please don't use this -_-
            # ---
            # control.append_message_area(f'ITERASI [{i}]')

            # Updating the progress bar status
            control.set_progress_bar(35 + round(45 * (i + 1) / (len(a2))))

            # Painting the scraped data to the output spreadsheet row
            # Section 1: 'IDENTITY'
            sheet[f'A{row_start}'] = a2[i]
            sheet[f'B{row_start}'] = b2[i]
            sheet[f'C{row_start}'] = c2[i]
            sheet[f'D{row_start}'] = d2[i]
            sheet[f'E{row_start}'] = e2[i]
            sheet[f'F{row_start}'] = f2[i]
            sheet[f'G{row_start}'] = g2[i]
            sheet[f'H{row_start}'] = h2[i]
            sheet[f'I{row_start}'] = i2[i]

            # Incrementing the 'row_start' iterator
            # Then continue the loop
            row_start += 1
            continue

        # Post-loop logging: successfully painted the output spreadsheet file
        control.append_message_area(f'+ Tabel sukses dipanen!')
        control.set_progress_bar(85)

        # Asking for the spreadsheet name to save as
        # ---
        # Logging and setting the progress bar
        control.append_message_area(f'+ Menyimpan spreadsheet luaran ...')
        control.set_progress_bar(90)
        # Dealing with file name prompt and saving
        # Using loop to mitigate the user clicking 'cancel'
        # in the file name dialog prompt
        while True:
            # Opening the dialog prompt
            output_spreadsheet = filedialog.asksaveasfilename(
                filetypes=[('Excel files', '*.xlsx')],
                initialfile='Sipesat - Lap Akhir Penelitian Ringkasan Risat.xlsx',
                title='Simpan sebagai ...'
            )

            # 'cancel' button in the dialog prompt is clicked
            if len(output_spreadsheet) == 0:
                # Showing confirmation
                x = messagebox.askyesno(
                    'Nama File Kosong',
                    'Apakah Anda yakin ingin melanjutkan tanpa menyimpan file spreadsheet hasil pemanenan?'
                )
                # Determining whether to break or to continue the loop
                # based on the inversed value of 'x'
                if x:
                    control.append_message_area(
                        f'+ Finalisasi pemanenan data tanpa menyimpan file spreadsheet luaran ...')
                    workbook.close()  # --- closing the workbook without saving
                    break
                else:
                    continue  # --- continuing the loop
            # File name does not end in spreadsheet extension
            elif output_spreadsheet[-5:] != '.xlsx':
                output_spreadsheet = output_spreadsheet + '.xlsx'

            # Saving the spreadsheet
            control.append_message_area(f'LOKASI_SPREADSHEET_LUARAN: {output_spreadsheet}')
            control.set_progress_bar(95)
            workbook.save(output_spreadsheet)

            # Closing the openpyxl workbook
            control.append_message_area(f'+ Menutup file spreadsheet ...')
            control.set_progress_bar(98)
            workbook.close()

            # Breaking the loop
            break

        # Notify for a successful scraping
        control.append_message_area(f'+ Pemanenan selesai pada: {str(dt.now())}')
        control.set_progress_bar(100)
        control.on_notify_successful_scraping()

    # This function harvests "Risat Usulan Penelitian > Data Detil Lengkap" data
    # and then store the harvested data as an excel file
    #
    # Required arguments:
    # - control             --> for updating the progress bar and
    #                           message area of the screen SipesatScrHarvest
    # - username, password  --> the Risat administrator username and password
    def run_harvest_r_1_usulan(self, control, username, password):

        # SipesatScrHarvest messenger
        control.set_header_desc('Panen Data "Risat Usulan Penelitian > Data Detil Lengkap"')
        control.set_help_label('Data sedang dipanen. Silahkan menunggu.')
        control.set_progress_bar(0)
        control.clear_message_area()

        # Preamble logging
        control.append_message_area(f'+ Memulai pemanenan data ...')
        control.append_message_area(f'+ Pemanenan dimulai pada: {str(dt.now())}')
        control.set_progress_bar(5)

        # Preparing the 'data_prompt' arrays
        control.append_message_area(f'+ Log masuk Risat sebagai [{username}] ...')
        control.set_progress_bar(10)
        data_prompt = self.get_risat_login(username, password)
        data_prompt = self.get_risat_penelitian(data_prompt)
        data_prompt = self.get_risat_penelitian_usulan_penelitian(data_prompt)

        # Parsing XML tree content
        control.append_message_area(f'+ Membaca halaman web ...')
        control.set_progress_bar(15)
        content = data_prompt['html_content']

        # Establishing the export spreadsheet file
        control.append_message_area(f'+ Mempersiapkan file spreadsheet luaran ...')
        control.set_progress_bar(20)
        workbook = xl.Workbook()
        sheet = workbook.active
        sheet.title = 'Usulan Penelitian Detil'

        # DEVELOPER'S NOTE (DO NOT REMOVE)
        '''
        Data yang perlu diambil:
        1. Judul
        2. Substansi usulan
        3. Identitas pengusul - ketua
        4. Identitas pengusul - anggota
        5. RAB
        6. Dokumen pendukung
        
        Urutan kolom excel:
        1. Judul
        2. Substansi usulan
        3. RAB
        4. Dokumen pendukung
        5. Identitas pengusul - ketua
        6. Identitas pengusul - anggota
        '''

        # Preparing the sheet header
        control.append_message_area(f'+ Mempersiapkan kepala lembar spreadsheet ...')
        control.set_progress_bar(25)
        # ---
        # Preparing the "NOMOR" header
        sheet.merge_cells(
            start_row=1,
            start_column=1,
            end_row=2,
            end_column=1
        )
        sheet['A1'].value = 'No.'
        sheet['A1'].alignment = Alignment(horizontal='center')
        # ---
        # Preparing the "JUDUL" header
        sheet.merge_cells(
            start_row=1,
            start_column=2,
            end_row=1,
            end_column=12
        )
        sheet['B1'].value = 'IDENTITAS'
        sheet['B1'].alignment = Alignment(horizontal='center')
        # Preparing the "JUDUL" sub-headers
        sheet['B2'].value = 'Judul'
        sheet['C2'].value = 'Tgl. Usulan'
        sheet['D2'].value = 'TKT Saat Ini'
        sheet['E2'].value = 'Level'
        sheet['F2'].value = 'Kategori'
        sheet['G2'].value = 'Skema'
        sheet['H2'].value = 'Rumpun Ilmu'
        sheet['I2'].value = 'Bidang Fokus'
        sheet['J2'].value = 'Tema'
        sheet['K2'].value = 'Topik'
        sheet['L2'].value = 'Lama Kegiatan'
        # ---
        # Preparing the "SUBSTANSI USULAN" header
        sheet.merge_cells(
            start_row=1,
            start_column=13,
            end_row=1,
            end_column=14
        )
        sheet['M1'].value = 'SUBSTANSI USULAN'
        sheet['M1'].alignment = Alignment(horizontal='center')
        # Preparing the "SUBSTANSI USULAN" sub-headers
        sheet['M2'].value = 'Kelompok Makro'
        sheet['N2'].value = 'File Proposal'
        # ---
        # Preparing the "RAB" header
        sheet.merge_cells(
            start_row=1,
            start_column=15,
            end_row=1,
            end_column=16
        )
        sheet['O1'].value = 'RENCANA ANGGARAN BIAYA'
        sheet['O1'].alignment = Alignment(horizontal='center')
        # Preparing the "RAB" sub-headers
        sheet['O2'].value = 'Biaya'
        sheet['P2'].value = 'File RAB'
        # ---
        # Preparing the "DOKUMEN PENDUKUNG" header
        sheet.merge_cells(
            start_row=1,
            start_column=17,
            end_row=1,
            end_column=19
        )
        sheet['Q1'].value = 'DOKUMEN PENDUKUNG'
        sheet['Q1'].alignment = Alignment(horizontal='center')
        # Preparing the "DOKUMEN PENDUKUNG" sub-headers
        sheet['Q2'].value = 'Mitra'
        sheet['R2'].value = 'Dukungan Biaya'
        sheet['S2'].value = 'Surat Dukungan Mitra'
        # ---
        # Preparing the "IDENTITAS PENGUSUL — KETUA" header
        sheet.merge_cells(
            start_row=1,
            start_column=20,
            end_row=1,
            end_column=24
        )
        sheet['T1'].value = 'IDENTITAS PENGUSUL — KETUA'
        sheet['T1'].alignment = Alignment(horizontal='center')
        # Preparing the "IDENTITAS PENGUSUL — KETUA" sub-headers
        sheet['T2'].value = 'N.I.P'
        sheet['U2'].value = 'N.I.K'
        sheet['V2'].value = 'N.I.D.N'
        sheet['W2'].value = 'Nama Lengkap'
        sheet['X2'].value = 'Jabatan Fungsional'

        # The base XPath location, pointing to each entry row
        base = '//div[@class="mw-100"]//div[@class="form-group f12"]/table[@width="100%"]//tr[@valign="top"]'

        # ---
        # Obtaining the data row values
        control.append_message_area(f'+ Mendapatkan data pada baris tabel ...')
        control.set_progress_bar(30)

        # Reading the HTML entry row hidden ASPX values
        # Copy-pasted from: /ssynthesia/ghostcity/ar/dumper-2/24__2023.02.13__requestsrisat.py
        all_kodetran_prop = content.xpath(base + '//input[1][@type="hidden"]/@name')
        all_kodetran_val = content.xpath(base + '//input[1][@type="hidden"]/@value')
        all_stat_prop = content.xpath(base + '//input[2][@type="hidden"]/@name')
        all_stat_val = content.xpath(base + '//input[2][@type="hidden"]/@value')
        all_submitbtn = content.xpath(base + '//tr[7]//input[@type="submit"]/@name')

        # The number of rows
        # Assumes the 'all_' array size equals the number of data rows
        number_of_row = len(all_kodetran_prop)

        # The starting row coordinate of the active sheet
        row_start = 3

        # The maximum number of 'DATA ANGGOTA' table data row
        y_max_row = 0

        # Iterating through each entry row element
        # Assumes all the 'all_' arrays in the previous code block
        # are of the same length/size
        # Copy-pasted from: /ssynthesia/ghostcity/ar/dumper-2/24__2023.02.13__requestsrisat.py
        control.append_message_area(f'+ Melakukan iterasi terhadap baris tabel dan menulis spreadsheet luaran ...')
        control.set_progress_bar(35)
        temporary_prompt = data_prompt
        for i in range(number_of_row):
            # Noisy preamble logging
            # Please don't use this -_-
            # ---
            # control.append_message_area(f'ITERASI [{i}]')

            # Updating the progress bar status
            control.set_progress_bar(35 + round(45 * (i + 1) / number_of_row))

            # Preparing the AJAX payload
            detail_prompt = {
                'viewstate' : temporary_prompt['viewstate'],
                'viewstategen' : temporary_prompt['viewstategen'],
                'eventvalidation' : temporary_prompt['eventvalidation'],
                'button_name' : all_submitbtn[i],
                'kodetran_prop' : all_kodetran_prop[i],
                'kodetran_val' : all_kodetran_val[i],
                'stat_prop' : all_stat_prop[i],
                'stat_val' : all_stat_val[i]
            }

            # Obtaining the response data of each individual entry row detail page
            data = self.get_risat_penelitian_usulan_penelitian_detil(detail_prompt)
            content = data['html_content']
            response = data['http_response']

            # Terminal logging for detecting errors
            log_string_1 = str(i+1)
            log_string_2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti1_kiri1_txjudul1"]/text()')[0].replace('\r','').replace('\n','').strip()
            control.append_message_area(f'+ Memanen data detil: [{log_string_1}] {log_string_2} ...')
            print(f'+ Harvesting detailed data: [{log_string_1}] {log_string_2}')

            # DEVELOPER'S NOTE (DO NOT REMOVE)
            '''
            Variables and their associated data
            
            'a1' ==> 'No.'
            'b2' ==> 'Judul'
            'c2' ==> 'Tgl. Usulan'
            'd2' ==> 'TKT Saat Ini'
            'e2' ==> 'Level'
            'f2' ==> 'Kategori'
            'g2' ==> 'Skema'
            'h2' ==> 'Rumpun Ilmu'
            'i2' ==> 'Bidang Fokus'
            'j2' ==> 'Tema'
            'k2' ==> 'Topik'
            'l2' ==> 'Lama Kegiatan'
            'm2' ==> 'Kelompok Makro'
            'n2' ==> 'File Proposal'
            'o2' ==> 'Biaya'
            'p2' ==> 'File RAB'
            'q2' ==> 'Mitra'
            'r2' ==> 'Dukungan Biaya'
            's2' ==> 'Surat Dukungan Mitra'
            't2' ==> 'N.I.P'
            'u2' ==> 'N.I.K'
            'v2' ==> 'N.I.D.N'
            'w2' ==> 'Nama Lengkap'
            'x2' ==> 'Jabatan Fungsional'
            
            In addition, variables starting with 'y' prefix are pertaining
            to the table data of "Data Anggota"
            '''

            # Getting the static (non-variable as in "Data Anggota") detail information
            # ---
            a1 = str(i+1)
            # Use try-except catching to mitigate empty data
            try:
                b2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti1_kiri1_txjudul1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                b2 = ''
            # Use try-except catching to mitigate empty data
            try:
                c2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti1_kiri1_tglusul1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                c2 = ''
            # Use try-except catching to mitigate empty data
            try:
                d2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti1_kiri1_ttkt1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                d2 = ''
            # Use try-except catching to mitigate empty data
            try:
                e2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti1_kiri1_ddllevel1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                e2 = ''
            # Use try-except catching to mitigate empty data
            try:
                f2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti1_kiri1_lkat1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                f2 = ''
            # Use try-except catching to mitigate empty data
            try:
                g2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti1_kiri1_ddlskema1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                g2 = ''
            # 'Rumpun Ilmu' scraped data contains multiple array of strings
            # This needs some extra tweaking
            h2 = ''
            # Use try-except catching to mitigate empty data
            try:
                h2_pre = content.xpath('//div[@id="ContentPlaceHolder1_teliti1_kiri1_updx"]/div[1][@class="panel panel-info"]/div[@class="panel-body f12"]/div[7][@class="row f12"]/div[@class="col-sm-4"]//text()')
                for l in h2_pre:
                    l = l.replace('\r','').replace('\n','').strip()
                    h2 += l +', '
                # Remove trailing ', ' characters
                h2 = h2[:-2]
            except IndexError:
                h2 = ''
            # ---
            # Use try-except catching to mitigate empty data
            try:
                i2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti1_kiri1_ddlfokus1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                i2 = ''
            # Use try-except catching to mitigate empty data
            try:
                j2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti1_kiri1_ddltema1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                j2 = ''
            # Use try-except catching to mitigate empty data
            try:
                k2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti1_kiri1_ddltopik1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                k2 = ''
            # Use try-except catching to mitigate empty data
            try:
                l2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti1_kiri1_ddllama1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                l2 = ''
            # Use try-except catching to mitigate empty data
            try:
                m2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti1_kiri1_ddlmakro1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                m2 = ''
            # Use try-except catching to mitigate empty data
            try:
                n2 = content.xpath('//a[@id="ContentPlaceHolder1_teliti1_kiri1_alblcatatan1"]/@href')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                n2 = ''
            # Use try-except catching to mitigate empty data
            try:
                o2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti1_kiri1_tbiayathn1"]/text()')[0].replace('\r','').replace('\n','').replace('Rp.', '').replace(',', '').strip()
            except IndexError:
                o2 = ''
            # Use try-except catching to mitigate empty data
            try:
                p2 = content.xpath('//a[@id="ContentPlaceHolder1_teliti1_kiri1_alblfilerab1"]/@href')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                p2 = ''
            # Some 'Data Detil' has no 'Mitra' data element
            # Use try-except catching to mitigate
            try:
                q2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti1_kiri1_tnmmitra1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                q2 = ''
            # Use try-except catching to mitigate empty data
            try:
                r2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti1_kiri1_tbiayadukung1"]/text()')[0].replace('\r','').replace('\n','').replace('Rp.', '').replace(',', '').strip()
            except IndexError:
                r2 = ''
            # Use try-except catching to mitigate empty data
            try:
                s2 = content.xpath('//a[@id="ContentPlaceHolder1_teliti1_kiri1_alblfilemitradukung1"]/@href')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                s2 = ''
            # Use try-except catching to mitigate empty data
            try:
                t2 = content.xpath('//div[@id="ContentPlaceHolder1_teliti1_kiri1_bio2_updk1"]/div[1]/div[@class="col-sm-4"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                t2 = ''
            # Use try-except catching to mitigate empty data
            try:
                u2 = content.xpath('//div[@id="ContentPlaceHolder1_teliti1_kiri1_bio2_updk1"]/div[2]/div[@class="col-sm-4"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                u2 = ''
            # Use try-except catching to mitigate empty data
            try:
                v2 = content.xpath('//div[@id="ContentPlaceHolder1_teliti1_kiri1_bio2_updk1"]/div[3]/div[@class="col-sm-4"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                v2 = ''
            # Use try-except catching to mitigate empty data
            try:
                w2 = content.xpath('//div[@id="ContentPlaceHolder1_teliti1_kiri1_bio2_updk1"]/div[4]/div[@class="col-sm-4"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                w2 = ''
            # Use try-except catching to mitigate empty data
            try:
                x2 = content.xpath('//div[@id="ContentPlaceHolder1_teliti1_kiri1_bio2_updk1"]/div[5]/div[@class="col-sm-4"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                x2 = ''

            # Painting the static (non-variable as in "Data Anggota") detail information
            sheet[f'A{row_start}'].value = a1
            sheet[f'B{row_start}'].value = b2
            sheet[f'C{row_start}'].value = c2
            sheet[f'D{row_start}'].value = d2
            sheet[f'E{row_start}'].value = e2
            sheet[f'F{row_start}'].value = f2
            sheet[f'G{row_start}'].value = g2
            sheet[f'H{row_start}'].value = h2
            sheet[f'I{row_start}'].value = i2
            sheet[f'J{row_start}'].value = j2
            sheet[f'K{row_start}'].value = k2
            sheet[f'L{row_start}'].value = l2
            sheet[f'M{row_start}'].value = m2
            sheet[f'N{row_start}'].value = n2
            sheet[f'O{row_start}'].value = o2
            sheet[f'P{row_start}'].value = p2
            sheet[f'Q{row_start}'].value = q2
            sheet[f'R{row_start}'].value = r2
            sheet[f'S{row_start}'].value = s2
            sheet[f'T{row_start}'].value = t2
            sheet[f'U{row_start}'].value = u2
            sheet[f'V{row_start}'].value = v2
            sheet[f'W{row_start}'].value = w2
            sheet[f'X{row_start}'].value = x2

            # The table which displays 'Identitas Pengusul - Anggota Peneliti'
            # Treated differently, specially
            # 'y_' is a DOM element representing a HTML table row (<tr>)
            y_ = content.xpath('//div[@id="ContentPlaceHolder1_teliti1_kiri1_updx"]/div[4]//table[@class="table"]//tr[position()>1]')

            # The base path
            y_base = '//div[@id="ContentPlaceHolder1_teliti1_kiri1_updx"]/div[4]//table[@class="table"]//tr[position()>1]'

            # Painting the scraped data to the output spreadsheet row
            # Section: 'DATA ANGGOTA'
            # ---
            # Checking if 'Identitas Pengusul - Anggota Peneliti' table data exists
            if len(y_) == 0:
                pass  # --- nope. the data does not exist
            else:

                # Checking if this table's 'LUARAN' data has the most rows
                y_all_rows = len(y_)
                y_max_row = max(y_max_row, y_all_rows)

                # ROW DATA LABEL CONVENTION
                # y_row_a -> "N.I.P"
                # y_row_b -> "Nama"
                # y_row_c -> "Bidang Keahlian"
                # y_row_d -> "Alamat"
                # y_row_e -> "Instansi"
                # y_row_f -> "Email"
                # y_row_g -> "No. HP"
                # y_row_h -> "Peran"
                # y_row_i -> "Tugas"

                y_row_a = [l.replace('\r','').replace('\n','').strip()
                      for l in content.xpath(y_base + '/td[2]/text()')]

                y_row_b = [l.replace('\r','').replace('\n','').strip()
                      for l in content.xpath(y_base + '/td[3]/text()')]

                y_row_c = [l.replace('\r','').replace('\n','').strip()
                      for l in content.xpath(y_base + '/td[4]/text()')]

                y_row_d = [l.replace('\r','').replace('\n','').strip()
                      for l in content.xpath(y_base + '/td[5]/text()')]

                y_row_e = [l.replace('\r','').replace('\n','').strip()
                      for l in content.xpath(y_base + '/td[6]/text()')]

                y_row_f = [l.replace('\r','').replace('\n','').strip()
                      for l in content.xpath(y_base + '/td[7]/text()')]

                y_row_g = [l.replace('\r','').replace('\n','').strip()
                      for l in content.xpath(y_base + '/td[8]/text()')]

                y_row_h = [l.replace('\r','').replace('\n','').strip()
                      for l in content.xpath(y_base + '/td[9]/text()')]

                y_row_i = [l.replace('\r','').replace('\n','').strip()
                      for l in content.xpath(y_base + '/td[10]/text()')]

                # The starting column coordinate for filling the 'LUARAN' table data
                col_start = 25

                # Iterating through each table row and write to the spreadsheet
                # Assumes the lists y_row_a, y_row_b, ... have the same array size
                for j in range(len(y_row_a)):

                    # Painting table data
                    sheet.cell(row=row_start, column=col_start).value = y_row_a[j]
                    sheet.cell(row=row_start, column=col_start+1).value = y_row_b[j]
                    sheet.cell(row=row_start, column=col_start+2).value = y_row_c[j]
                    sheet.cell(row=row_start, column=col_start+3).value = y_row_d[j]
                    sheet.cell(row=row_start, column=col_start+4).value = y_row_e[j]
                    sheet.cell(row=row_start, column=col_start+5).value = y_row_f[j]
                    sheet.cell(row=row_start, column=col_start+6).value = y_row_g[j]
                    sheet.cell(row=row_start, column=col_start+7).value = y_row_h[j]
                    sheet.cell(row=row_start, column=col_start+8).value = y_row_i[j]

                    # Incrementing the 'col_start' iterator before continuing the loop
                    col_start += 9
                    continue

            # Incrementing the value of 'row_start' before continuing
            row_start += 1

            # Reopening the "Dana Penelitian" list page,
            # then assign the AJAX response to the temporary array 'temporary_prompt'
            # The 'data' array is obtained from opening individual entry row detail page
            #
            # This is done only on the 7-multiple row iteration, because
            # a single set of 'viewstate', 'viewstategen', and 'eventvalidation' values of ASPX
            # can only be used to do at most 8 operations before having to be renewed.
            if ( i % 7 == 0 ) and ( i > 0 ):
                temporary_prompt = self.get_risat_penelitian_usulan_penelitian(data)
            continue

        # Post-loop logging: appending the header over the 'DATA ANGGOTA' columns
        control.append_message_area(f'+ Melengkapi kepala tabel pada bagian "Identitas Pengusul - Anggota Peneliti" ...')
        # ---
        # The starting column for the 'DATA ANGGOTA' data
        col_start = 25
        # Beginning the loop that detects the maximum number of 'LUARAN' rows
        # according to the variable 'y_max_row'
        for i in range(1, y_max_row+1):
            # Setting the top header
            sheet.merge_cells(
                start_row=1,
                start_column=col_start,
                end_row=1,
                end_column=col_start+8
            )
            sheet.cell(row=1, column=col_start).value = f'IDENTITAS PENGUSUL — ANGGOTA #{i}'
            sheet.cell(row=1, column=col_start).alignment = Alignment(horizontal='center')
            # Setting the sub headers
            sheet.cell(row=2, column=col_start).value = 'N.I.P'
            sheet.cell(row=2, column=col_start+1).value = 'Nama'
            sheet.cell(row=2, column=col_start+2).value = 'Bidang Keahlian'
            sheet.cell(row=2, column=col_start+3).value = 'Alamat'
            sheet.cell(row=2, column=col_start+4).value = 'Instansi'
            sheet.cell(row=2, column=col_start+5).value = 'Email'
            sheet.cell(row=2, column=col_start+6).value = 'No. HP'
            sheet.cell(row=2, column=col_start+7).value = 'Peran'
            sheet.cell(row=2, column=col_start+8).value = 'Tugas'
            # Incrementing the 'col_start' iterator before continuing the loop
            col_start += 9
            continue

        # Post-loop logging: successfully painted the output spreadsheet file
        control.append_message_area(f'+ Tabel sukses dipanen!')
        control.set_progress_bar(85)

        # Asking for the spreadsheet name to save as
        # ---
        # Logging and setting the progress bar
        control.append_message_area(f'+ Menyimpan spreadsheet luaran ...')
        control.set_progress_bar(90)
        # Dealing with file name prompt and saving
        # Using loop to mitigate the user clicking 'cancel'
        # in the file name dialog prompt
        while True:
            # Opening the dialog prompt
            output_spreadsheet = filedialog.asksaveasfilename(
                filetypes=[('Excel files', '*.xlsx')],
                initialfile='Sipesat - Usulan Penelitian Detil Risat.xlsx',
                title='Simpan sebagai ...'
            )

            # 'cancel' button in the dialog prompt is clicked
            if len(output_spreadsheet) == 0:
                # Showing confirmation
                x = messagebox.askyesno(
                    'Nama File Kosong',
                    'Apakah Anda yakin ingin melanjutkan tanpa menyimpan file spreadsheet hasil pemanenan?'
                )
                # Determining whether to break or to continue the loop
                # based on the inversed value of 'x'
                if x:
                    control.append_message_area(
                        f'+ Finalisasi pemanenan data tanpa menyimpan file spreadsheet luaran ...')
                    workbook.close()  # --- closing the workbook without saving
                    break
                else:
                    continue  # --- continuing the loop
            # File name does not end in spreadsheet extension
            elif output_spreadsheet[-5:] != '.xlsx':
                output_spreadsheet = output_spreadsheet + '.xlsx'

            # Saving the spreadsheet
            control.append_message_area(f'LOKASI_SPREADSHEET_LUARAN: {output_spreadsheet}')
            control.set_progress_bar(95)
            workbook.save(output_spreadsheet)

            # Closing the openpyxl workbook
            control.append_message_area(f'+ Menutup file spreadsheet ...')
            control.set_progress_bar(98)
            workbook.close()

            # Breaking the loop
            break

        # Notify for a successful scraping
        control.append_message_area(f'+ Pemanenan selesai pada: {str(dt.now())}')
        control.set_progress_bar(100)
        control.on_notify_successful_scraping()

    # This function harvests "Risat Berkas Disetujui Penelitian > Data Detil Lengkap" data
    # and then store the harvested data as an excel file
    #
    # Required arguments:
    # - control             --> for updating the progress bar and
    #                           message area of the screen SipesatScrHarvest
    # - username, password  --> the Risat administrator username and password
    def run_harvest_r_1_berdisetujui(self, control, username, password):

        # SipesatScrHarvest messenger
        control.set_header_desc('Panen Data "Risat Berkas Disetujui Penelitian > Data Detil Lengkap"')
        control.set_help_label('Data sedang dipanen. Silahkan menunggu.')
        control.set_progress_bar(0)
        control.clear_message_area()

        # Preamble logging
        control.append_message_area(f'+ Memulai pemanenan data ...')
        control.append_message_area(f'+ Pemanenan dimulai pada: {str(dt.now())}')
        control.set_progress_bar(5)

        # Preparing the 'data_prompt' arrays
        control.append_message_area(f'+ Log masuk Risat sebagai [{username}] ...')
        control.set_progress_bar(10)
        data_prompt = self.get_risat_login(username, password)
        data_prompt = self.get_risat_penelitian(data_prompt)
        data_prompt = self.get_risat_penelitian_berdisetujui_penelitian(data_prompt)

        # Parsing XML tree content
        control.append_message_area(f'+ Membaca halaman web ...')
        control.set_progress_bar(15)
        content = data_prompt['html_content']

        # Establishing the export spreadsheet file
        control.append_message_area(f'+ Mempersiapkan file spreadsheet luaran ...')
        control.set_progress_bar(20)
        workbook = xl.Workbook()
        sheet = workbook.active
        sheet.title = 'Berkas Disetujui Penelitian Detil'

        # DEVELOPER'S NOTE (DO NOT REMOVE)
        '''
        Data yang perlu diambil:
        1. Judul
        2. Substansi usulan
        3. Identitas pengusul - ketua
        4. Identitas pengusul - anggota
        5. RAB
        6. Dokumen pendukung
        
        Urutan kolom excel:
        1. Judul
        2. Substansi usulan
        3. RAB
        4. Dokumen pendukung
        5. Identitas pengusul - ketua
        6. Identitas pengusul - anggota
        '''

        # Preparing the sheet header
        control.append_message_area(f'+ Mempersiapkan kepala lembar spreadsheet ...')
        control.set_progress_bar(25)
        # ---
        # Preparing the "NOMOR" header
        sheet.merge_cells(
            start_row=1,
            start_column=1,
            end_row=2,
            end_column=1
        )
        sheet['A1'].value = 'No.'
        sheet['A1'].alignment = Alignment(horizontal='center')
        # ---
        # Preparing the "JUDUL" header
        sheet.merge_cells(
            start_row=1,
            start_column=2,
            end_row=1,
            end_column=12
        )
        sheet['B1'].value = 'IDENTITAS'
        sheet['B1'].alignment = Alignment(horizontal='center')
        # Preparing the "JUDUL" sub-headers
        sheet['B2'].value = 'Judul'
        sheet['C2'].value = 'Tgl. Usulan'
        sheet['D2'].value = 'TKT Saat Ini'
        sheet['E2'].value = 'Level'
        sheet['F2'].value = 'Kategori'
        sheet['G2'].value = 'Skema'
        sheet['H2'].value = 'Rumpun Ilmu'
        sheet['I2'].value = 'Bidang Fokus'
        sheet['J2'].value = 'Tema'
        sheet['K2'].value = 'Topik'
        sheet['L2'].value = 'Lama Kegiatan'
        # ---
        # Preparing the "SUBSTANSI USULAN" header
        sheet.merge_cells(
            start_row=1,
            start_column=13,
            end_row=1,
            end_column=14
        )
        sheet['M1'].value = 'SUBSTANSI USULAN'
        sheet['M1'].alignment = Alignment(horizontal='center')
        # Preparing the "SUBSTANSI USULAN" sub-headers
        sheet['M2'].value = 'Kelompok Makro'
        sheet['N2'].value = 'File Proposal'
        # ---
        # Preparing the "RAB" header
        sheet.merge_cells(
            start_row=1,
            start_column=15,
            end_row=1,
            end_column=16
        )
        sheet['O1'].value = 'RENCANA ANGGARAN BIAYA'
        sheet['O1'].alignment = Alignment(horizontal='center')
        # Preparing the "RAB" sub-headers
        sheet['O2'].value = 'Biaya'
        sheet['P2'].value = 'File RAB'
        # ---
        # Preparing the "DOKUMEN PENDUKUNG" header
        sheet.merge_cells(
            start_row=1,
            start_column=17,
            end_row=1,
            end_column=19
        )
        sheet['Q1'].value = 'DOKUMEN PENDUKUNG'
        sheet['Q1'].alignment = Alignment(horizontal='center')
        # Preparing the "DOKUMEN PENDUKUNG" sub-headers
        sheet['Q2'].value = 'Mitra'
        sheet['R2'].value = 'Dukungan Biaya'
        sheet['S2'].value = 'Surat Dukungan Mitra'
        # ---
        # Preparing the "IDENTITAS PENGUSUL — KETUA" header
        sheet.merge_cells(
            start_row=1,
            start_column=20,
            end_row=1,
            end_column=24
        )
        sheet['T1'].value = 'IDENTITAS PENGUSUL — KETUA'
        sheet['T1'].alignment = Alignment(horizontal='center')
        # Preparing the "IDENTITAS PENGUSUL — KETUA" sub-headers
        sheet['T2'].value = 'N.I.P'
        sheet['U2'].value = 'N.I.K'
        sheet['V2'].value = 'N.I.D.N'
        sheet['W2'].value = 'Nama Lengkap'
        sheet['X2'].value = 'Jabatan Fungsional'

        # The base XPath location, pointing to each entry row
        base = '//div[@class="mw-100"]//div[@class="form-group f12"]/table[@width="100%"]//tr[@valign="top"]'

        # ---
        # Obtaining the data row values
        control.append_message_area(f'+ Mendapatkan data pada baris tabel ...')
        control.set_progress_bar(30)

        # Reading the HTML entry row hidden ASPX values
        # Copy-pasted from: /ssynthesia/ghostcity/ar/dumper-2/24__2023.02.13__requestsrisat.py
        all_kodetran_prop = content.xpath(base + '//input[1][@type="hidden"]/@name')
        all_kodetran_val = content.xpath(base + '//input[1][@type="hidden"]/@value')
        all_stat_prop = content.xpath(base + '//input[2][@type="hidden"]/@name')
        all_stat_val = content.xpath(base + '//input[2][@type="hidden"]/@value')
        all_submitbtn = content.xpath(base + '//tr[7]//input[@type="submit"]/@name')

        # The number of rows
        # Assumes the 'all_' array size equals the number of data rows
        number_of_row = len(all_kodetran_prop)

        # The starting row coordinate of the active sheet
        row_start = 3

        # The maximum number of 'DATA ANGGOTA' table data row
        y_max_row = 0

        # Iterating through each entry row element
        # Assumes all the 'all_' arrays in the previous code block
        # are of the same length/size
        # Copy-pasted from: /ssynthesia/ghostcity/ar/dumper-2/24__2023.02.13__requestsrisat.py
        control.append_message_area(f'+ Melakukan iterasi terhadap baris tabel dan menulis spreadsheet luaran ...')
        control.set_progress_bar(35)
        temporary_prompt = data_prompt
        for i in range(number_of_row):
            # Noisy preamble logging
            # Please don't use this -_-
            # ---
            # control.append_message_area(f'ITERASI [{i}]')

            # Updating the progress bar status
            control.set_progress_bar(35 + round(45 * (i + 1) / number_of_row))

            # Preparing the AJAX payload
            detail_prompt = {
                'viewstate' : temporary_prompt['viewstate'],
                'viewstategen' : temporary_prompt['viewstategen'],
                'eventvalidation' : temporary_prompt['eventvalidation'],
                'button_name' : all_submitbtn[i],
                'kodetran_prop' : all_kodetran_prop[i],
                'kodetran_val' : all_kodetran_val[i],
                'stat_prop' : all_stat_prop[i],
                'stat_val' : all_stat_val[i]
            }

            # Obtaining the response data of each individual entry row detail page
            data = self.get_risat_penelitian_berdisetujui_penelitian_detil(detail_prompt)
            content = data['html_content']
            response = data['http_response']

            # Terminal logging for detecting errors
            log_string_1 = str(i+1)
            log_string_2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti2_kiri1_txjudul1"]/text()')[0].replace('\r','').replace('\n','').strip()
            control.append_message_area(f'+ Memanen data detil: [{log_string_1}] {log_string_2} ...')
            print(f'+ Harvesting detailed data: [{log_string_1}] {log_string_2}')

            # DEVELOPER'S NOTE (DO NOT REMOVE)
            '''
            Variables and their associated data
            
            'a1' ==> 'No.'
            'b2' ==> 'Judul'
            'c2' ==> 'Tgl. Usulan'
            'd2' ==> 'TKT Saat Ini'
            'e2' ==> 'Level'
            'f2' ==> 'Kategori'
            'g2' ==> 'Skema'
            'h2' ==> 'Rumpun Ilmu'
            'i2' ==> 'Bidang Fokus'
            'j2' ==> 'Tema'
            'k2' ==> 'Topik'
            'l2' ==> 'Lama Kegiatan'
            'm2' ==> 'Kelompok Makro'
            'n2' ==> 'File Proposal'
            'o2' ==> 'Biaya'
            'p2' ==> 'File RAB'
            'q2' ==> 'Mitra'
            'r2' ==> 'Dukungan Biaya'
            's2' ==> 'Surat Dukungan Mitra'
            't2' ==> 'N.I.P'
            'u2' ==> 'N.I.K'
            'v2' ==> 'N.I.D.N'
            'w2' ==> 'Nama Lengkap'
            'x2' ==> 'Jabatan Fungsional'
            
            In addition, variables starting with 'y' prefix are pertaining
            to the table data of "Data Anggota"
            '''

            # Getting the static (non-variable as in "Data Anggota") detail information
            # ---
            a1 = str(i+1)
            # Use try-except catching to mitigate empty data
            try:
                b2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti2_kiri1_txjudul1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                b2 = ''
            # Use try-except catching to mitigate empty data
            try:
                c2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti2_kiri1_tglusul1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                c2 = ''
            # Use try-except catching to mitigate empty data
            try:
                d2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti2_kiri1_ttkt1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                d2 = ''
            # Use try-except catching to mitigate empty data
            try:
                e2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti2_kiri1_ddllevel1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                e2 = ''
            # Use try-except catching to mitigate empty data
            try:
                f2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti2_kiri1_lkat1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                f2 = ''
            # Use try-except catching to mitigate empty data
            try:
                g2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti2_kiri1_ddlskema1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                g2 = ''
            # 'Rumpun Ilmu' scraped data contains multiple array of strings
            # This needs some extra tweaking
            h2 = ''
            # Use try-except catching to mitigate empty data
            try:
                h2_pre = content.xpath('//div[@id="ContentPlaceHolder1_teliti2_kiri1_updx"]/div[1][@class="panel panel-info"]/div[@class="panel-body f12"]/div[7][@class="row f12"]/div[@class="col-sm-4"]//text()')
                for l in h2_pre:
                    l = l.replace('\r','').replace('\n','').strip()
                    h2 += l +', '
                # Remove trailing ', ' characters
                h2 = h2[:-2]
            except IndexError:
                h2 = ''
            # ---
            # Use try-except catching to mitigate empty data
            try:
                i2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti2_kiri1_ddlfokus1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                i2 = ''
            # Use try-except catching to mitigate empty data
            try:
                j2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti2_kiri1_ddltema1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                j2 = ''
            # Use try-except catching to mitigate empty data
            try:
                k2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti2_kiri1_ddltopik1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                k2 = ''
            # Use try-except catching to mitigate empty data
            try:
                l2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti2_kiri1_ddllama1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                l2 = ''
            # Use try-except catching to mitigate empty data
            try:
                m2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti2_kiri1_ddlmakro1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                m2 = ''
            # Use try-except catching to mitigate empty data
            try:
                n2 = content.xpath('//a[@id="ContentPlaceHolder1_teliti2_kiri1_alblcatatan1"]/@href')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                n2 = ''
            # Use try-except catching to mitigate empty data
            try:
                o2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti2_kiri1_tbiayathn1"]/text()')[0].replace('\r','').replace('\n','').replace('Rp.', '').replace(',', '').strip()
            except IndexError:
                o2 = ''
            # Use try-except catching to mitigate empty data
            try:
                p2 = content.xpath('//a[@id="ContentPlaceHolder1_teliti2_kiri1_alblfilerab1"]/@href')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                p2 = ''
            # Some 'Data Detil' has no 'Mitra' data element
            # Use try-except catching to mitigate
            try:
                q2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti2_kiri1_tnmmitra1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                q2 = ''
            # Use try-except catching to mitigate empty data
            try:
                r2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti2_kiri1_tbiayadukung1"]/text()')[0].replace('\r','').replace('\n','').replace('Rp.', '').replace(',', '').strip()
            except IndexError:
                r2 = ''
            # Use try-except catching to mitigate empty data
            try:
                s2 = content.xpath('//a[@id="ContentPlaceHolder1_teliti2_kiri1_alblfilemitradukung1"]/@href')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                s2 = ''
            # Use try-except catching to mitigate empty data
            try:
                t2 = content.xpath('//div[@id="ContentPlaceHolder1_teliti2_kiri1_bio2_updk1"]/div[1]/div[@class="col-sm-4"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                t2 = ''
            # Use try-except catching to mitigate empty data
            try:
                u2 = content.xpath('//div[@id="ContentPlaceHolder1_teliti2_kiri1_bio2_updk1"]/div[2]/div[@class="col-sm-4"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                u2 = ''
            # Use try-except catching to mitigate empty data
            try:
                v2 = content.xpath('//div[@id="ContentPlaceHolder1_teliti2_kiri1_bio2_updk1"]/div[3]/div[@class="col-sm-4"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                v2 = ''
            # Use try-except catching to mitigate empty data
            try:
                w2 = content.xpath('//div[@id="ContentPlaceHolder1_teliti2_kiri1_bio2_updk1"]/div[4]/div[@class="col-sm-4"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                w2 = ''
            # Use try-except catching to mitigate empty data
            try:
                x2 = content.xpath('//div[@id="ContentPlaceHolder1_teliti2_kiri1_bio2_updk1"]/div[5]/div[@class="col-sm-4"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                x2 = ''

            # Painting the static (non-variable as in "Data Anggota") detail information
            sheet[f'A{row_start}'].value = a1
            sheet[f'B{row_start}'].value = b2
            sheet[f'C{row_start}'].value = c2
            sheet[f'D{row_start}'].value = d2
            sheet[f'E{row_start}'].value = e2
            sheet[f'F{row_start}'].value = f2
            sheet[f'G{row_start}'].value = g2
            sheet[f'H{row_start}'].value = h2
            sheet[f'I{row_start}'].value = i2
            sheet[f'J{row_start}'].value = j2
            sheet[f'K{row_start}'].value = k2
            sheet[f'L{row_start}'].value = l2
            sheet[f'M{row_start}'].value = m2
            sheet[f'N{row_start}'].value = n2
            sheet[f'O{row_start}'].value = o2
            sheet[f'P{row_start}'].value = p2
            sheet[f'Q{row_start}'].value = q2
            sheet[f'R{row_start}'].value = r2
            sheet[f'S{row_start}'].value = s2
            sheet[f'T{row_start}'].value = t2
            sheet[f'U{row_start}'].value = u2
            sheet[f'V{row_start}'].value = v2
            sheet[f'W{row_start}'].value = w2
            sheet[f'X{row_start}'].value = x2

            # The table which displays 'Identitas Pengusul - Anggota Peneliti'
            # Treated differently, specially
            # 'y_' is a DOM element representing a HTML table row (<tr>)
            y_ = content.xpath('//div[@id="ContentPlaceHolder1_teliti2_kiri1_updx"]/div[4]//table[@class="table"]//tr[position()>1]')

            # The base path
            y_base = '//div[@id="ContentPlaceHolder1_teliti2_kiri1_updx"]/div[4]//table[@class="table"]//tr[position()>1]'

            # Painting the scraped data to the output spreadsheet row
            # Section: 'DATA ANGGOTA'
            # ---
            # Checking if 'Identitas Pengusul - Anggota Peneliti' table data exists
            if len(y_) == 0:
                pass  # --- nope. the data does not exist
            else:

                # Checking if this table's 'LUARAN' data has the most rows
                y_all_rows = len(y_)
                y_max_row = max(y_max_row, y_all_rows)

                # ROW DATA LABEL CONVENTION
                # y_row_a -> "N.I.P"
                # y_row_b -> "Nama"
                # y_row_c -> "Bidang Keahlian"
                # y_row_d -> "Alamat"
                # y_row_e -> "Instansi"
                # y_row_f -> "Email"
                # y_row_g -> "No. HP"
                # y_row_h -> "Peran"
                # y_row_i -> "Tugas"

                y_row_a = [l.replace('\r','').replace('\n','').strip()
                      for l in content.xpath(y_base + '/td[2]/text()')]

                y_row_b = [l.replace('\r','').replace('\n','').strip()
                      for l in content.xpath(y_base + '/td[3]/text()')]

                y_row_c = [l.replace('\r','').replace('\n','').strip()
                      for l in content.xpath(y_base + '/td[4]/text()')]

                y_row_d = [l.replace('\r','').replace('\n','').strip()
                      for l in content.xpath(y_base + '/td[5]/text()')]

                y_row_e = [l.replace('\r','').replace('\n','').strip()
                      for l in content.xpath(y_base + '/td[6]/text()')]

                y_row_f = [l.replace('\r','').replace('\n','').strip()
                      for l in content.xpath(y_base + '/td[7]/text()')]

                y_row_g = [l.replace('\r','').replace('\n','').strip()
                      for l in content.xpath(y_base + '/td[8]/text()')]

                y_row_h = [l.replace('\r','').replace('\n','').strip()
                      for l in content.xpath(y_base + '/td[9]/text()')]

                y_row_i = [l.replace('\r','').replace('\n','').strip()
                      for l in content.xpath(y_base + '/td[10]/text()')]

                # The starting column coordinate for filling the 'LUARAN' table data
                col_start = 25

                # Iterating through each table row and write to the spreadsheet
                # Assumes the lists y_row_a, y_row_b, ... have the same array size
                for j in range(len(y_row_a)):

                    # Painting table data
                    sheet.cell(row=row_start, column=col_start).value = y_row_a[j]
                    sheet.cell(row=row_start, column=col_start+1).value = y_row_b[j]
                    sheet.cell(row=row_start, column=col_start+2).value = y_row_c[j]
                    sheet.cell(row=row_start, column=col_start+3).value = y_row_d[j]
                    sheet.cell(row=row_start, column=col_start+4).value = y_row_e[j]
                    sheet.cell(row=row_start, column=col_start+5).value = y_row_f[j]
                    sheet.cell(row=row_start, column=col_start+6).value = y_row_g[j]
                    sheet.cell(row=row_start, column=col_start+7).value = y_row_h[j]
                    sheet.cell(row=row_start, column=col_start+8).value = y_row_i[j]

                    # Incrementing the 'col_start' iterator before continuing the loop
                    col_start += 9
                    continue

            # Incrementing the value of 'row_start' before continuing
            row_start += 1

            # Reopening the "Dana Penelitian" list page,
            # then assign the AJAX response to the temporary array 'temporary_prompt'
            # The 'data' array is obtained from opening individual entry row detail page
            #
            # This is done only on the 7-multiple row iteration, because
            # a single set of 'viewstate', 'viewstategen', and 'eventvalidation' values of ASPX
            # can only be used to do at most 8 operations before having to be renewed.
            if ( i % 7 == 0 ) and ( i > 0 ):
                temporary_prompt = self.get_risat_penelitian_berdisetujui_penelitian(data)
            continue

        # Post-loop logging: appending the header over the 'DATA ANGGOTA' columns
        control.append_message_area(f'+ Melengkapi kepala tabel pada bagian "Identitas Pengusul - Anggota Peneliti" ...')
        # ---
        # The starting column for the 'DATA ANGGOTA' data
        col_start = 25
        # Beginning the loop that detects the maximum number of 'LUARAN' rows
        # according to the variable 'y_max_row'
        for i in range(1, y_max_row+1):
            # Setting the top header
            sheet.merge_cells(
                start_row=1,
                start_column=col_start,
                end_row=1,
                end_column=col_start+8
            )
            sheet.cell(row=1, column=col_start).value = f'IDENTITAS PENGUSUL — ANGGOTA #{i}'
            sheet.cell(row=1, column=col_start).alignment = Alignment(horizontal='center')
            # Setting the sub headers
            sheet.cell(row=2, column=col_start).value = 'N.I.P'
            sheet.cell(row=2, column=col_start+1).value = 'Nama'
            sheet.cell(row=2, column=col_start+2).value = 'Bidang Keahlian'
            sheet.cell(row=2, column=col_start+3).value = 'Alamat'
            sheet.cell(row=2, column=col_start+4).value = 'Instansi'
            sheet.cell(row=2, column=col_start+5).value = 'Email'
            sheet.cell(row=2, column=col_start+6).value = 'No. HP'
            sheet.cell(row=2, column=col_start+7).value = 'Peran'
            sheet.cell(row=2, column=col_start+8).value = 'Tugas'
            # Incrementing the 'col_start' iterator before continuing the loop
            col_start += 9
            continue

        # Post-loop logging: successfully painted the output spreadsheet file
        control.append_message_area(f'+ Tabel sukses dipanen!')
        control.set_progress_bar(85)

        # Asking for the spreadsheet name to save as
        # ---
        # Logging and setting the progress bar
        control.append_message_area(f'+ Menyimpan spreadsheet luaran ...')
        control.set_progress_bar(90)
        # Dealing with file name prompt and saving
        # Using loop to mitigate the user clicking 'cancel'
        # in the file name dialog prompt
        while True:
            # Opening the dialog prompt
            output_spreadsheet = filedialog.asksaveasfilename(
                filetypes=[('Excel files', '*.xlsx')],
                initialfile='Sipesat - Berkas Disetujui Penelitian Detil Risat.xlsx',
                title='Simpan sebagai ...'
            )

            # 'cancel' button in the dialog prompt is clicked
            if len(output_spreadsheet) == 0:
                # Showing confirmation
                x = messagebox.askyesno(
                    'Nama File Kosong',
                    'Apakah Anda yakin ingin melanjutkan tanpa menyimpan file spreadsheet hasil pemanenan?'
                )
                # Determining whether to break or to continue the loop
                # based on the inversed value of 'x'
                if x:
                    control.append_message_area(
                        f'+ Finalisasi pemanenan data tanpa menyimpan file spreadsheet luaran ...')
                    workbook.close()  # --- closing the workbook without saving
                    break
                else:
                    continue  # --- continuing the loop
            # File name does not end in spreadsheet extension
            elif output_spreadsheet[-5:] != '.xlsx':
                output_spreadsheet = output_spreadsheet + '.xlsx'

            # Saving the spreadsheet
            control.append_message_area(f'LOKASI_SPREADSHEET_LUARAN: {output_spreadsheet}')
            control.set_progress_bar(95)
            workbook.save(output_spreadsheet)

            # Closing the openpyxl workbook
            control.append_message_area(f'+ Menutup file spreadsheet ...')
            control.set_progress_bar(98)
            workbook.close()

            # Breaking the loop
            break

        # Notify for a successful scraping
        control.append_message_area(f'+ Pemanenan selesai pada: {str(dt.now())}')
        control.set_progress_bar(100)
        control.on_notify_successful_scraping()

    # This function harvests "Risat Berkas Direvisi Penelitian > Data Detil Lengkap" data
    # and then store the harvested data as an excel file
    #
    # Required arguments:
    # - control             --> for updating the progress bar and
    #                           message area of the screen SipesatScrHarvest
    # - username, password  --> the Risat administrator username and password
    def run_harvest_r_1_berdirevisi(self, control, username, password):

        # SipesatScrHarvest messenger
        control.set_header_desc('Panen Data "Risat Berkas Direvisi Penelitian > Data Detil Lengkap"')
        control.set_help_label('Data sedang dipanen. Silahkan menunggu.')
        control.set_progress_bar(0)
        control.clear_message_area()

        # Preamble logging
        control.append_message_area(f'+ Memulai pemanenan data ...')
        control.append_message_area(f'+ Pemanenan dimulai pada: {str(dt.now())}')
        control.set_progress_bar(5)

        # Preparing the 'data_prompt' arrays
        control.append_message_area(f'+ Log masuk Risat sebagai [{username}] ...')
        control.set_progress_bar(10)
        data_prompt = self.get_risat_login(username, password)
        data_prompt = self.get_risat_penelitian(data_prompt)
        data_prompt = self.get_risat_penelitian_berdirevisi_penelitian(data_prompt)

        # Parsing XML tree content
        control.append_message_area(f'+ Membaca halaman web ...')
        control.set_progress_bar(15)
        content = data_prompt['html_content']

        # Establishing the export spreadsheet file
        control.append_message_area(f'+ Mempersiapkan file spreadsheet luaran ...')
        control.set_progress_bar(20)
        workbook = xl.Workbook()
        sheet = workbook.active
        sheet.title = 'Berkas Direvisi Penelitian Detil'

        # DEVELOPER'S NOTE (DO NOT REMOVE)
        '''
        Data yang perlu diambil:
        1. Judul
        2. Substansi usulan
        3. Identitas pengusul - ketua
        4. Identitas pengusul - anggota
        5. RAB
        6. Dokumen pendukung
        
        Urutan kolom excel:
        1. Judul
        2. Substansi usulan
        3. RAB
        4. Dokumen pendukung
        5. Identitas pengusul - ketua
        6. Identitas pengusul - anggota
        '''

        # Preparing the sheet header
        control.append_message_area(f'+ Mempersiapkan kepala lembar spreadsheet ...')
        control.set_progress_bar(25)
        # ---
        # Preparing the "NOMOR" header
        sheet.merge_cells(
            start_row=1,
            start_column=1,
            end_row=2,
            end_column=1
        )
        sheet['A1'].value = 'No.'
        sheet['A1'].alignment = Alignment(horizontal='center')
        # ---
        # Preparing the "JUDUL" header
        sheet.merge_cells(
            start_row=1,
            start_column=2,
            end_row=1,
            end_column=12
        )
        sheet['B1'].value = 'IDENTITAS'
        sheet['B1'].alignment = Alignment(horizontal='center')
        # Preparing the "JUDUL" sub-headers
        sheet['B2'].value = 'Judul'
        sheet['C2'].value = 'Tgl. Usulan'
        sheet['D2'].value = 'TKT Saat Ini'
        sheet['E2'].value = 'Level'
        sheet['F2'].value = 'Kategori'
        sheet['G2'].value = 'Skema'
        sheet['H2'].value = 'Rumpun Ilmu'
        sheet['I2'].value = 'Bidang Fokus'
        sheet['J2'].value = 'Tema'
        sheet['K2'].value = 'Topik'
        sheet['L2'].value = 'Lama Kegiatan'
        # ---
        # Preparing the "SUBSTANSI USULAN" header
        sheet.merge_cells(
            start_row=1,
            start_column=13,
            end_row=1,
            end_column=14
        )
        sheet['M1'].value = 'SUBSTANSI USULAN'
        sheet['M1'].alignment = Alignment(horizontal='center')
        # Preparing the "SUBSTANSI USULAN" sub-headers
        sheet['M2'].value = 'Kelompok Makro'
        sheet['N2'].value = 'File Proposal'
        # ---
        # Preparing the "RAB" header
        sheet.merge_cells(
            start_row=1,
            start_column=15,
            end_row=1,
            end_column=16
        )
        sheet['O1'].value = 'RENCANA ANGGARAN BIAYA'
        sheet['O1'].alignment = Alignment(horizontal='center')
        # Preparing the "RAB" sub-headers
        sheet['O2'].value = 'Biaya'
        sheet['P2'].value = 'File RAB'
        # ---
        # Preparing the "DOKUMEN PENDUKUNG" header
        sheet.merge_cells(
            start_row=1,
            start_column=17,
            end_row=1,
            end_column=19
        )
        sheet['Q1'].value = 'DOKUMEN PENDUKUNG'
        sheet['Q1'].alignment = Alignment(horizontal='center')
        # Preparing the "DOKUMEN PENDUKUNG" sub-headers
        sheet['Q2'].value = 'Mitra'
        sheet['R2'].value = 'Dukungan Biaya'
        sheet['S2'].value = 'Surat Dukungan Mitra'
        # ---
        # Preparing the "IDENTITAS PENGUSUL — KETUA" header
        sheet.merge_cells(
            start_row=1,
            start_column=20,
            end_row=1,
            end_column=24
        )
        sheet['T1'].value = 'IDENTITAS PENGUSUL — KETUA'
        sheet['T1'].alignment = Alignment(horizontal='center')
        # Preparing the "IDENTITAS PENGUSUL — KETUA" sub-headers
        sheet['T2'].value = 'N.I.P'
        sheet['U2'].value = 'N.I.K'
        sheet['V2'].value = 'N.I.D.N'
        sheet['W2'].value = 'Nama Lengkap'
        sheet['X2'].value = 'Jabatan Fungsional'

        # The base XPath location, pointing to each entry row
        base = '//div[@class="mw-100"]//div[@class="form-group f12"]/table[@width="100%"]//tr[@valign="top"]'

        # ---
        # Obtaining the data row values
        control.append_message_area(f'+ Mendapatkan data pada baris tabel ...')
        control.set_progress_bar(30)

        # Reading the HTML entry row hidden ASPX values
        # Copy-pasted from: /ssynthesia/ghostcity/ar/dumper-2/24__2023.02.13__requestsrisat.py
        all_kodetran_prop = content.xpath(base + '//input[1][@type="hidden"]/@name')
        all_kodetran_val = content.xpath(base + '//input[1][@type="hidden"]/@value')
        all_stat_prop = content.xpath(base + '//input[2][@type="hidden"]/@name')
        all_stat_val = content.xpath(base + '//input[2][@type="hidden"]/@value')
        all_submitbtn = content.xpath(base + '//tr[7]//input[@type="submit"]/@name')

        # The number of rows
        # Assumes the 'all_' array size equals the number of data rows
        number_of_row = len(all_kodetran_prop)

        # The starting row coordinate of the active sheet
        row_start = 3

        # The maximum number of 'DATA ANGGOTA' table data row
        y_max_row = 0

        # Iterating through each entry row element
        # Assumes all the 'all_' arrays in the previous code block
        # are of the same length/size
        # Copy-pasted from: /ssynthesia/ghostcity/ar/dumper-2/24__2023.02.13__requestsrisat.py
        control.append_message_area(f'+ Melakukan iterasi terhadap baris tabel dan menulis spreadsheet luaran ...')
        control.set_progress_bar(35)
        temporary_prompt = data_prompt
        for i in range(number_of_row):
            # Noisy preamble logging
            # Please don't use this -_-
            # ---
            # control.append_message_area(f'ITERASI [{i}]')

            # Updating the progress bar status
            control.set_progress_bar(35 + round(45 * (i + 1) / number_of_row))

            # Preparing the AJAX payload
            detail_prompt = {
                'viewstate' : temporary_prompt['viewstate'],
                'viewstategen' : temporary_prompt['viewstategen'],
                'eventvalidation' : temporary_prompt['eventvalidation'],
                'button_name' : all_submitbtn[i],
                'kodetran_prop' : all_kodetran_prop[i],
                'kodetran_val' : all_kodetran_val[i],
                'stat_prop' : all_stat_prop[i],
                'stat_val' : all_stat_val[i]
            }

            # Obtaining the response data of each individual entry row detail page
            data = self.get_risat_penelitian_berdirevisi_penelitian_detil(detail_prompt)
            content = data['html_content']
            response = data['http_response']

            # Terminal logging for detecting errors
            log_string_1 = str(i+1)
            log_string_2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti3_kiri1_txjudul1"]/text()')[0].replace('\r','').replace('\n','').strip()
            control.append_message_area(f'+ Memanen data detil: [{log_string_1}] {log_string_2} ...')
            print(f'+ Harvesting detailed data: [{log_string_1}] {log_string_2}')

            # DEVELOPER'S NOTE (DO NOT REMOVE)
            '''
            Variables and their associated data
            
            'a1' ==> 'No.'
            'b2' ==> 'Judul'
            'c2' ==> 'Tgl. Usulan'
            'd2' ==> 'TKT Saat Ini'
            'e2' ==> 'Level'
            'f2' ==> 'Kategori'
            'g2' ==> 'Skema'
            'h2' ==> 'Rumpun Ilmu'
            'i2' ==> 'Bidang Fokus'
            'j2' ==> 'Tema'
            'k2' ==> 'Topik'
            'l2' ==> 'Lama Kegiatan'
            'm2' ==> 'Kelompok Makro'
            'n2' ==> 'File Proposal'
            'o2' ==> 'Biaya'
            'p2' ==> 'File RAB'
            'q2' ==> 'Mitra'
            'r2' ==> 'Dukungan Biaya'
            's2' ==> 'Surat Dukungan Mitra'
            't2' ==> 'N.I.P'
            'u2' ==> 'N.I.K'
            'v2' ==> 'N.I.D.N'
            'w2' ==> 'Nama Lengkap'
            'x2' ==> 'Jabatan Fungsional'
            
            In addition, variables starting with 'y' prefix are pertaining
            to the table data of "Data Anggota"
            '''

            # Getting the static (non-variable as in "Data Anggota") detail information
            # ---
            a1 = str(i+1)
            # Use try-except catching to mitigate empty data
            try:
                b2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti3_kiri1_txjudul1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                b2 = ''
            # Use try-except catching to mitigate empty data
            try:
                c2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti3_kiri1_tglusul1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                c2 = ''
            # Use try-except catching to mitigate empty data
            try:
                d2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti3_kiri1_ttkt1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                d2 = ''
            # Use try-except catching to mitigate empty data
            try:
                e2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti3_kiri1_ddllevel1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                e2 = ''
            # Use try-except catching to mitigate empty data
            try:
                f2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti3_kiri1_lkat1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                f2 = ''
            # Use try-except catching to mitigate empty data
            try:
                g2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti3_kiri1_ddlskema1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                g2 = ''
            # 'Rumpun Ilmu' scraped data contains multiple array of strings
            # This needs some extra tweaking
            h2 = ''
            # Use try-except catching to mitigate empty data
            try:
                h2_pre = content.xpath('//div[@id="ContentPlaceHolder1_teliti3_kiri1_updx"]/div[1][@class="panel panel-info"]/div[@class="panel-body f12"]/div[7][@class="row f12"]/div[@class="col-sm-4"]//text()')
                for l in h2_pre:
                    l = l.replace('\r','').replace('\n','').strip()
                    h2 += l +', '
                # Remove trailing ', ' characters
                h2 = h2[:-2]
            except IndexError:
                h2 = ''
            # ---
            # Use try-except catching to mitigate empty data
            try:
                i2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti3_kiri1_ddlfokus1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                i2 = ''
            # Use try-except catching to mitigate empty data
            try:
                j2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti3_kiri1_ddltema1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                j2 = ''
            # Use try-except catching to mitigate empty data
            try:
                k2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti3_kiri1_ddltopik1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                k2 = ''
            # Use try-except catching to mitigate empty data
            try:
                l2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti3_kiri1_ddllama1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                l2 = ''
            # Use try-except catching to mitigate empty data
            try:
                m2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti3_kiri1_ddlmakro1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                m2 = ''
            # Use try-except catching to mitigate empty data
            try:
                n2 = content.xpath('//a[@id="ContentPlaceHolder1_teliti3_kiri1_alblcatatan1"]/@href')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                n2 = ''
            # Use try-except catching to mitigate empty data
            try:
                o2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti3_kiri1_tbiayathn1"]/text()')[0].replace('\r','').replace('\n','').replace('Rp.', '').replace(',', '').strip()
            except IndexError:
                o2 = ''
            # Use try-except catching to mitigate empty data
            try:
                p2 = content.xpath('//a[@id="ContentPlaceHolder1_teliti3_kiri1_alblfilerab1"]/@href')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                p2 = ''
            # Some 'Data Detil' has no 'Mitra' data element
            # Use try-except catching to mitigate
            try:
                q2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti3_kiri1_tnmmitra1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                q2 = ''
            # Use try-except catching to mitigate empty data
            try:
                r2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti3_kiri1_tbiayadukung1"]/text()')[0].replace('\r','').replace('\n','').replace('Rp.', '').replace(',', '').strip()
            except IndexError:
                r2 = ''
            # Use try-except catching to mitigate empty data
            try:
                s2 = content.xpath('//a[@id="ContentPlaceHolder1_teliti3_kiri1_alblfilemitradukung1"]/@href')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                s2 = ''
            # Use try-except catching to mitigate empty data
            try:
                t2 = content.xpath('//div[@id="ContentPlaceHolder1_teliti3_kiri1_bio2_updk1"]/div[1]/div[@class="col-sm-4"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                t2 = ''
            # Use try-except catching to mitigate empty data
            try:
                u2 = content.xpath('//div[@id="ContentPlaceHolder1_teliti3_kiri1_bio2_updk1"]/div[2]/div[@class="col-sm-4"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                u2 = ''
            # Use try-except catching to mitigate empty data
            try:
                v2 = content.xpath('//div[@id="ContentPlaceHolder1_teliti3_kiri1_bio2_updk1"]/div[3]/div[@class="col-sm-4"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                v2 = ''
            # Use try-except catching to mitigate empty data
            try:
                w2 = content.xpath('//div[@id="ContentPlaceHolder1_teliti3_kiri1_bio2_updk1"]/div[4]/div[@class="col-sm-4"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                w2 = ''
            # Use try-except catching to mitigate empty data
            try:
                x2 = content.xpath('//div[@id="ContentPlaceHolder1_teliti3_kiri1_bio2_updk1"]/div[5]/div[@class="col-sm-4"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                x2 = ''

            # Painting the static (non-variable as in "Data Anggota") detail information
            sheet[f'A{row_start}'].value = a1
            sheet[f'B{row_start}'].value = b2
            sheet[f'C{row_start}'].value = c2
            sheet[f'D{row_start}'].value = d2
            sheet[f'E{row_start}'].value = e2
            sheet[f'F{row_start}'].value = f2
            sheet[f'G{row_start}'].value = g2
            sheet[f'H{row_start}'].value = h2
            sheet[f'I{row_start}'].value = i2
            sheet[f'J{row_start}'].value = j2
            sheet[f'K{row_start}'].value = k2
            sheet[f'L{row_start}'].value = l2
            sheet[f'M{row_start}'].value = m2
            sheet[f'N{row_start}'].value = n2
            sheet[f'O{row_start}'].value = o2
            sheet[f'P{row_start}'].value = p2
            sheet[f'Q{row_start}'].value = q2
            sheet[f'R{row_start}'].value = r2
            sheet[f'S{row_start}'].value = s2
            sheet[f'T{row_start}'].value = t2
            sheet[f'U{row_start}'].value = u2
            sheet[f'V{row_start}'].value = v2
            sheet[f'W{row_start}'].value = w2
            sheet[f'X{row_start}'].value = x2

            # The table which displays 'Identitas Pengusul - Anggota Peneliti'
            # Treated differently, specially
            # 'y_' is a DOM element representing a HTML table row (<tr>)
            y_ = content.xpath('//div[@id="ContentPlaceHolder1_teliti3_kiri1_updx"]/div[4]//table[@class="table"]//tr[position()>1]')

            # The base path
            y_base = '//div[@id="ContentPlaceHolder1_teliti3_kiri1_updx"]/div[4]//table[@class="table"]//tr[position()>1]'

            # Painting the scraped data to the output spreadsheet row
            # Section: 'DATA ANGGOTA'
            # ---
            # Checking if 'Identitas Pengusul - Anggota Peneliti' table data exists
            if len(y_) == 0:
                pass  # --- nope. the data does not exist
            else:

                # Checking if this table's 'LUARAN' data has the most rows
                y_all_rows = len(y_)
                y_max_row = max(y_max_row, y_all_rows)

                # ROW DATA LABEL CONVENTION
                # y_row_a -> "N.I.P"
                # y_row_b -> "Nama"
                # y_row_c -> "Bidang Keahlian"
                # y_row_d -> "Alamat"
                # y_row_e -> "Instansi"
                # y_row_f -> "Email"
                # y_row_g -> "No. HP"
                # y_row_h -> "Peran"
                # y_row_i -> "Tugas"

                y_row_a = [l.replace('\r','').replace('\n','').strip()
                      for l in content.xpath(y_base + '/td[2]/text()')]

                y_row_b = [l.replace('\r','').replace('\n','').strip()
                      for l in content.xpath(y_base + '/td[3]/text()')]

                y_row_c = [l.replace('\r','').replace('\n','').strip()
                      for l in content.xpath(y_base + '/td[4]/text()')]

                y_row_d = [l.replace('\r','').replace('\n','').strip()
                      for l in content.xpath(y_base + '/td[5]/text()')]

                y_row_e = [l.replace('\r','').replace('\n','').strip()
                      for l in content.xpath(y_base + '/td[6]/text()')]

                y_row_f = [l.replace('\r','').replace('\n','').strip()
                      for l in content.xpath(y_base + '/td[7]/text()')]

                y_row_g = [l.replace('\r','').replace('\n','').strip()
                      for l in content.xpath(y_base + '/td[8]/text()')]

                y_row_h = [l.replace('\r','').replace('\n','').strip()
                      for l in content.xpath(y_base + '/td[9]/text()')]

                y_row_i = [l.replace('\r','').replace('\n','').strip()
                      for l in content.xpath(y_base + '/td[10]/text()')]

                # The starting column coordinate for filling the 'LUARAN' table data
                col_start = 25

                # Iterating through each table row and write to the spreadsheet
                # Assumes the lists y_row_a, y_row_b, ... have the same array size
                for j in range(len(y_row_a)):

                    # Painting table data
                    sheet.cell(row=row_start, column=col_start).value = y_row_a[j]
                    sheet.cell(row=row_start, column=col_start+1).value = y_row_b[j]
                    sheet.cell(row=row_start, column=col_start+2).value = y_row_c[j]
                    sheet.cell(row=row_start, column=col_start+3).value = y_row_d[j]
                    sheet.cell(row=row_start, column=col_start+4).value = y_row_e[j]
                    sheet.cell(row=row_start, column=col_start+5).value = y_row_f[j]
                    sheet.cell(row=row_start, column=col_start+6).value = y_row_g[j]
                    sheet.cell(row=row_start, column=col_start+7).value = y_row_h[j]
                    sheet.cell(row=row_start, column=col_start+8).value = y_row_i[j]

                    # Incrementing the 'col_start' iterator before continuing the loop
                    col_start += 9
                    continue

            # Incrementing the value of 'row_start' before continuing
            row_start += 1

            # Reopening the "Dana Penelitian" list page,
            # then assign the AJAX response to the temporary array 'temporary_prompt'
            # The 'data' array is obtained from opening individual entry row detail page
            #
            # This is done only on the 7-multiple row iteration, because
            # a single set of 'viewstate', 'viewstategen', and 'eventvalidation' values of ASPX
            # can only be used to do at most 8 operations before having to be renewed.
            if ( i % 7 == 0 ) and ( i > 0 ):
                temporary_prompt = self.get_risat_penelitian_berdirevisi_penelitian(data)
            continue

        # Post-loop logging: appending the header over the 'DATA ANGGOTA' columns
        control.append_message_area(f'+ Melengkapi kepala tabel pada bagian "Identitas Pengusul - Anggota Peneliti" ...')
        # ---
        # The starting column for the 'DATA ANGGOTA' data
        col_start = 25
        # Beginning the loop that detects the maximum number of 'LUARAN' rows
        # according to the variable 'y_max_row'
        for i in range(1, y_max_row+1):
            # Setting the top header
            sheet.merge_cells(
                start_row=1,
                start_column=col_start,
                end_row=1,
                end_column=col_start+8
            )
            sheet.cell(row=1, column=col_start).value = f'IDENTITAS PENGUSUL — ANGGOTA #{i}'
            sheet.cell(row=1, column=col_start).alignment = Alignment(horizontal='center')
            # Setting the sub headers
            sheet.cell(row=2, column=col_start).value = 'N.I.P'
            sheet.cell(row=2, column=col_start+1).value = 'Nama'
            sheet.cell(row=2, column=col_start+2).value = 'Bidang Keahlian'
            sheet.cell(row=2, column=col_start+3).value = 'Alamat'
            sheet.cell(row=2, column=col_start+4).value = 'Instansi'
            sheet.cell(row=2, column=col_start+5).value = 'Email'
            sheet.cell(row=2, column=col_start+6).value = 'No. HP'
            sheet.cell(row=2, column=col_start+7).value = 'Peran'
            sheet.cell(row=2, column=col_start+8).value = 'Tugas'
            # Incrementing the 'col_start' iterator before continuing the loop
            col_start += 9
            continue

        # Post-loop logging: successfully painted the output spreadsheet file
        control.append_message_area(f'+ Tabel sukses dipanen!')
        control.set_progress_bar(85)

        # Asking for the spreadsheet name to save as
        # ---
        # Logging and setting the progress bar
        control.append_message_area(f'+ Menyimpan spreadsheet luaran ...')
        control.set_progress_bar(90)
        # Dealing with file name prompt and saving
        # Using loop to mitigate the user clicking 'cancel'
        # in the file name dialog prompt
        while True:
            # Opening the dialog prompt
            output_spreadsheet = filedialog.asksaveasfilename(
                filetypes=[('Excel files', '*.xlsx')],
                initialfile='Sipesat - Berkas Direvisi Penelitian Detil Risat.xlsx',
                title='Simpan sebagai ...'
            )

            # 'cancel' button in the dialog prompt is clicked
            if len(output_spreadsheet) == 0:
                # Showing confirmation
                x = messagebox.askyesno(
                    'Nama File Kosong',
                    'Apakah Anda yakin ingin melanjutkan tanpa menyimpan file spreadsheet hasil pemanenan?'
                )
                # Determining whether to break or to continue the loop
                # based on the inversed value of 'x'
                if x:
                    control.append_message_area(
                        f'+ Finalisasi pemanenan data tanpa menyimpan file spreadsheet luaran ...')
                    workbook.close()  # --- closing the workbook without saving
                    break
                else:
                    continue  # --- continuing the loop
            # File name does not end in spreadsheet extension
            elif output_spreadsheet[-5:] != '.xlsx':
                output_spreadsheet = output_spreadsheet + '.xlsx'

            # Saving the spreadsheet
            control.append_message_area(f'LOKASI_SPREADSHEET_LUARAN: {output_spreadsheet}')
            control.set_progress_bar(95)
            workbook.save(output_spreadsheet)

            # Closing the openpyxl workbook
            control.append_message_area(f'+ Menutup file spreadsheet ...')
            control.set_progress_bar(98)
            workbook.close()

            # Breaking the loop
            break

        # Notify for a successful scraping
        control.append_message_area(f'+ Pemanenan selesai pada: {str(dt.now())}')
        control.set_progress_bar(100)
        control.on_notify_successful_scraping()

    # This function harvests "Risat Telah Direview Penelitian > Data Detil Lengkap" data
    # and then store the harvested data as an excel file
    #
    # Required arguments:
    # - control             --> for updating the progress bar and
    #                           message area of the screen SipesatScrHarvest
    # - username, password  --> the Risat administrator username and password
    def run_harvest_r_1_teldireview(self, control, username, password):
        # SipesatScrHarvest messenger
        control.set_header_desc('Panen Data "Risat Telah Direview Penelitian > Data Detil Lengkap"')
        control.set_help_label('Data sedang dipanen. Silahkan menunggu.')
        control.set_progress_bar(0)
        control.clear_message_area()

        # Preamble logging
        control.append_message_area(f'+ Memulai pemanenan data ...')
        control.append_message_area(f'+ Pemanenan dimulai pada: {str(dt.now())}')
        control.set_progress_bar(5)

        # Preparing the 'data_prompt' arrays
        control.append_message_area(f'+ Log masuk Risat sebagai [{username}] ...')
        control.set_progress_bar(10)
        data_prompt = self.get_risat_login(username, password)
        data_prompt = self.get_risat_penelitian(data_prompt)
        data_prompt = self.get_risat_penelitian_teldireview_penelitian(data_prompt)

        # Parsing XML tree content
        control.append_message_area(f'+ Membaca halaman web ...')
        control.set_progress_bar(15)
        content = data_prompt['html_content']

        # Establishing the export spreadsheet file
        control.append_message_area(f'+ Mempersiapkan file spreadsheet luaran ...')
        control.set_progress_bar(20)
        workbook = xl.Workbook()
        sheet = workbook.active
        sheet.title = 'Telah Direview Penelitian Detil'

        # DEVELOPER'S NOTE (DO NOT REMOVE)
        '''
        Data yang perlu diambil:
        1. Judul
        2. Substansi usulan
        3. Identitas pengusul - ketua
        4. Identitas pengusul - anggota
        5. RAB
        6. Dokumen pendukung
        7. Reviewer 1
        8. Reviewer 2

        Urutan kolom excel:
        1. Judul
        2. Substansi usulan
        3. RAB
        4. Dokumen pendukung
        5. Reviewer 1
        6. Reviewer 2
        7. Identitas pengusul - ketua
        8. Identitas pengusul - anggota
        '''

        # Preparing the sheet header
        control.append_message_area(f'+ Mempersiapkan kepala lembar spreadsheet ...')
        control.set_progress_bar(25)
        # ---
        # Preparing the "NOMOR" header
        sheet.merge_cells(
            start_row=1, start_column=1, end_row=2, end_column=1
        )
        sheet['A1'].value = 'No.'
        sheet['A1'].alignment = Alignment(horizontal='center')
        # ---
        # Preparing the "JUDUL" header
        sheet.merge_cells(
            start_row=1, start_column=2, end_row=1, end_column=15
        )
        sheet['B1'].value = 'IDENTITAS'
        sheet['B1'].alignment = Alignment(horizontal='center')
        # Preparing the "JUDUL" sub-headers
        sheet['B2'].value = 'Judul'
        sheet['C2'].value = 'Tgl. Usulan'
        sheet['D2'].value = 'TKT Saat Ini'
        sheet['E2'].value = 'Level'
        sheet['F2'].value = 'Kategori'
        sheet['G2'].value = 'Skema'
        sheet['H2'].value = 'Rumpun Ilmu'
        sheet['I2'].value = 'Bidang Fokus'
        sheet['J2'].value = 'Tema'
        sheet['K2'].value = 'Topik'
        sheet['L2'].value = 'Lama Kegiatan'
        sheet['M2'].value = 'Biaya Setelah Revisi'
        sheet['N2'].value = 'Catatan Revisi'
        sheet['O2'].value = 'File Revisi'
        # ---
        # Preparing the "SUBSTANSI USULAN" header
        sheet.merge_cells(
            start_row=1, start_column=16, end_row=1, end_column=17
        )
        sheet['P1'].value = 'SUBSTANSI USULAN'
        sheet['P1'].alignment = Alignment(horizontal='center')
        # Preparing the "SUBSTANSI USULAN" sub-headers
        sheet['P2'].value = 'Kelompok Makro'
        sheet['Q2'].value = 'File Proposal'
        # ---
        # Preparing the "RAB" header
        sheet.merge_cells(
            start_row=1, start_column=18, end_row=1, end_column=19
        )
        sheet['R1'].value = 'RENCANA ANGGARAN BIAYA'
        sheet['R1'].alignment = Alignment(horizontal='center')
        # Preparing the "RAB" sub-headers
        sheet['R2'].value = 'Biaya'
        sheet['S2'].value = 'File RAB'
        # ---
        # Preparing the "DOKUMEN PENDUKUNG" header
        sheet.merge_cells(
            start_row=1, start_column=20, end_row=1, end_column=22
        )
        sheet['T1'].value = 'DOKUMEN PENDUKUNG'
        sheet['T1'].alignment = Alignment(horizontal='center')
        # Preparing the "DOKUMEN PENDUKUNG" sub-headers
        sheet['T2'].value = 'Mitra'
        sheet['U2'].value = 'Dukungan Biaya'
        sheet['V2'].value = 'Surat Dukungan Mitra'
        # ---
        # Preparing the "REVIEWER 1" header
        sheet.merge_cells(
            start_row=1, start_column=23, end_row=1, end_column=26
        )
        sheet['W1'].value = 'REVIEWER 1'
        sheet['W1'].alignment = Alignment(horizontal='center')
        # Preparing the "REVIEWER 1" sub-headers
        sheet['W2'].value = 'Nama Reviewer'
        sheet['X2'].value = 'Nilai'
        sheet['Y2'].value = 'Rekomendasi Dana'
        sheet['Z2'].value = 'Komentar'
        # ---
        # Preparing the "REVIEWER 2" header
        sheet.merge_cells(
            start_row=1, start_column=27, end_row=1, end_column=30
        )
        sheet['AA1'].value = 'REVIEWER 2'
        sheet['AA1'].alignment = Alignment(horizontal='center')
        # Preparing the "REVIEWER 2" sub-headers
        sheet['AA2'].value = 'Nama Reviewer'
        sheet['AB2'].value = 'Nilai'
        sheet['AC2'].value = 'Rekomendasi Dana'
        sheet['AD2'].value = 'Komentar'
        # ---
        # Preparing the "IDENTITAS PENGUSUL — KETUA" header
        sheet.merge_cells(
            start_row=1, start_column=31, end_row=1, end_column=35
        )
        sheet['AE1'].value = 'IDENTITAS PENGUSUL — KETUA'
        sheet['AE1'].alignment = Alignment(horizontal='center')
        # Preparing the "IDENTITAS PENGUSUL — KETUA" sub-headers
        sheet['AE2'].value = 'N.I.P'
        sheet['AF2'].value = 'N.I.K'
        sheet['AG2'].value = 'N.I.D.N'
        sheet['AH2'].value = 'Nama Lengkap'
        sheet['AI2'].value = 'Jabatan Fungsional'

        # The base XPath location, pointing to each entry row
        base = '//div[@class="mw-100"]//div[@class="form-group f12"]/table[@width="100%"]//tr[@valign="top"]'

        # ---
        # Obtaining the data row values ("Ringkasan")
        control.append_message_area(f'+ Mendapatkan data pada baris tabel (ringkasan) ...')
        control.set_progress_bar(27)

        # HYPOTHESIS:
        # Xpath cannot detect 'tbody' element.
        # So instead of using 'table/tbody/tr', use 'table//tr' instead
        #
        # RESULT:
        # The hypothesis is correct.
        # Therefore, don't mention 'tbody' in any of the following Xpath paths

        m2 = [l.replace('Rp.', '').replace(',', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[10]/td/table//tr[1]/td[3]/text()')]

        n2 = [l.strip()
              for l in content.xpath(base + '/td[2]/table//tr[10]/td/table//tr[2]/td[3]/text()')]

        o2 = [l.strip()
              for l in content.xpath(base + '/td[2]/table//tr[10]/td/table//tr[3]/td[3]/a/@href')]

        # The 'Reviewer 1' all-content stripper
        rev_1 = [l.strip() for l in content.xpath(base + '/td[2]/table//tr[8]/td/table[@width="100%"]//tr[1]/td[1]/text()')]

        # Calculation for data pattern search
        # This is equal to the number of entries
        # 'rev_1_n' is always a multiple of 5
        rev_1_n = int(len(rev_1) / 5)

        w2 = []
        for i in range(rev_1_n):
            loc = 1 + (i * 5)  # --- the location of the data in the array
            w2.append(rev_1[loc].strip())

        x2 = []
        for i in range(rev_1_n):
            loc = 2 + (i * 5)  # --- the location of the data in the array
            x2.append(rev_1[loc].strip())

        y2 = []
        for i in range(rev_1_n):
            loc = 3 + (i * 5)  # --- the location of the data in the array
            y2.append(rev_1[loc].replace('Rp.', '').replace(',', '').strip())

        z2 = []
        for i in range(rev_1_n):
            loc = 4 + (i * 5)  # --- the location of the data in the array
            z2.append(rev_1[loc].strip())

        # The 'Reviewer 2' all-content stripper
        rev_2 = [l.strip() for l in content.xpath(base + '/td[2]/table//tr[8]/td/table[@width="100%"]//tr[1]/td[3]/text()')]

        # Calculation for data pattern search
        # This is equal to the number of entries
        # 'rev_2_n' is always a multiple of 5
        rev_2_n = int(len(rev_2) / 5)

        aa2 = []
        for i in range(rev_2_n):
            loc = 1 + (i * 5)  # --- the location of the data in the array
            aa2.append(rev_2[loc].strip())

        ab2 = []
        for i in range(rev_2_n):
            loc = 2 + (i * 5)  # --- the location of the data in the array
            ab2.append(rev_2[loc].strip())

        ac2 = []
        for i in range(rev_2_n):
            loc = 3 + (i * 5)  # --- the location of the data in the array
            ac2.append(rev_2[loc].replace('Rp.', '').replace(',', '').strip())

        ad2 = []
        for i in range(rev_2_n):
            loc = 4 + (i * 5)  # --- the location of the data in the array
            ad2.append(rev_2[loc].strip())

        # ---
        # Obtaining the data row values ("Detil")
        control.append_message_area(f'+ Mendapatkan data pada baris tabel (detil) ...')
        control.set_progress_bar(30)

        # Reading the HTML entry row hidden ASPX values
        # Copy-pasted from: /ssynthesia/ghostcity/ar/dumper-2/24__2023.02.13__requestsrisat.py
        all_kodetran_prop = content.xpath(base + '//input[1][@type="hidden"]/@name')
        all_kodetran_val = content.xpath(base + '//input[1][@type="hidden"]/@value')
        all_stat_prop = content.xpath(base + '//input[2][@type="hidden"]/@name')
        all_stat_val = content.xpath(base + '//input[2][@type="hidden"]/@value')
        all_submitbtn = content.xpath(base + '//input[@type="submit"][@value="Detil"]/@name')

        # DEBUG
        # Please comment out after use
        # ---
        # print('LENGTH_ALL_ASPX_VALUES', len(all_kodetran_prop), len(all_kodetran_val), len(all_stat_prop), len(all_stat_val), len(all_submitbtn))

        # The number of rows
        # Assumes the 'all_' array size equals the number of data rows
        number_of_row = len(all_kodetran_prop)

        # The starting row coordinate of the active sheet
        row_start = 3

        # The maximum number of 'DATA ANGGOTA' table data row
        y_max_row = 0

        # Iterating through each entry row element
        # Assumes all the 'all_' arrays in the previous code block
        # are of the same length/size
        # Copy-pasted from: /ssynthesia/ghostcity/ar/dumper-2/24__2023.02.13__requestsrisat.py
        control.append_message_area(f'+ Melakukan iterasi terhadap baris tabel dan menulis spreadsheet luaran ...')
        control.set_progress_bar(35)
        temporary_prompt = data_prompt
        for i in range(number_of_row):
            # Noisy preamble logging
            # Please don't use this -_-
            # ---
            # control.append_message_area(f'ITERASI [{i}]')

            # Updating the progress bar status
            control.set_progress_bar(35 + round(45 * (i + 1) / number_of_row))

            # Preparing the AJAX payload
            detail_prompt = {
                'viewstate': temporary_prompt['viewstate'],
                'viewstategen': temporary_prompt['viewstategen'],
                'eventvalidation': temporary_prompt['eventvalidation'],
                'button_name': all_submitbtn[i],
                'kodetran_prop': all_kodetran_prop[i],
                'kodetran_val': all_kodetran_val[i],
                'stat_prop': all_stat_prop[i],
                'stat_val': all_stat_val[i]
            }

            # Obtaining the response data of each individual entry row detail page
            data = self.get_risat_penelitian_teldireview_penelitian_detil(detail_prompt)
            content = data['html_content']
            response = data['http_response']

            # Terminal logging for detecting errors
            log_string_1 = str(i + 1)
            log_string_2 = content.xpath('//span[@id="ContentPlaceHolder1_daftar1_kiri1_txjudul1"]/text()')[0].replace('\r', '').replace('\n', '').strip()
            control.append_message_area(f'+ Memanen data detil: [{log_string_1}] {log_string_2} ...')
            print(f'+ Harvesting detailed data: [{log_string_1}] {log_string_2}')

            # DEVELOPER'S NOTE (DO NOT REMOVE)
            '''
            Variables and their associated data

            'a1' ==> 'No.'
            'b2' ==> 'Judul'
            'c2' ==> 'Tgl. Usulan'
            'd2' ==> 'TKT Saat Ini'
            'e2' ==> 'Level'
            'f2' ==> 'Kategori'
            'g2' ==> 'Skema'
            'h2' ==> 'Rumpun Ilmu'
            'i2' ==> 'Bidang Fokus'
            'j2' ==> 'Tema'
            'k2' ==> 'Topik'
            'l2' ==> 'Lama Kegiatan'
            'm2' ==> 'Biaya Setelah Revisi'
            'n2' ==> 'Catatan Revisi'
            'o2' ==> 'File Revisi'
            'p2' ==> 'Kelompok Makro'
            'q2' ==> 'File Proposal'
            'r2' ==> 'Biaya'
            's2' ==> 'File RAB'
            't2' ==> 'Mitra'
            'u2' ==> 'Dukungan Biaya'
            'v2' ==> 'Surat Dukungan Mitra'
            'w2' ==> 'Reviewer 1 - Nama Reviewer'
            'x2' ==> 'Reviewer 1 - Nilai'
            'y2' ==> 'Reviewer 1 - Rekomendasi Dana'
            'z2' ==> 'Reviewer 1 - Komentar'
            'aa2' ==> 'Reviewer 2 - Nama Reviewer'
            'ab2' ==> 'Reviewer 2 - Nilai'
            'ac2' ==> 'Reviewer 2 - Rekomendasi Dana'
            'ad2' ==> 'Reviewer 2 - Komentar'
            'ae2' ==> 'N.I.P'
            'af2' ==> 'N.I.K'
            'ag2' ==> 'N.I.D.N'
            'ah2' ==> 'Nama Lengkap'
            'ai2' ==> 'Jabatan Fungsional'

            In addition, variables starting with 'y' prefix are pertaining
            to the table data of "Data Anggota"
            '''

            # Getting the static (non-variable as in "Data Anggota") detail information
            # ---
            a1 = str(i+1)
            # Use try-except catching to mitigate empty data
            try:
                b2 = content.xpath('//span[@id="ContentPlaceHolder1_daftar1_kiri1_txjudul1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                b2 = ''
            # Use try-except catching to mitigate empty data
            try:
                c2 = content.xpath('//span[@id="ContentPlaceHolder1_daftar1_kiri1_tglusul1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                c2 = ''
            # Use try-except catching to mitigate empty data
            try:
                d2 = content.xpath('//span[@id="ContentPlaceHolder1_daftar1_kiri1_ttkt1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                d2 = ''
            # Use try-except catching to mitigate empty data
            try:
                e2 = content.xpath('//span[@id="ContentPlaceHolder1_daftar1_kiri1_ddllevel1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                e2 = ''
            # Use try-except catching to mitigate empty data
            try:
                f2 = content.xpath('//span[@id="ContentPlaceHolder1_daftar1_kiri1_lkat1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                f2 = ''
            # Use try-except catching to mitigate empty data
            try:
                g2 = content.xpath('//span[@id="ContentPlaceHolder1_daftar1_kiri1_ddlskema1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                g2 = ''
            # 'Rumpun Ilmu' scraped data contains multiple array of strings
            # This needs some extra tweaking
            h2 = ''
            # Use try-except catching to mitigate empty data
            try:
                h2_pre = content.xpath('//div[@id="ContentPlaceHolder1_daftar1_kiri1_updx"]/div[1][@class="panel panel-info"]/div[@class="panel-body f12"]/div[7][@class="row f12"]/div[@class="col-sm-4"]//text()')
                for l in h2_pre:
                    l = l.replace('\r','').replace('\n','').strip()
                    h2 += l +', '
                # Remove trailing ', ' characters
                h2 = h2[:-2]
            except IndexError:
                h2 = ''
            # ---
            # Use try-except catching to mitigate empty data
            try:
                i2 = content.xpath('//span[@id="ContentPlaceHolder1_daftar1_kiri1_ddlfokus1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                i2 = ''
            # Use try-except catching to mitigate empty data
            try:
                j2 = content.xpath('//span[@id="ContentPlaceHolder1_daftar1_kiri1_ddltema1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                j2 = ''
            # Use try-except catching to mitigate empty data
            try:
                k2 = content.xpath('//span[@id="ContentPlaceHolder1_daftar1_kiri1_ddltopik1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                k2 = ''
            # Use try-except catching to mitigate empty data
            try:
                l2 = content.xpath('//span[@id="ContentPlaceHolder1_daftar1_kiri1_ddllama1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                l2 = ''
            # Use try-except catching to mitigate empty data
            try:
                p2 = content.xpath('//span[@id="ContentPlaceHolder1_daftar1_kiri1_ddlmakro1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                p2 = ''
            # Use try-except catching to mitigate empty data
            try:
                q2 = content.xpath('//a[@id="ContentPlaceHolder1_daftar1_kiri1_alblcatatan1"]/@href')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                q2 = ''
            # Use try-except catching to mitigate empty data
            try:
                r2 = content.xpath('//span[@id="ContentPlaceHolder1_daftar1_kiri1_tbiayathn1"]/text()')[0].replace('\r','').replace('\n','').replace('Rp.', '').replace(',', '').strip()
            except IndexError:
                r2 = ''
            # Use try-except catching to mitigate empty data
            try:
                s2 = content.xpath('//a[@id="ContentPlaceHolder1_daftar1_kiri1_alblfilerab1"]/@href')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                s2 = ''
            # Some 'Data Detil' has no 'Mitra' data element
            # Use try-except catching to mitigate
            try:
                t2 = content.xpath('//span[@id="ContentPlaceHolder1_daftar1_kiri1_tnmmitra1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                t2 = ''
            # Use try-except catching to mitigate empty data
            try:
                u2 = content.xpath('//span[@id="ContentPlaceHolder1_daftar1_kiri1_tbiayadukung1"]/text()')[0].replace('\r','').replace('\n','').replace('Rp.', '').replace(',', '').strip()
            except IndexError:
                u2 = ''
            # Use try-except catching to mitigate empty data
            try:
                v2 = content.xpath('//a[@id="ContentPlaceHolder1_daftar1_kiri1_alblfilemitradukung1"]/@href')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                v2 = ''
            # Use try-except catching to mitigate empty data
            try:
                ae2 = content.xpath('//div[@id="ContentPlaceHolder1_daftar1_kiri1_bio2_updk1"]/div[1]/div[@class="col-sm-4"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                ae2 = ''
            # Use try-except catching to mitigate empty data
            try:
                af2 = content.xpath('//div[@id="ContentPlaceHolder1_daftar1_kiri1_bio2_updk1"]/div[2]/div[@class="col-sm-4"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                af2 = ''
            # Use try-except catching to mitigate empty data
            try:
                ag2 = content.xpath('//div[@id="ContentPlaceHolder1_daftar1_kiri1_bio2_updk1"]/div[3]/div[@class="col-sm-4"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                ag2 = ''
            # Use try-except catching to mitigate empty data
            try:
                ah2 = content.xpath('//div[@id="ContentPlaceHolder1_daftar1_kiri1_bio2_updk1"]/div[4]/div[@class="col-sm-4"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                ah2 = ''
            # Use try-except catching to mitigate empty data
            try:
                ai2 = content.xpath('//div[@id="ContentPlaceHolder1_daftar1_kiri1_bio2_updk1"]/div[5]/div[@class="col-sm-4"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                ai2 = ''

            # DEBUG
            # Please comment out after use
            # ---
            # print(a1, b2, c2, d2, e2, f2, g2, h2, i2, j2, k2, l2, m2[i], n2[i], o2[i], p2, q2, r2, s2, t2, u2, v2, w2[i], x2[i], y2[i], z2[i], aa2[i], ab2[i], ac2[i], ad2[i], ae2, af2, ag2, ah2, ai2)

            # Painting the static (non-variable as in "Data Anggota") detail information
            sheet[f'A{row_start}'].value = a1
            sheet[f'B{row_start}'].value = b2
            sheet[f'C{row_start}'].value = c2
            sheet[f'D{row_start}'].value = d2
            sheet[f'E{row_start}'].value = e2
            sheet[f'F{row_start}'].value = f2
            sheet[f'G{row_start}'].value = g2
            sheet[f'H{row_start}'].value = h2
            sheet[f'I{row_start}'].value = i2
            sheet[f'J{row_start}'].value = j2
            sheet[f'K{row_start}'].value = k2
            sheet[f'L{row_start}'].value = l2
            sheet[f'M{row_start}'].value = m2[i]
            sheet[f'N{row_start}'].value = n2[i]
            sheet[f'O{row_start}'].value = o2[i]
            sheet[f'P{row_start}'].value = p2
            sheet[f'Q{row_start}'].value = q2
            sheet[f'R{row_start}'].value = r2
            sheet[f'S{row_start}'].value = s2
            sheet[f'T{row_start}'].value = t2
            sheet[f'U{row_start}'].value = u2
            sheet[f'V{row_start}'].value = v2
            sheet[f'W{row_start}'].value = w2[i]
            sheet[f'X{row_start}'].value = x2[i]
            sheet[f'Y{row_start}'].value = y2[i]
            sheet[f'Z{row_start}'].value = z2[i]
            sheet[f'AA{row_start}'].value = aa2[i]
            sheet[f'AB{row_start}'].value = ab2[i]
            sheet[f'AC{row_start}'].value = ac2[i]
            sheet[f'AD{row_start}'].value = ad2[i]
            sheet[f'AE{row_start}'].value = ae2
            sheet[f'AF{row_start}'].value = af2
            sheet[f'AG{row_start}'].value = ag2
            sheet[f'AH{row_start}'].value = ah2
            sheet[f'AI{row_start}'].value = ai2

            # The table which displays 'Identitas Pengusul - Anggota Peneliti'
            # Treated differently, specially
            # 'y_' is a DOM element representing a HTML table row (<tr>)
            y_ = content.xpath(
                '//div[@id="ContentPlaceHolder1_daftar1_kiri1_updx"]/div[4]//table[@class="table"]//tr[position()>1]')

            # The base path
            y_base = '//div[@id="ContentPlaceHolder1_daftar1_kiri1_updx"]/div[4]//table[@class="table"]//tr[position()>1]'

            # Painting the scraped data to the output spreadsheet row
            # Section: 'DATA ANGGOTA'
            # ---
            # Checking if 'Identitas Pengusul - Anggota Peneliti' table data exists
            if len(y_) == 0:
                pass  # --- nope. the data does not exist
            else:

                # Checking if this table's 'LUARAN' data has the most rows
                y_all_rows = len(y_)
                y_max_row = max(y_max_row, y_all_rows)

                # ROW DATA LABEL CONVENTION
                # y_row_a -> "N.I.P"
                # y_row_b -> "Nama"
                # y_row_c -> "Bidang Keahlian"
                # y_row_d -> "Alamat"
                # y_row_e -> "Instansi"
                # y_row_f -> "Email"
                # y_row_g -> "No. HP"
                # y_row_h -> "Peran"
                # y_row_i -> "Tugas"

                y_row_a = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[2]/text()')]

                y_row_b = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[3]/text()')]

                y_row_c = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[4]/text()')]

                y_row_d = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[5]/text()')]

                y_row_e = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[6]/text()')]

                y_row_f = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[7]/text()')]

                y_row_g = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[8]/text()')]

                y_row_h = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[9]/text()')]

                y_row_i = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[10]/text()')]

                # The starting column coordinate for filling the 'LUARAN' table data
                col_start = 36

                # Iterating through each table row and write to the spreadsheet
                # Assumes the lists y_row_a, y_row_b, ... have the same array size
                for j in range(len(y_row_a)):
                    # Painting table data
                    sheet.cell(row=row_start, column=col_start).value = y_row_a[j]
                    sheet.cell(row=row_start, column=col_start + 1).value = y_row_b[j]
                    sheet.cell(row=row_start, column=col_start + 2).value = y_row_c[j]
                    sheet.cell(row=row_start, column=col_start + 3).value = y_row_d[j]
                    sheet.cell(row=row_start, column=col_start + 4).value = y_row_e[j]
                    sheet.cell(row=row_start, column=col_start + 5).value = y_row_f[j]
                    sheet.cell(row=row_start, column=col_start + 6).value = y_row_g[j]
                    sheet.cell(row=row_start, column=col_start + 7).value = y_row_h[j]
                    sheet.cell(row=row_start, column=col_start + 8).value = y_row_i[j]

                    # Incrementing the 'col_start' iterator before continuing the loop
                    col_start += 9
                    continue

            # Incrementing the value of 'row_start' before continuing
            row_start += 1

            # Reopening the "Berkas Telah Direview Penelitian" list page,
            # then assign the AJAX response to the temporary array 'temporary_prompt'
            # The 'data' array is obtained from opening individual entry row detail page
            #
            # This is done only on the 7-multiple row iteration, because
            # a single set of 'viewstate', 'viewstategen', and 'eventvalidation' values of ASPX
            # can only be used to do at most 8 operations before having to be renewed.
            if (i % 7 == 0) and (i > 0):
                temporary_prompt = self.get_risat_penelitian_teldireview_penelitian(data)
            continue

        # Post-loop logging: appending the header over the 'DATA ANGGOTA' columns
        control.append_message_area(
            f'+ Melengkapi kepala tabel pada bagian "Identitas Pengusul - Anggota Peneliti" ...')
        # ---
        # The starting column for the 'DATA ANGGOTA' data
        col_start = 36
        # Beginning the loop that detects the maximum number of 'LUARAN' rows
        # according to the variable 'y_max_row'
        for i in range(1, y_max_row + 1):
            # Setting the top header
            sheet.merge_cells(
                start_row=1,
                start_column=col_start,
                end_row=1,
                end_column=col_start + 8
            )
            sheet.cell(row=1, column=col_start).value = f'IDENTITAS PENGUSUL — ANGGOTA #{i}'
            sheet.cell(row=1, column=col_start).alignment = Alignment(horizontal='center')
            # Setting the sub headers
            sheet.cell(row=2, column=col_start).value = 'N.I.P'
            sheet.cell(row=2, column=col_start + 1).value = 'Nama'
            sheet.cell(row=2, column=col_start + 2).value = 'Bidang Keahlian'
            sheet.cell(row=2, column=col_start + 3).value = 'Alamat'
            sheet.cell(row=2, column=col_start + 4).value = 'Instansi'
            sheet.cell(row=2, column=col_start + 5).value = 'Email'
            sheet.cell(row=2, column=col_start + 6).value = 'No. HP'
            sheet.cell(row=2, column=col_start + 7).value = 'Peran'
            sheet.cell(row=2, column=col_start + 8).value = 'Tugas'
            # Incrementing the 'col_start' iterator before continuing the loop
            col_start += 9
            continue

        # Post-loop logging: successfully painted the output spreadsheet file
        control.append_message_area(f'+ Tabel sukses dipanen!')
        control.set_progress_bar(85)

        # Asking for the spreadsheet name to save as
        # ---
        # Logging and setting the progress bar
        control.append_message_area(f'+ Menyimpan spreadsheet luaran ...')
        control.set_progress_bar(90)
        # Dealing with file name prompt and saving
        # Using loop to mitigate the user clicking 'cancel'
        # in the file name dialog prompt
        while True:
            # Opening the dialog prompt
            output_spreadsheet = filedialog.asksaveasfilename(
                filetypes=[('Excel files', '*.xlsx')],
                initialfile='Sipesat - Telah Direview Penelitian Detil Risat.xlsx',
                title='Simpan sebagai ...'
            )

            # 'cancel' button in the dialog prompt is clicked
            if len(output_spreadsheet) == 0:
                # Showing confirmation
                x = messagebox.askyesno(
                    'Nama File Kosong',
                    'Apakah Anda yakin ingin melanjutkan tanpa menyimpan file spreadsheet hasil pemanenan?'
                )
                # Determining whether to break or to continue the loop
                # based on the inversed value of 'x'
                if x:
                    control.append_message_area(
                        f'+ Finalisasi pemanenan data tanpa menyimpan file spreadsheet luaran ...')
                    workbook.close()  # --- closing the workbook without saving
                    break
                else:
                    continue  # --- continuing the loop
            # File name does not end in spreadsheet extension
            elif output_spreadsheet[-5:] != '.xlsx':
                output_spreadsheet = output_spreadsheet + '.xlsx'

            # Saving the spreadsheet
            control.append_message_area(f'LOKASI_SPREADSHEET_LUARAN: {output_spreadsheet}')
            control.set_progress_bar(95)
            workbook.save(output_spreadsheet)

            # Closing the openpyxl workbook
            control.append_message_area(f'+ Menutup file spreadsheet ...')
            control.set_progress_bar(98)
            workbook.close()

            # Breaking the loop
            break

        # Notify for a successful scraping
        control.append_message_area(f'+ Pemanenan selesai pada: {str(dt.now())}')
        control.set_progress_bar(100)
        control.on_notify_successful_scraping()

    # This function harvests "Risat Disetujui Penelitian > Data Detil Lengkap" data
    # and then store the harvested data as an excel file
    #
    # Required arguments:
    # - control             --> for updating the progress bar and
    #                           message area of the screen SipesatScrHarvest
    # - username, password  --> the Risat administrator username and password
    def run_harvest_r_1_disetujui(self, control, username, password):
        # SipesatScrHarvest messenger
        control.set_header_desc('Panen Data "Risat Disetujui Penelitian > Data Detil Lengkap"')
        control.set_help_label('Data sedang dipanen. Silahkan menunggu.')
        control.set_progress_bar(0)
        control.clear_message_area()

        # Preamble logging
        control.append_message_area(f'+ Memulai pemanenan data ...')
        control.append_message_area(f'+ Pemanenan dimulai pada: {str(dt.now())}')
        control.set_progress_bar(5)

        # Preparing the 'data_prompt' arrays
        control.append_message_area(f'+ Log masuk Risat sebagai [{username}] ...')
        control.set_progress_bar(10)
        data_prompt = self.get_risat_login(username, password)
        data_prompt = self.get_risat_penelitian(data_prompt)
        data_prompt = self.get_risat_penelitian_disetujui_penelitian(data_prompt)

        # Parsing XML tree content
        control.append_message_area(f'+ Membaca halaman web ...')
        control.set_progress_bar(15)
        content = data_prompt['html_content']

        # Establishing the export spreadsheet file
        control.append_message_area(f'+ Mempersiapkan file spreadsheet luaran ...')
        control.set_progress_bar(20)
        workbook = xl.Workbook()
        sheet = workbook.active
        sheet.title = 'Disetujui Penelitian Detil'

        # DEVELOPER'S NOTE (DO NOT REMOVE)
        '''
        Data yang perlu diambil:
        1. Judul
        2. Substansi usulan
        3. Identitas pengusul - ketua
        4. Identitas pengusul - anggota
        5. RAB
        6. Dokumen pendukung
        7. Reviewer 1
        8. Reviewer 2

        Urutan kolom excel:
        1. Judul
        2. Substansi usulan
        3. RAB
        4. Dokumen pendukung
        5. Reviewer 1
        6. Reviewer 2
        7. Identitas pengusul - ketua
        8. Identitas pengusul - anggota
        '''

        # Preparing the sheet header
        control.append_message_area(f'+ Mempersiapkan kepala lembar spreadsheet ...')
        control.set_progress_bar(25)
        # ---
        # Preparing the "NOMOR" header
        sheet.merge_cells(
            start_row=1, start_column=1, end_row=2, end_column=1
        )
        sheet['A1'].value = 'No.'
        sheet['A1'].alignment = Alignment(horizontal='center')
        # ---
        # Preparing the "JUDUL" header
        sheet.merge_cells(
            start_row=1, start_column=2, end_row=1, end_column=15
        )
        sheet['B1'].value = 'IDENTITAS'
        sheet['B1'].alignment = Alignment(horizontal='center')
        # Preparing the "JUDUL" sub-headers
        sheet['B2'].value = 'Judul'
        sheet['C2'].value = 'Tgl. Usulan'
        sheet['D2'].value = 'TKT Saat Ini'
        sheet['E2'].value = 'Level'
        sheet['F2'].value = 'Kategori'
        sheet['G2'].value = 'Skema'
        sheet['H2'].value = 'Rumpun Ilmu'
        sheet['I2'].value = 'Bidang Fokus'
        sheet['J2'].value = 'Tema'
        sheet['K2'].value = 'Topik'
        sheet['L2'].value = 'Lama Kegiatan'
        sheet['M2'].value = 'Biaya Setelah Revisi'
        sheet['N2'].value = 'Catatan Revisi'
        sheet['O2'].value = 'File Revisi'
        # ---
        # Preparing the "SUBSTANSI USULAN" header
        sheet.merge_cells(
            start_row=1, start_column=16, end_row=1, end_column=17
        )
        sheet['P1'].value = 'SUBSTANSI USULAN'
        sheet['P1'].alignment = Alignment(horizontal='center')
        # Preparing the "SUBSTANSI USULAN" sub-headers
        sheet['P2'].value = 'Kelompok Makro'
        sheet['Q2'].value = 'File Proposal'
        # ---
        # Preparing the "RAB" header
        sheet.merge_cells(
            start_row=1, start_column=18, end_row=1, end_column=19
        )
        sheet['R1'].value = 'RENCANA ANGGARAN BIAYA'
        sheet['R1'].alignment = Alignment(horizontal='center')
        # Preparing the "RAB" sub-headers
        sheet['R2'].value = 'Biaya'
        sheet['S2'].value = 'File RAB'
        # ---
        # Preparing the "DOKUMEN PENDUKUNG" header
        sheet.merge_cells(
            start_row=1, start_column=20, end_row=1, end_column=22
        )
        sheet['T1'].value = 'DOKUMEN PENDUKUNG'
        sheet['T1'].alignment = Alignment(horizontal='center')
        # Preparing the "DOKUMEN PENDUKUNG" sub-headers
        sheet['T2'].value = 'Mitra'
        sheet['U2'].value = 'Dukungan Biaya'
        sheet['V2'].value = 'Surat Dukungan Mitra'
        # ---
        # Preparing the "REVIEWER 1" header
        sheet.merge_cells(
            start_row=1, start_column=23, end_row=1, end_column=26
        )
        sheet['W1'].value = 'REVIEWER 1'
        sheet['W1'].alignment = Alignment(horizontal='center')
        # Preparing the "REVIEWER 1" sub-headers
        sheet['W2'].value = 'Nama Reviewer'
        sheet['X2'].value = 'Nilai'
        sheet['Y2'].value = 'Rekomendasi Dana'
        sheet['Z2'].value = 'Komentar'
        # ---
        # Preparing the "REVIEWER 2" header
        sheet.merge_cells(
            start_row=1, start_column=27, end_row=1, end_column=30
        )
        sheet['AA1'].value = 'REVIEWER 2'
        sheet['AA1'].alignment = Alignment(horizontal='center')
        # Preparing the "REVIEWER 2" sub-headers
        sheet['AA2'].value = 'Nama Reviewer'
        sheet['AB2'].value = 'Nilai'
        sheet['AC2'].value = 'Rekomendasi Dana'
        sheet['AD2'].value = 'Komentar'
        # ---
        # Preparing the "IDENTITAS PENGUSUL — KETUA" header
        sheet.merge_cells(
            start_row=1, start_column=31, end_row=1, end_column=35
        )
        sheet['AE1'].value = 'IDENTITAS PENGUSUL — KETUA'
        sheet['AE1'].alignment = Alignment(horizontal='center')
        # Preparing the "IDENTITAS PENGUSUL — KETUA" sub-headers
        sheet['AE2'].value = 'N.I.P'
        sheet['AF2'].value = 'N.I.K'
        sheet['AG2'].value = 'N.I.D.N'
        sheet['AH2'].value = 'Nama Lengkap'
        sheet['AI2'].value = 'Jabatan Fungsional'

        # The base XPath location, pointing to each entry row
        base = '//div[@class="mw-100"]//div[@class="form-group f12"]/table[@width="100%"]//tr[@valign="top"]'

        # ---
        # Obtaining the data row values ("Ringkasan")
        control.append_message_area(f'+ Mendapatkan data pada baris tabel (ringkasan) ...')
        control.set_progress_bar(27)

        # HYPOTHESIS:
        # Xpath cannot detect 'tbody' element.
        # So instead of using 'table/tbody/tr', use 'table//tr' instead
        #
        # RESULT:
        # The hypothesis is correct.
        # Therefore, don't mention 'tbody' in any of the following Xpath paths

        m2 = [l.replace('Rp.', '').replace(',', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[10]/td/table//tr[1]/td[3]/text()')]

        n2 = [l.strip()
              for l in content.xpath(base + '/td[2]/table//tr[10]/td/table//tr[2]/td[3]/text()')]

        o2 = [l.strip()
              for l in content.xpath(base + '/td[2]/table//tr[10]/td/table//tr[3]/td[3]/a/@href')]

        # The 'Reviewer 1' all-content stripper
        rev_1 = [l.strip() for l in content.xpath(base + '/td[2]/table//tr[8]/td/table[@width="100%"]//tr[1]/td[1]/text()')]

        # Calculation for data pattern search
        # This is equal to the number of entries
        # 'rev_1_n' is always a multiple of 5
        rev_1_n = int(len(rev_1) / 5)

        w2 = []
        for i in range(rev_1_n):
            loc = 1 + (i * 5)  # --- the location of the data in the array
            w2.append(rev_1[loc].strip())

        x2 = []
        for i in range(rev_1_n):
            loc = 2 + (i * 5)  # --- the location of the data in the array
            x2.append(rev_1[loc].strip())

        y2 = []
        for i in range(rev_1_n):
            loc = 3 + (i * 5)  # --- the location of the data in the array
            y2.append(rev_1[loc].replace('Rp.', '').replace(',', '').strip())

        z2 = []
        for i in range(rev_1_n):
            loc = 4 + (i * 5)  # --- the location of the data in the array
            z2.append(rev_1[loc].strip())

        # The 'Reviewer 2' all-content stripper
        rev_2 = [l.strip() for l in content.xpath(base + '/td[2]/table//tr[8]/td/table[@width="100%"]//tr[1]/td[3]/text()')]

        # Calculation for data pattern search
        # This is equal to the number of entries
        # 'rev_2_n' is always a multiple of 5
        rev_2_n = int(len(rev_2) / 5)

        aa2 = []
        for i in range(rev_2_n):
            loc = 1 + (i * 5)  # --- the location of the data in the array
            aa2.append(rev_2[loc].strip())

        ab2 = []
        for i in range(rev_2_n):
            loc = 2 + (i * 5)  # --- the location of the data in the array
            ab2.append(rev_2[loc].strip())

        ac2 = []
        for i in range(rev_2_n):
            loc = 3 + (i * 5)  # --- the location of the data in the array
            ac2.append(rev_2[loc].replace('Rp.', '').replace(',', '').strip())

        ad2 = []
        for i in range(rev_2_n):
            loc = 4 + (i * 5)  # --- the location of the data in the array
            ad2.append(rev_2[loc].strip())

        # ---
        # Obtaining the data row values ("Detil")
        control.append_message_area(f'+ Mendapatkan data pada baris tabel (detil) ...')
        control.set_progress_bar(30)

        # Reading the HTML entry row hidden ASPX values
        # Copy-pasted from: /ssynthesia/ghostcity/ar/dumper-2/24__2023.02.13__requestsrisat.py
        all_kodetran_prop = content.xpath(base + '//input[1][@type="hidden"]/@name')
        all_kodetran_val = content.xpath(base + '//input[1][@type="hidden"]/@value')
        all_stat_prop = content.xpath(base + '//input[2][@type="hidden"]/@name')
        all_stat_val = content.xpath(base + '//input[2][@type="hidden"]/@value')
        all_submitbtn = content.xpath(base + '//input[@type="submit"][@value="Detil"]/@name')

        # DEBUG
        # Please comment out after use
        # ---
        # print('LENGTH_ALL_ASPX_VALUES', len(all_kodetran_prop), len(all_kodetran_val), len(all_stat_prop), len(all_stat_val), len(all_submitbtn))

        # The number of rows
        # Assumes the 'all_' array size equals the number of data rows
        number_of_row = len(all_kodetran_prop)

        # The starting row coordinate of the active sheet
        row_start = 3

        # The maximum number of 'DATA ANGGOTA' table data row
        y_max_row = 0

        # Iterating through each entry row element
        # Assumes all the 'all_' arrays in the previous code block
        # are of the same length/size
        # Copy-pasted from: /ssynthesia/ghostcity/ar/dumper-2/24__2023.02.13__requestsrisat.py
        control.append_message_area(f'+ Melakukan iterasi terhadap baris tabel dan menulis spreadsheet luaran ...')
        control.set_progress_bar(35)
        temporary_prompt = data_prompt
        for i in range(number_of_row):
            # Noisy preamble logging
            # Please don't use this -_-
            # ---
            # control.append_message_area(f'ITERASI [{i}]')

            # Updating the progress bar status
            control.set_progress_bar(35 + round(45 * (i + 1) / number_of_row))

            # Preparing the AJAX payload
            detail_prompt = {
                'viewstate': temporary_prompt['viewstate'],
                'viewstategen': temporary_prompt['viewstategen'],
                'eventvalidation': temporary_prompt['eventvalidation'],
                'button_name': all_submitbtn[i],
                'kodetran_prop': all_kodetran_prop[i],
                'kodetran_val': all_kodetran_val[i],
                'stat_prop': all_stat_prop[i],
                'stat_val': all_stat_val[i]
            }

            # Obtaining the response data of each individual entry row detail page
            data = self.get_risat_penelitian_disetujui_penelitian_detil(detail_prompt)
            content = data['html_content']
            response = data['http_response']

            # Terminal logging for detecting errors
            log_string_1 = str(i + 1)
            log_string_2 = content.xpath('//span[@id="ContentPlaceHolder1_daftar2_kiri1_txjudul1"]/text()')[0].replace('\r', '').replace('\n', '').strip()
            control.append_message_area(f'+ Memanen data detil: [{log_string_1}] {log_string_2} ...')
            print(f'+ Harvesting detailed data: [{log_string_1}] {log_string_2}')

            # DEVELOPER'S NOTE (DO NOT REMOVE)
            '''
            Variables and their associated data

            'a1' ==> 'No.'
            'b2' ==> 'Judul'
            'c2' ==> 'Tgl. Usulan'
            'd2' ==> 'TKT Saat Ini'
            'e2' ==> 'Level'
            'f2' ==> 'Kategori'
            'g2' ==> 'Skema'
            'h2' ==> 'Rumpun Ilmu'
            'i2' ==> 'Bidang Fokus'
            'j2' ==> 'Tema'
            'k2' ==> 'Topik'
            'l2' ==> 'Lama Kegiatan'
            'm2' ==> 'Biaya Setelah Revisi'
            'n2' ==> 'Catatan Revisi'
            'o2' ==> 'File Revisi'
            'p2' ==> 'Kelompok Makro'
            'q2' ==> 'File Proposal'
            'r2' ==> 'Biaya'
            's2' ==> 'File RAB'
            't2' ==> 'Mitra'
            'u2' ==> 'Dukungan Biaya'
            'v2' ==> 'Surat Dukungan Mitra'
            'w2' ==> 'Reviewer 1 - Nama Reviewer'
            'x2' ==> 'Reviewer 1 - Nilai'
            'y2' ==> 'Reviewer 1 - Rekomendasi Dana'
            'z2' ==> 'Reviewer 1 - Komentar'
            'aa2' ==> 'Reviewer 2 - Nama Reviewer'
            'ab2' ==> 'Reviewer 2 - Nilai'
            'ac2' ==> 'Reviewer 2 - Rekomendasi Dana'
            'ad2' ==> 'Reviewer 2 - Komentar'
            'ae2' ==> 'N.I.P'
            'af2' ==> 'N.I.K'
            'ag2' ==> 'N.I.D.N'
            'ah2' ==> 'Nama Lengkap'
            'ai2' ==> 'Jabatan Fungsional'

            In addition, variables starting with 'y' prefix are pertaining
            to the table data of "Data Anggota"
            '''

            # Getting the static (non-variable as in "Data Anggota") detail information
            # ---
            a1 = str(i+1)
            # Use try-except catching to mitigate empty data
            try:
                b2 = content.xpath('//span[@id="ContentPlaceHolder1_daftar2_kiri1_txjudul1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                b2 = ''
            # Use try-except catching to mitigate empty data
            try:
                c2 = content.xpath('//span[@id="ContentPlaceHolder1_daftar2_kiri1_tglusul1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                c2 = ''
            # Use try-except catching to mitigate empty data
            try:
                d2 = content.xpath('//span[@id="ContentPlaceHolder1_daftar2_kiri1_ttkt1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                d2 = ''
            # Use try-except catching to mitigate empty data
            try:
                e2 = content.xpath('//span[@id="ContentPlaceHolder1_daftar2_kiri1_ddllevel1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                e2 = ''
            # Use try-except catching to mitigate empty data
            try:
                f2 = content.xpath('//span[@id="ContentPlaceHolder1_daftar2_kiri1_lkat1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                f2 = ''
            # Use try-except catching to mitigate empty data
            try:
                g2 = content.xpath('//span[@id="ContentPlaceHolder1_daftar2_kiri1_ddlskema1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                g2 = ''
            # 'Rumpun Ilmu' scraped data contains multiple array of strings
            # This needs some extra tweaking
            h2 = ''
            # Use try-except catching to mitigate empty data
            try:
                h2_pre = content.xpath('//div[@id="ContentPlaceHolder1_daftar2_kiri1_updx"]/div[1][@class="panel panel-info"]/div[@class="panel-body f12"]/div[7][@class="row f12"]/div[@class="col-sm-4"]//text()')
                for l in h2_pre:
                    l = l.replace('\r','').replace('\n','').strip()
                    h2 += l +', '
                # Remove trailing ', ' characters
                h2 = h2[:-2]
            except IndexError:
                h2 = ''
            # ---
            # Use try-except catching to mitigate empty data
            try:
                i2 = content.xpath('//span[@id="ContentPlaceHolder1_daftar2_kiri1_ddlfokus1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                i2 = ''
            # Use try-except catching to mitigate empty data
            try:
                j2 = content.xpath('//span[@id="ContentPlaceHolder1_daftar2_kiri1_ddltema1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                j2 = ''
            # Use try-except catching to mitigate empty data
            try:
                k2 = content.xpath('//span[@id="ContentPlaceHolder1_daftar2_kiri1_ddltopik1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                k2 = ''
            # Use try-except catching to mitigate empty data
            try:
                l2 = content.xpath('//span[@id="ContentPlaceHolder1_daftar2_kiri1_ddllama1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                l2 = ''
            # Use try-except catching to mitigate empty data
            try:
                p2 = content.xpath('//span[@id="ContentPlaceHolder1_daftar2_kiri1_ddlmakro1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                p2 = ''
            # Use try-except catching to mitigate empty data
            try:
                q2 = content.xpath('//a[@id="ContentPlaceHolder1_daftar2_kiri1_alblcatatan1"]/@href')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                q2 = ''
            # Use try-except catching to mitigate empty data
            try:
                r2 = content.xpath('//span[@id="ContentPlaceHolder1_daftar2_kiri1_tbiayathn1"]/text()')[0].replace('\r','').replace('\n','').replace('Rp.', '').replace(',', '').strip()
            except IndexError:
                r2 = ''
            # Use try-except catching to mitigate empty data
            try:
                s2 = content.xpath('//a[@id="ContentPlaceHolder1_daftar2_kiri1_alblfilerab1"]/@href')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                s2 = ''
            # Some 'Data Detil' has no 'Mitra' data element
            # Use try-except catching to mitigate
            try:
                t2 = content.xpath('//span[@id="ContentPlaceHolder1_daftar2_kiri1_tnmmitra1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                t2 = ''
            # Use try-except catching to mitigate empty data
            try:
                u2 = content.xpath('//span[@id="ContentPlaceHolder1_daftar2_kiri1_tbiayadukung1"]/text()')[0].replace('\r','').replace('\n','').replace('Rp.', '').replace(',', '').strip()
            except IndexError:
                u2 = ''
            # Use try-except catching to mitigate empty data
            try:
                v2 = content.xpath('//a[@id="ContentPlaceHolder1_daftar2_kiri1_alblfilemitradukung1"]/@href')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                v2 = ''
            # Use try-except catching to mitigate empty data
            try:
                ae2 = content.xpath('//div[@id="ContentPlaceHolder1_daftar2_kiri1_bio2_updk1"]/div[1]/div[@class="col-sm-4"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                ae2 = ''
            # Use try-except catching to mitigate empty data
            try:
                af2 = content.xpath('//div[@id="ContentPlaceHolder1_daftar2_kiri1_bio2_updk1"]/div[2]/div[@class="col-sm-4"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                af2 = ''
            # Use try-except catching to mitigate empty data
            try:
                ag2 = content.xpath('//div[@id="ContentPlaceHolder1_daftar2_kiri1_bio2_updk1"]/div[3]/div[@class="col-sm-4"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                ag2 = ''
            # Use try-except catching to mitigate empty data
            try:
                ah2 = content.xpath('//div[@id="ContentPlaceHolder1_daftar2_kiri1_bio2_updk1"]/div[4]/div[@class="col-sm-4"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                ah2 = ''
            # Use try-except catching to mitigate empty data
            try:
                ai2 = content.xpath('//div[@id="ContentPlaceHolder1_daftar2_kiri1_bio2_updk1"]/div[5]/div[@class="col-sm-4"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                ai2 = ''

            # DEBUG
            # Please comment out after use
            # ---
            # print(a1, b2, c2, d2, e2, f2, g2, h2, i2, j2, k2, l2, m2[i], n2[i], o2[i], p2, q2, r2, s2, t2, u2, v2, w2[i], x2[i], y2[i], z2[i], aa2[i], ab2[i], ac2[i], ad2[i], ae2, af2, ag2, ah2, ai2)

            # Painting the static (non-variable as in "Data Anggota") detail information
            sheet[f'A{row_start}'].value = a1
            sheet[f'B{row_start}'].value = b2
            sheet[f'C{row_start}'].value = c2
            sheet[f'D{row_start}'].value = d2
            sheet[f'E{row_start}'].value = e2
            sheet[f'F{row_start}'].value = f2
            sheet[f'G{row_start}'].value = g2
            sheet[f'H{row_start}'].value = h2
            sheet[f'I{row_start}'].value = i2
            sheet[f'J{row_start}'].value = j2
            sheet[f'K{row_start}'].value = k2
            sheet[f'L{row_start}'].value = l2
            sheet[f'M{row_start}'].value = m2[i]
            sheet[f'N{row_start}'].value = n2[i]
            sheet[f'O{row_start}'].value = o2[i]
            sheet[f'P{row_start}'].value = p2
            sheet[f'Q{row_start}'].value = q2
            sheet[f'R{row_start}'].value = r2
            sheet[f'S{row_start}'].value = s2
            sheet[f'T{row_start}'].value = t2
            sheet[f'U{row_start}'].value = u2
            sheet[f'V{row_start}'].value = v2
            sheet[f'W{row_start}'].value = w2[i]
            sheet[f'X{row_start}'].value = x2[i]
            sheet[f'Y{row_start}'].value = y2[i]
            sheet[f'Z{row_start}'].value = z2[i]
            sheet[f'AA{row_start}'].value = aa2[i]
            sheet[f'AB{row_start}'].value = ab2[i]
            sheet[f'AC{row_start}'].value = ac2[i]
            sheet[f'AD{row_start}'].value = ad2[i]
            sheet[f'AE{row_start}'].value = ae2
            sheet[f'AF{row_start}'].value = af2
            sheet[f'AG{row_start}'].value = ag2
            sheet[f'AH{row_start}'].value = ah2
            sheet[f'AI{row_start}'].value = ai2

            # The table which displays 'Identitas Pengusul - Anggota Peneliti'
            # Treated differently, specially
            # 'y_' is a DOM element representing a HTML table row (<tr>)
            y_ = content.xpath(
                '//div[@id="ContentPlaceHolder1_daftar2_kiri1_updx"]/div[4]//table[@class="table"]//tr[position()>1]')

            # The base path
            y_base = '//div[@id="ContentPlaceHolder1_daftar2_kiri1_updx"]/div[4]//table[@class="table"]//tr[position()>1]'

            # Painting the scraped data to the output spreadsheet row
            # Section: 'DATA ANGGOTA'
            # ---
            # Checking if 'Identitas Pengusul - Anggota Peneliti' table data exists
            if len(y_) == 0:
                pass  # --- nope. the data does not exist
            else:

                # Checking if this table's 'LUARAN' data has the most rows
                y_all_rows = len(y_)
                y_max_row = max(y_max_row, y_all_rows)

                # ROW DATA LABEL CONVENTION
                # y_row_a -> "N.I.P"
                # y_row_b -> "Nama"
                # y_row_c -> "Bidang Keahlian"
                # y_row_d -> "Alamat"
                # y_row_e -> "Instansi"
                # y_row_f -> "Email"
                # y_row_g -> "No. HP"
                # y_row_h -> "Peran"
                # y_row_i -> "Tugas"

                y_row_a = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[2]/text()')]

                y_row_b = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[3]/text()')]

                y_row_c = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[4]/text()')]

                y_row_d = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[5]/text()')]

                y_row_e = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[6]/text()')]

                y_row_f = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[7]/text()')]

                y_row_g = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[8]/text()')]

                y_row_h = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[9]/text()')]

                y_row_i = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[10]/text()')]

                # The starting column coordinate for filling the 'LUARAN' table data
                col_start = 36

                # Iterating through each table row and write to the spreadsheet
                # Assumes the lists y_row_a, y_row_b, ... have the same array size
                for j in range(len(y_row_a)):
                    # Painting table data
                    sheet.cell(row=row_start, column=col_start).value = y_row_a[j]
                    sheet.cell(row=row_start, column=col_start + 1).value = y_row_b[j]
                    sheet.cell(row=row_start, column=col_start + 2).value = y_row_c[j]
                    sheet.cell(row=row_start, column=col_start + 3).value = y_row_d[j]
                    sheet.cell(row=row_start, column=col_start + 4).value = y_row_e[j]
                    sheet.cell(row=row_start, column=col_start + 5).value = y_row_f[j]
                    sheet.cell(row=row_start, column=col_start + 6).value = y_row_g[j]
                    sheet.cell(row=row_start, column=col_start + 7).value = y_row_h[j]
                    sheet.cell(row=row_start, column=col_start + 8).value = y_row_i[j]

                    # Incrementing the 'col_start' iterator before continuing the loop
                    col_start += 9
                    continue

            # Incrementing the value of 'row_start' before continuing
            row_start += 1

            # Reopening the "Berkas Disetujui Penelitian" list page,
            # then assign the AJAX response to the temporary array 'temporary_prompt'
            # The 'data' array is obtained from opening individual entry row detail page
            #
            # This is done only on the 7-multiple row iteration, because
            # a single set of 'viewstate', 'viewstategen', and 'eventvalidation' values of ASPX
            # can only be used to do at most 8 operations before having to be renewed.
            if (i % 7 == 0) and (i > 0):
                temporary_prompt = self.get_risat_penelitian_disetujui_penelitian(data)
            continue

        # Post-loop logging: appending the header over the 'DATA ANGGOTA' columns
        control.append_message_area(
            f'+ Melengkapi kepala tabel pada bagian "Identitas Pengusul - Anggota Peneliti" ...')
        # ---
        # The starting column for the 'DATA ANGGOTA' data
        col_start = 36
        # Beginning the loop that detects the maximum number of 'LUARAN' rows
        # according to the variable 'y_max_row'
        for i in range(1, y_max_row + 1):
            # Setting the top header
            sheet.merge_cells(
                start_row=1,
                start_column=col_start,
                end_row=1,
                end_column=col_start + 8
            )
            sheet.cell(row=1, column=col_start).value = f'IDENTITAS PENGUSUL — ANGGOTA #{i}'
            sheet.cell(row=1, column=col_start).alignment = Alignment(horizontal='center')
            # Setting the sub headers
            sheet.cell(row=2, column=col_start).value = 'N.I.P'
            sheet.cell(row=2, column=col_start + 1).value = 'Nama'
            sheet.cell(row=2, column=col_start + 2).value = 'Bidang Keahlian'
            sheet.cell(row=2, column=col_start + 3).value = 'Alamat'
            sheet.cell(row=2, column=col_start + 4).value = 'Instansi'
            sheet.cell(row=2, column=col_start + 5).value = 'Email'
            sheet.cell(row=2, column=col_start + 6).value = 'No. HP'
            sheet.cell(row=2, column=col_start + 7).value = 'Peran'
            sheet.cell(row=2, column=col_start + 8).value = 'Tugas'
            # Incrementing the 'col_start' iterator before continuing the loop
            col_start += 9
            continue

        # Post-loop logging: successfully painted the output spreadsheet file
        control.append_message_area(f'+ Tabel sukses dipanen!')
        control.set_progress_bar(85)

        # Asking for the spreadsheet name to save as
        # ---
        # Logging and setting the progress bar
        control.append_message_area(f'+ Menyimpan spreadsheet luaran ...')
        control.set_progress_bar(90)
        # Dealing with file name prompt and saving
        # Using loop to mitigate the user clicking 'cancel'
        # in the file name dialog prompt
        while True:
            # Opening the dialog prompt
            output_spreadsheet = filedialog.asksaveasfilename(
                filetypes=[('Excel files', '*.xlsx')],
                initialfile='Sipesat - Disetujui Penelitian Detil Risat.xlsx',
                title='Simpan sebagai ...'
            )

            # 'cancel' button in the dialog prompt is clicked
            if len(output_spreadsheet) == 0:
                # Showing confirmation
                x = messagebox.askyesno(
                    'Nama File Kosong',
                    'Apakah Anda yakin ingin melanjutkan tanpa menyimpan file spreadsheet hasil pemanenan?'
                )
                # Determining whether to break or to continue the loop
                # based on the inversed value of 'x'
                if x:
                    control.append_message_area(
                        f'+ Finalisasi pemanenan data tanpa menyimpan file spreadsheet luaran ...')
                    workbook.close()  # --- closing the workbook without saving
                    break
                else:
                    continue  # --- continuing the loop
            # File name does not end in spreadsheet extension
            elif output_spreadsheet[-5:] != '.xlsx':
                output_spreadsheet = output_spreadsheet + '.xlsx'

            # Saving the spreadsheet
            control.append_message_area(f'LOKASI_SPREADSHEET_LUARAN: {output_spreadsheet}')
            control.set_progress_bar(95)
            workbook.save(output_spreadsheet)

            # Closing the openpyxl workbook
            control.append_message_area(f'+ Menutup file spreadsheet ...')
            control.set_progress_bar(98)
            workbook.close()

            # Breaking the loop
            break

        # Notify for a successful scraping
        control.append_message_area(f'+ Pemanenan selesai pada: {str(dt.now())}')
        control.set_progress_bar(100)
        control.on_notify_successful_scraping()

    # This function harvests "Risat Disetujui Dgn Revisi (DDR) Penelitian > Data Detil Lengkap" data
    # and then store the harvested data as an excel file
    #
    # Required arguments:
    # - control             --> for updating the progress bar and
    #                           message area of the screen SipesatScrHarvest
    # - username, password  --> the Risat administrator username and password
    def run_harvest_r_1_ddr(self, control, username, password):
        # SipesatScrHarvest messenger
        control.set_header_desc('Panen Data "Risat Disetujui Dgn Revisi (DDR) Penelitian > Data Detil Lengkap"')
        control.set_help_label('Data sedang dipanen. Silahkan menunggu.')
        control.set_progress_bar(0)
        control.clear_message_area()

        # Preamble logging
        control.append_message_area(f'+ Memulai pemanenan data ...')
        control.append_message_area(f'+ Pemanenan dimulai pada: {str(dt.now())}')
        control.set_progress_bar(5)

        # Preparing the 'data_prompt' arrays
        control.append_message_area(f'+ Log masuk Risat sebagai [{username}] ...')
        control.set_progress_bar(10)
        data_prompt = self.get_risat_login(username, password)
        data_prompt = self.get_risat_penelitian(data_prompt)
        data_prompt = self.get_risat_penelitian_ddr_penelitian(data_prompt)

        # Parsing XML tree content
        control.append_message_area(f'+ Membaca halaman web ...')
        control.set_progress_bar(15)
        content = data_prompt['html_content']

        # Establishing the export spreadsheet file
        control.append_message_area(f'+ Mempersiapkan file spreadsheet luaran ...')
        control.set_progress_bar(20)
        workbook = xl.Workbook()
        sheet = workbook.active
        sheet.title = 'DDR Penelitian Detil'

        # DEVELOPER'S NOTE (DO NOT REMOVE)
        '''
        Data yang perlu diambil:
        1. Judul
        2. Substansi usulan
        3. Identitas pengusul - ketua
        4. Identitas pengusul - anggota
        5. RAB
        6. Dokumen pendukung
        7. Reviewer 1
        8. Reviewer 2

        Urutan kolom excel:
        1. Judul
        2. Substansi usulan
        3. RAB
        4. Dokumen pendukung
        5. Reviewer 1
        6. Reviewer 2
        7. Identitas pengusul - ketua
        8. Identitas pengusul - anggota
        '''

        # Preparing the sheet header
        control.append_message_area(f'+ Mempersiapkan kepala lembar spreadsheet ...')
        control.set_progress_bar(25)
        # ---
        # Preparing the "NOMOR" header
        sheet.merge_cells(
            start_row=1, start_column=1, end_row=2, end_column=1
        )
        sheet['A1'].value = 'No.'
        sheet['A1'].alignment = Alignment(horizontal='center')
        # ---
        # Preparing the "JUDUL" header
        sheet.merge_cells(
            start_row=1, start_column=2, end_row=1, end_column=15
        )
        sheet['B1'].value = 'IDENTITAS'
        sheet['B1'].alignment = Alignment(horizontal='center')
        # Preparing the "JUDUL" sub-headers
        sheet['B2'].value = 'Judul'
        sheet['C2'].value = 'Tgl. Usulan'
        sheet['D2'].value = 'TKT Saat Ini'
        sheet['E2'].value = 'Level'
        sheet['F2'].value = 'Kategori'
        sheet['G2'].value = 'Skema'
        sheet['H2'].value = 'Rumpun Ilmu'
        sheet['I2'].value = 'Bidang Fokus'
        sheet['J2'].value = 'Tema'
        sheet['K2'].value = 'Topik'
        sheet['L2'].value = 'Lama Kegiatan'
        sheet['M2'].value = 'Biaya Setelah Revisi'
        sheet['N2'].value = 'Catatan Revisi'
        sheet['O2'].value = 'File Revisi'
        # ---
        # Preparing the "SUBSTANSI USULAN" header
        sheet.merge_cells(
            start_row=1, start_column=16, end_row=1, end_column=17
        )
        sheet['P1'].value = 'SUBSTANSI USULAN'
        sheet['P1'].alignment = Alignment(horizontal='center')
        # Preparing the "SUBSTANSI USULAN" sub-headers
        sheet['P2'].value = 'Kelompok Makro'
        sheet['Q2'].value = 'File Proposal'
        # ---
        # Preparing the "RAB" header
        sheet.merge_cells(
            start_row=1, start_column=18, end_row=1, end_column=19
        )
        sheet['R1'].value = 'RENCANA ANGGARAN BIAYA'
        sheet['R1'].alignment = Alignment(horizontal='center')
        # Preparing the "RAB" sub-headers
        sheet['R2'].value = 'Biaya'
        sheet['S2'].value = 'File RAB'
        # ---
        # Preparing the "DOKUMEN PENDUKUNG" header
        sheet.merge_cells(
            start_row=1, start_column=20, end_row=1, end_column=22
        )
        sheet['T1'].value = 'DOKUMEN PENDUKUNG'
        sheet['T1'].alignment = Alignment(horizontal='center')
        # Preparing the "DOKUMEN PENDUKUNG" sub-headers
        sheet['T2'].value = 'Mitra'
        sheet['U2'].value = 'Dukungan Biaya'
        sheet['V2'].value = 'Surat Dukungan Mitra'
        # ---
        # Preparing the "REVIEWER 1" header
        sheet.merge_cells(
            start_row=1, start_column=23, end_row=1, end_column=26
        )
        sheet['W1'].value = 'REVIEWER 1'
        sheet['W1'].alignment = Alignment(horizontal='center')
        # Preparing the "REVIEWER 1" sub-headers
        sheet['W2'].value = 'Nama Reviewer'
        sheet['X2'].value = 'Nilai'
        sheet['Y2'].value = 'Rekomendasi Dana'
        sheet['Z2'].value = 'Komentar'
        # ---
        # Preparing the "REVIEWER 2" header
        sheet.merge_cells(
            start_row=1, start_column=27, end_row=1, end_column=30
        )
        sheet['AA1'].value = 'REVIEWER 2'
        sheet['AA1'].alignment = Alignment(horizontal='center')
        # Preparing the "REVIEWER 2" sub-headers
        sheet['AA2'].value = 'Nama Reviewer'
        sheet['AB2'].value = 'Nilai'
        sheet['AC2'].value = 'Rekomendasi Dana'
        sheet['AD2'].value = 'Komentar'
        # ---
        # Preparing the "IDENTITAS PENGUSUL — KETUA" header
        sheet.merge_cells(
            start_row=1, start_column=31, end_row=1, end_column=35
        )
        sheet['AE1'].value = 'IDENTITAS PENGUSUL — KETUA'
        sheet['AE1'].alignment = Alignment(horizontal='center')
        # Preparing the "IDENTITAS PENGUSUL — KETUA" sub-headers
        sheet['AE2'].value = 'N.I.P'
        sheet['AF2'].value = 'N.I.K'
        sheet['AG2'].value = 'N.I.D.N'
        sheet['AH2'].value = 'Nama Lengkap'
        sheet['AI2'].value = 'Jabatan Fungsional'

        # The base XPath location, pointing to each entry row
        base = '//div[@class="mw-100"]//div[@class="form-group f12"]/table[@width="100%"]//tr[@valign="top"]'

        # ---
        # Obtaining the data row values ("Ringkasan")
        control.append_message_area(f'+ Mendapatkan data pada baris tabel (ringkasan) ...')
        control.set_progress_bar(27)

        # HYPOTHESIS:
        # Xpath cannot detect 'tbody' element.
        # So instead of using 'table/tbody/tr', use 'table//tr' instead
        #
        # RESULT:
        # The hypothesis is correct.
        # Therefore, don't mention 'tbody' in any of the following Xpath paths

        m2 = [l.replace('Rp.', '').replace(',', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[10]/td/table//tr[1]/td[3]/text()')]

        n2 = [l.strip()
              for l in content.xpath(base + '/td[2]/table//tr[10]/td/table//tr[2]/td[3]/text()')]

        o2 = [l.strip()
              for l in content.xpath(base + '/td[2]/table//tr[10]/td/table//tr[3]/td[3]/a/@href')]

        # The 'Reviewer 1' all-content stripper
        rev_1 = [l.strip() for l in content.xpath(base + '/td[2]/table//tr[8]/td/table[@width="100%"]//tr[1]/td[1]/text()')]

        # Calculation for data pattern search
        # This is equal to the number of entries
        # 'rev_1_n' is always a multiple of 5
        rev_1_n = int(len(rev_1) / 5)

        w2 = []
        for i in range(rev_1_n):
            loc = 1 + (i * 5)  # --- the location of the data in the array
            w2.append(rev_1[loc].strip())

        x2 = []
        for i in range(rev_1_n):
            loc = 2 + (i * 5)  # --- the location of the data in the array
            x2.append(rev_1[loc].strip())

        y2 = []
        for i in range(rev_1_n):
            loc = 3 + (i * 5)  # --- the location of the data in the array
            y2.append(rev_1[loc].replace('Rp.', '').replace(',', '').strip())

        z2 = []
        for i in range(rev_1_n):
            loc = 4 + (i * 5)  # --- the location of the data in the array
            z2.append(rev_1[loc].strip())

        # The 'Reviewer 2' all-content stripper
        rev_2 = [l.strip() for l in content.xpath(base + '/td[2]/table//tr[8]/td/table[@width="100%"]//tr[1]/td[3]/text()')]

        # Calculation for data pattern search
        # This is equal to the number of entries
        # 'rev_2_n' is always a multiple of 5
        rev_2_n = int(len(rev_2) / 5)

        aa2 = []
        for i in range(rev_2_n):
            loc = 1 + (i * 5)  # --- the location of the data in the array
            aa2.append(rev_2[loc].strip())

        ab2 = []
        for i in range(rev_2_n):
            loc = 2 + (i * 5)  # --- the location of the data in the array
            ab2.append(rev_2[loc].strip())

        ac2 = []
        for i in range(rev_2_n):
            loc = 3 + (i * 5)  # --- the location of the data in the array
            ac2.append(rev_2[loc].replace('Rp.', '').replace(',', '').strip())

        ad2 = []
        for i in range(rev_2_n):
            loc = 4 + (i * 5)  # --- the location of the data in the array
            ad2.append(rev_2[loc].strip())

        # ---
        # Obtaining the data row values ("Detil")
        control.append_message_area(f'+ Mendapatkan data pada baris tabel (detil) ...')
        control.set_progress_bar(30)

        # Reading the HTML entry row hidden ASPX values
        # Copy-pasted from: /ssynthesia/ghostcity/ar/dumper-2/24__2023.02.13__requestsrisat.py
        all_kodetran_prop = content.xpath(base + '//input[1][@type="hidden"]/@name')
        all_kodetran_val = content.xpath(base + '//input[1][@type="hidden"]/@value')
        all_stat_prop = content.xpath(base + '//input[2][@type="hidden"]/@name')
        all_stat_val = content.xpath(base + '//input[2][@type="hidden"]/@value')
        all_submitbtn = content.xpath(base + '//input[@type="submit"][@value="Detil"]/@name')

        # DEBUG
        # Please comment out after use
        # ---
        # print('LENGTH_ALL_ASPX_VALUES', len(all_kodetran_prop), len(all_kodetran_val), len(all_stat_prop), len(all_stat_val), len(all_submitbtn))

        # The number of rows
        # Assumes the 'all_' array size equals the number of data rows
        number_of_row = len(all_kodetran_prop)

        # The starting row coordinate of the active sheet
        row_start = 3

        # The maximum number of 'DATA ANGGOTA' table data row
        y_max_row = 0

        # Iterating through each entry row element
        # Assumes all the 'all_' arrays in the previous code block
        # are of the same length/size
        # Copy-pasted from: /ssynthesia/ghostcity/ar/dumper-2/24__2023.02.13__requestsrisat.py
        control.append_message_area(f'+ Melakukan iterasi terhadap baris tabel dan menulis spreadsheet luaran ...')
        control.set_progress_bar(35)
        temporary_prompt = data_prompt
        for i in range(number_of_row):
            # Noisy preamble logging
            # Please don't use this -_-
            # ---
            # control.append_message_area(f'ITERASI [{i}]')

            # Updating the progress bar status
            control.set_progress_bar(35 + round(45 * (i + 1) / number_of_row))

            # Preparing the AJAX payload
            detail_prompt = {
                'viewstate': temporary_prompt['viewstate'],
                'viewstategen': temporary_prompt['viewstategen'],
                'eventvalidation': temporary_prompt['eventvalidation'],
                'button_name': all_submitbtn[i],
                'kodetran_prop': all_kodetran_prop[i],
                'kodetran_val': all_kodetran_val[i],
                'stat_prop': all_stat_prop[i],
                'stat_val': all_stat_val[i]
            }

            # Obtaining the response data of each individual entry row detail page
            data = self.get_risat_penelitian_ddr_penelitian_detil(detail_prompt)
            content = data['html_content']
            response = data['http_response']

            # Terminal logging for detecting errors
            log_string_1 = str(i + 1)
            log_string_2 = content.xpath('//span[@id="ContentPlaceHolder1_daftar3_kiri1_txjudul1"]/text()')[0].replace('\r', '').replace('\n', '').strip()
            control.append_message_area(f'+ Memanen data detil: [{log_string_1}] {log_string_2} ...')
            print(f'+ Harvesting detailed data: [{log_string_1}] {log_string_2}')

            # DEVELOPER'S NOTE (DO NOT REMOVE)
            '''
            Variables and their associated data

            'a1' ==> 'No.'
            'b2' ==> 'Judul'
            'c2' ==> 'Tgl. Usulan'
            'd2' ==> 'TKT Saat Ini'
            'e2' ==> 'Level'
            'f2' ==> 'Kategori'
            'g2' ==> 'Skema'
            'h2' ==> 'Rumpun Ilmu'
            'i2' ==> 'Bidang Fokus'
            'j2' ==> 'Tema'
            'k2' ==> 'Topik'
            'l2' ==> 'Lama Kegiatan'
            'm2' ==> 'Biaya Setelah Revisi'
            'n2' ==> 'Catatan Revisi'
            'o2' ==> 'File Revisi'
            'p2' ==> 'Kelompok Makro'
            'q2' ==> 'File Proposal'
            'r2' ==> 'Biaya'
            's2' ==> 'File RAB'
            't2' ==> 'Mitra'
            'u2' ==> 'Dukungan Biaya'
            'v2' ==> 'Surat Dukungan Mitra'
            'w2' ==> 'Reviewer 1 - Nama Reviewer'
            'x2' ==> 'Reviewer 1 - Nilai'
            'y2' ==> 'Reviewer 1 - Rekomendasi Dana'
            'z2' ==> 'Reviewer 1 - Komentar'
            'aa2' ==> 'Reviewer 2 - Nama Reviewer'
            'ab2' ==> 'Reviewer 2 - Nilai'
            'ac2' ==> 'Reviewer 2 - Rekomendasi Dana'
            'ad2' ==> 'Reviewer 2 - Komentar'
            'ae2' ==> 'N.I.P'
            'af2' ==> 'N.I.K'
            'ag2' ==> 'N.I.D.N'
            'ah2' ==> 'Nama Lengkap'
            'ai2' ==> 'Jabatan Fungsional'

            In addition, variables starting with 'y' prefix are pertaining
            to the table data of "Data Anggota"
            '''

            # Getting the static (non-variable as in "Data Anggota") detail information
            # ---
            a1 = str(i+1)
            # Use try-except catching to mitigate empty data
            try:
                b2 = content.xpath('//span[@id="ContentPlaceHolder1_daftar3_kiri1_txjudul1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                b2 = ''
            # Use try-except catching to mitigate empty data
            try:
                c2 = content.xpath('//span[@id="ContentPlaceHolder1_daftar3_kiri1_tglusul1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                c2 = ''
            # Use try-except catching to mitigate empty data
            try:
                d2 = content.xpath('//span[@id="ContentPlaceHolder1_daftar3_kiri1_ttkt1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                d2 = ''
            # Use try-except catching to mitigate empty data
            try:
                e2 = content.xpath('//span[@id="ContentPlaceHolder1_daftar3_kiri1_ddllevel1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                e2 = ''
            # Use try-except catching to mitigate empty data
            try:
                f2 = content.xpath('//span[@id="ContentPlaceHolder1_daftar3_kiri1_lkat1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                f2 = ''
            # Use try-except catching to mitigate empty data
            try:
                g2 = content.xpath('//span[@id="ContentPlaceHolder1_daftar3_kiri1_ddlskema1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                g2 = ''
            # 'Rumpun Ilmu' scraped data contains multiple array of strings
            # This needs some extra tweaking
            h2 = ''
            # Use try-except catching to mitigate empty data
            try:
                h2_pre = content.xpath('//div[@id="ContentPlaceHolder1_daftar3_kiri1_updx"]/div[1][@class="panel panel-info"]/div[@class="panel-body f12"]/div[7][@class="row f12"]/div[@class="col-sm-4"]//text()')
                for l in h2_pre:
                    l = l.replace('\r','').replace('\n','').strip()
                    h2 += l +', '
                # Remove trailing ', ' characters
                h2 = h2[:-2]
            except IndexError:
                h2 = ''
            # ---
            # Use try-except catching to mitigate empty data
            try:
                i2 = content.xpath('//span[@id="ContentPlaceHolder1_daftar3_kiri1_ddlfokus1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                i2 = ''
            # Use try-except catching to mitigate empty data
            try:
                j2 = content.xpath('//span[@id="ContentPlaceHolder1_daftar3_kiri1_ddltema1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                j2 = ''
            # Use try-except catching to mitigate empty data
            try:
                k2 = content.xpath('//span[@id="ContentPlaceHolder1_daftar3_kiri1_ddltopik1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                k2 = ''
            # Use try-except catching to mitigate empty data
            try:
                l2 = content.xpath('//span[@id="ContentPlaceHolder1_daftar3_kiri1_ddllama1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                l2 = ''
            # Use try-except catching to mitigate empty data
            try:
                p2 = content.xpath('//span[@id="ContentPlaceHolder1_daftar3_kiri1_ddlmakro1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                p2 = ''
            # Use try-except catching to mitigate empty data
            try:
                q2 = content.xpath('//a[@id="ContentPlaceHolder1_daftar3_kiri1_alblcatatan1"]/@href')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                q2 = ''
            # Use try-except catching to mitigate empty data
            try:
                r2 = content.xpath('//span[@id="ContentPlaceHolder1_daftar3_kiri1_tbiayathn1"]/text()')[0].replace('\r','').replace('\n','').replace('Rp.', '').replace(',', '').strip()
            except IndexError:
                r2 = ''
            # Use try-except catching to mitigate empty data
            try:
                s2 = content.xpath('//a[@id="ContentPlaceHolder1_daftar3_kiri1_alblfilerab1"]/@href')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                s2 = ''
            # Some 'Data Detil' has no 'Mitra' data element
            # Use try-except catching to mitigate
            try:
                t2 = content.xpath('//span[@id="ContentPlaceHolder1_daftar3_kiri1_tnmmitra1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                t2 = ''
            # Use try-except catching to mitigate empty data
            try:
                u2 = content.xpath('//span[@id="ContentPlaceHolder1_daftar3_kiri1_tbiayadukung1"]/text()')[0].replace('\r','').replace('\n','').replace('Rp.', '').replace(',', '').strip()
            except IndexError:
                u2 = ''
            # Use try-except catching to mitigate empty data
            try:
                v2 = content.xpath('//a[@id="ContentPlaceHolder1_daftar3_kiri1_alblfilemitradukung1"]/@href')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                v2 = ''
            # Use try-except catching to mitigate empty data
            try:
                ae2 = content.xpath('//div[@id="ContentPlaceHolder1_daftar3_kiri1_bio2_updk1"]/div[1]/div[@class="col-sm-4"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                ae2 = ''
            # Use try-except catching to mitigate empty data
            try:
                af2 = content.xpath('//div[@id="ContentPlaceHolder1_daftar3_kiri1_bio2_updk1"]/div[2]/div[@class="col-sm-4"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                af2 = ''
            # Use try-except catching to mitigate empty data
            try:
                ag2 = content.xpath('//div[@id="ContentPlaceHolder1_daftar3_kiri1_bio2_updk1"]/div[3]/div[@class="col-sm-4"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                ag2 = ''
            # Use try-except catching to mitigate empty data
            try:
                ah2 = content.xpath('//div[@id="ContentPlaceHolder1_daftar3_kiri1_bio2_updk1"]/div[4]/div[@class="col-sm-4"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                ah2 = ''
            # Use try-except catching to mitigate empty data
            try:
                ai2 = content.xpath('//div[@id="ContentPlaceHolder1_daftar3_kiri1_bio2_updk1"]/div[5]/div[@class="col-sm-4"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                ai2 = ''

            # DEBUG
            # Please comment out after use
            # ---
            # print(a1, b2, c2, d2, e2, f2, g2, h2, i2, j2, k2, l2, m2[i], n2[i], o2[i], p2, q2, r2, s2, t2, u2, v2, w2[i], x2[i], y2[i], z2[i], aa2[i], ab2[i], ac2[i], ad2[i], ae2, af2, ag2, ah2, ai2)

            # Painting the static (non-variable as in "Data Anggota") detail information
            sheet[f'A{row_start}'].value = a1
            sheet[f'B{row_start}'].value = b2
            sheet[f'C{row_start}'].value = c2
            sheet[f'D{row_start}'].value = d2
            sheet[f'E{row_start}'].value = e2
            sheet[f'F{row_start}'].value = f2
            sheet[f'G{row_start}'].value = g2
            sheet[f'H{row_start}'].value = h2
            sheet[f'I{row_start}'].value = i2
            sheet[f'J{row_start}'].value = j2
            sheet[f'K{row_start}'].value = k2
            sheet[f'L{row_start}'].value = l2
            sheet[f'M{row_start}'].value = m2[i]
            sheet[f'N{row_start}'].value = n2[i]
            sheet[f'O{row_start}'].value = o2[i]
            sheet[f'P{row_start}'].value = p2
            sheet[f'Q{row_start}'].value = q2
            sheet[f'R{row_start}'].value = r2
            sheet[f'S{row_start}'].value = s2
            sheet[f'T{row_start}'].value = t2
            sheet[f'U{row_start}'].value = u2
            sheet[f'V{row_start}'].value = v2
            sheet[f'W{row_start}'].value = w2[i]
            sheet[f'X{row_start}'].value = x2[i]
            sheet[f'Y{row_start}'].value = y2[i]
            sheet[f'Z{row_start}'].value = z2[i]
            sheet[f'AA{row_start}'].value = aa2[i]
            sheet[f'AB{row_start}'].value = ab2[i]
            sheet[f'AC{row_start}'].value = ac2[i]
            sheet[f'AD{row_start}'].value = ad2[i]
            sheet[f'AE{row_start}'].value = ae2
            sheet[f'AF{row_start}'].value = af2
            sheet[f'AG{row_start}'].value = ag2
            sheet[f'AH{row_start}'].value = ah2
            sheet[f'AI{row_start}'].value = ai2

            # The table which displays 'Identitas Pengusul - Anggota Peneliti'
            # Treated differently, specially
            # 'y_' is a DOM element representing a HTML table row (<tr>)
            y_ = content.xpath(
                '//div[@id="ContentPlaceHolder1_daftar3_kiri1_updx"]/div[4]//table[@class="table"]//tr[position()>1]')

            # The base path
            y_base = '//div[@id="ContentPlaceHolder1_daftar3_kiri1_updx"]/div[4]//table[@class="table"]//tr[position()>1]'

            # Painting the scraped data to the output spreadsheet row
            # Section: 'DATA ANGGOTA'
            # ---
            # Checking if 'Identitas Pengusul - Anggota Peneliti' table data exists
            if len(y_) == 0:
                pass  # --- nope. the data does not exist
            else:

                # Checking if this table's 'LUARAN' data has the most rows
                y_all_rows = len(y_)
                y_max_row = max(y_max_row, y_all_rows)

                # ROW DATA LABEL CONVENTION
                # y_row_a -> "N.I.P"
                # y_row_b -> "Nama"
                # y_row_c -> "Bidang Keahlian"
                # y_row_d -> "Alamat"
                # y_row_e -> "Instansi"
                # y_row_f -> "Email"
                # y_row_g -> "No. HP"
                # y_row_h -> "Peran"
                # y_row_i -> "Tugas"

                y_row_a = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[2]/text()')]

                y_row_b = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[3]/text()')]

                y_row_c = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[4]/text()')]

                y_row_d = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[5]/text()')]

                y_row_e = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[6]/text()')]

                y_row_f = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[7]/text()')]

                y_row_g = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[8]/text()')]

                y_row_h = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[9]/text()')]

                y_row_i = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[10]/text()')]

                # The starting column coordinate for filling the 'LUARAN' table data
                col_start = 36

                # Iterating through each table row and write to the spreadsheet
                # Assumes the lists y_row_a, y_row_b, ... have the same array size
                for j in range(len(y_row_a)):
                    # Painting table data
                    sheet.cell(row=row_start, column=col_start).value = y_row_a[j]
                    sheet.cell(row=row_start, column=col_start + 1).value = y_row_b[j]
                    sheet.cell(row=row_start, column=col_start + 2).value = y_row_c[j]
                    sheet.cell(row=row_start, column=col_start + 3).value = y_row_d[j]
                    sheet.cell(row=row_start, column=col_start + 4).value = y_row_e[j]
                    sheet.cell(row=row_start, column=col_start + 5).value = y_row_f[j]
                    sheet.cell(row=row_start, column=col_start + 6).value = y_row_g[j]
                    sheet.cell(row=row_start, column=col_start + 7).value = y_row_h[j]
                    sheet.cell(row=row_start, column=col_start + 8).value = y_row_i[j]

                    # Incrementing the 'col_start' iterator before continuing the loop
                    col_start += 9
                    continue

            # Incrementing the value of 'row_start' before continuing
            row_start += 1

            # Reopening the "Berkas Disetujui Dgn Revisi (DDR) Penelitian" list page,
            # then assign the AJAX response to the temporary array 'temporary_prompt'
            # The 'data' array is obtained from opening individual entry row detail page
            #
            # This is done only on the 7-multiple row iteration, because
            # a single set of 'viewstate', 'viewstategen', and 'eventvalidation' values of ASPX
            # can only be used to do at most 8 operations before having to be renewed.
            if (i % 7 == 0) and (i > 0):
                temporary_prompt = self.get_risat_penelitian_ddr_penelitian(data)
            continue

        # Post-loop logging: appending the header over the 'DATA ANGGOTA' columns
        control.append_message_area(
            f'+ Melengkapi kepala tabel pada bagian "Identitas Pengusul - Anggota Peneliti" ...')
        # ---
        # The starting column for the 'DATA ANGGOTA' data
        col_start = 36
        # Beginning the loop that detects the maximum number of 'LUARAN' rows
        # according to the variable 'y_max_row'
        for i in range(1, y_max_row + 1):
            # Setting the top header
            sheet.merge_cells(
                start_row=1,
                start_column=col_start,
                end_row=1,
                end_column=col_start + 8
            )
            sheet.cell(row=1, column=col_start).value = f'IDENTITAS PENGUSUL — ANGGOTA #{i}'
            sheet.cell(row=1, column=col_start).alignment = Alignment(horizontal='center')
            # Setting the sub headers
            sheet.cell(row=2, column=col_start).value = 'N.I.P'
            sheet.cell(row=2, column=col_start + 1).value = 'Nama'
            sheet.cell(row=2, column=col_start + 2).value = 'Bidang Keahlian'
            sheet.cell(row=2, column=col_start + 3).value = 'Alamat'
            sheet.cell(row=2, column=col_start + 4).value = 'Instansi'
            sheet.cell(row=2, column=col_start + 5).value = 'Email'
            sheet.cell(row=2, column=col_start + 6).value = 'No. HP'
            sheet.cell(row=2, column=col_start + 7).value = 'Peran'
            sheet.cell(row=2, column=col_start + 8).value = 'Tugas'
            # Incrementing the 'col_start' iterator before continuing the loop
            col_start += 9
            continue

        # Post-loop logging: successfully painted the output spreadsheet file
        control.append_message_area(f'+ Tabel sukses dipanen!')
        control.set_progress_bar(85)

        # Asking for the spreadsheet name to save as
        # ---
        # Logging and setting the progress bar
        control.append_message_area(f'+ Menyimpan spreadsheet luaran ...')
        control.set_progress_bar(90)
        # Dealing with file name prompt and saving
        # Using loop to mitigate the user clicking 'cancel'
        # in the file name dialog prompt
        while True:
            # Opening the dialog prompt
            output_spreadsheet = filedialog.asksaveasfilename(
                filetypes=[('Excel files', '*.xlsx')],
                initialfile='Sipesat - Disetujui Dgn Revisi (DDR) Penelitian Detil Risat.xlsx',
                title='Simpan sebagai ...'
            )

            # 'cancel' button in the dialog prompt is clicked
            if len(output_spreadsheet) == 0:
                # Showing confirmation
                x = messagebox.askyesno(
                    'Nama File Kosong',
                    'Apakah Anda yakin ingin melanjutkan tanpa menyimpan file spreadsheet hasil pemanenan?'
                )
                # Determining whether to break or to continue the loop
                # based on the inversed value of 'x'
                if x:
                    control.append_message_area(
                        f'+ Finalisasi pemanenan data tanpa menyimpan file spreadsheet luaran ...')
                    workbook.close()  # --- closing the workbook without saving
                    break
                else:
                    continue  # --- continuing the loop
            # File name does not end in spreadsheet extension
            elif output_spreadsheet[-5:] != '.xlsx':
                output_spreadsheet = output_spreadsheet + '.xlsx'

            # Saving the spreadsheet
            control.append_message_area(f'LOKASI_SPREADSHEET_LUARAN: {output_spreadsheet}')
            control.set_progress_bar(95)
            workbook.save(output_spreadsheet)

            # Closing the openpyxl workbook
            control.append_message_area(f'+ Menutup file spreadsheet ...')
            control.set_progress_bar(98)
            workbook.close()

            # Breaking the loop
            break

        # Notify for a successful scraping
        control.append_message_area(f'+ Pemanenan selesai pada: {str(dt.now())}')
        control.set_progress_bar(100)
        control.on_notify_successful_scraping()

    # This function harvests "Risat Ditolak Penelitian > Data Detil Lengkap" data
    # and then store the harvested data as an excel file
    #
    # Required arguments:
    # - control             --> for updating the progress bar and
    #                           message area of the screen SipesatScrHarvest
    # - username, password  --> the Risat administrator username and password
    def run_harvest_r_1_ditolak(self, control, username, password):
        # SipesatScrHarvest messenger
        control.set_header_desc('Panen Data "Risat Ditolak Penelitian > Data Detil Lengkap"')
        control.set_help_label('Data sedang dipanen. Silahkan menunggu.')
        control.set_progress_bar(0)
        control.clear_message_area()

        # Preamble logging
        control.append_message_area(f'+ Memulai pemanenan data ...')
        control.append_message_area(f'+ Pemanenan dimulai pada: {str(dt.now())}')
        control.set_progress_bar(5)

        # Preparing the 'data_prompt' arrays
        control.append_message_area(f'+ Log masuk Risat sebagai [{username}] ...')
        control.set_progress_bar(10)
        data_prompt = self.get_risat_login(username, password)
        data_prompt = self.get_risat_penelitian(data_prompt)
        data_prompt = self.get_risat_penelitian_ditolak_penelitian(data_prompt)

        # Parsing XML tree content
        control.append_message_area(f'+ Membaca halaman web ...')
        control.set_progress_bar(15)
        content = data_prompt['html_content']

        # Establishing the export spreadsheet file
        control.append_message_area(f'+ Mempersiapkan file spreadsheet luaran ...')
        control.set_progress_bar(20)
        workbook = xl.Workbook()
        sheet = workbook.active
        sheet.title = 'Ditolak Penelitian Detil'

        # DEVELOPER'S NOTE (DO NOT REMOVE)
        '''
        Data yang perlu diambil:
        1. Judul
        2. Substansi usulan
        3. Identitas pengusul - ketua
        4. Identitas pengusul - anggota
        5. RAB
        6. Dokumen pendukung
        7. Reviewer 1
        8. Reviewer 2

        Urutan kolom excel:
        1. Judul
        2. Substansi usulan
        3. RAB
        4. Dokumen pendukung
        5. Reviewer 1
        6. Reviewer 2
        7. Identitas pengusul - ketua
        8. Identitas pengusul - anggota
        '''

        # Preparing the sheet header
        control.append_message_area(f'+ Mempersiapkan kepala lembar spreadsheet ...')
        control.set_progress_bar(25)
        # ---
        # Preparing the "NOMOR" header
        sheet.merge_cells(
            start_row=1, start_column=1, end_row=2, end_column=1
        )
        sheet['A1'].value = 'No.'
        sheet['A1'].alignment = Alignment(horizontal='center')
        # ---
        # Preparing the "JUDUL" header
        sheet.merge_cells(
            start_row=1, start_column=2, end_row=1, end_column=15
        )
        sheet['B1'].value = 'IDENTITAS'
        sheet['B1'].alignment = Alignment(horizontal='center')
        # Preparing the "JUDUL" sub-headers
        sheet['B2'].value = 'Judul'
        sheet['C2'].value = 'Tgl. Usulan'
        sheet['D2'].value = 'TKT Saat Ini'
        sheet['E2'].value = 'Level'
        sheet['F2'].value = 'Kategori'
        sheet['G2'].value = 'Skema'
        sheet['H2'].value = 'Rumpun Ilmu'
        sheet['I2'].value = 'Bidang Fokus'
        sheet['J2'].value = 'Tema'
        sheet['K2'].value = 'Topik'
        sheet['L2'].value = 'Lama Kegiatan'
        sheet['M2'].value = 'Biaya Setelah Revisi'
        sheet['N2'].value = 'Catatan Revisi'
        sheet['O2'].value = 'File Revisi'
        # ---
        # Preparing the "SUBSTANSI USULAN" header
        sheet.merge_cells(
            start_row=1, start_column=16, end_row=1, end_column=17
        )
        sheet['P1'].value = 'SUBSTANSI USULAN'
        sheet['P1'].alignment = Alignment(horizontal='center')
        # Preparing the "SUBSTANSI USULAN" sub-headers
        sheet['P2'].value = 'Kelompok Makro'
        sheet['Q2'].value = 'File Proposal'
        # ---
        # Preparing the "RAB" header
        sheet.merge_cells(
            start_row=1, start_column=18, end_row=1, end_column=19
        )
        sheet['R1'].value = 'RENCANA ANGGARAN BIAYA'
        sheet['R1'].alignment = Alignment(horizontal='center')
        # Preparing the "RAB" sub-headers
        sheet['R2'].value = 'Biaya'
        sheet['S2'].value = 'File RAB'
        # ---
        # Preparing the "DOKUMEN PENDUKUNG" header
        sheet.merge_cells(
            start_row=1, start_column=20, end_row=1, end_column=22
        )
        sheet['T1'].value = 'DOKUMEN PENDUKUNG'
        sheet['T1'].alignment = Alignment(horizontal='center')
        # Preparing the "DOKUMEN PENDUKUNG" sub-headers
        sheet['T2'].value = 'Mitra'
        sheet['U2'].value = 'Dukungan Biaya'
        sheet['V2'].value = 'Surat Dukungan Mitra'
        # ---
        # Preparing the "REVIEWER 1" header
        sheet.merge_cells(
            start_row=1, start_column=23, end_row=1, end_column=26
        )
        sheet['W1'].value = 'REVIEWER 1'
        sheet['W1'].alignment = Alignment(horizontal='center')
        # Preparing the "REVIEWER 1" sub-headers
        sheet['W2'].value = 'Nama Reviewer'
        sheet['X2'].value = 'Nilai'
        sheet['Y2'].value = 'Rekomendasi Dana'
        sheet['Z2'].value = 'Komentar'
        # ---
        # Preparing the "REVIEWER 2" header
        sheet.merge_cells(
            start_row=1, start_column=27, end_row=1, end_column=30
        )
        sheet['AA1'].value = 'REVIEWER 2'
        sheet['AA1'].alignment = Alignment(horizontal='center')
        # Preparing the "REVIEWER 2" sub-headers
        sheet['AA2'].value = 'Nama Reviewer'
        sheet['AB2'].value = 'Nilai'
        sheet['AC2'].value = 'Rekomendasi Dana'
        sheet['AD2'].value = 'Komentar'
        # ---
        # Preparing the "IDENTITAS PENGUSUL — KETUA" header
        sheet.merge_cells(
            start_row=1, start_column=31, end_row=1, end_column=35
        )
        sheet['AE1'].value = 'IDENTITAS PENGUSUL — KETUA'
        sheet['AE1'].alignment = Alignment(horizontal='center')
        # Preparing the "IDENTITAS PENGUSUL — KETUA" sub-headers
        sheet['AE2'].value = 'N.I.P'
        sheet['AF2'].value = 'N.I.K'
        sheet['AG2'].value = 'N.I.D.N'
        sheet['AH2'].value = 'Nama Lengkap'
        sheet['AI2'].value = 'Jabatan Fungsional'

        # The base XPath location, pointing to each entry row
        base = '//div[@class="mw-100"]//div[@class="form-group f12"]/table[@width="100%"]//tr[@valign="top"]'

        # ---
        # Obtaining the data row values ("Ringkasan")
        control.append_message_area(f'+ Mendapatkan data pada baris tabel (ringkasan) ...')
        control.set_progress_bar(27)

        # HYPOTHESIS:
        # Xpath cannot detect 'tbody' element.
        # So instead of using 'table/tbody/tr', use 'table//tr' instead
        #
        # RESULT:
        # The hypothesis is correct.
        # Therefore, don't mention 'tbody' in any of the following Xpath paths

        m2 = [l.replace('Rp.', '').replace(',', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[10]/td/table//tr[1]/td[3]/text()')]

        n2 = [l.strip()
              for l in content.xpath(base + '/td[2]/table//tr[10]/td/table//tr[2]/td[3]/text()')]

        o2 = [l.strip()
              for l in content.xpath(base + '/td[2]/table//tr[10]/td/table//tr[3]/td[3]/a/@href')]

        # The 'Reviewer 1' all-content stripper
        rev_1 = [l.strip() for l in content.xpath(base + '/td[2]/table//tr[8]/td/table[@width="100%"]//tr[1]/td[1]/text()')]

        # Calculation for data pattern search
        # This is equal to the number of entries
        # 'rev_1_n' is always a multiple of 5
        rev_1_n = int(len(rev_1) / 5)

        w2 = []
        for i in range(rev_1_n):
            loc = 1 + (i * 5)  # --- the location of the data in the array
            w2.append(rev_1[loc].strip())

        x2 = []
        for i in range(rev_1_n):
            loc = 2 + (i * 5)  # --- the location of the data in the array
            x2.append(rev_1[loc].strip())

        y2 = []
        for i in range(rev_1_n):
            loc = 3 + (i * 5)  # --- the location of the data in the array
            y2.append(rev_1[loc].replace('Rp.', '').replace(',', '').strip())

        z2 = []
        for i in range(rev_1_n):
            loc = 4 + (i * 5)  # --- the location of the data in the array
            z2.append(rev_1[loc].strip())

        # The 'Reviewer 2' all-content stripper
        rev_2 = [l.strip() for l in content.xpath(base + '/td[2]/table//tr[8]/td/table[@width="100%"]//tr[1]/td[3]/text()')]

        # Calculation for data pattern search
        # This is equal to the number of entries
        # 'rev_2_n' is always a multiple of 5
        rev_2_n = int(len(rev_2) / 5)

        aa2 = []
        for i in range(rev_2_n):
            loc = 1 + (i * 5)  # --- the location of the data in the array
            aa2.append(rev_2[loc].strip())

        ab2 = []
        for i in range(rev_2_n):
            loc = 2 + (i * 5)  # --- the location of the data in the array
            ab2.append(rev_2[loc].strip())

        ac2 = []
        for i in range(rev_2_n):
            loc = 3 + (i * 5)  # --- the location of the data in the array
            ac2.append(rev_2[loc].replace('Rp.', '').replace(',', '').strip())

        ad2 = []
        for i in range(rev_2_n):
            loc = 4 + (i * 5)  # --- the location of the data in the array
            ad2.append(rev_2[loc].strip())

        # ---
        # Obtaining the data row values ("Detil")
        control.append_message_area(f'+ Mendapatkan data pada baris tabel (detil) ...')
        control.set_progress_bar(30)

        # Reading the HTML entry row hidden ASPX values
        # Copy-pasted from: /ssynthesia/ghostcity/ar/dumper-2/24__2023.02.13__requestsrisat.py
        all_kodetran_prop = content.xpath(base + '//input[1][@type="hidden"]/@name')
        all_kodetran_val = content.xpath(base + '//input[1][@type="hidden"]/@value')
        all_stat_prop = content.xpath(base + '//input[2][@type="hidden"]/@name')
        all_stat_val = content.xpath(base + '//input[2][@type="hidden"]/@value')
        all_submitbtn = content.xpath(base + '//input[@type="submit"][@value="Detil"]/@name')

        # DEBUG
        # Please comment out after use
        # ---
        # print('LENGTH_ALL_ASPX_VALUES', len(all_kodetran_prop), len(all_kodetran_val), len(all_stat_prop), len(all_stat_val), len(all_submitbtn))

        # The number of rows
        # Assumes the 'all_' array size equals the number of data rows
        number_of_row = len(all_kodetran_prop)

        # The starting row coordinate of the active sheet
        row_start = 3

        # The maximum number of 'DATA ANGGOTA' table data row
        y_max_row = 0

        # Iterating through each entry row element
        # Assumes all the 'all_' arrays in the previous code block
        # are of the same length/size
        # Copy-pasted from: /ssynthesia/ghostcity/ar/dumper-2/24__2023.02.13__requestsrisat.py
        control.append_message_area(f'+ Melakukan iterasi terhadap baris tabel dan menulis spreadsheet luaran ...')
        control.set_progress_bar(35)
        temporary_prompt = data_prompt
        for i in range(number_of_row):
            # Noisy preamble logging
            # Please don't use this -_-
            # ---
            # control.append_message_area(f'ITERASI [{i}]')

            # Updating the progress bar status
            control.set_progress_bar(35 + round(45 * (i + 1) / number_of_row))

            # Preparing the AJAX payload
            detail_prompt = {
                'viewstate': temporary_prompt['viewstate'],
                'viewstategen': temporary_prompt['viewstategen'],
                'eventvalidation': temporary_prompt['eventvalidation'],
                'button_name': all_submitbtn[i],
                'kodetran_prop': all_kodetran_prop[i],
                'kodetran_val': all_kodetran_val[i],
                'stat_prop': all_stat_prop[i],
                'stat_val': all_stat_val[i]
            }

            # Obtaining the response data of each individual entry row detail page
            data = self.get_risat_penelitian_ditolak_penelitian_detil(detail_prompt)
            content = data['html_content']
            response = data['http_response']

            # Terminal logging for detecting errors
            log_string_1 = str(i + 1)
            log_string_2 = content.xpath('//span[@id="ContentPlaceHolder1_daftar4_kiri1_txjudul1"]/text()')[0].replace('\r', '').replace('\n', '').strip()
            control.append_message_area(f'+ Memanen data detil: [{log_string_1}] {log_string_2} ...')
            print(f'+ Harvesting detailed data: [{log_string_1}] {log_string_2}')

            # DEVELOPER'S NOTE (DO NOT REMOVE)
            '''
            Variables and their associated data

            'a1' ==> 'No.'
            'b2' ==> 'Judul'
            'c2' ==> 'Tgl. Usulan'
            'd2' ==> 'TKT Saat Ini'
            'e2' ==> 'Level'
            'f2' ==> 'Kategori'
            'g2' ==> 'Skema'
            'h2' ==> 'Rumpun Ilmu'
            'i2' ==> 'Bidang Fokus'
            'j2' ==> 'Tema'
            'k2' ==> 'Topik'
            'l2' ==> 'Lama Kegiatan'
            'm2' ==> 'Biaya Setelah Revisi'
            'n2' ==> 'Catatan Revisi'
            'o2' ==> 'File Revisi'
            'p2' ==> 'Kelompok Makro'
            'q2' ==> 'File Proposal'
            'r2' ==> 'Biaya'
            's2' ==> 'File RAB'
            't2' ==> 'Mitra'
            'u2' ==> 'Dukungan Biaya'
            'v2' ==> 'Surat Dukungan Mitra'
            'w2' ==> 'Reviewer 1 - Nama Reviewer'
            'x2' ==> 'Reviewer 1 - Nilai'
            'y2' ==> 'Reviewer 1 - Rekomendasi Dana'
            'z2' ==> 'Reviewer 1 - Komentar'
            'aa2' ==> 'Reviewer 2 - Nama Reviewer'
            'ab2' ==> 'Reviewer 2 - Nilai'
            'ac2' ==> 'Reviewer 2 - Rekomendasi Dana'
            'ad2' ==> 'Reviewer 2 - Komentar'
            'ae2' ==> 'N.I.P'
            'af2' ==> 'N.I.K'
            'ag2' ==> 'N.I.D.N'
            'ah2' ==> 'Nama Lengkap'
            'ai2' ==> 'Jabatan Fungsional'

            In addition, variables starting with 'y' prefix are pertaining
            to the table data of "Data Anggota"
            '''

            # Getting the static (non-variable as in "Data Anggota") detail information
            # ---
            a1 = str(i+1)
            # Use try-except catching to mitigate empty data
            try:
                b2 = content.xpath('//span[@id="ContentPlaceHolder1_daftar4_kiri1_txjudul1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                b2 = ''
            # Use try-except catching to mitigate empty data
            try:
                c2 = content.xpath('//span[@id="ContentPlaceHolder1_daftar4_kiri1_tglusul1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                c2 = ''
            # Use try-except catching to mitigate empty data
            try:
                d2 = content.xpath('//span[@id="ContentPlaceHolder1_daftar4_kiri1_ttkt1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                d2 = ''
            # Use try-except catching to mitigate empty data
            try:
                e2 = content.xpath('//span[@id="ContentPlaceHolder1_daftar4_kiri1_ddllevel1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                e2 = ''
            # Use try-except catching to mitigate empty data
            try:
                f2 = content.xpath('//span[@id="ContentPlaceHolder1_daftar4_kiri1_lkat1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                f2 = ''
            # Use try-except catching to mitigate empty data
            try:
                g2 = content.xpath('//span[@id="ContentPlaceHolder1_daftar4_kiri1_ddlskema1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                g2 = ''
            # 'Rumpun Ilmu' scraped data contains multiple array of strings
            # This needs some extra tweaking
            h2 = ''
            # Use try-except catching to mitigate empty data
            try:
                h2_pre = content.xpath('//div[@id="ContentPlaceHolder1_daftar4_kiri1_updx"]/div[1][@class="panel panel-info"]/div[@class="panel-body f12"]/div[7][@class="row f12"]/div[@class="col-sm-4"]//text()')
                for l in h2_pre:
                    l = l.replace('\r','').replace('\n','').strip()
                    h2 += l +', '
                # Remove trailing ', ' characters
                h2 = h2[:-2]
            except IndexError:
                h2 = ''
            # ---
            # Use try-except catching to mitigate empty data
            try:
                i2 = content.xpath('//span[@id="ContentPlaceHolder1_daftar4_kiri1_ddlfokus1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                i2 = ''
            # Use try-except catching to mitigate empty data
            try:
                j2 = content.xpath('//span[@id="ContentPlaceHolder1_daftar4_kiri1_ddltema1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                j2 = ''
            # Use try-except catching to mitigate empty data
            try:
                k2 = content.xpath('//span[@id="ContentPlaceHolder1_daftar4_kiri1_ddltopik1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                k2 = ''
            # Use try-except catching to mitigate empty data
            try:
                l2 = content.xpath('//span[@id="ContentPlaceHolder1_daftar4_kiri1_ddllama1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                l2 = ''
            # Use try-except catching to mitigate empty data
            try:
                p2 = content.xpath('//span[@id="ContentPlaceHolder1_daftar4_kiri1_ddlmakro1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                p2 = ''
            # Use try-except catching to mitigate empty data
            try:
                q2 = content.xpath('//a[@id="ContentPlaceHolder1_daftar4_kiri1_alblcatatan1"]/@href')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                q2 = ''
            # Use try-except catching to mitigate empty data
            try:
                r2 = content.xpath('//span[@id="ContentPlaceHolder1_daftar4_kiri1_tbiayathn1"]/text()')[0].replace('\r','').replace('\n','').replace('Rp.', '').replace(',', '').strip()
            except IndexError:
                r2 = ''
            # Use try-except catching to mitigate empty data
            try:
                s2 = content.xpath('//a[@id="ContentPlaceHolder1_daftar4_kiri1_alblfilerab1"]/@href')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                s2 = ''
            # Some 'Data Detil' has no 'Mitra' data element
            # Use try-except catching to mitigate
            try:
                t2 = content.xpath('//span[@id="ContentPlaceHolder1_daftar4_kiri1_tnmmitra1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                t2 = ''
            # Use try-except catching to mitigate empty data
            try:
                u2 = content.xpath('//span[@id="ContentPlaceHolder1_daftar4_kiri1_tbiayadukung1"]/text()')[0].replace('\r','').replace('\n','').replace('Rp.', '').replace(',', '').strip()
            except IndexError:
                u2 = ''
            # Use try-except catching to mitigate empty data
            try:
                v2 = content.xpath('//a[@id="ContentPlaceHolder1_daftar4_kiri1_alblfilemitradukung1"]/@href')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                v2 = ''
            # Use try-except catching to mitigate empty data
            try:
                ae2 = content.xpath('//div[@id="ContentPlaceHolder1_daftar4_kiri1_bio2_updk1"]/div[1]/div[@class="col-sm-4"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                ae2 = ''
            # Use try-except catching to mitigate empty data
            try:
                af2 = content.xpath('//div[@id="ContentPlaceHolder1_daftar4_kiri1_bio2_updk1"]/div[2]/div[@class="col-sm-4"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                af2 = ''
            # Use try-except catching to mitigate empty data
            try:
                ag2 = content.xpath('//div[@id="ContentPlaceHolder1_daftar4_kiri1_bio2_updk1"]/div[3]/div[@class="col-sm-4"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                ag2 = ''
            # Use try-except catching to mitigate empty data
            try:
                ah2 = content.xpath('//div[@id="ContentPlaceHolder1_daftar4_kiri1_bio2_updk1"]/div[4]/div[@class="col-sm-4"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                ah2 = ''
            # Use try-except catching to mitigate empty data
            try:
                ai2 = content.xpath('//div[@id="ContentPlaceHolder1_daftar4_kiri1_bio2_updk1"]/div[5]/div[@class="col-sm-4"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                ai2 = ''

            # DEBUG
            # Please comment out after use
            # ---
            # print(a1, b2, c2, d2, e2, f2, g2, h2, i2, j2, k2, l2, m2[i], n2[i], o2[i], p2, q2, r2, s2, t2, u2, v2, w2[i], x2[i], y2[i], z2[i], aa2[i], ab2[i], ac2[i], ad2[i], ae2, af2, ag2, ah2, ai2)

            # Painting the static (non-variable as in "Data Anggota") detail information
            sheet[f'A{row_start}'].value = a1
            sheet[f'B{row_start}'].value = b2
            sheet[f'C{row_start}'].value = c2
            sheet[f'D{row_start}'].value = d2
            sheet[f'E{row_start}'].value = e2
            sheet[f'F{row_start}'].value = f2
            sheet[f'G{row_start}'].value = g2
            sheet[f'H{row_start}'].value = h2
            sheet[f'I{row_start}'].value = i2
            sheet[f'J{row_start}'].value = j2
            sheet[f'K{row_start}'].value = k2
            sheet[f'L{row_start}'].value = l2
            sheet[f'M{row_start}'].value = m2[i]
            sheet[f'N{row_start}'].value = n2[i]
            sheet[f'O{row_start}'].value = o2[i]
            sheet[f'P{row_start}'].value = p2
            sheet[f'Q{row_start}'].value = q2
            sheet[f'R{row_start}'].value = r2
            sheet[f'S{row_start}'].value = s2
            sheet[f'T{row_start}'].value = t2
            sheet[f'U{row_start}'].value = u2
            sheet[f'V{row_start}'].value = v2
            sheet[f'W{row_start}'].value = w2[i]
            sheet[f'X{row_start}'].value = x2[i]
            sheet[f'Y{row_start}'].value = y2[i]
            sheet[f'Z{row_start}'].value = z2[i]
            sheet[f'AA{row_start}'].value = aa2[i]
            sheet[f'AB{row_start}'].value = ab2[i]
            sheet[f'AC{row_start}'].value = ac2[i]
            sheet[f'AD{row_start}'].value = ad2[i]
            sheet[f'AE{row_start}'].value = ae2
            sheet[f'AF{row_start}'].value = af2
            sheet[f'AG{row_start}'].value = ag2
            sheet[f'AH{row_start}'].value = ah2
            sheet[f'AI{row_start}'].value = ai2

            # The table which displays 'Identitas Pengusul - Anggota Peneliti'
            # Treated differently, specially
            # 'y_' is a DOM element representing a HTML table row (<tr>)
            y_ = content.xpath(
                '//div[@id="ContentPlaceHolder1_daftar4_kiri1_updx"]/div[4]//table[@class="table"]//tr[position()>1]')

            # The base path
            y_base = '//div[@id="ContentPlaceHolder1_daftar4_kiri1_updx"]/div[4]//table[@class="table"]//tr[position()>1]'

            # Painting the scraped data to the output spreadsheet row
            # Section: 'DATA ANGGOTA'
            # ---
            # Checking if 'Identitas Pengusul - Anggota Peneliti' table data exists
            if len(y_) == 0:
                pass  # --- nope. the data does not exist
            else:

                # Checking if this table's 'LUARAN' data has the most rows
                y_all_rows = len(y_)
                y_max_row = max(y_max_row, y_all_rows)

                # ROW DATA LABEL CONVENTION
                # y_row_a -> "N.I.P"
                # y_row_b -> "Nama"
                # y_row_c -> "Bidang Keahlian"
                # y_row_d -> "Alamat"
                # y_row_e -> "Instansi"
                # y_row_f -> "Email"
                # y_row_g -> "No. HP"
                # y_row_h -> "Peran"
                # y_row_i -> "Tugas"

                y_row_a = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[2]/text()')]

                y_row_b = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[3]/text()')]

                y_row_c = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[4]/text()')]

                y_row_d = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[5]/text()')]

                y_row_e = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[6]/text()')]

                y_row_f = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[7]/text()')]

                y_row_g = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[8]/text()')]

                y_row_h = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[9]/text()')]

                y_row_i = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[10]/text()')]

                # The starting column coordinate for filling the 'LUARAN' table data
                col_start = 36

                # Iterating through each table row and write to the spreadsheet
                # Assumes the lists y_row_a, y_row_b, ... have the same array size
                for j in range(len(y_row_a)):
                    # Painting table data
                    sheet.cell(row=row_start, column=col_start).value = y_row_a[j]
                    sheet.cell(row=row_start, column=col_start + 1).value = y_row_b[j]
                    sheet.cell(row=row_start, column=col_start + 2).value = y_row_c[j]
                    sheet.cell(row=row_start, column=col_start + 3).value = y_row_d[j]
                    sheet.cell(row=row_start, column=col_start + 4).value = y_row_e[j]
                    sheet.cell(row=row_start, column=col_start + 5).value = y_row_f[j]
                    sheet.cell(row=row_start, column=col_start + 6).value = y_row_g[j]
                    sheet.cell(row=row_start, column=col_start + 7).value = y_row_h[j]
                    sheet.cell(row=row_start, column=col_start + 8).value = y_row_i[j]

                    # Incrementing the 'col_start' iterator before continuing the loop
                    col_start += 9
                    continue

            # Incrementing the value of 'row_start' before continuing
            row_start += 1

            # Reopening the "Berkas Ditolak Penelitian" list page,
            # then assign the AJAX response to the temporary array 'temporary_prompt'
            # The 'data' array is obtained from opening individual entry row detail page
            #
            # This is done only on the 7-multiple row iteration, because
            # a single set of 'viewstate', 'viewstategen', and 'eventvalidation' values of ASPX
            # can only be used to do at most 8 operations before having to be renewed.
            if (i % 7 == 0) and (i > 0):
                temporary_prompt = self.get_risat_penelitian_ditolak_penelitian(data)
            continue

        # Post-loop logging: appending the header over the 'DATA ANGGOTA' columns
        control.append_message_area(
            f'+ Melengkapi kepala tabel pada bagian "Identitas Pengusul - Anggota Peneliti" ...')
        # ---
        # The starting column for the 'DATA ANGGOTA' data
        col_start = 36
        # Beginning the loop that detects the maximum number of 'LUARAN' rows
        # according to the variable 'y_max_row'
        for i in range(1, y_max_row + 1):
            # Setting the top header
            sheet.merge_cells(
                start_row=1,
                start_column=col_start,
                end_row=1,
                end_column=col_start + 8
            )
            sheet.cell(row=1, column=col_start).value = f'IDENTITAS PENGUSUL — ANGGOTA #{i}'
            sheet.cell(row=1, column=col_start).alignment = Alignment(horizontal='center')
            # Setting the sub headers
            sheet.cell(row=2, column=col_start).value = 'N.I.P'
            sheet.cell(row=2, column=col_start + 1).value = 'Nama'
            sheet.cell(row=2, column=col_start + 2).value = 'Bidang Keahlian'
            sheet.cell(row=2, column=col_start + 3).value = 'Alamat'
            sheet.cell(row=2, column=col_start + 4).value = 'Instansi'
            sheet.cell(row=2, column=col_start + 5).value = 'Email'
            sheet.cell(row=2, column=col_start + 6).value = 'No. HP'
            sheet.cell(row=2, column=col_start + 7).value = 'Peran'
            sheet.cell(row=2, column=col_start + 8).value = 'Tugas'
            # Incrementing the 'col_start' iterator before continuing the loop
            col_start += 9
            continue

        # Post-loop logging: successfully painted the output spreadsheet file
        control.append_message_area(f'+ Tabel sukses dipanen!')
        control.set_progress_bar(85)

        # Asking for the spreadsheet name to save as
        # ---
        # Logging and setting the progress bar
        control.append_message_area(f'+ Menyimpan spreadsheet luaran ...')
        control.set_progress_bar(90)
        # Dealing with file name prompt and saving
        # Using loop to mitigate the user clicking 'cancel'
        # in the file name dialog prompt
        while True:
            # Opening the dialog prompt
            output_spreadsheet = filedialog.asksaveasfilename(
                filetypes=[('Excel files', '*.xlsx')],
                initialfile='Sipesat - Ditolak Penelitian Detil Risat.xlsx',
                title='Simpan sebagai ...'
            )

            # 'cancel' button in the dialog prompt is clicked
            if len(output_spreadsheet) == 0:
                # Showing confirmation
                x = messagebox.askyesno(
                    'Nama File Kosong',
                    'Apakah Anda yakin ingin melanjutkan tanpa menyimpan file spreadsheet hasil pemanenan?'
                )
                # Determining whether to break or to continue the loop
                # based on the inversed value of 'x'
                if x:
                    control.append_message_area(
                        f'+ Finalisasi pemanenan data tanpa menyimpan file spreadsheet luaran ...')
                    workbook.close()  # --- closing the workbook without saving
                    break
                else:
                    continue  # --- continuing the loop
            # File name does not end in spreadsheet extension
            elif output_spreadsheet[-5:] != '.xlsx':
                output_spreadsheet = output_spreadsheet + '.xlsx'

            # Saving the spreadsheet
            control.append_message_area(f'LOKASI_SPREADSHEET_LUARAN: {output_spreadsheet}')
            control.set_progress_bar(95)
            workbook.save(output_spreadsheet)

            # Closing the openpyxl workbook
            control.append_message_area(f'+ Menutup file spreadsheet ...')
            control.set_progress_bar(98)
            workbook.close()

            # Breaking the loop
            break

        # Notify for a successful scraping
        control.append_message_area(f'+ Pemanenan selesai pada: {str(dt.now())}')
        control.set_progress_bar(100)
        control.on_notify_successful_scraping()

    # This function harvests "Risat Dana Penelitian > Data Detil Lengkap" data
    # and then store the harvested data as an excel file
    #
    # Required arguments:
    # - control             --> for updating the progress bar and
    #                           message area of the screen SipesatScrHarvest
    # - username, password  --> the Risat administrator username and password
    def run_harvest_r_1_dana(self, control, username, password):

        # SipesatScrHarvest messenger
        control.set_header_desc('Panen Data "Risat Dana Penelitian > Data Detil Lengkap"')
        control.set_help_label('Data sedang dipanen. Silahkan menunggu.')
        control.set_progress_bar(0)
        control.clear_message_area()

        # Preamble logging
        control.append_message_area(f'+ Memulai pemanenan data ...')
        control.append_message_area(f'+ Pemanenan dimulai pada: {str(dt.now())}')
        control.set_progress_bar(5)

        # Preparing the 'data_prompt' arrays
        control.append_message_area(f'+ Log masuk Risat sebagai [{username}] ...')
        control.set_progress_bar(10)
        data_prompt = self.get_risat_login(username, password)
        data_prompt = self.get_risat_penelitian(data_prompt)
        data_prompt = self.get_risat_penelitian_dana_penelitian(data_prompt)

        # Parsing XML tree content
        control.append_message_area(f'+ Membaca halaman web ...')
        control.set_progress_bar(15)
        content = data_prompt['html_content']

        # Establishing the export spreadsheet file
        control.append_message_area(f'+ Mempersiapkan file spreadsheet luaran ...')
        control.set_progress_bar(20)
        workbook = xl.Workbook()
        sheet = workbook.active
        sheet.title = 'Dana Penelitian Detil'

        # DEVELOPER'S NOTE (DO NOT REMOVE)
        '''
        Data yang perlu diambil:
        1. Judul
        2. Substansi usulan
        3. Identitas pengusul - ketua
        4. Identitas pengusul - anggota
        5. RAB
        6. Dokumen pendukung
        
        Urutan kolom excel:
        1. Judul
        2. Substansi usulan
        3. RAB
        4. Dokumen pendukung
        5. Identitas pengusul - ketua
        6. Identitas pengusul - anggota
        '''

        # Preparing the sheet header
        control.append_message_area(f'+ Mempersiapkan kepala lembar spreadsheet ...')
        control.set_progress_bar(25)
        # ---
        # Preparing the "NOMOR" header
        sheet.merge_cells(
            start_row=1,
            start_column=1,
            end_row=2,
            end_column=1
        )
        sheet['A1'].value = 'No.'
        sheet['A1'].alignment = Alignment(horizontal='center')
        # ---
        # Preparing the "JUDUL" header
        sheet.merge_cells(
            start_row=1,
            start_column=2,
            end_row=1,
            end_column=12
        )
        sheet['B1'].value = 'IDENTITAS'
        sheet['B1'].alignment = Alignment(horizontal='center')
        # Preparing the "JUDUL" sub-headers
        sheet['B2'].value = 'Judul'
        sheet['C2'].value = 'Tgl. Usulan'
        sheet['D2'].value = 'TKT Saat Ini'
        sheet['E2'].value = 'Level'
        sheet['F2'].value = 'Kategori'
        sheet['G2'].value = 'Skema'
        sheet['H2'].value = 'Rumpun Ilmu'
        sheet['I2'].value = 'Bidang Fokus'
        sheet['J2'].value = 'Tema'
        sheet['K2'].value = 'Topik'
        sheet['L2'].value = 'Lama Kegiatan'
        # ---
        # Preparing the "SUBSTANSI USULAN" header
        sheet.merge_cells(
            start_row=1,
            start_column=13,
            end_row=1,
            end_column=14
        )
        sheet['M1'].value = 'SUBSTANSI USULAN'
        sheet['M1'].alignment = Alignment(horizontal='center')
        # Preparing the "SUBSTANSI USULAN" sub-headers
        sheet['M2'].value = 'Kelompok Makro'
        sheet['N2'].value = 'File Proposal'
        # ---
        # Preparing the "RAB" header
        sheet.merge_cells(
            start_row=1,
            start_column=15,
            end_row=1,
            end_column=16
        )
        sheet['O1'].value = 'RENCANA ANGGARAN BIAYA'
        sheet['O1'].alignment = Alignment(horizontal='center')
        # Preparing the "RAB" sub-headers
        sheet['O2'].value = 'Biaya'
        sheet['P2'].value = 'File RAB'
        # ---
        # Preparing the "DOKUMEN PENDUKUNG" header
        sheet.merge_cells(
            start_row=1,
            start_column=17,
            end_row=1,
            end_column=19
        )
        sheet['Q1'].value = 'DOKUMEN PENDUKUNG'
        sheet['Q1'].alignment = Alignment(horizontal='center')
        # Preparing the "DOKUMEN PENDUKUNG" sub-headers
        sheet['Q2'].value = 'Mitra'
        sheet['R2'].value = 'Dukungan Biaya'
        sheet['S2'].value = 'Surat Dukungan Mitra'
        # ---
        # Preparing the "IDENTITAS PENGUSUL — KETUA" header
        sheet.merge_cells(
            start_row=1,
            start_column=20,
            end_row=1,
            end_column=24
        )
        sheet['T1'].value = 'IDENTITAS PENGUSUL — KETUA'
        sheet['T1'].alignment = Alignment(horizontal='center')
        # Preparing the "IDENTITAS PENGUSUL — KETUA" sub-headers
        sheet['T2'].value = 'N.I.P'
        sheet['U2'].value = 'N.I.K'
        sheet['V2'].value = 'N.I.D.N'
        sheet['W2'].value = 'Nama Lengkap'
        sheet['X2'].value = 'Jabatan Fungsional'

        # The base XPath location, pointing to each entry row
        base = '//div[@class="mw-100"]//div[@class="form-group f12"]/table[@width="100%"]//tr[@valign="top"]'

        # ---
        # Obtaining the data row values
        control.append_message_area(f'+ Mendapatkan data pada baris tabel ...')
        control.set_progress_bar(30)

        # Reading the HTML entry row hidden ASPX values
        # Copy-pasted from: /ssynthesia/ghostcity/ar/dumper-2/24__2023.02.13__requestsrisat.py
        all_kodetran_prop = content.xpath(base + '//input[1][@type="hidden"]/@name')
        all_kodetran_val = content.xpath(base + '//input[1][@type="hidden"]/@value')
        all_stat_prop = content.xpath(base + '//input[2][@type="hidden"]/@name')
        all_stat_val = content.xpath(base + '//input[2][@type="hidden"]/@value')
        all_submitbtn = content.xpath(base + '//input[@type="submit"]/@name')

        # The number of rows
        # Assumes the 'all_' array size equals the number of data rows
        number_of_row = len(all_kodetran_prop)
        
        # The starting row coordinate of the active sheet
        row_start = 3
        
        # The maximum number of 'DATA ANGGOTA' table data row
        y_max_row = 0

        # Iterating through each entry row element
        # Assumes all the 'all_' arrays in the previous code block
        # are of the same length/size
        # Copy-pasted from: /ssynthesia/ghostcity/ar/dumper-2/24__2023.02.13__requestsrisat.py
        control.append_message_area(f'+ Melakukan iterasi terhadap baris tabel dan menulis spreadsheet luaran ...')
        control.set_progress_bar(35)
        temporary_prompt = data_prompt
        for i in range(number_of_row):
            # Noisy preamble logging
            # Please don't use this -_-
            # ---
            # control.append_message_area(f'ITERASI [{i}]')

            # Updating the progress bar status
            control.set_progress_bar(35 + round(45 * (i + 1) / number_of_row))

            # Preparing the AJAX payload
            detail_prompt = {
                'viewstate' : temporary_prompt['viewstate'],
                'viewstategen' : temporary_prompt['viewstategen'],
                'eventvalidation' : temporary_prompt['eventvalidation'],
                'button_name' : all_submitbtn[i],
                'kodetran_prop' : all_kodetran_prop[i],
                'kodetran_val' : all_kodetran_val[i],
                'stat_prop' : all_stat_prop[i],
                'stat_val' : all_stat_val[i]
            }

            # Obtaining the response data of each individual entry row detail page
            data = self.get_risat_penelitian_dana_penelitian_detil(detail_prompt)
            content = data['html_content']
            response = data['http_response']
            
            # Terminal logging for detecting errors
            log_string_1 = str(i+1)
            log_string_2 = content.xpath('//span[@id="ContentPlaceHolder1_danacair1_kiri1_txjudul1"]/text()')[0].replace('\r','').replace('\n','').strip()
            control.append_message_area(f'+ Memanen data detil: [{log_string_1}] {log_string_2} ...')
            print(f'+ Harvesting detailed data: [{log_string_1}] {log_string_2}')
            
            # DEVELOPER'S NOTE (DO NOT REMOVE)
            '''
            Variables and their associated data
            
            'a1' ==> 'No.'
            'b2' ==> 'Judul'
            'c2' ==> 'Tgl. Usulan'
            'd2' ==> 'TKT Saat Ini'
            'e2' ==> 'Level'
            'f2' ==> 'Kategori'
            'g2' ==> 'Skema'
            'h2' ==> 'Rumpun Ilmu'
            'i2' ==> 'Bidang Fokus'
            'j2' ==> 'Tema'
            'k2' ==> 'Topik'
            'l2' ==> 'Lama Kegiatan'
            'm2' ==> 'Kelompok Makro'
            'n2' ==> 'File Proposal'
            'o2' ==> 'Biaya'
            'p2' ==> 'File RAB'
            'q2' ==> 'Mitra'
            'r2' ==> 'Dukungan Biaya'
            's2' ==> 'Surat Dukungan Mitra'
            't2' ==> 'N.I.P'
            'u2' ==> 'N.I.K'
            'v2' ==> 'N.I.D.N'
            'w2' ==> 'Nama Lengkap'
            'x2' ==> 'Jabatan Fungsional'
            
            In addition, variables starting with 'y' prefix are pertaining
            to the table data of "Data Anggota"
            '''

            # Getting the static (non-variable as in "Data Anggota") detail information
            # ---
            a1 = str(i+1)
            # Use try-except catching to mitigate empty data
            try:
                b2 = content.xpath('//span[@id="ContentPlaceHolder1_danacair1_kiri1_txjudul1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                b2 = ''
            # Use try-except catching to mitigate empty data
            try:
                c2 = content.xpath('//span[@id="ContentPlaceHolder1_danacair1_kiri1_tglusul1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                c2 = ''
            # Use try-except catching to mitigate empty data
            try:
                d2 = content.xpath('//span[@id="ContentPlaceHolder1_danacair1_kiri1_ttkt1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                d2 = ''
            # Use try-except catching to mitigate empty data
            try:
                e2 = content.xpath('//span[@id="ContentPlaceHolder1_danacair1_kiri1_ddllevel1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                e2 = ''
            # Use try-except catching to mitigate empty data
            try:
                f2 = content.xpath('//span[@id="ContentPlaceHolder1_danacair1_kiri1_lkat1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                f2 = ''
            # Use try-except catching to mitigate empty data
            try:
                g2 = content.xpath('//span[@id="ContentPlaceHolder1_danacair1_kiri1_ddlskema1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                g2 = ''
            # 'Rumpun Ilmu' scraped data contains multiple array of strings
            # This needs some extra tweaking
            h2 = ''
            # Use try-except catching to mitigate empty data
            try:
                h2_pre = content.xpath('//div[@id="ContentPlaceHolder1_danacair1_kiri1_updx"]/div[1][@class="panel panel-info"]/div[@class="panel-body f12"]/div[7][@class="row f12"]/div[@class="col-sm-4"]//text()')
                for l in h2_pre:
                    l = l.replace('\r','').replace('\n','').strip()
                    h2 += l +', '
                # Remove trailing ', ' characters
                h2 = h2[:-2]
            except IndexError:
                h2 = ''
            # ---
            # Use try-except catching to mitigate empty data
            try:
                i2 = content.xpath('//span[@id="ContentPlaceHolder1_danacair1_kiri1_ddlfokus1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                i2 = ''
            # Use try-except catching to mitigate empty data
            try:
                j2 = content.xpath('//span[@id="ContentPlaceHolder1_danacair1_kiri1_ddltema1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                j2 = ''
            # Use try-except catching to mitigate empty data
            try:
                k2 = content.xpath('//span[@id="ContentPlaceHolder1_danacair1_kiri1_ddltopik1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                k2 = ''
            # Use try-except catching to mitigate empty data
            try:
                l2 = content.xpath('//span[@id="ContentPlaceHolder1_danacair1_kiri1_ddllama1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                l2 = ''
            # Use try-except catching to mitigate empty data
            try:
                m2 = content.xpath('//span[@id="ContentPlaceHolder1_danacair1_kiri1_ddlmakro1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                m2 = ''
            # Use try-except catching to mitigate empty data
            try:
                n2 = content.xpath('//a[@id="ContentPlaceHolder1_danacair1_kiri1_alblcatatan1"]/@href')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                n2 = ''
            # Use try-except catching to mitigate empty data
            try:
                o2 = content.xpath('//span[@id="ContentPlaceHolder1_danacair1_kiri1_tbiayathn1"]/text()')[0].replace('\r','').replace('\n','').replace('Rp.', '').replace(',', '').strip()
            except IndexError:
                o2 = ''
            # Use try-except catching to mitigate empty data
            try:
                p2 = content.xpath('//a[@id="ContentPlaceHolder1_danacair1_kiri1_alblfilerab1"]/@href')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                p2 = ''
            # Some 'Data Detil' has no 'Mitra' data element
            # Use try-except catching to mitigate
            try:
                q2 = content.xpath('//span[@id="ContentPlaceHolder1_danacair1_kiri1_tnmmitra1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                q2 = ''
            # Use try-except catching to mitigate empty data
            try:
                r2 = content.xpath('//span[@id="ContentPlaceHolder1_danacair1_kiri1_tbiayadukung1"]/text()')[0].replace('\r','').replace('\n','').replace('Rp.', '').replace(',', '').strip()
            except IndexError:
                r2 = ''
            # Use try-except catching to mitigate empty data
            try:
                s2 = content.xpath('//a[@id="ContentPlaceHolder1_danacair1_kiri1_alblfilemitradukung1"]/@href')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                s2 = ''
            # Use try-except catching to mitigate empty data
            try:
                t2 = content.xpath('//div[@id="ContentPlaceHolder1_danacair1_kiri1_bio2_updk1"]/div[1]/div[@class="col-sm-4"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                t2 = ''
            # Use try-except catching to mitigate empty data
            try:
                u2 = content.xpath('//div[@id="ContentPlaceHolder1_danacair1_kiri1_bio2_updk1"]/div[2]/div[@class="col-sm-4"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                u2 = ''
            # Use try-except catching to mitigate empty data
            try:
                v2 = content.xpath('//div[@id="ContentPlaceHolder1_danacair1_kiri1_bio2_updk1"]/div[3]/div[@class="col-sm-4"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                v2 = ''
            # Use try-except catching to mitigate empty data
            try:
                w2 = content.xpath('//div[@id="ContentPlaceHolder1_danacair1_kiri1_bio2_updk1"]/div[4]/div[@class="col-sm-4"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                w2 = ''
            # Use try-except catching to mitigate empty data
            try:
                x2 = content.xpath('//div[@id="ContentPlaceHolder1_danacair1_kiri1_bio2_updk1"]/div[5]/div[@class="col-sm-4"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                x2 = ''
            
            # Painting the static (non-variable as in "Data Anggota") detail information
            sheet[f'A{row_start}'].value = a1
            sheet[f'B{row_start}'].value = b2
            sheet[f'C{row_start}'].value = c2
            sheet[f'D{row_start}'].value = d2
            sheet[f'E{row_start}'].value = e2
            sheet[f'F{row_start}'].value = f2
            sheet[f'G{row_start}'].value = g2
            sheet[f'H{row_start}'].value = h2
            sheet[f'I{row_start}'].value = i2
            sheet[f'J{row_start}'].value = j2
            sheet[f'K{row_start}'].value = k2
            sheet[f'L{row_start}'].value = l2
            sheet[f'M{row_start}'].value = m2
            sheet[f'N{row_start}'].value = n2
            sheet[f'O{row_start}'].value = o2
            sheet[f'P{row_start}'].value = p2
            sheet[f'Q{row_start}'].value = q2
            sheet[f'R{row_start}'].value = r2
            sheet[f'S{row_start}'].value = s2
            sheet[f'T{row_start}'].value = t2
            sheet[f'U{row_start}'].value = u2
            sheet[f'V{row_start}'].value = v2
            sheet[f'W{row_start}'].value = w2
            sheet[f'X{row_start}'].value = x2
            
            # The table which displays 'Identitas Pengusul - Anggota Peneliti'
            # Treated differently, specially
            # 'y_' is a DOM element representing a HTML table row (<tr>)
            y_ = content.xpath('//div[@id="ContentPlaceHolder1_danacair1_kiri1_updx"]/div[4]//table[@class="table"]//tr[position()>1]')
            
            # The base path
            y_base = '//div[@id="ContentPlaceHolder1_danacair1_kiri1_updx"]/div[4]//table[@class="table"]//tr[position()>1]'
            
            # Painting the scraped data to the output spreadsheet row
            # Section: 'DATA ANGGOTA'
            # ---
            # Checking if 'Identitas Pengusul - Anggota Peneliti' table data exists
            if len(y_) == 0:
                pass  # --- nope. the data does not exist
            else:
                
                # Checking if this table's 'LUARAN' data has the most rows
                y_all_rows = len(y_)
                y_max_row = max(y_max_row, y_all_rows)
                
                # ROW DATA LABEL CONVENTION
                # y_row_a -> "N.I.P"
                # y_row_b -> "Nama"
                # y_row_c -> "Bidang Keahlian"
                # y_row_d -> "Alamat"
                # y_row_e -> "Instansi"
                # y_row_f -> "Email"
                # y_row_g -> "No. HP"
                # y_row_h -> "Peran"
                # y_row_i -> "Tugas"
                
                y_row_a = [l.replace('\r','').replace('\n','').strip()
                      for l in content.xpath(y_base + '/td[2]/text()')]

                y_row_b = [l.replace('\r','').replace('\n','').strip()
                      for l in content.xpath(y_base + '/td[3]/text()')]

                y_row_c = [l.replace('\r','').replace('\n','').strip()
                      for l in content.xpath(y_base + '/td[4]/text()')]

                y_row_d = [l.replace('\r','').replace('\n','').strip()
                      for l in content.xpath(y_base + '/td[5]/text()')]

                y_row_e = [l.replace('\r','').replace('\n','').strip()
                      for l in content.xpath(y_base + '/td[6]/text()')]

                y_row_f = [l.replace('\r','').replace('\n','').strip()
                      for l in content.xpath(y_base + '/td[7]/text()')]

                y_row_g = [l.replace('\r','').replace('\n','').strip()
                      for l in content.xpath(y_base + '/td[8]/text()')]

                y_row_h = [l.replace('\r','').replace('\n','').strip()
                      for l in content.xpath(y_base + '/td[9]/text()')]

                y_row_i = [l.replace('\r','').replace('\n','').strip()
                      for l in content.xpath(y_base + '/td[10]/text()')]

                # The starting column coordinate for filling the 'LUARAN' table data
                col_start = 25
                
                # Iterating through each table row and write to the spreadsheet
                # Assumes the lists y_row_a, y_row_b, ... have the same array size
                for j in range(len(y_row_a)):

                    # Painting table data
                    sheet.cell(row=row_start, column=col_start).value = y_row_a[j]
                    sheet.cell(row=row_start, column=col_start+1).value = y_row_b[j]
                    sheet.cell(row=row_start, column=col_start+2).value = y_row_c[j]
                    sheet.cell(row=row_start, column=col_start+3).value = y_row_d[j]
                    sheet.cell(row=row_start, column=col_start+4).value = y_row_e[j]
                    sheet.cell(row=row_start, column=col_start+5).value = y_row_f[j]
                    sheet.cell(row=row_start, column=col_start+6).value = y_row_g[j]
                    sheet.cell(row=row_start, column=col_start+7).value = y_row_h[j]
                    sheet.cell(row=row_start, column=col_start+8).value = y_row_i[j]

                    # Incrementing the 'col_start' iterator before continuing the loop
                    col_start += 9
                    continue
                
            # Incrementing the value of 'row_start' before continuing
            row_start += 1

            # Reopening the "Dana Penelitian" list page,
            # then assign the AJAX response to the temporary array 'temporary_prompt'
            # The 'data' array is obtained from opening individual entry row detail page
            #
            # This is done only on the 7-multiple row iteration, because
            # a single set of 'viewstate', 'viewstategen', and 'eventvalidation' values of ASPX
            # can only be used to do at most 8 operations before having to be renewed.
            if ( i % 7 == 0 ) and ( i > 0 ):
                temporary_prompt = self.get_risat_penelitian_dana_penelitian(data)
            continue
        
        # Post-loop logging: appending the header over the 'DATA ANGGOTA' columns
        control.append_message_area(f'+ Melengkapi kepala tabel pada bagian "Identitas Pengusul - Anggota Peneliti" ...')
        # ---
        # The starting column for the 'DATA ANGGOTA' data
        col_start = 25
        # Beginning the loop that detects the maximum number of 'LUARAN' rows
        # according to the variable 'y_max_row'
        for i in range(1, y_max_row+1):
            # Setting the top header
            sheet.merge_cells(
                start_row=1,
                start_column=col_start,
                end_row=1,
                end_column=col_start+8
            )
            sheet.cell(row=1, column=col_start).value = f'IDENTITAS PENGUSUL — ANGGOTA #{i}'
            sheet.cell(row=1, column=col_start).alignment = Alignment(horizontal='center')
            # Setting the sub headers
            sheet.cell(row=2, column=col_start).value = 'N.I.P'
            sheet.cell(row=2, column=col_start+1).value = 'Nama'
            sheet.cell(row=2, column=col_start+2).value = 'Bidang Keahlian'
            sheet.cell(row=2, column=col_start+3).value = 'Alamat'
            sheet.cell(row=2, column=col_start+4).value = 'Instansi'
            sheet.cell(row=2, column=col_start+5).value = 'Email'
            sheet.cell(row=2, column=col_start+6).value = 'No. HP'
            sheet.cell(row=2, column=col_start+7).value = 'Peran'
            sheet.cell(row=2, column=col_start+8).value = 'Tugas'
            # Incrementing the 'col_start' iterator before continuing the loop
            col_start += 9
            continue

        # Post-loop logging: successfully painted the output spreadsheet file
        control.append_message_area(f'+ Tabel sukses dipanen!')
        control.set_progress_bar(85)

        # Asking for the spreadsheet name to save as
        # ---
        # Logging and setting the progress bar
        control.append_message_area(f'+ Menyimpan spreadsheet luaran ...')
        control.set_progress_bar(90)
        # Dealing with file name prompt and saving
        # Using loop to mitigate the user clicking 'cancel'
        # in the file name dialog prompt
        while True:
            # Opening the dialog prompt
            output_spreadsheet = filedialog.asksaveasfilename(
                filetypes=[('Excel files', '*.xlsx')],
                initialfile='Sipesat - Dana Penelitian Detil Risat.xlsx',
                title='Simpan sebagai ...'
            )

            # 'cancel' button in the dialog prompt is clicked
            if len(output_spreadsheet) == 0:
                # Showing confirmation
                x = messagebox.askyesno(
                    'Nama File Kosong',
                    'Apakah Anda yakin ingin melanjutkan tanpa menyimpan file spreadsheet hasil pemanenan?'
                )
                # Determining whether to break or to continue the loop
                # based on the inversed value of 'x'
                if x:
                    control.append_message_area(
                        f'+ Finalisasi pemanenan data tanpa menyimpan file spreadsheet luaran ...')
                    workbook.close()  # --- closing the workbook without saving
                    break
                else:
                    continue  # --- continuing the loop
            # File name does not end in spreadsheet extension
            elif output_spreadsheet[-5:] != '.xlsx':
                output_spreadsheet = output_spreadsheet + '.xlsx'

            # Saving the spreadsheet
            control.append_message_area(f'LOKASI_SPREADSHEET_LUARAN: {output_spreadsheet}')
            control.set_progress_bar(95)
            workbook.save(output_spreadsheet)

            # Closing the openpyxl workbook
            control.append_message_area(f'+ Menutup file spreadsheet ...')
            control.set_progress_bar(98)
            workbook.close()

            # Breaking the loop
            break

        # Notify for a successful scraping
        control.append_message_area(f'+ Pemanenan selesai pada: {str(dt.now())}')
        control.set_progress_bar(100)
        control.on_notify_successful_scraping()
        
    # This function harvests "Risat Arsip Penelitian > Data Detil Lengkap" data
    # and then store the harvested data as an excel file
    #
    # Required arguments:
    # - control             --> for updating the progress bar and
    #                           message area of the screen SipesatScrHarvest
    # - username, password  --> the Risat administrator username and password
    def run_harvest_r_1_arsip(self, control, username, password):

        # Dodge Python function documentation
        '''
        '''

        # Give a warning about bugs in 'Arsip' data harvester
        # This bug is due to Risat's server lag, not because of this program by itself
        # (Just in case. As per 2023-06-12, turns out the 'Arsip' page of Risat
        # does not always load in 30 seconds -- depending on circumstances)
        '''
        messagebox.showerror(
            title='Pesan Pengembang',
            message='' +
                    'Dikarenakan kendala teknis pada server risat.uksw.edu, ' +
                    'data arsip Risat membutuhkan waktu sekitar 30 detik untuk dapat termuat. ' +
                    'Setiap pemuatan dari server, sebanyak 7 data dapat dipanen sekaligus. ' +
                    'Sehingga, secara keseluruhan dibutuhkan waktu sekitar 4 menit (dengan asumsi jumlah data sebanyak 50) ' +
                    'untuk dapat memanen semua data arsip Risat. ' +
                    'Antarmuka program SiPe.Sat akan sesekali membeku dan tombol-tombolnya hilang ' +
                    'selama pemanenan berlangsung karena menunggu pemuatan data arsip dari server risat.uksw.edu. ' +
                    '\n\n' +
                    'Pengembang SiPe.Sat meminta maaf atas ketidaknyamanan ini.'
        )
        '''

        # SipesatScrHarvest messenger
        control.set_header_desc('Panen Data "Risat Arsip Penelitian > Data Detil Lengkap"')
        control.set_help_label('Data sedang dipanen. Silahkan menunggu.')
        control.set_progress_bar(0)
        control.clear_message_area()

        # Preamble logging
        control.append_message_area(f'+ Memulai pemanenan data ...')
        control.append_message_area(f'+ Pemanenan dimulai pada: {str(dt.now())}')
        control.set_progress_bar(5)

        # Preparing the 'data_prompt' arrays
        control.append_message_area(f'+ Log masuk Risat sebagai [{username}] ...')
        control.set_progress_bar(10)
        data_prompt = self.get_risat_login(username, password)
        data_prompt = self.get_risat_penelitian(data_prompt)
        data_prompt = self.get_risat_penelitian_arsip(data_prompt)

        # Parsing XML tree content
        control.append_message_area(f'+ Membaca halaman web ...')
        control.set_progress_bar(15)
        content = data_prompt['html_content']

        # Establishing the export spreadsheet file
        control.append_message_area(f'+ Mempersiapkan file spreadsheet luaran ...')
        control.set_progress_bar(20)
        workbook = xl.Workbook()
        sheet = workbook.active
        sheet.title = 'Arsip Penelitian Detil'

        # DEVELOPER'S NOTE (DO NOT REMOVE)
        '''
        Data yang perlu diambil:
        1. Judul
        2. Substansi usulan
        3. Identitas pengusul - ketua
        4. Identitas pengusul - anggota
        5. RAB
        6. Dokumen pendukung
        
        Urutan kolom excel:
        1. Judul
        2. Substansi usulan
        3. RAB
        4. Dokumen pendukung
        5. Identitas pengusul - ketua
        6. Identitas pengusul - anggota
        '''

        # Preparing the sheet header
        control.append_message_area(f'+ Mempersiapkan kepala lembar spreadsheet ...')
        control.set_progress_bar(25)
        # ---
        # Preparing the "NOMOR" header
        sheet.merge_cells(
            start_row=1,
            start_column=1,
            end_row=2,
            end_column=1
        )
        sheet['A1'].value = 'No.'
        sheet['A1'].alignment = Alignment(horizontal='center')
        # ---
        # Preparing the "JUDUL" header
        sheet.merge_cells(
            start_row=1,
            start_column=2,
            end_row=1,
            end_column=12
        )
        sheet['B1'].value = 'IDENTITAS'
        sheet['B1'].alignment = Alignment(horizontal='center')
        # Preparing the "JUDUL" sub-headers
        sheet['B2'].value = 'Judul'
        sheet['C2'].value = 'Tgl. Usulan'
        sheet['D2'].value = 'TKT Saat Ini'
        sheet['E2'].value = 'Level'
        sheet['F2'].value = 'Kategori'
        sheet['G2'].value = 'Skema'
        sheet['H2'].value = 'Rumpun Ilmu'
        sheet['I2'].value = 'Bidang Fokus'
        sheet['J2'].value = 'Tema'
        sheet['K2'].value = 'Topik'
        sheet['L2'].value = 'Lama Kegiatan'
        # ---
        # Preparing the "SUBSTANSI USULAN" header
        sheet.merge_cells(
            start_row=1,
            start_column=13,
            end_row=1,
            end_column=14
        )
        sheet['M1'].value = 'SUBSTANSI USULAN'
        sheet['M1'].alignment = Alignment(horizontal='center')
        # Preparing the "SUBSTANSI USULAN" sub-headers
        sheet['M2'].value = 'Kelompok Makro'
        sheet['N2'].value = 'File Proposal'
        # ---
        # Preparing the "RAB" header
        sheet.merge_cells(
            start_row=1,
            start_column=15,
            end_row=1,
            end_column=16
        )
        sheet['O1'].value = 'RENCANA ANGGARAN BIAYA'
        sheet['O1'].alignment = Alignment(horizontal='center')
        # Preparing the "RAB" sub-headers
        sheet['O2'].value = 'Biaya'
        sheet['P2'].value = 'File RAB'
        # ---
        # Preparing the "DOKUMEN PENDUKUNG" header
        sheet.merge_cells(
            start_row=1,
            start_column=17,
            end_row=1,
            end_column=19
        )
        sheet['Q1'].value = 'DOKUMEN PENDUKUNG'
        sheet['Q1'].alignment = Alignment(horizontal='center')
        # Preparing the "DOKUMEN PENDUKUNG" sub-headers
        sheet['Q2'].value = 'Mitra'
        sheet['R2'].value = 'Dukungan Biaya'
        sheet['S2'].value = 'Surat Dukungan Mitra'
        # ---
        # Preparing the "IDENTITAS PENGUSUL — KETUA" header
        sheet.merge_cells(
            start_row=1,
            start_column=20,
            end_row=1,
            end_column=24
        )
        sheet['T1'].value = 'IDENTITAS PENGUSUL — KETUA'
        sheet['T1'].alignment = Alignment(horizontal='center')
        # Preparing the "IDENTITAS PENGUSUL — KETUA" sub-headers
        sheet['T2'].value = 'N.I.P'
        sheet['U2'].value = 'N.I.K'
        sheet['V2'].value = 'N.I.D.N'
        sheet['W2'].value = 'Nama Lengkap'
        sheet['X2'].value = 'Jabatan Fungsional'

        # The base XPath location, pointing to each entry row
        base = '//div[@id="ContentPlaceHolder1_arsip1_updg1"]//div[@class="panel-body f12"]/table[@width="100%"]'

        # ---
        # Obtaining the data row values
        control.append_message_area(f'+ Mendapatkan data pada baris tabel ...')
        control.set_progress_bar(30)

        # Reading the HTML entry row hidden ASPX values
        # Copy-pasted from: /ssynthesia/ghostcity/ar/dumper-2/24__2023.02.13__requestsrisat.py
        all_idat_prop = content.xpath(base + '//input[1][@type="hidden"]/@name')
        all_idat_val = content.xpath(base + '//input[1][@type="hidden"]/@value')
        all_itgl_prop = content.xpath(base + '//input[2][@type="hidden"]/@name')
        all_itgl_val = content.xpath(base + '//input[2][@type="hidden"]/@value')
        all_submitbtn = content.xpath(base + '//input[@type="submit"]/@name')

        # The number of rows
        # Assumes the 'all_' array size equals the number of data rows
        number_of_row = len(all_idat_prop)

        # DEBUG
        # Please comment out after use
        # ---
        # print(f'NUMBER_OF_ROW_LEN: {number_of_row}')
        
        # The starting row coordinate of the active sheet
        row_start = 3
        
        # The maximum number of 'DATA ANGGOTA' table data row
        y_max_row = 0

        # Iterating through each entry row element
        # Assumes all the 'all_' arrays in the previous code block
        # are of the same length/size
        # Copy-pasted from: /ssynthesia/ghostcity/ar/dumper-2/24__2023.02.13__requestsrisat.py
        control.append_message_area(f'+ Melakukan iterasi terhadap baris tabel dan menulis spreadsheet luaran ...')
        control.set_progress_bar(35)
        temporary_prompt = data_prompt
        for i in range(number_of_row):
            # Noisy preamble logging
            # Please don't use this -_-
            # ---
            # control.append_message_area(f'ITERASI [{i}]')

            # Updating the progress bar status
            control.set_progress_bar(35 + round(45 * (i + 1) / number_of_row))

            # Preparing the AJAX payload
            detail_prompt = {
                'viewstate' : temporary_prompt['viewstate'],
                'viewstategen' : temporary_prompt['viewstategen'],
                'eventvalidation' : temporary_prompt['eventvalidation'],
                'button_name' : all_submitbtn[i],
                'idat_prop' : all_idat_prop[i],
                'idat_val' : all_idat_val[i],
                'itgl_prop' : all_itgl_prop[i],
                'itgl_val' : all_itgl_val[i]
            }

            # Obtaining the response data of each individual entry row detail page
            data = self.get_risat_penelitian_arsip_detil(detail_prompt)
            content = data['html_content']
            response = data['http_response']
            
            # Terminal logging for detecting errors
            log_string_1 = str(i+1)
            log_string_2 = content.xpath('//span[@id="ContentPlaceHolder1_arsip1_kiri1_txjudul1"]/text()')[0].replace('\r','').replace('\n','').strip()
            control.append_message_area(f'+ Memanen data detil: [{log_string_1}] {log_string_2} ...')
            print(f'+ Harvesting detailed data: [{log_string_1}] {log_string_2}')
            
            # DEVELOPER'S NOTE (DO NOT REMOVE)
            '''
            Variables and their associated data
            
            'a1' ==> 'No.'
            'b2' ==> 'Judul'
            'c2' ==> 'Tgl. Usulan'
            'd2' ==> 'TKT Saat Ini'
            'e2' ==> 'Level'
            'f2' ==> 'Kategori'
            'g2' ==> 'Skema'
            'h2' ==> 'Rumpun Ilmu'
            'i2' ==> 'Bidang Fokus'
            'j2' ==> 'Tema'
            'k2' ==> 'Topik'
            'l2' ==> 'Lama Kegiatan'
            'm2' ==> 'Kelompok Makro'
            'n2' ==> 'File Proposal'
            'o2' ==> 'Biaya'
            'p2' ==> 'File RAB'
            'q2' ==> 'Mitra'
            'r2' ==> 'Dukungan Biaya'
            's2' ==> 'Surat Dukungan Mitra'
            't2' ==> 'N.I.P'
            'u2' ==> 'N.I.K'
            'v2' ==> 'N.I.D.N'
            'w2' ==> 'Nama Lengkap'
            'x2' ==> 'Jabatan Fungsional'
            
            In addition, variables starting with 'y' prefix are pertaining
            to the table data of "Data Anggota"
            '''

            # Getting the static (non-variable as in "Data Anggota") detail information
            # ---
            a1 = str(i+1)
            # Use try-except catching to mitigate empty data
            try:
                b2 = content.xpath('//span[@id="ContentPlaceHolder1_arsip1_kiri1_txjudul1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                b2 = ''
            # Use try-except catching to mitigate empty data
            try:
                c2 = content.xpath('//span[@id="ContentPlaceHolder1_arsip1_kiri1_tglusul1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                c2 = ''
            # Use try-except catching to mitigate empty data
            try:
                d2 = content.xpath('//span[@id="ContentPlaceHolder1_arsip1_kiri1_ttkt1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                d2 = ''
            # Use try-except catching to mitigate empty data
            try:
                e2 = content.xpath('//span[@id="ContentPlaceHolder1_arsip1_kiri1_ddllevel1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                e2 = ''
            # Use try-except catching to mitigate empty data
            try:
                f2 = content.xpath('//span[@id="ContentPlaceHolder1_arsip1_kiri1_lkat1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                f2 = ''
            # Use try-except catching to mitigate empty data
            try:
                g2 = content.xpath('//span[@id="ContentPlaceHolder1_arsip1_kiri1_ddlskema1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                g2 = ''
            # 'Rumpun Ilmu' scraped data contains multiple array of strings
            # This needs some extra tweaking
            h2 = ''
            # Use try-except catching to mitigate empty data
            try:
                h2_pre = content.xpath('//div[@id="ContentPlaceHolder1_arsip1_kiri1_updx"]/div[1][@class="panel panel-info"]/div[@class="panel-body f12"]/div[7][@class="row f12"]/div[@class="col-sm-4"]//text()')
                for l in h2_pre:
                    l = l.replace('\r','').replace('\n','').strip()
                    h2 += l +', '
                # Remove trailing ', ' characters
                h2 = h2[:-2]
            except IndexError:
                h2 = ''
            # ---
            # Use try-except catching to mitigate empty data
            try:
                i2 = content.xpath('//span[@id="ContentPlaceHolder1_arsip1_kiri1_ddlfokus1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                i2 = ''
            # Use try-except catching to mitigate empty data
            try:
                j2 = content.xpath('//span[@id="ContentPlaceHolder1_arsip1_kiri1_ddltema1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                j2 = ''
            # Use try-except catching to mitigate empty data
            try:
                k2 = content.xpath('//span[@id="ContentPlaceHolder1_arsip1_kiri1_ddltopik1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                k2 = ''
            # Use try-except catching to mitigate empty data
            try:
                l2 = content.xpath('//span[@id="ContentPlaceHolder1_arsip1_kiri1_ddllama1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                l2 = ''
            # Use try-except catching to mitigate empty data
            try:
                m2 = content.xpath('//span[@id="ContentPlaceHolder1_arsip1_kiri1_ddlmakro1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                m2 = ''
            # Use try-except catching to mitigate empty data
            try:
                n2 = content.xpath('//a[@id="ContentPlaceHolder1_arsip1_kiri1_alblcatatan1"]/@href')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                n2 = ''
            # Use try-except catching to mitigate empty data
            try:
                o2 = content.xpath('//span[@id="ContentPlaceHolder1_arsip1_kiri1_tbiayathn1"]/text()')[0].replace('\r','').replace('\n','').replace('Rp.', '').replace(',', '').strip()
            except IndexError:
                o2 = ''
            # Use try-except catching to mitigate empty data
            try:
                p2 = content.xpath('//a[@id="ContentPlaceHolder1_arsip1_kiri1_alblfilerab1"]/@href')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                p2 = ''
            # Some 'Data Detil' has no 'Mitra' data element
            # Use try-except catching to mitigate
            try:
                q2 = content.xpath('//span[@id="ContentPlaceHolder1_arsip1_kiri1_tnmmitra1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                q2 = ''
            # Use try-except catching to mitigate empty data
            try:
                r2 = content.xpath('//span[@id="ContentPlaceHolder1_arsip1_kiri1_tbiayadukung1"]/text()')[0].replace('\r','').replace('\n','').replace('Rp.', '').replace(',', '').strip()
            except IndexError:
                r2 = ''
            # Use try-except catching to mitigate empty data
            try:
                s2 = content.xpath('//a[@id="ContentPlaceHolder1_arsip1_kiri1_alblfilemitradukung1"]/@href')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                s2 = ''
            # Use try-except catching to mitigate empty data
            try:
                t2 = content.xpath('//div[@id="ContentPlaceHolder1_arsip1_kiri1_bio2_updk1"]/div[1]/div[@class="col-sm-4"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                t2 = ''
            # Use try-except catching to mitigate empty data
            try:
                u2 = content.xpath('//div[@id="ContentPlaceHolder1_arsip1_kiri1_bio2_updk1"]/div[2]/div[@class="col-sm-4"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                u2 = ''
            # Use try-except catching to mitigate empty data
            try:
                v2 = content.xpath('//div[@id="ContentPlaceHolder1_arsip1_kiri1_bio2_updk1"]/div[3]/div[@class="col-sm-4"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                v2 = ''
            # Use try-except catching to mitigate empty data
            try:
                w2 = content.xpath('//div[@id="ContentPlaceHolder1_arsip1_kiri1_bio2_updk1"]/div[4]/div[@class="col-sm-4"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                w2 = ''
            # Use try-except catching to mitigate empty data
            try:
                x2 = content.xpath('//div[@id="ContentPlaceHolder1_arsip1_kiri1_bio2_updk1"]/div[5]/div[@class="col-sm-4"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                x2 = ''
            
            # Painting the static (non-variable as in "Data Anggota") detail information
            sheet[f'A{row_start}'].value = a1
            sheet[f'B{row_start}'].value = b2
            sheet[f'C{row_start}'].value = c2
            sheet[f'D{row_start}'].value = d2
            sheet[f'E{row_start}'].value = e2
            sheet[f'F{row_start}'].value = f2
            sheet[f'G{row_start}'].value = g2
            sheet[f'H{row_start}'].value = h2
            sheet[f'I{row_start}'].value = i2
            sheet[f'J{row_start}'].value = j2
            sheet[f'K{row_start}'].value = k2
            sheet[f'L{row_start}'].value = l2
            sheet[f'M{row_start}'].value = m2
            sheet[f'N{row_start}'].value = n2
            sheet[f'O{row_start}'].value = o2
            sheet[f'P{row_start}'].value = p2
            sheet[f'Q{row_start}'].value = q2
            sheet[f'R{row_start}'].value = r2
            sheet[f'S{row_start}'].value = s2
            sheet[f'T{row_start}'].value = t2
            sheet[f'U{row_start}'].value = u2
            sheet[f'V{row_start}'].value = v2
            sheet[f'W{row_start}'].value = w2
            sheet[f'X{row_start}'].value = x2

            # DEBUG
            # Please comment out after use
            # ---
            '''
            print()
            print('-------------------------------------------------------------------------------------------------------')
            print(f' > DEBUG LOGGING FOR DATA ITERATION NO. {i}: NON-ANGGOTA DATA')
            print('-------------------------------------------------------------------------------------------------------')
            print(f'A1 = {a1}')
            print(f'B2 = {b2}')
            print(f'C2 = {c2}')
            print(f'D2 = {d2}')
            print(f'E2 = {e2}')
            print(f'F2 = {f2}')
            print(f'G2 = {g2}')
            print(f'H2 = {h2}')
            print(f'I2 = {i2}')
            print(f'J2 = {j2}')
            print(f'K2 = {k2}')
            print(f'L2 = {l2}')
            print(f'M2 = {m2}')
            print(f'N2 = {n2}')
            print(f'O2 = {o2}')
            print(f'P2 = {p2}')
            print(f'Q2 = {q2}')
            print(f'R2 = {r2}')
            print(f'S2 = {s2}')
            print(f'T2 = {t2}')
            print(f'U2 = {u2}')
            print(f'V2 = {v2}')
            print(f'W2 = {w2}')
            print(f'X2 = {x2}')
            print('-------------------------------------------------------------------------------------------------------', end='')
            print()
            '''
            
            # The table which displays 'Identitas Pengusul - Anggota Peneliti'
            # Treated differently, specially
            # 'y_' is a DOM element representing a HTML table row (<tr>)
            y_ = content.xpath('//div[@id="ContentPlaceHolder1_arsip1_kiri1_updx"]/div[4]//table[@class="table"]//tr[position()>1]')
            
            # The base path
            y_base = '//div[@id="ContentPlaceHolder1_arsip1_kiri1_updx"]/div[4]//table[@class="table"]//tr[position()>1]'
            
            # Painting the scraped data to the output spreadsheet row
            # Section: 'DATA ANGGOTA'
            # ---
            # Checking if 'Identitas Pengusul - Anggota Peneliti' table data exists
            if len(y_) == 0:
                pass  # --- nope. the data does not exist
            else:
                
                # Checking if this table's 'LUARAN' data has the most rows
                y_all_rows = len(y_)
                y_max_row = max(y_max_row, y_all_rows)
                
                # ROW DATA LABEL CONVENTION
                # y_row_a -> "N.I.P"
                # y_row_b -> "Nama"
                # y_row_c -> "Bidang Keahlian"
                # y_row_d -> "Alamat"
                # y_row_e -> "Instansi"
                # y_row_f -> "Email"
                # y_row_g -> "No. HP"
                # y_row_h -> "Peran"
                # y_row_i -> "Tugas"
                
                y_row_a = [l.replace('\r','').replace('\n','').strip()
                      for l in content.xpath(y_base + '/td[2]/text()')]

                y_row_b = [l.replace('\r','').replace('\n','').strip()
                      for l in content.xpath(y_base + '/td[3]/text()')]

                y_row_c = [l.replace('\r','').replace('\n','').strip()
                      for l in content.xpath(y_base + '/td[4]/text()')]

                y_row_d = [l.replace('\r','').replace('\n','').strip()
                      for l in content.xpath(y_base + '/td[5]/text()')]

                y_row_e = [l.replace('\r','').replace('\n','').strip()
                      for l in content.xpath(y_base + '/td[6]/text()')]

                y_row_f = [l.replace('\r','').replace('\n','').strip()
                      for l in content.xpath(y_base + '/td[7]/text()')]

                y_row_g = [l.replace('\r','').replace('\n','').strip()
                      for l in content.xpath(y_base + '/td[8]/text()')]

                y_row_h = [l.replace('\r','').replace('\n','').strip()
                      for l in content.xpath(y_base + '/td[9]/text()')]

                y_row_i = [l.replace('\r','').replace('\n','').strip()
                      for l in content.xpath(y_base + '/td[10]/text()')]

                # The starting column coordinate for filling the 'LUARAN' table data
                col_start = 25
                
                # Iterating through each table row and write to the spreadsheet
                # Assumes the lists y_row_a, y_row_b, ... have the same array size
                for j in range(len(y_row_a)):

                    # Painting table data
                    sheet.cell(row=row_start, column=col_start).value = y_row_a[j]
                    sheet.cell(row=row_start, column=col_start+1).value = y_row_b[j]
                    sheet.cell(row=row_start, column=col_start+2).value = y_row_c[j]
                    sheet.cell(row=row_start, column=col_start+3).value = y_row_d[j]
                    sheet.cell(row=row_start, column=col_start+4).value = y_row_e[j]
                    sheet.cell(row=row_start, column=col_start+5).value = y_row_f[j]
                    sheet.cell(row=row_start, column=col_start+6).value = y_row_g[j]
                    sheet.cell(row=row_start, column=col_start+7).value = y_row_h[j]
                    sheet.cell(row=row_start, column=col_start+8).value = y_row_i[j]

                    # DEBUG
                    # Please comment out after use
                    # ---
                    '''
                    print()
                    print('-------------------------------------------------------------------------------------------------------')
                    print(f' > DEBUG LOGGING FOR DATA ITERATION NO. {i}: ANGGOTA DATA {j}')
                    print('-------------------------------------------------------------------------------------------------------')
                    print(f'y_row_a[{j}] = {y_row_a[j]}')
                    print(f'y_row_b[{j}] = {y_row_b[j]}')
                    print(f'y_row_c[{j}] = {y_row_c[j]}')
                    print(f'y_row_d[{j}] = {y_row_d[j]}')
                    print(f'y_row_e[{j}] = {y_row_e[j]}')
                    print(f'y_row_f[{j}] = {y_row_f[j]}')
                    print(f'y_row_g[{j}] = {y_row_g[j]}')
                    print(f'y_row_h[{j}] = {y_row_h[j]}')
                    print(f'y_row_i[{j}] = {y_row_i[j]}')
                    print('-------------------------------------------------------------------------------------------------------', end='')
                    print()
                    '''

                    # Incrementing the 'col_start' iterator before continuing the loop
                    col_start += 9
                    continue
                
            # Incrementing the value of 'row_start' before continuing
            row_start += 1

            # Reopening the "Arsip Penelitian" list page,
            # then assign the AJAX response to the temporary array 'temporary_prompt'
            # The 'data' array is obtained from opening individual entry row detail page
            #
            # This is done only on the 7-multiple row iteration, because
            # a single set of 'viewstate', 'viewstategen', and 'eventvalidation' values of ASPX
            # can only be used to do at most 8 operations before having to be renewed.
            if ( i % 7 == 0 ) and ( i > 0 ):
                temporary_prompt = self.get_risat_penelitian_arsip(data)
            continue
        
        # Post-loop logging: appending the header over the 'DATA ANGGOTA' columns
        control.append_message_area(f'+ Melengkapi kepala tabel pada bagian "Identitas Pengusul - Anggota Peneliti" ...')
        # ---
        # The starting column for the 'DATA ANGGOTA' data
        col_start = 25
        # Beginning the loop that detects the maximum number of 'LUARAN' rows
        # according to the variable 'y_max_row'
        for i in range(1, y_max_row+1):
            # Setting the top header
            sheet.merge_cells(
                start_row=1,
                start_column=col_start,
                end_row=1,
                end_column=col_start+8
            )
            sheet.cell(row=1, column=col_start).value = f'IDENTITAS PENGUSUL — ANGGOTA #{i}'
            sheet.cell(row=1, column=col_start).alignment = Alignment(horizontal='center')
            # Setting the sub headers
            sheet.cell(row=2, column=col_start).value = 'N.I.P'
            sheet.cell(row=2, column=col_start+1).value = 'Nama'
            sheet.cell(row=2, column=col_start+2).value = 'Bidang Keahlian'
            sheet.cell(row=2, column=col_start+3).value = 'Alamat'
            sheet.cell(row=2, column=col_start+4).value = 'Instansi'
            sheet.cell(row=2, column=col_start+5).value = 'Email'
            sheet.cell(row=2, column=col_start+6).value = 'No. HP'
            sheet.cell(row=2, column=col_start+7).value = 'Peran'
            sheet.cell(row=2, column=col_start+8).value = 'Tugas'
            # Incrementing the 'col_start' iterator before continuing the loop
            col_start += 9
            continue

        # Post-loop logging: successfully painted the output spreadsheet file
        control.append_message_area(f'+ Tabel sukses dipanen!')
        control.set_progress_bar(85)

        # Asking for the spreadsheet name to save as
        # ---
        # Logging and setting the progress bar
        control.append_message_area(f'+ Menyimpan spreadsheet luaran ...')
        control.set_progress_bar(90)
        # Dealing with file name prompt and saving
        # Using loop to mitigate the user clicking 'cancel'
        # in the file name dialog prompt
        while True:
            # Opening the dialog prompt
            output_spreadsheet = filedialog.asksaveasfilename(
                filetypes=[('Excel files', '*.xlsx')],
                initialfile='Sipesat - Arsip Penelitian Detil Risat.xlsx',
                title='Simpan sebagai ...'
            )

            # 'cancel' button in the dialog prompt is clicked
            if len(output_spreadsheet) == 0:
                # Showing confirmation
                x = messagebox.askyesno(
                    'Nama File Kosong',
                    'Apakah Anda yakin ingin melanjutkan tanpa menyimpan file spreadsheet hasil pemanenan?'
                )
                # Determining whether to break or to continue the loop
                # based on the inversed value of 'x'
                if x:
                    control.append_message_area(
                        f'+ Finalisasi pemanenan data tanpa menyimpan file spreadsheet luaran ...')
                    workbook.close()  # --- closing the workbook without saving
                    break
                else:
                    continue  # --- continuing the loop
            # File name does not end in spreadsheet extension
            elif output_spreadsheet[-5:] != '.xlsx':
                output_spreadsheet = output_spreadsheet + '.xlsx'

            # Saving the spreadsheet
            control.append_message_area(f'LOKASI_SPREADSHEET_LUARAN: {output_spreadsheet}')
            control.set_progress_bar(95)
            workbook.save(output_spreadsheet)

            # Closing the openpyxl workbook
            control.append_message_area(f'+ Menutup file spreadsheet ...')
            control.set_progress_bar(98)
            workbook.close()

            # Breaking the loop
            break

        # Notify for a successful scraping
        control.append_message_area(f'+ Pemanenan selesai pada: {str(dt.now())}')
        control.set_progress_bar(100)
        control.on_notify_successful_scraping()

    # This function harvests "Risat Laporan Akhir Penelitian > Data Detil Lengkap" data
    # and then store the harvested data as an excel file
    #
    # Required arguments:
    # - control             --> for updating the progress bar and
    #                           message area of the screen SipesatScrHarvest
    # - username, password  --> the Risat administrator username and password
    def run_harvest_r_1_lapakhir(self, control, username, password):

        # SipesatScrHarvest messenger
        control.set_header_desc('Panen Data "Risat Laporan Akhir Penelitian > Data Detil Lengkap"')
        control.set_help_label('Data sedang dipanen. Silahkan menunggu.')
        control.set_progress_bar(0)
        control.clear_message_area()

        # Preamble logging
        control.append_message_area(f'+ Memulai pemanenan data ...')
        control.append_message_area(f'+ Pemanenan dimulai pada: {str(dt.now())}')
        control.set_progress_bar(5)

        # Preparing the 'data_prompt' arrays
        control.append_message_area(f'+ Log masuk Risat sebagai [{username}] ...')
        control.set_progress_bar(10)
        data_prompt = self.get_risat_login(username, password)
        data_prompt = self.get_risat_penelitian(data_prompt)
        data_prompt = self.get_risat_penelitian_pelak_kegi(data_prompt)
        data_prompt = self.get_risat_penelitian_pelak_kegi_lapakhir(data_prompt)

        # Parsing XML tree content
        control.append_message_area(f'+ Membaca halaman web ...')
        control.set_progress_bar(15)
        content = data_prompt['html_content']

        # Establishing the export spreadsheet file
        control.append_message_area(f'+ Mempersiapkan file spreadsheet luaran ...')
        control.set_progress_bar(20)
        workbook = xl.Workbook()
        sheet = workbook.active
        sheet.title = 'Lap. Akhir Detil Penelitian'

        # Preparing the sheet header
        control.append_message_area(f'+ Mempersiapkan kepala lembar spreadsheet ...')
        control.set_progress_bar(25)
        # ---
        # Preparing the "IDENTITY" header
        sheet.merge_cells('A1:I1')
        sheet['A1'].value = 'IDENTITAS'
        sheet['A1'].alignment = Alignment(horizontal='center')
        # Preparing the "IDENTITY" sub-headers
        sheet['A2'].value = 'No.'
        sheet['B2'].value = 'Judul'
        sheet['C2'].value = 'Ketua'
        sheet['D2'].value = 'Bidang Fokus'
        sheet['E2'].value = 'Tgl. Usulan'
        sheet['F2'].value = 'Jumlah Anggota'
        sheet['G2'].value = 'Biaya'
        sheet['H2'].value = 'Dana Disetujui'
        sheet['I2'].value = 'Lama Kegiatan'

        # The base XPath location, pointing to each entry row
        base = '//div[@id="ContentPlaceHolder1_lapakhirbp3m1_updg1"]//div[@class="panel-body f12"]/table[@width="100%"]'

        # ---
        # Obtaining the data row values
        control.append_message_area(f'+ Mendapatkan data pada baris tabel ...')
        control.set_progress_bar(30)

        # HYPOTHESIS:
        # Xpath cannot detect 'tbody' element.
        # So instead of using 'table/tbody/tr', use 'table//tr' instead
        #
        # RESULT:
        # The hypothesis is correct.
        # Therefore, don't mention 'tbody' in any of the following Xpath paths

        a2 = [str(i)
              for i in range(1, len(content.xpath(base)) + 1)]

        b2 = [l.strip()
              for l in content.xpath(base + '//tr[@valign="top"]/td[@colspan="3"]/b/text()')]

        c2 = [l.replace('Ketua:', '').strip()
              for l in content.xpath(base + '//tr[2][not(descendant::table) and not(ancestor::table[@class="table"])]/td[2]/text()')]

        d2 = [l.replace('Bidang Fokus:', '').strip()
              for l in content.xpath(base + '//tr[3][not(descendant::table) and not(ancestor::table[@class="table"])]/td[@colspan="3"]/text()')]

        e2 = [l.replace('Tgl Usulan:', '').strip()
              for l in content.xpath(base + '//tr[4][not(descendant::table) and not(ancestor::table[@class="table"])]/td[@colspan="3"]/text()')]

        f2 = [l.replace('Jml Anggota:', '').strip()
              for l in content.xpath(base + '//tr[5][not(descendant::table) and not(ancestor::table[@class="table"])]/td[@colspan="3"]/text()')]

        g2 = [l.replace('Biaya:', '').replace('Rp.', '').replace(',', '').split('disetujui')[0].strip()
              for l in content.xpath(base + '//tr[6][not(descendant::table) and not(ancestor::table[@class="table"])]/td[@colspan="3"]/text()')]

        h2 = [l.replace('Biaya:', '').replace('Rp.', '').replace(',', '').split('disetujui')[1].strip()
              for l in content.xpath(base + '//tr[6][not(descendant::table) and not(ancestor::table[@class="table"])]/td[@colspan="3"]/text()')]

        i2 = [l.replace('Lama Kegiatan:', '').strip()
              for l in content.xpath(base + '//tr[7][not(descendant::table) and not(ancestor::table[@class="table"])]/td[@colspan="3"]/text()')]

        # The table which displays 'Luaran Laporan Akhir'
        # Treated differently, specially
        # 'j_' is a DOM element representing a HTML table
        j_ = content.xpath(base + '//tr[9]/td[@colspan="3"]/table[@class="table"]')

        # The starting row coordinate of the active sheet
        row_start = 3

        # DEBUG
        # Please comment out after use
        # ---
        '''
        print()
        print('-------------------------------------------------------------------------------------------------------')
        for a in (a2, b2, c2, d2, e2, f2, g2, h2, i2, j_):
            print('LEN', len(a))
            print('CONTENT', a)
            print()
        print('-------------------------------------------------------------------------------------------------------', end='')
        print()
        '''

        # The maximum number of 'LUARAN' table data row
        j_max_row = 0

        # Iterating through each table row and write to the table
        # Assumes the lists a1, b1, c1, ... have the same array size
        control.append_message_area(f'+ Melakukan iterasi terhadap baris tabel dan menulis spreadsheet luaran ...')
        control.set_progress_bar(35)
        for i in range(len(a2)):
            # Noisy preamble logging
            # Please don't use this -_-
            # ---
            # control.append_message_area(f'ITERASI [{i}]')

            # Updating the progress bar status
            control.set_progress_bar(35 + round(45 * (i + 1) / (len(a2))))

            # Painting the scraped data to the output spreadsheet row
            # Section 1: 'IDENTITY'
            sheet[f'A{row_start}'] = a2[i]
            sheet[f'B{row_start}'] = b2[i]
            sheet[f'C{row_start}'] = c2[i]
            sheet[f'D{row_start}'] = d2[i]
            sheet[f'E{row_start}'] = e2[i]
            sheet[f'F{row_start}'] = f2[i]
            sheet[f'G{row_start}'] = g2[i]
            sheet[f'H{row_start}'] = h2[i]
            sheet[f'I{row_start}'] = i2[i]

            # Painting the scraped data to the output spreadsheet row
            # Section 2: 'LUARAN'
            # ---
            # Checking if 'luaran' table data exists
            data_luaran = j_[i].xpath('.//td')
            if len(data_luaran) == 0:
                pass  # --- nope. the data does not exist
            else:

                # Checking if this table's 'LUARAN' data has the most rows
                j_all_rows = j_[i].xpath('.//tr[not(descendant::th)]')
                j_max_row = max(j_max_row, len(j_all_rows))

                # ROW DATA LABEL CONVENTION
                # j_row_a -> "No. Luaran"
                # j_row_b -> "Jenis Luaran"
                # j_row_c -> "Status Luaran"
                # j_row_d -> "File Laporan Akhir"
                # j_row_e -> "Alamat URL"
                # j_row_f -> "Status Akhir Luaran"
                # j_row_g -> "File PTJ Keuangan"

                j_row_a = [l.strip()
                      for l in j_[i].xpath('.//tr[not(descendant::th)]/td[1]/text()')]

                j_row_b = [l.strip()
                      for l in j_[i].xpath('.//tr[not(descendant::th)]/td[2]/text()')]

                j_row_c = [l.strip()
                      for l in j_[i].xpath('.//tr[not(descendant::th)]/td[3]/text()')]

                j_row_d = [l.strip()
                      for l in j_[i].xpath('.//tr[not(descendant::th)]/td[4]/a/@href')]

                j_row_e = [l.strip()
                      for l in j_[i].xpath('.//tr[not(descendant::th)]/td[5]/a/@href')]

                j_row_f = [l.strip()
                      for l in j_[i].xpath('.//tr[not(descendant::th)]/td[6]/text()')]

                j_row_g = [l.strip()
                      for l in j_[i].xpath('.//tr[not(descendant::th)]/td[7]/a/@href')]

                # The starting column coordinate for filling the 'LUARAN' table data
                col_start = 10

                # DEBUG
                # Please comment out after use
                # ---
                '''
                print()
                print(
                    '-------------------------------------------------------------------------------------------------------')
                for b in (j_row_a, j_row_b, j_row_c, j_row_d, j_row_e, j_row_f, j_row_g):
                    print('LEN', len(b))
                    print('CONTENT', b)
                    print()
                print(
                    '-------------------------------------------------------------------------------------------------------',
                    end='')
                print()
                '''

                # Iterating through each table row and write to the spreadsheet
                # Assumes the lists j_row_a, j_row_b, ... have the same array size
                for j in range(len(j_row_a)):

                    # Painting table data
                    sheet.cell(row=row_start, column=col_start).value = j_row_b[j]
                    sheet.cell(row=row_start, column=col_start+1).value = j_row_c[j]
                    sheet.cell(row=row_start, column=col_start+2).value = j_row_d[j]
                    sheet.cell(row=row_start, column=col_start+3).value = j_row_e[j]
                    sheet.cell(row=row_start, column=col_start+4).value = j_row_f[j]
                    sheet.cell(row=row_start, column=col_start+5).value = j_row_g[j]

                    # Incrementing the 'col_start' iterator before continuing the loop
                    col_start += 6
                    continue

            # Incrementing the 'row_start' iterator
            # Then continue the loop
            row_start += 1
            continue

        # Post-loop logging: appending the header over the 'LUARAN' columns
        control.append_message_area(f'+ Melengkapi kepala tabel pada bagian "LUARAN" ...')
        # ---
        # The starting column for the 'LUARAN' data
        col_start = 10
        # Beginning the loop that detects the maximum number of 'LUARAN' rows
        # according to the variable 'j_max_row'
        for i in range(1, j_max_row+1):
            # Setting the top header
            sheet.merge_cells(
                start_row=1,
                start_column=col_start,
                end_row=1,
                end_column=col_start+5
            )
            sheet.cell(row=1, column=col_start).value = f'LUARAN #{i}'
            sheet.cell(row=1, column=col_start).alignment = Alignment(horizontal='center')
            # Setting the sub headers
            sheet.cell(row=2, column=col_start).value = 'Jenis Luaran'
            sheet.cell(row=2, column=col_start+1).value = 'Status Luaran'
            sheet.cell(row=2, column=col_start+2).value = 'File Lap. Akhir'
            sheet.cell(row=2, column=col_start+3).value = 'Alamat URL'
            sheet.cell(row=2, column=col_start+4).value = 'Status Akhir Luaran'
            sheet.cell(row=2, column=col_start+5).value = 'File PTJ Keuangan'
            # Incrementing the 'col_start' iterator before continuing the loop
            col_start += 6
            continue

        # Post-loop logging: successfully painted the output spreadsheet file
        control.append_message_area(f'+ Tabel sukses dipanen!')
        control.set_progress_bar(85)

        # Asking for the spreadsheet name to save as
        # ---
        # Logging and setting the progress bar
        control.append_message_area(f'+ Menyimpan spreadsheet luaran ...')
        control.set_progress_bar(90)
        # Dealing with file name prompt and saving
        # Using loop to mitigate the user clicking 'cancel'
        # in the file name dialog prompt
        while True:
            # Opening the dialog prompt
            output_spreadsheet = filedialog.asksaveasfilename(
                filetypes=[('Excel files', '*.xlsx')],
                initialfile='Sipesat - Lap Akhir Penelitian Detil Risat.xlsx',
                title='Simpan sebagai ...'
            )

            # 'cancel' button in the dialog prompt is clicked
            if len(output_spreadsheet) == 0:
                # Showing confirmation
                x = messagebox.askyesno(
                    'Nama File Kosong',
                    'Apakah Anda yakin ingin melanjutkan tanpa menyimpan file spreadsheet hasil pemanenan?'
                )
                # Determining whether to break or to continue the loop
                # based on the inversed value of 'x'
                if x:
                    control.append_message_area(
                        f'+ Finalisasi pemanenan data tanpa menyimpan file spreadsheet luaran ...')
                    workbook.close()  # --- closing the workbook without saving
                    break
                else:
                    continue  # --- continuing the loop
            # File name does not end in spreadsheet extension
            elif output_spreadsheet[-5:] != '.xlsx':
                output_spreadsheet = output_spreadsheet + '.xlsx'

            # Saving the spreadsheet
            control.append_message_area(f'LOKASI_SPREADSHEET_LUARAN: {output_spreadsheet}')
            control.set_progress_bar(95)
            workbook.save(output_spreadsheet)

            # Closing the openpyxl workbook
            control.append_message_area(f'+ Menutup file spreadsheet ...')
            control.set_progress_bar(98)
            workbook.close()

            # Breaking the loop
            break

        # Notify for a successful scraping
        control.append_message_area(f'+ Pemanenan selesai pada: {str(dt.now())}')
        control.set_progress_bar(100)
        control.on_notify_successful_scraping()

    # This function harvests "Risat Usulan Pengabdian > Ringkasan Data" data
    # and then store the harvested data as an excel file
    #
    # Required arguments:
    # - control             --> for updating the progress bar and
    #                           message area of the screen SipesatScrHarvest
    # - username, password  --> the Risat administrator username and password
    def run_harvest_c_0_usulan(self, control, username, password):

        # SipesatScrHarvest messenger
        control.set_header_desc('Panen Data "Risat Usulan Pengabdian > Ringkasan Data"')
        control.set_help_label('Data sedang dipanen. Silahkan menunggu.')
        control.set_progress_bar(0)
        control.clear_message_area()

        # Preamble logging
        control.append_message_area(f'+ Memulai pemanenan data ...')
        control.append_message_area(f'+ Pemanenan dimulai pada: {str(dt.now())}')
        control.set_progress_bar(5)

        # Preparing the 'data_prompt' arrays
        control.append_message_area(f'+ Log masuk Risat sebagai [{username}] ...')
        control.set_progress_bar(10)
        data_prompt = self.get_risat_login(username, password)
        data_prompt = self.get_risat_pengabdian(data_prompt)
        data_prompt = self.get_risat_pengabdian_usulan_pengabdian(data_prompt)

        # Parsing XML tree content
        control.append_message_area(f'+ Membaca halaman web ...')
        control.set_progress_bar(15)
        content = data_prompt['html_content']

        # Establishing the export spreadsheet file
        control.append_message_area(f'+ Mempersiapkan file spreadsheet luaran ...')
        control.set_progress_bar(20)
        workbook = xl.Workbook()
        sheet = workbook.active
        sheet.title = 'Usulan Pengabdian Ringkasan'

        # Preparing the sheet header
        control.append_message_area(f'+ Mempersiapkan kepala lembar spreadsheet ...')
        control.set_progress_bar(25)
        sheet['A1'].value = 'No.'
        sheet['B1'].value = 'Judul'
        sheet['C1'].value = 'Ketua'
        sheet['D1'].value = 'Jml. Anggota'
        sheet['E1'].value = 'Tgl. Usulan'
        sheet['F1'].value = 'Bidang Fokus'
        sheet['G1'].value = 'Rencana Biaya'
        sheet['H1'].value = 'Lama Kegiatan'

        # The base XPath location, pointing to each entry row
        base = '//div[@class="mw-100"]//div[@class="form-group f12"]/table[@width="100%"]//tr[@valign="top"]'

        # ---
        # Obtaining the data row values
        control.append_message_area(f'+ Mendapatkan data pada baris tabel ...')
        control.set_progress_bar(30)

        # HYPOTHESIS:
        # Xpath cannot detect 'tbody' element.
        # So instead of using 'table/tbody/tr', use 'table//tr' instead
        #
        # RESULT:
        # The hypothesis is correct.
        # Therefore, don't mention 'tbody' in any of the following Xpath paths

        a1 = [str(i)
              for i in range(1, len(content.xpath(base))+1)]

        b1 = [l.strip()
              for l in content.xpath(base + '//span[@class="hijau"]/text()')]

        c1 = [l.replace('Ketua:', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[2]/td/table//tr/td[1]/text()')]

        d1 = [l.replace('Jumlah Anggota:', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[2]/td/table//tr/td[3]/text()')]

        e1 = [l.replace('Tgl Usulan:', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[3]/td[1]/text()')]

        f1 = [l.replace('Bidang Fokus:', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[4]/td[1]/text()')]

        g1 = [l.replace('Rencana Biaya:', '').replace('Rp.', '').replace(',', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[5]/td[1]/text()')]

        h1 = [l.replace('Lama Kegiatan:', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[6]/td[1]/text()[1]')]

        # The starting row coordinate of the active sheet
        row_start = 2

        # DEBUG
        # Please comment out after use
        # ---
        # print(a1, b1, c1, d1, e1, f1, g1, h1, i1, j1, k1, l1)

        # Iterating through each table row and write to the table
        # Assumes the lists a1, b1, c1, ... have the same array size
        control.append_message_area(f'+ Melakukan iterasi terhadap baris tabel dan menulis spreadsheet luaran ...')
        control.set_progress_bar(35)
        for i in range(len(a1)):

            # Noisy preamble logging
            # Please don't use this -_-
            # ---
            # control.append_message_area(f'ITERASI [{i}]')

            # Updating the progress bar status
            control.set_progress_bar(35 + round(45*(i+1)/(len(a1))))

            # Painting the scraped data to the output spreadsheet row
            sheet[f'A{row_start}'] = a1[i]
            sheet[f'B{row_start}'] = b1[i]
            sheet[f'C{row_start}'] = c1[i]
            sheet[f'D{row_start}'] = d1[i]
            sheet[f'E{row_start}'] = e1[i]
            sheet[f'F{row_start}'] = f1[i]
            sheet[f'G{row_start}'] = g1[i]
            sheet[f'H{row_start}'] = h1[i]

            # Incrementing the 'row_start' iterator
            # Then continue the loop
            row_start += 1
            continue

        # Post-loop logging: successfully painted the output spreadsheet file
        control.append_message_area(f'+ Tabel sukses dipanen!')
        control.set_progress_bar(85)

        # Asking for the spreadsheet name to save as
        # ---
        # Logging and setting the progress bar
        control.append_message_area(f'+ Menyimpan spreadsheet luaran ...')
        control.set_progress_bar(90)
        # Dealing with file name prompt and saving
        # Using loop to mitigate the user clicking 'cancel'
        # in the file name dialog prompt
        while True:
            # Opening the dialog prompt
            output_spreadsheet = filedialog.asksaveasfilename(
                filetypes=[('Excel files', '*.xlsx')],
                initialfile='Sipesat - Usulan Pengabdian Ringkasan Risat.xlsx',
                title='Simpan sebagai ...'
            )

            # 'cancel' button in the dialog prompt is clicked
            if len(output_spreadsheet) == 0:
                # Showing confirmation
                x = messagebox.askyesno(
                    'Nama File Kosong',
                    'Apakah Anda yakin ingin melanjutkan tanpa menyimpan file spreadsheet hasil pemanenan?'
                )
                # Determining whether to break or to continue the loop
                # based on the inversed value of 'x'
                if x:
                    control.append_message_area(f'+ Finalisasi pemanenan data tanpa menyimpan file spreadsheet luaran ...')
                    workbook.close()  # --- closing the workbook without saving
                    break
                else:
                    continue  # --- continuing the loop
            # File name does not end in spreadsheet extension
            elif output_spreadsheet[-5:] != '.xlsx':
                output_spreadsheet = output_spreadsheet + '.xlsx'

            # Saving the spreadsheet
            control.append_message_area(f'LOKASI_SPREADSHEET_LUARAN: {output_spreadsheet}')
            control.set_progress_bar(95)
            workbook.save(output_spreadsheet)

            # Closing the openpyxl workbook
            control.append_message_area(f'+ Menutup file spreadsheet ...')
            control.set_progress_bar(98)
            workbook.close()

            # Breaking the loop
            break

        # Notify for a successful scraping
        control.append_message_area(f'+ Pemanenan selesai pada: {str(dt.now())}')
        control.set_progress_bar(100)
        control.on_notify_successful_scraping()

    # This function harvests "Risat Berkas Disetujui Pengabdian > Ringkasan Data" data
    # and then store the harvested data as an excel file
    #
    # Required arguments:
    # - control             --> for updating the progress bar and
    #                           message area of the screen SipesatScrHarvest
    # - username, password  --> the Risat administrator username and password
    def run_harvest_c_0_berdisetujui(self, control, username, password):

        # SipesatScrHarvest messenger
        control.set_header_desc('Panen Data "Risat Berkas Disetujui Pengabdian > Ringkasan Data"')
        control.set_help_label('Data sedang dipanen. Silahkan menunggu.')
        control.set_progress_bar(0)
        control.clear_message_area()

        # Preamble logging
        control.append_message_area(f'+ Memulai pemanenan data ...')
        control.append_message_area(f'+ Pemanenan dimulai pada: {str(dt.now())}')
        control.set_progress_bar(5)

        # Preparing the 'data_prompt' arrays
        control.append_message_area(f'+ Log masuk Risat sebagai [{username}] ...')
        control.set_progress_bar(10)
        data_prompt = self.get_risat_login(username, password)
        data_prompt = self.get_risat_pengabdian(data_prompt)
        data_prompt = self.get_risat_pengabdian_berdisetujui_pengabdian(data_prompt)

        # Parsing XML tree content
        control.append_message_area(f'+ Membaca halaman web ...')
        control.set_progress_bar(15)
        content = data_prompt['html_content']

        # Establishing the export spreadsheet file
        control.append_message_area(f'+ Mempersiapkan file spreadsheet luaran ...')
        control.set_progress_bar(20)
        workbook = xl.Workbook()
        sheet = workbook.active
        sheet.title = 'Berkas Disetujui Pengabdian Ringkasan'

        # Preparing the sheet header
        control.append_message_area(f'+ Mempersiapkan kepala lembar spreadsheet ...')
        control.set_progress_bar(25)
        sheet['A1'].value = 'No.'
        sheet['B1'].value = 'Judul'
        sheet['C1'].value = 'Ketua'
        sheet['D1'].value = 'Jml. Anggota'
        sheet['E1'].value = 'Tgl. Usulan'
        sheet['F1'].value = 'Bidang Fokus'
        sheet['G1'].value = 'Rencana Biaya'
        sheet['H1'].value = 'Lama Kegiatan'

        # The base XPath location, pointing to each entry row
        base = '//div[@class="mw-100"]//div[@class="form-group f12"]/table[@width="100%"]//tr[@valign="top"]'

        # ---
        # Obtaining the data row values
        control.append_message_area(f'+ Mendapatkan data pada baris tabel ...')
        control.set_progress_bar(30)

        # HYPOTHESIS:
        # Xpath cannot detect 'tbody' element.
        # So instead of using 'table/tbody/tr', use 'table//tr' instead
        #
        # RESULT:
        # The hypothesis is correct.
        # Therefore, don't mention 'tbody' in any of the following Xpath paths

        a1 = [str(i)
              for i in range(1, len(content.xpath(base))+1)]

        b1 = [l.strip()
              for l in content.xpath(base + '//span[@class="hijau"]/text()')]

        c1 = [l.replace('Ketua:', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[2]/td/table//tr/td[1]/text()')]

        d1 = [l.replace('Jumlah Anggota:', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[2]/td/table//tr/td[3]/text()')]

        e1 = [l.replace('Tgl Usulan:', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[3]/td[1]/text()')]

        f1 = [l.replace('Bidang Fokus:', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[4]/td[1]/text()')]

        g1 = [l.replace('Rencana Biaya:', '').replace('Rp.', '').replace(',', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[5]/td[1]/text()')]

        h1 = [l.replace('Lama Kegiatan:', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[6]/td[1]/text()[1]')]

        # The starting row coordinate of the active sheet
        row_start = 2

        # DEBUG
        # Please comment out after use
        # ---
        # print(a1, b1, c1, d1, e1, f1, g1, h1, i1, j1, k1, l1)

        # Iterating through each table row and write to the table
        # Assumes the lists a1, b1, c1, ... have the same array size
        control.append_message_area(f'+ Melakukan iterasi terhadap baris tabel dan menulis spreadsheet luaran ...')
        control.set_progress_bar(35)
        for i in range(len(a1)):

            # Noisy preamble logging
            # Please don't use this -_-
            # ---
            # control.append_message_area(f'ITERASI [{i}]')

            # Updating the progress bar status
            control.set_progress_bar(35 + round(45*(i+1)/(len(a1))))

            # Painting the scraped data to the output spreadsheet row
            sheet[f'A{row_start}'] = a1[i]
            sheet[f'B{row_start}'] = b1[i]
            sheet[f'C{row_start}'] = c1[i]
            sheet[f'D{row_start}'] = d1[i]
            sheet[f'E{row_start}'] = e1[i]
            sheet[f'F{row_start}'] = f1[i]
            sheet[f'G{row_start}'] = g1[i]
            sheet[f'H{row_start}'] = h1[i]

            # Incrementing the 'row_start' iterator
            # Then continue the loop
            row_start += 1
            continue

        # Post-loop logging: successfully painted the output spreadsheet file
        control.append_message_area(f'+ Tabel sukses dipanen!')
        control.set_progress_bar(85)

        # Asking for the spreadsheet name to save as
        # ---
        # Logging and setting the progress bar
        control.append_message_area(f'+ Menyimpan spreadsheet luaran ...')
        control.set_progress_bar(90)
        # Dealing with file name prompt and saving
        # Using loop to mitigate the user clicking 'cancel'
        # in the file name dialog prompt
        while True:
            # Opening the dialog prompt
            output_spreadsheet = filedialog.asksaveasfilename(
                filetypes=[('Excel files', '*.xlsx')],
                initialfile='Sipesat - Berkas Disetujui Pengabdian Ringkasan Risat.xlsx',
                title='Simpan sebagai ...'
            )

            # 'cancel' button in the dialog prompt is clicked
            if len(output_spreadsheet) == 0:
                # Showing confirmation
                x = messagebox.askyesno(
                    'Nama File Kosong',
                    'Apakah Anda yakin ingin melanjutkan tanpa menyimpan file spreadsheet hasil pemanenan?'
                )
                # Determining whether to break or to continue the loop
                # based on the inversed value of 'x'
                if x:
                    control.append_message_area(f'+ Finalisasi pemanenan data tanpa menyimpan file spreadsheet luaran ...')
                    workbook.close()  # --- closing the workbook without saving
                    break
                else:
                    continue  # --- continuing the loop
            # File name does not end in spreadsheet extension
            elif output_spreadsheet[-5:] != '.xlsx':
                output_spreadsheet = output_spreadsheet + '.xlsx'

            # Saving the spreadsheet
            control.append_message_area(f'LOKASI_SPREADSHEET_LUARAN: {output_spreadsheet}')
            control.set_progress_bar(95)
            workbook.save(output_spreadsheet)

            # Closing the openpyxl workbook
            control.append_message_area(f'+ Menutup file spreadsheet ...')
            control.set_progress_bar(98)
            workbook.close()

            # Breaking the loop
            break

        # Notify for a successful scraping
        control.append_message_area(f'+ Pemanenan selesai pada: {str(dt.now())}')
        control.set_progress_bar(100)
        control.on_notify_successful_scraping()

    # This function harvests "Risat Berkas Direvisi Pengabdian > Ringkasan Data" data
    # and then store the harvested data as an excel file
    #
    # Required arguments:
    # - control             --> for updating the progress bar and
    #                           message area of the screen SipesatScrHarvest
    # - username, password  --> the Risat administrator username and password
    def run_harvest_c_0_berdirevisi(self, control, username, password):

        # SipesatScrHarvest messenger
        control.set_header_desc('Panen Data "Risat Berkas Direvisi Pengabdian > Ringkasan Data"')
        control.set_help_label('Data sedang dipanen. Silahkan menunggu.')
        control.set_progress_bar(0)
        control.clear_message_area()

        # Preamble logging
        control.append_message_area(f'+ Memulai pemanenan data ...')
        control.append_message_area(f'+ Pemanenan dimulai pada: {str(dt.now())}')
        control.set_progress_bar(5)

        # Preparing the 'data_prompt' arrays
        control.append_message_area(f'+ Log masuk Risat sebagai [{username}] ...')
        control.set_progress_bar(10)
        data_prompt = self.get_risat_login(username, password)
        data_prompt = self.get_risat_pengabdian(data_prompt)
        data_prompt = self.get_risat_pengabdian_berdirevisi_pengabdian(data_prompt)

        # Parsing XML tree content
        control.append_message_area(f'+ Membaca halaman web ...')
        control.set_progress_bar(15)
        content = data_prompt['html_content']

        # Establishing the export spreadsheet file
        control.append_message_area(f'+ Mempersiapkan file spreadsheet luaran ...')
        control.set_progress_bar(20)
        workbook = xl.Workbook()
        sheet = workbook.active
        sheet.title = 'Berkas Direvisi Pengabdian Ringkasan'

        # Preparing the sheet header
        control.append_message_area(f'+ Mempersiapkan kepala lembar spreadsheet ...')
        control.set_progress_bar(25)
        sheet['A1'].value = 'No.'
        sheet['B1'].value = 'Judul'
        sheet['C1'].value = 'Ketua'
        sheet['D1'].value = 'Jml. Anggota'
        sheet['E1'].value = 'Tgl. Usulan'
        sheet['F1'].value = 'Bidang Fokus'
        sheet['G1'].value = 'Rencana Biaya'
        sheet['H1'].value = 'Lama Kegiatan'

        # The base XPath location, pointing to each entry row
        base = '//div[@class="mw-100"]//div[@class="form-group f12"]/table[@width="100%"]//tr[@valign="top"]'

        # ---
        # Obtaining the data row values
        control.append_message_area(f'+ Mendapatkan data pada baris tabel ...')
        control.set_progress_bar(30)

        # HYPOTHESIS:
        # Xpath cannot detect 'tbody' element.
        # So instead of using 'table/tbody/tr', use 'table//tr' instead
        #
        # RESULT:
        # The hypothesis is correct.
        # Therefore, don't mention 'tbody' in any of the following Xpath paths

        a1 = [str(i)
              for i in range(1, len(content.xpath(base))+1)]

        b1 = [l.strip()
              for l in content.xpath(base + '//span[@class="hijau"]/text()')]

        c1 = [l.replace('Ketua:', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[2]/td/table//tr/td[1]/text()')]

        d1 = [l.replace('Jumlah Anggota:', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[2]/td/table//tr/td[3]/text()')]

        e1 = [l.replace('Tgl Usulan:', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[3]/td[1]/text()')]

        f1 = [l.replace('Bidang Fokus:', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[4]/td[1]/text()')]

        g1 = [l.replace('Rencana Biaya:', '').replace('Rp.', '').replace(',', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[5]/td[1]/text()')]

        h1 = [l.replace('Lama Kegiatan:', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[6]/td[1]/text()[1]')]

        # The starting row coordinate of the active sheet
        row_start = 2

        # DEBUG
        # Please comment out after use
        # ---
        # print(a1, b1, c1, d1, e1, f1, g1, h1, i1, j1, k1, l1)

        # Iterating through each table row and write to the table
        # Assumes the lists a1, b1, c1, ... have the same array size
        control.append_message_area(f'+ Melakukan iterasi terhadap baris tabel dan menulis spreadsheet luaran ...')
        control.set_progress_bar(35)
        for i in range(len(a1)):

            # Noisy preamble logging
            # Please don't use this -_-
            # ---
            # control.append_message_area(f'ITERASI [{i}]')

            # Updating the progress bar status
            control.set_progress_bar(35 + round(45*(i+1)/(len(a1))))

            # Painting the scraped data to the output spreadsheet row
            sheet[f'A{row_start}'] = a1[i]
            sheet[f'B{row_start}'] = b1[i]
            sheet[f'C{row_start}'] = c1[i]
            sheet[f'D{row_start}'] = d1[i]
            sheet[f'E{row_start}'] = e1[i]
            sheet[f'F{row_start}'] = f1[i]
            sheet[f'G{row_start}'] = g1[i]
            sheet[f'H{row_start}'] = h1[i]

            # Incrementing the 'row_start' iterator
            # Then continue the loop
            row_start += 1
            continue

        # Post-loop logging: successfully painted the output spreadsheet file
        control.append_message_area(f'+ Tabel sukses dipanen!')
        control.set_progress_bar(85)

        # Asking for the spreadsheet name to save as
        # ---
        # Logging and setting the progress bar
        control.append_message_area(f'+ Menyimpan spreadsheet luaran ...')
        control.set_progress_bar(90)
        # Dealing with file name prompt and saving
        # Using loop to mitigate the user clicking 'cancel'
        # in the file name dialog prompt
        while True:
            # Opening the dialog prompt
            output_spreadsheet = filedialog.asksaveasfilename(
                filetypes=[('Excel files', '*.xlsx')],
                initialfile='Sipesat - Berkas Direvisi Pengabdian Ringkasan Risat.xlsx',
                title='Simpan sebagai ...'
            )

            # 'cancel' button in the dialog prompt is clicked
            if len(output_spreadsheet) == 0:
                # Showing confirmation
                x = messagebox.askyesno(
                    'Nama File Kosong',
                    'Apakah Anda yakin ingin melanjutkan tanpa menyimpan file spreadsheet hasil pemanenan?'
                )
                # Determining whether to break or to continue the loop
                # based on the inversed value of 'x'
                if x:
                    control.append_message_area(f'+ Finalisasi pemanenan data tanpa menyimpan file spreadsheet luaran ...')
                    workbook.close()  # --- closing the workbook without saving
                    break
                else:
                    continue  # --- continuing the loop
            # File name does not end in spreadsheet extension
            elif output_spreadsheet[-5:] != '.xlsx':
                output_spreadsheet = output_spreadsheet + '.xlsx'

            # Saving the spreadsheet
            control.append_message_area(f'LOKASI_SPREADSHEET_LUARAN: {output_spreadsheet}')
            control.set_progress_bar(95)
            workbook.save(output_spreadsheet)

            # Closing the openpyxl workbook
            control.append_message_area(f'+ Menutup file spreadsheet ...')
            control.set_progress_bar(98)
            workbook.close()

            # Breaking the loop
            break

        # Notify for a successful scraping
        control.append_message_area(f'+ Pemanenan selesai pada: {str(dt.now())}')
        control.set_progress_bar(100)
        control.on_notify_successful_scraping()

    # This function harvests "Risat Ditolak Pengabdian > Ringkasan Data" data
    # and then store the harvested data as an excel file
    #
    # Required arguments:
    # - control             --> for updating the progress bar and
    #                           message area of the screen SipesatScrHarvest
    # - username, password  --> the Risat administrator username and password
    def run_harvest_c_0_ditolak(self, control, username, password):
        # SipesatScrHarvest messenger
        control.set_header_desc('Panen Data "Risat Ditolak Pengabdian > Ringkasan Data"')
        control.set_help_label('Data sedang dipanen. Silahkan menunggu.')
        control.set_progress_bar(0)
        control.clear_message_area()

        # Preamble logging
        control.append_message_area(f'+ Memulai pemanenan data ...')
        control.append_message_area(f'+ Pemanenan dimulai pada: {str(dt.now())}')
        control.set_progress_bar(5)

        # Preparing the 'data_prompt' arrays
        control.append_message_area(f'+ Log masuk Risat sebagai [{username}] ...')
        control.set_progress_bar(10)
        data_prompt = self.get_risat_login(username, password)
        data_prompt = self.get_risat_pengabdian(data_prompt)
        data_prompt = self.get_risat_pengabdian_ditolak_pengabdian(data_prompt)

        # Parsing XML tree content
        control.append_message_area(f'+ Membaca halaman web ...')
        control.set_progress_bar(15)
        content = data_prompt['html_content']

        # Establishing the export spreadsheet file
        control.append_message_area(f'+ Mempersiapkan file spreadsheet luaran ...')
        control.set_progress_bar(20)
        workbook = xl.Workbook()
        sheet = workbook.active
        sheet.title = 'Ditolak Pengabdian Ringkasan'

        # Preparing the sheet header
        control.append_message_area(f'+ Mempersiapkan kepala lembar spreadsheet ...')
        control.set_progress_bar(25)
        # ---
        # Preparing the "NOMOR" header
        sheet.merge_cells(
            start_row=1, start_column=1, end_row=2, end_column=1
        )
        sheet['A1'].value = 'No.'
        sheet['A1'].alignment = Alignment(horizontal='center')
        # ---
        # Preparing the "JUDUL" header
        sheet.merge_cells(
            start_row=1, start_column=2, end_row=1, end_column=11
        )
        sheet['B1'].value = 'IDENTITAS'
        sheet['B1'].alignment = Alignment(horizontal='center')
        # Preparing the "JUDUL" sub-headers
        sheet['B2'].value = 'Judul'
        sheet['C2'].value = 'Ketua'
        sheet['D2'].value = 'Jml. Anggota'
        sheet['E2'].value = 'Tgl. Usulan'
        sheet['F2'].value = 'Bidang Fokus'
        sheet['G2'].value = 'Rencana Biaya'
        sheet['H2'].value = 'Lama Kegiatan'
        sheet['I2'].value = 'Biaya Setelah Revisi'
        sheet['J2'].value = 'Catatan Revisi'
        sheet['K2'].value = 'File Revisi'
        # ---
        # Preparing the "REVIEWER 1" header
        sheet.merge_cells(
            start_row=1, start_column=12, end_row=1, end_column=15
        )
        sheet['L1'].value = 'REVIEWER 1'
        sheet['L1'].alignment = Alignment(horizontal='center')
        # Preparing the "REVIEWER 1" sub-headers
        sheet['L2'].value = 'Nama Reviewer'
        sheet['M2'].value = 'Nilai'
        sheet['N2'].value = 'Rekomendasi Dana'
        sheet['O2'].value = 'Komentar'
        # ---
        # Preparing the "REVIEWER 2" header
        sheet.merge_cells(
            start_row=1, start_column=16, end_row=1, end_column=19
        )
        sheet['P1'].value = 'REVIEWER 2'
        sheet['P1'].alignment = Alignment(horizontal='center')
        # Preparing the "REVIEWER 2" sub-headers
        sheet['P2'].value = 'Nama Reviewer'
        sheet['Q2'].value = 'Nilai'
        sheet['R2'].value = 'Rekomendasi Dana'
        sheet['S2'].value = 'Komentar'

        # The base XPath location, pointing to each entry row
        base = '//div[@class="mw-100"]//div[@class="form-group f12"]/table[@width="100%"]//tr[@valign="top"]'

        # ---
        # Obtaining the data row values
        control.append_message_area(f'+ Mendapatkan data pada baris tabel ...')
        control.set_progress_bar(30)

        # HYPOTHESIS:
        # Xpath cannot detect 'tbody' element.
        # So instead of using 'table/tbody/tr', use 'table//tr' instead
        #
        # RESULT:
        # The hypothesis is correct.
        # Therefore, don't mention 'tbody' in any of the following Xpath paths

        a1 = [str(i)
              for i in range(1, len(content.xpath(base))+1)]

        b2 = [l.strip()
              for l in content.xpath(base + '//span[@class="hijau"]/text()')]

        c2 = [l.replace('Ketua:', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[2]/td/table//tr/td[1]/text()')]

        d2 = [l.replace('Jumlah Anggota:', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[2]/td/table//tr/td[3]/text()')]

        e2 = [l.replace('Tgl Usulan:', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[3]/td[1]/text()')]

        f2 = [l.replace('Bidang Fokus:', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[4]/td[1]/text()')]

        g2 = [l.replace('Rencana Biaya:', '').replace('Rp.', '').replace(',', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[5]/td[1]/text()')]

        h2 = [l.replace('Lama Kegiatan:', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[6]/td[1]/text()[1]')]

        i2 = [l.replace('Rp.', '').replace(',', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[10]/td/table//tr[1]/td[3]/text()')]

        j2 = [l.strip()
              for l in content.xpath(base + '/td[2]/table//tr[10]/td/table//tr[2]/td[3]/text()')]

        k2 = [l.strip()
              for l in content.xpath(base + '/td[2]/table//tr[10]/td/table//tr[3]/td[3]/a/@href')]

        # The 'Reviewer 1' all-content stripper
        rev_1 = [l.strip() for l in content.xpath(base + '/td[2]/table//tr[8]/td/table[@width="100%"]//tr[1]/td[1]/text()')]

        # Calculation for data pattern search
        # This is equal to the number of entries
        # 'rev_1_n' is always a multiple of 5
        rev_1_n = int(len(rev_1) / 5)

        l2 = []
        for i in range(rev_1_n):
            loc = 1 + (i * 5)  # --- the location of the data in the array
            l2.append(rev_1[loc].strip())

        m2 = []
        for i in range(rev_1_n):
            loc = 2 + (i * 5)  # --- the location of the data in the array
            m2.append(rev_1[loc].strip())

        n2 = []
        for i in range(rev_1_n):
            loc = 3 + (i * 5)  # --- the location of the data in the array
            n2.append(rev_1[loc].replace('Rp.', '').replace(',', '').strip())

        o2 = []
        for i in range(rev_1_n):
            loc = 4 + (i * 5)  # --- the location of the data in the array
            o2.append(rev_1[loc].strip())

        # The 'Reviewer 2' all-content stripper
        rev_2 = [l.strip() for l in content.xpath(base + '/td[2]/table//tr[8]/td/table[@width="100%"]//tr[1]/td[3]/text()')]

        # Calculation for data pattern search
        # This is equal to the number of entries
        # 'rev_2_n' is always a multiple of 5
        rev_2_n = int(len(rev_2) / 5)

        p2 = []
        for i in range(rev_2_n):
            loc = 1 + (i * 5)  # --- the location of the data in the array
            p2.append(rev_2[loc].strip())

        q2 = []
        for i in range(rev_2_n):
            loc = 2 + (i * 5)  # --- the location of the data in the array
            q2.append(rev_2[loc].strip())

        r2 = []
        for i in range(rev_2_n):
            loc = 3 + (i * 5)  # --- the location of the data in the array
            r2.append(rev_2[loc].replace('Rp.', '').replace(',', '').strip())

        s2 = []
        for i in range(rev_2_n):
            loc = 4 + (i * 5)  # --- the location of the data in the array
            s2.append(rev_2[loc].strip())

        # The starting row coordinate of the active sheet
        row_start = 3

        # DEBUG
        # Please comment out after use
        # ---
        # print(a1, b2, c2, d2, e2, f2, g2, h2, i2, j2, k2, l2, m2, n2, o2, p2, q2, r2, s2)

        # Iterating through each table row and write to the table
        # Assumes the lists a1, b2, c2, ... have the same array size
        control.append_message_area(f'+ Melakukan iterasi terhadap baris tabel dan menulis spreadsheet luaran ...')
        control.set_progress_bar(35)
        for i in range(len(a1)):

            # Noisy preamble logging
            # Please don't use this -_-
            # ---
            # control.append_message_area(f'ITERASI [{i}]')

            # Updating the progress bar status
            control.set_progress_bar(35 + round(45*(i+1)/(len(a1))))

            # Painting the scraped data to the output spreadsheet row
            sheet[f'A{row_start}'] = a1[i]
            sheet[f'B{row_start}'] = b2[i]
            sheet[f'C{row_start}'] = c2[i]
            sheet[f'D{row_start}'] = d2[i]
            sheet[f'E{row_start}'] = e2[i]
            sheet[f'F{row_start}'] = f2[i]
            sheet[f'G{row_start}'] = g2[i]
            sheet[f'H{row_start}'] = h2[i]
            sheet[f'I{row_start}'] = i2[i]
            sheet[f'J{row_start}'] = j2[i]
            sheet[f'K{row_start}'] = k2[i]
            sheet[f'L{row_start}'] = l2[i]
            sheet[f'M{row_start}'] = m2[i]
            sheet[f'N{row_start}'] = n2[i]
            sheet[f'O{row_start}'] = o2[i]
            sheet[f'P{row_start}'] = p2[i]
            sheet[f'Q{row_start}'] = q2[i]
            sheet[f'R{row_start}'] = r2[i]
            sheet[f'S{row_start}'] = s2[i]

            # Incrementing the 'row_start' iterator
            # Then continue the loop
            row_start += 1
            continue

        # Post-loop logging: successfully painted the output spreadsheet file
        control.append_message_area(f'+ Tabel sukses dipanen!')
        control.set_progress_bar(85)

        # Asking for the spreadsheet name to save as
        # ---
        # Logging and setting the progress bar
        control.append_message_area(f'+ Menyimpan spreadsheet luaran ...')
        control.set_progress_bar(90)
        # Dealing with file name prompt and saving
        # Using loop to mitigate the user clicking 'cancel'
        # in the file name dialog prompt
        while True:
            # Opening the dialog prompt
            output_spreadsheet = filedialog.asksaveasfilename(
                filetypes=[('Excel files', '*.xlsx')],
                initialfile='Sipesat - Ditolak Pengabdian Ringkasan Risat.xlsx',
                title='Simpan sebagai ...'
            )

            # 'cancel' button in the dialog prompt is clicked
            if len(output_spreadsheet) == 0:
                # Showing confirmation
                x = messagebox.askyesno(
                    'Nama File Kosong',
                    'Apakah Anda yakin ingin melanjutkan tanpa menyimpan file spreadsheet hasil pemanenan?'
                )
                # Determining whether to break or to continue the loop
                # based on the inversed value of 'x'
                if x:
                    control.append_message_area(f'+ Finalisasi pemanenan data tanpa menyimpan file spreadsheet luaran ...')
                    workbook.close()  # --- closing the workbook without saving
                    break
                else:
                    continue  # --- continuing the loop
            # File name does not end in spreadsheet extension
            elif output_spreadsheet[-5:] != '.xlsx':
                output_spreadsheet = output_spreadsheet + '.xlsx'

            # Saving the spreadsheet
            control.append_message_area(f'LOKASI_SPREADSHEET_LUARAN: {output_spreadsheet}')
            control.set_progress_bar(95)
            workbook.save(output_spreadsheet)

            # Closing the openpyxl workbook
            control.append_message_area(f'+ Menutup file spreadsheet ...')
            control.set_progress_bar(98)
            workbook.close()

            # Breaking the loop
            break

        # Notify for a successful scraping
        control.append_message_area(f'+ Pemanenan selesai pada: {str(dt.now())}')
        control.set_progress_bar(100)
        control.on_notify_successful_scraping()

    # This function harvests "Risat Dana Pengabdian > Ringkasan Data" data
    # and then store the harvested data as an excel file
    #
    # Required arguments:
    # - control             --> for updating the progress bar and
    #                           message area of the screen SipesatScrHarvest
    # - username, password  --> the Risat administrator username and password
    def run_harvest_c_0_dana(self, control, username, password):

        # SipesatScrHarvest messenger
        control.set_header_desc('Panen Data "Risat Dana Pengabdian > Ringkasan Data"')
        control.set_help_label('Data sedang dipanen. Silahkan menunggu.')
        control.set_progress_bar(0)
        control.clear_message_area()

        # Preamble logging
        control.append_message_area(f'+ Memulai pemanenan data ...')
        control.append_message_area(f'+ Pemanenan dimulai pada: {str(dt.now())}')
        control.set_progress_bar(5)

        # Preparing the 'data_prompt' arrays
        control.append_message_area(f'+ Log masuk Risat sebagai [{username}] ...')
        control.set_progress_bar(10)
        data_prompt = self.get_risat_login(username, password)
        data_prompt = self.get_risat_pengabdian(data_prompt)
        data_prompt = self.get_risat_pengabdian_dana_pengabdian(data_prompt)

        # Parsing XML tree content
        control.append_message_area(f'+ Membaca halaman web ...')
        control.set_progress_bar(15)
        content = data_prompt['html_content']

        # Establishing the export spreadsheet file
        control.append_message_area(f'+ Mempersiapkan file spreadsheet luaran ...')
        control.set_progress_bar(20)
        workbook = xl.Workbook()
        sheet = workbook.active
        sheet.title = 'Dana Pengabdian Ringkasan'

        # Preparing the sheet header
        control.append_message_area(f'+ Mempersiapkan kepala lembar spreadsheet ...')
        control.set_progress_bar(25)
        sheet['A1'].value = 'No.'
        sheet['B1'].value = 'Judul'
        sheet['C1'].value = 'Ketua'
        sheet['D1'].value = 'Jml. Anggota'
        sheet['E1'].value = 'Tgl. Usulan'
        sheet['F1'].value = 'Bidang Fokus'
        sheet['G1'].value = 'Rencana Biaya'
        sheet['H1'].value = 'Lama Kegiatan'
        sheet['I1'].value = 'Dana Disetujui'
        sheet['J1'].value = 'Tgl. Persetujuan'
        sheet['K1'].value = 'Dana Ditransfer'
        sheet['L1'].value = 'Tgl. Ditransfer'

        # The base XPath location, pointing to each entry row
        base = '//div[@class="mw-100"]//div[@class="form-group f12"]/table[@width="100%"]//tr[@valign="top"]'

        # ---
        # Obtaining the data row values
        control.append_message_area(f'+ Mendapatkan data pada baris tabel ...')
        control.set_progress_bar(30)

        # HYPOTHESIS:
        # Xpath cannot detect 'tbody' element.
        # So instead of using 'table/tbody/tr', use 'table//tr' instead
        #
        # RESULT:
        # The hypothesis is correct.
        # Therefore, don't mention 'tbody' in any of the following Xpath paths

        a1 = [str(i)
              for i in range(1, len(content.xpath(base)) + 1)]

        b1 = [l.strip()
              for l in content.xpath(base + '//span[@class="hijau"]/text()')]

        c1 = [l.replace('Ketua:', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[2]/td/table//tr/td[1]/text()')]

        d1 = [l.replace('Jumlah Anggota:', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[2]/td/table//tr/td[3]/text()')]

        e1 = [l.replace('Tgl Usulan:', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[3]/td[1]/text()')]

        f1 = [l.replace('Bidang Fokus:', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[4]/td[1]/text()')]

        g1 = [l.replace('Rencana Biaya:', '').replace('Rp.', '').replace(',', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[5]/td[1]/text()')]

        h1 = [l.replace('Lama Kegiatan:', '').strip()
              for l in content.xpath(base + '/td[2]/table//tr[6]/td[1]/text()[1]')]

        i1 = [l.replace('Rp.', '').replace(',', '').strip()
              for l in content.xpath(base + '/td[3]/text()')]

        j1 = [l.strip()
              for l in content.xpath(base + '/td[4]/text()')]

        k1 = [l.replace('Rp.', '').replace(',', '').strip()
              for l in content.xpath(base + '/td[5]/text()')]

        l1 = [l.replace('Terealisasi', '').strip()
              for l in content.xpath(base + '/td[6]/text()')]

        # The starting row coordinate of the active sheet
        row_start = 2

        # DEBUG
        # Please comment out after use
        # ---
        # print(a1, b1, c1, d1, e1, f1, g1, h1, i1, j1, k1, l1)

        # Iterating through each table row and write to the table
        # Assumes the lists a1, b1, c1, ... have the same array size
        control.append_message_area(f'+ Melakukan iterasi terhadap baris tabel dan menulis spreadsheet luaran ...')
        control.set_progress_bar(35)
        for i in range(len(a1)):
            # Noisy preamble logging
            # Please don't use this -_-
            # ---
            # control.append_message_area(f'ITERASI [{i}]')

            # Updating the progress bar status
            control.set_progress_bar(35 + round(45 * (i + 1) / (len(a1))))

            # Painting the scraped data to the output spreadsheet row
            sheet[f'A{row_start}'] = a1[i]
            sheet[f'B{row_start}'] = b1[i]
            sheet[f'C{row_start}'] = c1[i]
            sheet[f'D{row_start}'] = d1[i]
            sheet[f'E{row_start}'] = e1[i]
            sheet[f'F{row_start}'] = f1[i]
            sheet[f'G{row_start}'] = g1[i]
            sheet[f'H{row_start}'] = h1[i]
            sheet[f'I{row_start}'] = i1[i]
            sheet[f'J{row_start}'] = j1[i]
            sheet[f'K{row_start}'] = k1[i]
            sheet[f'L{row_start}'] = l1[i]

            # Incrementing the 'row_start' iterator
            # Then continue the loop
            row_start += 1
            continue

        # Post-loop logging: successfully painted the output spreadsheet file
        control.append_message_area(f'+ Tabel sukses dipanen!')
        control.set_progress_bar(85)

        # Asking for the spreadsheet name to save as
        # ---
        # Logging and setting the progress bar
        control.append_message_area(f'+ Menyimpan spreadsheet luaran ...')
        control.set_progress_bar(90)
        # Dealing with file name prompt and saving
        # Using loop to mitigate the user clicking 'cancel'
        # in the file name dialog prompt
        while True:
            # Opening the dialog prompt
            output_spreadsheet = filedialog.asksaveasfilename(
                filetypes=[('Excel files', '*.xlsx')],
                initialfile='Sipesat - Dana Pengabdian Ringkasan Risat.xlsx',
                title='Simpan sebagai ...'
            )

            # 'cancel' button in the dialog prompt is clicked
            if len(output_spreadsheet) == 0:
                # Showing confirmation
                x = messagebox.askyesno(
                    'Nama File Kosong',
                    'Apakah Anda yakin ingin melanjutkan tanpa menyimpan file spreadsheet hasil pemanenan?'
                )
                # Determining whether to break or to continue the loop
                # based on the inversed value of 'x'
                if x:
                    control.append_message_area(
                        f'+ Finalisasi pemanenan data tanpa menyimpan file spreadsheet luaran ...')
                    workbook.close()  # --- closing the workbook without saving
                    break
                else:
                    continue  # --- continuing the loop
            # File name does not end in spreadsheet extension
            elif output_spreadsheet[-5:] != '.xlsx':
                output_spreadsheet = output_spreadsheet + '.xlsx'

            # Saving the spreadsheet
            control.append_message_area(f'LOKASI_SPREADSHEET_LUARAN: {output_spreadsheet}')
            control.set_progress_bar(95)
            workbook.save(output_spreadsheet)

            # Closing the openpyxl workbook
            control.append_message_area(f'+ Menutup file spreadsheet ...')
            control.set_progress_bar(98)
            workbook.close()

            # Breaking the loop
            break

        # Notify for a successful scraping
        control.append_message_area(f'+ Pemanenan selesai pada: {str(dt.now())}')
        control.set_progress_bar(100)
        control.on_notify_successful_scraping()

    # This function harvests "Risat Arsip Pengabdian > Ringkasan Data" data
    # and then store the harvested data as an excel file
    #
    # Required arguments:
    # - control             --> for updating the progress bar and
    #                           message area of the screen SipesatScrHarvest
    # - username, password  --> the Risat administrator username and password
    def run_harvest_c_0_arsip(self, control, username, password):
        # SipesatScrHarvest messenger
        control.set_header_desc('Panen Data "Risat Arsip Pengabdian > Ringkasan Data"')
        control.set_help_label('Data sedang dipanen. Silahkan menunggu.')
        control.set_progress_bar(0)
        control.clear_message_area()

        # Preamble logging
        control.append_message_area(f'+ Memulai pemanenan data ...')
        control.append_message_area(f'+ Pemanenan dimulai pada: {str(dt.now())}')
        control.set_progress_bar(5)

        # Preparing the 'data_prompt' arrays
        control.append_message_area(f'+ Log masuk Risat sebagai [{username}] ...')
        control.set_progress_bar(10)
        data_prompt = self.get_risat_login(username, password)
        data_prompt = self.get_risat_pengabdian(data_prompt)
        data_prompt = self.get_risat_pengabdian_arsip(data_prompt)

        # Parsing XML tree content
        control.append_message_area(f'+ Membaca halaman web ...')
        control.set_progress_bar(15)
        content = data_prompt['html_content']

        # Establishing the export spreadsheet file
        control.append_message_area(f'+ Mempersiapkan file spreadsheet luaran ...')
        control.set_progress_bar(20)
        workbook = xl.Workbook()
        sheet = workbook.active
        sheet.title = 'Arsip Pengabdian Ringkasan'

        # Preparing the sheet header
        control.append_message_area(f'+ Mempersiapkan kepala lembar spreadsheet ...')
        control.set_progress_bar(25)
        sheet['A1'].value = 'No.'
        sheet['B1'].value = 'Judul'
        sheet['C1'].value = 'Ketua'
        sheet['D1'].value = 'Bidang Fokus'
        sheet['E1'].value = 'Tgl. Usulan'
        sheet['F1'].value = 'Lama Kegiatan'

        # The base XPath location, pointing to each entry row
        base = '//div[@id="ContentPlaceHolder1_arsip1_updg1"]//div[@class="panel-body f12"]/table[@width="100%"]'

        # ---
        # Obtaining the data row values
        control.append_message_area(f'+ Mendapatkan data pada baris tabel ...')
        control.set_progress_bar(30)

        # HYPOTHESIS:
        # Xpath cannot detect 'tbody' element.
        # So instead of using 'table/tbody/tr', use 'table//tr' instead
        #
        # RESULT:
        # The hypothesis is correct.
        # Therefore, don't mention 'tbody' in any of the following Xpath paths

        a1 = [str(i)
              for i in range(1, len(content.xpath(base)) + 1)]

        b1 = [l.strip()
              for l in content.xpath(base + '//tr[1][@valign="top"]/td[@colspan="4"]/b/text()')]

        c1 = [l.replace('Ketua:', '').strip()
              for l in content.xpath(base + '//tr[3][@valign="top"]/td[2]/table//tr[1]/td/text()')]

        d1 = [l.replace('Bidang Fokus:', '').strip()
              for l in content.xpath(base + '//tr[3][@valign="top"]/td[2]/table//tr[2]/td/text()')]

        e1 = [l.replace('Tgl Usulan:', '').strip()
              for l in content.xpath(base + '//tr[3][@valign="top"]/td[2]/table//tr[3]/td/text()')]

        f1 = [l.replace('Lama Kegiatan:', '').strip()
              for l in content.xpath(base + '//tr[3][@valign="top"]/td[2]/table//tr[4]/td/text()')]

        # The starting row coordinate of the active sheet
        row_start = 2

        # DEBUG
        # Please comment out after use
        # ---
        '''
        print()
        print('-------------------------------------------------------------------------------------------------------')
        for a in (a1, b1, c1, d1, e1, f1):
            print('LEN', len(a))
            print('CONTENT', a)
            print()
        print('-------------------------------------------------------------------------------------------------------', end='')
        print()
        '''

        # Iterating through each table row and write to the table
        # Assumes the lists a1, b1, c1, ... have the same array size
        control.append_message_area(f'+ Melakukan iterasi terhadap baris tabel dan menulis spreadsheet luaran ...')
        control.set_progress_bar(35)
        for i in range(len(a1)):
            # Noisy preamble logging
            # Please don't use this -_-
            # ---
            # control.append_message_area(f'ITERASI [{i}]')

            # Updating the progress bar status
            control.set_progress_bar(35 + round(45 * (i + 1) / (len(a1))))

            # Painting the scraped data to the output spreadsheet row
            sheet[f'A{row_start}'] = a1[i]
            sheet[f'B{row_start}'] = b1[i]
            sheet[f'C{row_start}'] = c1[i]
            sheet[f'D{row_start}'] = d1[i]
            sheet[f'E{row_start}'] = e1[i]
            sheet[f'F{row_start}'] = f1[i]

            # Incrementing the 'row_start' iterator
            # Then continue the loop
            row_start += 1
            continue

        # Post-loop logging: successfully painted the output spreadsheet file
        control.append_message_area(f'+ Tabel sukses dipanen!')
        control.set_progress_bar(85)

        # Asking for the spreadsheet name to save as
        # ---
        # Logging and setting the progress bar
        control.append_message_area(f'+ Menyimpan spreadsheet luaran ...')
        control.set_progress_bar(90)
        # Dealing with file name prompt and saving
        # Using loop to mitigate the user clicking 'cancel'
        # in the file name dialog prompt
        while True:
            # Opening the dialog prompt
            output_spreadsheet = filedialog.asksaveasfilename(
                filetypes=[('Excel files', '*.xlsx')],
                initialfile='Sipesat - Arsip Pengabdian Ringkasan Risat.xlsx',
                title='Simpan sebagai ...'
            )

            # 'cancel' button in the dialog prompt is clicked
            if len(output_spreadsheet) == 0:
                # Showing confirmation
                x = messagebox.askyesno(
                    'Nama File Kosong',
                    'Apakah Anda yakin ingin melanjutkan tanpa menyimpan file spreadsheet hasil pemanenan?'
                )
                # Determining whether to break or to continue the loop
                # based on the inversed value of 'x'
                if x:
                    control.append_message_area(
                        f'+ Finalisasi pemanenan data tanpa menyimpan file spreadsheet luaran ...')
                    workbook.close()  # --- closing the workbook without saving
                    break
                else:
                    continue  # --- continuing the loop
            # File name does not end in spreadsheet extension
            elif output_spreadsheet[-5:] != '.xlsx':
                output_spreadsheet = output_spreadsheet + '.xlsx'

            # Saving the spreadsheet
            control.append_message_area(f'LOKASI_SPREADSHEET_LUARAN: {output_spreadsheet}')
            control.set_progress_bar(95)
            workbook.save(output_spreadsheet)

            # Closing the openpyxl workbook
            control.append_message_area(f'+ Menutup file spreadsheet ...')
            control.set_progress_bar(98)
            workbook.close()

            # Breaking the loop
            break

        # Notify for a successful scraping
        control.append_message_area(f'+ Pemanenan selesai pada: {str(dt.now())}')
        control.set_progress_bar(100)
        control.on_notify_successful_scraping()

    # This function harvests "Risat Laporan Akhir Pengabdian > Data Ringkasan" data
    # and then store the harvested data as an excel file
    #
    # Required arguments:
    # - control             --> for updating the progress bar and
    #                           message area of the screen SipesatScrHarvest
    # - username, password  --> the Risat administrator username and password
    def run_harvest_c_0_lapakhir(self, control, username, password):

        # SipesatScrHarvest messenger
        control.set_header_desc('Panen Data "Risat Laporan Akhir Pengabdian > Data Ringkasan"')
        control.set_help_label('Data sedang dipanen. Silahkan menunggu.')
        control.set_progress_bar(0)
        control.clear_message_area()

        # Preamble logging
        control.append_message_area(f'+ Memulai pemanenan data ...')
        control.append_message_area(f'+ Pemanenan dimulai pada: {str(dt.now())}')
        control.set_progress_bar(5)

        # Preparing the 'data_prompt' arrays
        control.append_message_area(f'+ Log masuk Risat sebagai [{username}] ...')
        control.set_progress_bar(10)
        data_prompt = self.get_risat_login(username, password)
        data_prompt = self.get_risat_pengabdian(data_prompt)
        data_prompt = self.get_risat_pengabdian_pelak_kegi(data_prompt)
        data_prompt = self.get_risat_pengabdian_pelak_kegi_lapakhir(data_prompt)

        # Parsing XML tree content
        control.append_message_area(f'+ Membaca halaman web ...')
        control.set_progress_bar(15)
        content = data_prompt['html_content']

        # Establishing the export spreadsheet file
        control.append_message_area(f'+ Mempersiapkan file spreadsheet luaran ...')
        control.set_progress_bar(20)
        workbook = xl.Workbook()
        sheet = workbook.active
        sheet.title = 'Lap. Akhir Ringkasan Pengabdian'

        # Preparing the sheet header
        control.append_message_area(f'+ Mempersiapkan kepala lembar spreadsheet ...')
        control.set_progress_bar(25)
        # ---
        # Preparing the headers
        sheet['A1'].value = 'No.'
        sheet['B1'].value = 'Judul'
        sheet['C1'].value = 'Ketua'
        sheet['D1'].value = 'Bidang Fokus'
        sheet['E1'].value = 'Tgl. Usulan'
        sheet['F1'].value = 'Jumlah Anggota'
        sheet['G1'].value = 'Biaya'
        sheet['H1'].value = 'Dana Disetujui'
        sheet['I1'].value = 'Lama Kegiatan'

        # The base XPath location, pointing to each entry row
        base = '//div[@id="ContentPlaceHolder1_lapakhirbp3m1_updg1"]//div[@class="panel-body f12"]/table[@width="100%"]'

        # ---
        # Obtaining the data row values
        control.append_message_area(f'+ Mendapatkan data pada baris tabel ...')
        control.set_progress_bar(30)

        # HYPOTHESIS:
        # Xpath cannot detect 'tbody' element.
        # So instead of using 'table/tbody/tr', use 'table//tr' instead
        #
        # RESULT:
        # The hypothesis is correct.
        # Therefore, don't mention 'tbody' in any of the following Xpath paths

        a2 = [str(i)
              for i in range(1, len(content.xpath(base)) + 1)]

        b2 = [l.strip()
              for l in content.xpath(base + '//tr[@valign="top"]/td[@colspan="3"]/b/text()')]

        c2 = [l.replace('Ketua:', '').strip()
              for l in content.xpath(base + '//tr[2][not(descendant::table) and not(ancestor::table[@class="table"])]/td[2]/text()')]

        d2 = [l.replace('Bidang Fokus:', '').strip()
              for l in content.xpath(base + '//tr[3][not(descendant::table) and not(ancestor::table[@class="table"])]/td[@colspan="3"]/text()')]

        e2 = [l.replace('Tgl Usulan:', '').strip()
              for l in content.xpath(base + '//tr[4][not(descendant::table) and not(ancestor::table[@class="table"])]/td[@colspan="3"]/text()')]

        f2 = [l.replace('Jml Anggota:', '').strip()
              for l in content.xpath(base + '//tr[5][not(descendant::table) and not(ancestor::table[@class="table"])]/td[@colspan="3"]/text()')]

        g2 = [l.replace('Biaya:', '').replace('Rp.', '').replace(',', '').split('disetujui')[0].strip()
              for l in content.xpath(base + '//tr[6][not(descendant::table) and not(ancestor::table[@class="table"])]/td[@colspan="3"]/text()')]

        h2 = [l.replace('Biaya:', '').replace('Rp.', '').replace(',', '').split('disetujui')[1].strip()
              for l in content.xpath(base + '//tr[6][not(descendant::table) and not(ancestor::table[@class="table"])]/td[@colspan="3"]/text()')]

        i2 = [l.replace('Lama Kegiatan:', '').strip()
              for l in content.xpath(base + '//tr[7][not(descendant::table) and not(ancestor::table[@class="table"])]/td[@colspan="3"]/text()')]

        # The starting row coordinate of the active sheet
        row_start = 2

        # DEBUG
        # Please comment out after use
        # ---
        '''
        print()
        print('-------------------------------------------------------------------------------------------------------')
        for a in (a2, b2, c2, d2, e2, f2, g2, h2, i2, j_):
            print('LEN', len(a))
            print('CONTENT', a)
            print()
        print('-------------------------------------------------------------------------------------------------------', end='')
        print()
        '''

        # Iterating through each table row and write to the table
        # Assumes the lists a1, b1, c1, ... have the same array size
        control.append_message_area(f'+ Melakukan iterasi terhadap baris tabel dan menulis spreadsheet luaran ...')
        control.set_progress_bar(35)
        for i in range(len(a2)):
            # Noisy preamble logging
            # Please don't use this -_-
            # ---
            # control.append_message_area(f'ITERASI [{i}]')

            # Updating the progress bar status
            control.set_progress_bar(35 + round(45 * (i + 1) / (len(a2))))

            # Painting the scraped data to the output spreadsheet row
            # Section 1: 'IDENTITY'
            sheet[f'A{row_start}'] = a2[i]
            sheet[f'B{row_start}'] = b2[i]
            sheet[f'C{row_start}'] = c2[i]
            sheet[f'D{row_start}'] = d2[i]
            sheet[f'E{row_start}'] = e2[i]
            sheet[f'F{row_start}'] = f2[i]
            sheet[f'G{row_start}'] = g2[i]
            sheet[f'H{row_start}'] = h2[i]
            sheet[f'I{row_start}'] = i2[i]

            # Incrementing the 'row_start' iterator
            # Then continue the loop
            row_start += 1
            continue

        # Post-loop logging: successfully painted the output spreadsheet file
        control.append_message_area(f'+ Tabel sukses dipanen!')
        control.set_progress_bar(85)

        # Asking for the spreadsheet name to save as
        # ---
        # Logging and setting the progress bar
        control.append_message_area(f'+ Menyimpan spreadsheet luaran ...')
        control.set_progress_bar(90)
        # Dealing with file name prompt and saving
        # Using loop to mitigate the user clicking 'cancel'
        # in the file name dialog prompt
        while True:
            # Opening the dialog prompt
            output_spreadsheet = filedialog.asksaveasfilename(
                filetypes=[('Excel files', '*.xlsx')],
                initialfile='Sipesat - Lap Akhir Pengabdian Ringkasan Risat.xlsx',
                title='Simpan sebagai ...'
            )

            # 'cancel' button in the dialog prompt is clicked
            if len(output_spreadsheet) == 0:
                # Showing confirmation
                x = messagebox.askyesno(
                    'Nama File Kosong',
                    'Apakah Anda yakin ingin melanjutkan tanpa menyimpan file spreadsheet hasil pemanenan?'
                )
                # Determining whether to break or to continue the loop
                # based on the inversed value of 'x'
                if x:
                    control.append_message_area(
                        f'+ Finalisasi pemanenan data tanpa menyimpan file spreadsheet luaran ...')
                    workbook.close()  # --- closing the workbook without saving
                    break
                else:
                    continue  # --- continuing the loop
            # File name does not end in spreadsheet extension
            elif output_spreadsheet[-5:] != '.xlsx':
                output_spreadsheet = output_spreadsheet + '.xlsx'

            # Saving the spreadsheet
            control.append_message_area(f'LOKASI_SPREADSHEET_LUARAN: {output_spreadsheet}')
            control.set_progress_bar(95)
            workbook.save(output_spreadsheet)

            # Closing the openpyxl workbook
            control.append_message_area(f'+ Menutup file spreadsheet ...')
            control.set_progress_bar(98)
            workbook.close()

            # Breaking the loop
            break

        # Notify for a successful scraping
        control.append_message_area(f'+ Pemanenan selesai pada: {str(dt.now())}')
        control.set_progress_bar(100)
        control.on_notify_successful_scraping()

    # This function harvests "Risat Usulan Pengabdian > Data Detil Lengkap" data
    # and then store the harvested data as an excel file
    #
    # Required arguments:
    # - control             --> for updating the progress bar and
    #                           message area of the screen SipesatScrHarvest
    # - username, password  --> the Risat administrator username and password
    def run_harvest_c_1_usulan(self, control, username, password):

        # SipesatScrHarvest messenger
        control.set_header_desc('Panen Data "Risat Usulan Pengabdian > Data Detil Lengkap"')
        control.set_help_label('Data sedang dipanen. Silahkan menunggu.')
        control.set_progress_bar(0)
        control.clear_message_area()

        # Preamble logging
        control.append_message_area(f'+ Memulai pemanenan data ...')
        control.append_message_area(f'+ Pemanenan dimulai pada: {str(dt.now())}')
        control.set_progress_bar(5)

        # Preparing the 'data_prompt' arrays
        control.append_message_area(f'+ Log masuk Risat sebagai [{username}] ...')
        control.set_progress_bar(10)
        data_prompt = self.get_risat_login(username, password)
        data_prompt = self.get_risat_pengabdian(data_prompt)
        data_prompt = self.get_risat_pengabdian_usulan_pengabdian(data_prompt)

        # Parsing XML tree content
        control.append_message_area(f'+ Membaca halaman web ...')
        control.set_progress_bar(15)
        content = data_prompt['html_content']

        # Establishing the export spreadsheet file
        control.append_message_area(f'+ Mempersiapkan file spreadsheet luaran ...')
        control.set_progress_bar(20)
        workbook = xl.Workbook()
        sheet = workbook.active
        sheet.title = 'Usulan Pengabdian Detil'

        # DEVELOPER'S NOTE (DO NOT REMOVE)
        '''
        Data yang perlu diambil:
        1. Judul
        2. Substansi usulan
        3. Identitas pengusul - ketua
        4. Identitas pengusul - anggota
        5. RAB
        6. Dokumen pendukung

        Urutan kolom excel:
        1. Judul
        2. Substansi usulan
        3. RAB
        4. Dokumen pendukung
        5. Identitas pengusul - ketua
        6. Identitas pengusul - anggota
        '''

        # Preparing the sheet header
        control.append_message_area(f'+ Mempersiapkan kepala lembar spreadsheet ...')
        control.set_progress_bar(25)
        # ---
        # Preparing the "NOMOR" header
        sheet.merge_cells(
            start_row=1,
            start_column=1,
            end_row=2,
            end_column=1
        )
        sheet['A1'].value = 'No.'
        sheet['A1'].alignment = Alignment(horizontal='center')
        # ---
        # Preparing the "JUDUL" header
        sheet.merge_cells(
            start_row=1,
            start_column=2,
            end_row=1,
            end_column=12
        )
        sheet['B1'].value = 'IDENTITAS'
        sheet['B1'].alignment = Alignment(horizontal='center')
        # Preparing the "JUDUL" sub-headers
        sheet['B2'].value = 'Judul'
        sheet['C2'].value = 'Tgl. Usulan'
        sheet['D2'].value = 'TKT Saat Ini'
        sheet['E2'].value = 'Level'
        sheet['F2'].value = 'Kategori'
        sheet['G2'].value = 'Skema'
        sheet['H2'].value = 'Rumpun Ilmu'
        sheet['I2'].value = 'Bidang Fokus'
        sheet['J2'].value = 'Tema'
        sheet['K2'].value = 'Topik'
        sheet['L2'].value = 'Lama Kegiatan'
        # ---
        # Preparing the "SUBSTANSI USULAN" header
        sheet.merge_cells(
            start_row=1,
            start_column=13,
            end_row=1,
            end_column=14
        )
        sheet['M1'].value = 'SUBSTANSI USULAN'
        sheet['M1'].alignment = Alignment(horizontal='center')
        # Preparing the "SUBSTANSI USULAN" sub-headers
        sheet['M2'].value = 'Kelompok Makro'
        sheet['N2'].value = 'File Proposal'
        # ---
        # Preparing the "RAB" header
        sheet.merge_cells(
            start_row=1,
            start_column=15,
            end_row=1,
            end_column=16
        )
        sheet['O1'].value = 'RENCANA ANGGARAN BIAYA'
        sheet['O1'].alignment = Alignment(horizontal='center')
        # Preparing the "RAB" sub-headers
        sheet['O2'].value = 'Biaya'
        sheet['P2'].value = 'File RAB'
        # ---
        # Preparing the "DOKUMEN PENDUKUNG" header
        sheet.merge_cells(
            start_row=1,
            start_column=17,
            end_row=1,
            end_column=19
        )
        sheet['Q1'].value = 'DOKUMEN PENDUKUNG'
        sheet['Q1'].alignment = Alignment(horizontal='center')
        # Preparing the "DOKUMEN PENDUKUNG" sub-headers
        sheet['Q2'].value = 'Mitra'
        sheet['R2'].value = 'Dukungan Biaya'
        sheet['S2'].value = 'Surat Dukungan Mitra'
        # ---
        # Preparing the "IDENTITAS PENGUSUL — KETUA" header
        sheet.merge_cells(
            start_row=1,
            start_column=20,
            end_row=1,
            end_column=24
        )
        sheet['T1'].value = 'IDENTITAS PENGUSUL — KETUA'
        sheet['T1'].alignment = Alignment(horizontal='center')
        # Preparing the "IDENTITAS PENGUSUL — KETUA" sub-headers
        sheet['T2'].value = 'N.I.P'
        sheet['U2'].value = 'N.I.K'
        sheet['V2'].value = 'N.I.D.N'
        sheet['W2'].value = 'Nama Lengkap'
        sheet['X2'].value = 'Jabatan Fungsional'

        # The base XPath location, pointing to each entry row
        base = '//div[@class="mw-100"]//div[@class="form-group f12"]/table[@width="100%"]//tr[@valign="top"]'

        # ---
        # Obtaining the data row values
        control.append_message_area(f'+ Mendapatkan data pada baris tabel ...')
        control.set_progress_bar(30)

        # Reading the HTML entry row hidden ASPX values
        # Copy-pasted from: /ssynthesia/ghostcity/ar/dumper-2/24__2023.02.13__requestsrisat.py
        all_kodetran_prop = content.xpath(base + '//input[1][@type="hidden"]/@name')
        all_kodetran_val = content.xpath(base + '//input[1][@type="hidden"]/@value')
        all_stat_prop = content.xpath(base + '//input[2][@type="hidden"]/@name')
        all_stat_val = content.xpath(base + '//input[2][@type="hidden"]/@value')
        all_submitbtn = content.xpath(base + '//tr[7]//input[@type="submit"]/@name')

        # The number of rows
        # Assumes the 'all_' array size equals the number of data rows
        number_of_row = len(all_kodetran_prop)

        # The starting row coordinate of the active sheet
        row_start = 3

        # The maximum number of 'DATA ANGGOTA' table data row
        y_max_row = 0

        # Iterating through each entry row element
        # Assumes all the 'all_' arrays in the previous code block
        # are of the same length/size
        # Copy-pasted from: /ssynthesia/ghostcity/ar/dumper-2/24__2023.02.13__requestsrisat.py
        control.append_message_area(f'+ Melakukan iterasi terhadap baris tabel dan menulis spreadsheet luaran ...')
        control.set_progress_bar(35)
        temporary_prompt = data_prompt
        for i in range(number_of_row):
            # Noisy preamble logging
            # Please don't use this -_-
            # ---
            # control.append_message_area(f'ITERASI [{i}]')

            # Updating the progress bar status
            control.set_progress_bar(35 + round(45 * (i + 1) / number_of_row))

            # Preparing the AJAX payload
            detail_prompt = {
                'viewstate': temporary_prompt['viewstate'],
                'viewstategen': temporary_prompt['viewstategen'],
                'eventvalidation': temporary_prompt['eventvalidation'],
                'button_name': all_submitbtn[i],
                'kodetran_prop': all_kodetran_prop[i],
                'kodetran_val': all_kodetran_val[i],
                'stat_prop': all_stat_prop[i],
                'stat_val': all_stat_val[i]
            }

            # Obtaining the response data of each individual entry row detail page
            data = self.get_risat_pengabdian_usulan_pengabdian_detil(detail_prompt)
            content = data['html_content']
            response = data['http_response']

            # Terminal logging for detecting errors
            log_string_1 = str(i + 1)
            log_string_2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti1_kiri1_txjudul1"]/text()')[
                0].replace('\r', '').replace('\n', '').strip()
            control.append_message_area(f'+ Memanen data detil: [{log_string_1}] {log_string_2} ...')
            print(f'+ Harvesting detailed data: [{log_string_1}] {log_string_2}')

            # DEVELOPER'S NOTE (DO NOT REMOVE)
            '''
            Variables and their associated data

            'a1' ==> 'No.'
            'b2' ==> 'Judul'
            'c2' ==> 'Tgl. Usulan'
            'd2' ==> 'TKT Saat Ini'
            'e2' ==> 'Level'
            'f2' ==> 'Kategori'
            'g2' ==> 'Skema'
            'h2' ==> 'Rumpun Ilmu'
            'i2' ==> 'Bidang Fokus'
            'j2' ==> 'Tema'
            'k2' ==> 'Topik'
            'l2' ==> 'Lama Kegiatan'
            'm2' ==> 'Kelompok Makro'
            'n2' ==> 'File Proposal'
            'o2' ==> 'Biaya'
            'p2' ==> 'File RAB'
            'q2' ==> 'Mitra'
            'r2' ==> 'Dukungan Biaya'
            's2' ==> 'Surat Dukungan Mitra'
            't2' ==> 'N.I.P'
            'u2' ==> 'N.I.K'
            'v2' ==> 'N.I.D.N'
            'w2' ==> 'Nama Lengkap'
            'x2' ==> 'Jabatan Fungsional'

            In addition, variables starting with 'y' prefix are pertaining
            to the table data of "Data Anggota"
            '''

            # Getting the static (non-variable as in "Data Anggota") detail information
            # ---
            a1 = str(i+1)
            # Use try-except catching to mitigate empty data
            try:
                b2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti1_kiri1_txjudul1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                b2 = ''
            # Use try-except catching to mitigate empty data
            try:
                c2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti1_kiri1_tglusul1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                c2 = ''
            # 'Pengabdian Masyarakat' ain't have TKT (d2) and level (e2) data!
            # ---
            d2 = '-'
            e2 = '-'
            # Use try-except catching to mitigate empty data
            try:
                f2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti1_kiri1_lkat1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                f2 = ''
            # Use try-except catching to mitigate empty data
            try:
                g2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti1_kiri1_ddlskema1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                g2 = ''
            # 'Rumpun Ilmu' scraped data contains multiple array of strings
            # This needs some extra tweaking
            h2 = ''
            # Use try-except catching to mitigate empty data
            try:
                h2_pre = content.xpath('//div[@id="ContentPlaceHolder1_teliti1_kiri1_updx"]/div[1][@class="panel panel-info"]/div[@class="panel-body f12"]/div[5][@class="row f12"]/div[@class="col-sm-4"]//text()')
                for l in h2_pre:
                    l = l.replace('\r','').replace('\n','').strip()
                    h2 += l +', '
                # Remove trailing ', ' characters
                h2 = h2[:-2]
            except IndexError:
                h2 = ''
            # ---
            # Use try-except catching to mitigate empty data
            try:
                i2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti1_kiri1_ddlfokus1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                i2 = ''
            # Use try-except catching to mitigate empty data
            try:
                j2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti1_kiri1_ddltema1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                j2 = ''
            # Some 'Topik' has no data element
            # Use try-except catching to mitigate
            try:
                k2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti1_kiri1_ddltopik1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                k2 = ''
            # Use try-except catching to mitigate empty data
            try:
                l2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti1_kiri1_ddllama1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                l2 = ''
            # Use try-except catching to mitigate empty data
            try:
                m2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti1_kiri1_ddlmakro1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                m2 = ''
            # Use try-except catching to mitigate empty data
            try:
                n2 = content.xpath('//a[@id="ContentPlaceHolder1_teliti1_kiri1_alblcatatan1"]/@href')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                n2 = ''
            # Use try-except catching to mitigate empty data
            try:
                o2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti1_kiri1_tbiayathn1"]/text()')[0].replace('\r','').replace('\n','').replace('Rp.', '').replace(',', '').strip()
            except IndexError:
                o2 = ''
            # Use try-except catching to mitigate empty data
            try:
                p2 = content.xpath('//a[@id="ContentPlaceHolder1_teliti1_kiri1_alblfilerab1"]/@href')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                p2 = ''
            # Some 'Data Detil' has no 'Mitra' data element
            # Use try-except catching to mitigate
            try:
                q2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti1_kiri1_tnmmitra1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                q2 = ''
            # Use try-except catching to mitigate empty data
            try:
                r2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti1_kiri1_tbiayadukung1"]/text()')[0].replace('\r','').replace('\n','').replace('Rp.', '').replace(',', '').strip()
            except IndexError:
                r2 = ''
            # Use try-except catching to mitigate empty data
            try:
                s2 = content.xpath('//a[@id="ContentPlaceHolder1_teliti1_kiri1_alblfilemitradukung1"]/@href')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                s2 = ''
            # Use try-except catching to mitigate empty data
            try:
                t2 = content.xpath('//div[@id="ContentPlaceHolder1_teliti1_kiri1_updx"]/div[3][@class="panel panel-info"]/div[@class="panel-body f12"]/div[1][@class="row"]/div[@class="col-sm-4"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                t2 = ''
            # Use try-except catching to mitigate empty data
            try:
                u2 = content.xpath('//div[@id="ContentPlaceHolder1_teliti1_kiri1_updx"]/div[3][@class="panel panel-info"]/div[@class="panel-body f12"]/div[2][@class="row"]/div[@class="col-sm-4"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                u2 = ''
            # Use try-except catching to mitigate empty data
            try:
                v2 = content.xpath('//div[@id="ContentPlaceHolder1_teliti1_kiri1_updx"]/div[3][@class="panel panel-info"]/div[@class="panel-body f12"]/div[3][@class="row"]/div[@class="col-sm-4"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                v2 = ''
            # Use try-except catching to mitigate empty data
            try:
                w2 = content.xpath('//div[@id="ContentPlaceHolder1_teliti1_kiri1_updx"]/div[3][@class="panel panel-info"]/div[@class="panel-body f12"]/div[4][@class="row"]/div[@class="col-sm-4"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                w2 = ''
            # Use try-except catching to mitigate empty data
            try:
                x2 = content.xpath('//div[@id="ContentPlaceHolder1_teliti1_kiri1_updx"]/div[3][@class="panel panel-info"]/div[@class="panel-body f12"]/div[5][@class="row"]/div[@class="col-sm-4"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                x2 = ''

            # Painting the static (non-variable as in "Data Anggota") detail information
            sheet[f'A{row_start}'].value = a1
            sheet[f'B{row_start}'].value = b2
            sheet[f'C{row_start}'].value = c2
            sheet[f'D{row_start}'].value = d2
            sheet[f'E{row_start}'].value = e2
            sheet[f'F{row_start}'].value = f2
            sheet[f'G{row_start}'].value = g2
            sheet[f'H{row_start}'].value = h2
            sheet[f'I{row_start}'].value = i2
            sheet[f'J{row_start}'].value = j2
            sheet[f'K{row_start}'].value = k2
            sheet[f'L{row_start}'].value = l2
            sheet[f'M{row_start}'].value = m2
            sheet[f'N{row_start}'].value = n2
            sheet[f'O{row_start}'].value = o2
            sheet[f'P{row_start}'].value = p2
            sheet[f'Q{row_start}'].value = q2
            sheet[f'R{row_start}'].value = r2
            sheet[f'S{row_start}'].value = s2
            sheet[f'T{row_start}'].value = t2
            sheet[f'U{row_start}'].value = u2
            sheet[f'V{row_start}'].value = v2
            sheet[f'W{row_start}'].value = w2
            sheet[f'X{row_start}'].value = x2

            # The table which displays 'Identitas Pengusul - Anggota Peneliti'
            # Treated differently, specially
            # 'y_' is a DOM element representing a HTML table row (<tr>)
            y_ = content.xpath(
                '//div[@id="ContentPlaceHolder1_teliti1_kiri1_updx"]/div[4]//table[@class="table"]//tr[position()>1]')

            # The base path
            y_base = '//div[@id="ContentPlaceHolder1_teliti1_kiri1_updx"]/div[4]//table[@class="table"]//tr[position()>1]'

            # Painting the scraped data to the output spreadsheet row
            # Section: 'DATA ANGGOTA'
            # ---
            # Checking if 'Identitas Pengusul - Anggota Peneliti' table data exists
            if len(y_) == 0:
                pass  # --- nope. the data does not exist
            else:

                # Checking if this table's 'LUARAN' data has the most rows
                y_all_rows = len(y_)
                y_max_row = max(y_max_row, y_all_rows)

                # ROW DATA LABEL CONVENTION
                # y_row_a -> "N.I.P"
                # y_row_b -> "Nama"
                # y_row_c -> "Bidang Keahlian"
                # y_row_d -> "Alamat"
                # y_row_e -> "Instansi"
                # y_row_f -> "Email"
                # y_row_g -> "No. HP"
                # y_row_h -> "Peran"
                # y_row_i -> "Tugas"

                y_row_a = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[2]/text()')]

                y_row_b = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[3]/text()')]

                y_row_c = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[4]/text()')]

                y_row_d = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[5]/text()')]

                y_row_e = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[6]/text()')]

                y_row_f = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[7]/text()')]

                y_row_g = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[8]/text()')]

                y_row_h = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[9]/text()')]

                y_row_i = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[10]/text()')]

                # The starting column coordinate for filling the 'LUARAN' table data
                col_start = 25

                # Iterating through each table row and write to the spreadsheet
                # Assumes the lists y_row_a, y_row_b, ... have the same array size
                for j in range(len(y_row_a)):
                    # Painting table data
                    sheet.cell(row=row_start, column=col_start).value = y_row_a[j]
                    sheet.cell(row=row_start, column=col_start + 1).value = y_row_b[j]
                    sheet.cell(row=row_start, column=col_start + 2).value = y_row_c[j]
                    sheet.cell(row=row_start, column=col_start + 3).value = y_row_d[j]
                    sheet.cell(row=row_start, column=col_start + 4).value = y_row_e[j]
                    sheet.cell(row=row_start, column=col_start + 5).value = y_row_f[j]
                    sheet.cell(row=row_start, column=col_start + 6).value = y_row_g[j]
                    sheet.cell(row=row_start, column=col_start + 7).value = y_row_h[j]
                    sheet.cell(row=row_start, column=col_start + 8).value = y_row_i[j]

                    # Incrementing the 'col_start' iterator before continuing the loop
                    col_start += 9
                    continue

            # Incrementing the value of 'row_start' before continuing
            row_start += 1

            # Reopening the "Dana Pengabdian" list page,
            # then assign the AJAX response to the temporary array 'temporary_prompt'
            # The 'data' array is obtained from opening individual entry row detail page
            #
            # This is done only on the 7-multiple row iteration, because
            # a single set of 'viewstate', 'viewstategen', and 'eventvalidation' values of ASPX
            # can only be used to do at most 8 operations before having to be renewed.
            if ( i % 7 == 0 ) and ( i > 0 ):
                temporary_prompt = self.get_risat_pengabdian_usulan_pengabdian(data)
            continue

        # Post-loop logging: appending the header over the 'DATA ANGGOTA' columns
        control.append_message_area(
            f'+ Melengkapi kepala tabel pada bagian "Identitas Pengusul - Anggota Peneliti" ...')
        # ---
        # The starting column for the 'DATA ANGGOTA' data
        col_start = 25
        # Beginning the loop that detects the maximum number of 'LUARAN' rows
        # according to the variable 'y_max_row'
        for i in range(1, y_max_row + 1):
            # Setting the top header
            sheet.merge_cells(
                start_row=1,
                start_column=col_start,
                end_row=1,
                end_column=col_start + 8
            )
            sheet.cell(row=1, column=col_start).value = f'IDENTITAS PENGUSUL — ANGGOTA #{i}'
            sheet.cell(row=1, column=col_start).alignment = Alignment(horizontal='center')
            # Setting the sub headers
            sheet.cell(row=2, column=col_start).value = 'N.I.P'
            sheet.cell(row=2, column=col_start + 1).value = 'Nama'
            sheet.cell(row=2, column=col_start + 2).value = 'Bidang Keahlian'
            sheet.cell(row=2, column=col_start + 3).value = 'Alamat'
            sheet.cell(row=2, column=col_start + 4).value = 'Instansi'
            sheet.cell(row=2, column=col_start + 5).value = 'Email'
            sheet.cell(row=2, column=col_start + 6).value = 'No. HP'
            sheet.cell(row=2, column=col_start + 7).value = 'Peran'
            sheet.cell(row=2, column=col_start + 8).value = 'Tugas'
            # Incrementing the 'col_start' iterator before continuing the loop
            col_start += 9
            continue

        # Post-loop logging: successfully painted the output spreadsheet file
        control.append_message_area(f'+ Tabel sukses dipanen!')
        control.set_progress_bar(85)

        # Asking for the spreadsheet name to save as
        # ---
        # Logging and setting the progress bar
        control.append_message_area(f'+ Menyimpan spreadsheet luaran ...')
        control.set_progress_bar(90)
        # Dealing with file name prompt and saving
        # Using loop to mitigate the user clicking 'cancel'
        # in the file name dialog prompt
        while True:
            # Opening the dialog prompt
            output_spreadsheet = filedialog.asksaveasfilename(
                filetypes=[('Excel files', '*.xlsx')],
                initialfile='Sipesat - Usulan Pengabdian Detil Risat.xlsx',
                title='Simpan sebagai ...'
            )

            # 'cancel' button in the dialog prompt is clicked
            if len(output_spreadsheet) == 0:
                # Showing confirmation
                x = messagebox.askyesno(
                    'Nama File Kosong',
                    'Apakah Anda yakin ingin melanjutkan tanpa menyimpan file spreadsheet hasil pemanenan?'
                )
                # Determining whether to break or to continue the loop
                # based on the inversed value of 'x'
                if x:
                    control.append_message_area(
                        f'+ Finalisasi pemanenan data tanpa menyimpan file spreadsheet luaran ...')
                    workbook.close()  # --- closing the workbook without saving
                    break
                else:
                    continue  # --- continuing the loop
            # File name does not end in spreadsheet extension
            elif output_spreadsheet[-5:] != '.xlsx':
                output_spreadsheet = output_spreadsheet + '.xlsx'

            # Saving the spreadsheet
            control.append_message_area(f'LOKASI_SPREADSHEET_LUARAN: {output_spreadsheet}')
            control.set_progress_bar(95)
            workbook.save(output_spreadsheet)

            # Closing the openpyxl workbook
            control.append_message_area(f'+ Menutup file spreadsheet ...')
            control.set_progress_bar(98)
            workbook.close()

            # Breaking the loop
            break

        # Notify for a successful scraping
        control.append_message_area(f'+ Pemanenan selesai pada: {str(dt.now())}')
        control.set_progress_bar(100)
        control.on_notify_successful_scraping()

    # This function harvests "Risat Berkas Disetujui Pengabdian > Data Detil Lengkap" data
    # and then store the harvested data as an excel file
    #
    # Required arguments:
    # - control             --> for updating the progress bar and
    #                           message area of the screen SipesatScrHarvest
    # - username, password  --> the Risat administrator username and password
    def run_harvest_c_1_berdisetujui(self, control, username, password):

        # SipesatScrHarvest messenger
        control.set_header_desc('Panen Data "Risat Berkas Disetujui Pengabdian > Data Detil Lengkap"')
        control.set_help_label('Data sedang dipanen. Silahkan menunggu.')
        control.set_progress_bar(0)
        control.clear_message_area()

        # Preamble logging
        control.append_message_area(f'+ Memulai pemanenan data ...')
        control.append_message_area(f'+ Pemanenan dimulai pada: {str(dt.now())}')
        control.set_progress_bar(5)

        # Preparing the 'data_prompt' arrays
        control.append_message_area(f'+ Log masuk Risat sebagai [{username}] ...')
        control.set_progress_bar(10)
        data_prompt = self.get_risat_login(username, password)
        data_prompt = self.get_risat_pengabdian(data_prompt)
        data_prompt = self.get_risat_pengabdian_berdisetujui_pengabdian(data_prompt)

        # Parsing XML tree content
        control.append_message_area(f'+ Membaca halaman web ...')
        control.set_progress_bar(15)
        content = data_prompt['html_content']

        # Establishing the export spreadsheet file
        control.append_message_area(f'+ Mempersiapkan file spreadsheet luaran ...')
        control.set_progress_bar(20)
        workbook = xl.Workbook()
        sheet = workbook.active
        sheet.title = 'Berkas Disetujui Pengabdian Detil'

        # DEVELOPER'S NOTE (DO NOT REMOVE)
        '''
        Data yang perlu diambil:
        1. Judul
        2. Substansi usulan
        3. Identitas pengusul - ketua
        4. Identitas pengusul - anggota
        5. RAB
        6. Dokumen pendukung

        Urutan kolom excel:
        1. Judul
        2. Substansi usulan
        3. RAB
        4. Dokumen pendukung
        5. Identitas pengusul - ketua
        6. Identitas pengusul - anggota
        '''

        # Preparing the sheet header
        control.append_message_area(f'+ Mempersiapkan kepala lembar spreadsheet ...')
        control.set_progress_bar(25)
        # ---
        # Preparing the "NOMOR" header
        sheet.merge_cells(
            start_row=1,
            start_column=1,
            end_row=2,
            end_column=1
        )
        sheet['A1'].value = 'No.'
        sheet['A1'].alignment = Alignment(horizontal='center')
        # ---
        # Preparing the "JUDUL" header
        sheet.merge_cells(
            start_row=1,
            start_column=2,
            end_row=1,
            end_column=12
        )
        sheet['B1'].value = 'IDENTITAS'
        sheet['B1'].alignment = Alignment(horizontal='center')
        # Preparing the "JUDUL" sub-headers
        sheet['B2'].value = 'Judul'
        sheet['C2'].value = 'Tgl. Usulan'
        sheet['D2'].value = 'TKT Saat Ini'
        sheet['E2'].value = 'Level'
        sheet['F2'].value = 'Kategori'
        sheet['G2'].value = 'Skema'
        sheet['H2'].value = 'Rumpun Ilmu'
        sheet['I2'].value = 'Bidang Fokus'
        sheet['J2'].value = 'Tema'
        sheet['K2'].value = 'Topik'
        sheet['L2'].value = 'Lama Kegiatan'
        # ---
        # Preparing the "SUBSTANSI USULAN" header
        sheet.merge_cells(
            start_row=1,
            start_column=13,
            end_row=1,
            end_column=14
        )
        sheet['M1'].value = 'SUBSTANSI USULAN'
        sheet['M1'].alignment = Alignment(horizontal='center')
        # Preparing the "SUBSTANSI USULAN" sub-headers
        sheet['M2'].value = 'Kelompok Makro'
        sheet['N2'].value = 'File Proposal'
        # ---
        # Preparing the "RAB" header
        sheet.merge_cells(
            start_row=1,
            start_column=15,
            end_row=1,
            end_column=16
        )
        sheet['O1'].value = 'RENCANA ANGGARAN BIAYA'
        sheet['O1'].alignment = Alignment(horizontal='center')
        # Preparing the "RAB" sub-headers
        sheet['O2'].value = 'Biaya'
        sheet['P2'].value = 'File RAB'
        # ---
        # Preparing the "DOKUMEN PENDUKUNG" header
        sheet.merge_cells(
            start_row=1,
            start_column=17,
            end_row=1,
            end_column=19
        )
        sheet['Q1'].value = 'DOKUMEN PENDUKUNG'
        sheet['Q1'].alignment = Alignment(horizontal='center')
        # Preparing the "DOKUMEN PENDUKUNG" sub-headers
        sheet['Q2'].value = 'Mitra'
        sheet['R2'].value = 'Dukungan Biaya'
        sheet['S2'].value = 'Surat Dukungan Mitra'
        # ---
        # Preparing the "IDENTITAS PENGUSUL — KETUA" header
        sheet.merge_cells(
            start_row=1,
            start_column=20,
            end_row=1,
            end_column=24
        )
        sheet['T1'].value = 'IDENTITAS PENGUSUL — KETUA'
        sheet['T1'].alignment = Alignment(horizontal='center')
        # Preparing the "IDENTITAS PENGUSUL — KETUA" sub-headers
        sheet['T2'].value = 'N.I.P'
        sheet['U2'].value = 'N.I.K'
        sheet['V2'].value = 'N.I.D.N'
        sheet['W2'].value = 'Nama Lengkap'
        sheet['X2'].value = 'Jabatan Fungsional'

        # The base XPath location, pointing to each entry row
        base = '//div[@class="mw-100"]//div[@class="form-group f12"]/table[@width="100%"]//tr[@valign="top"]'

        # ---
        # Obtaining the data row values
        control.append_message_area(f'+ Mendapatkan data pada baris tabel ...')
        control.set_progress_bar(30)

        # Reading the HTML entry row hidden ASPX values
        # Copy-pasted from: /ssynthesia/ghostcity/ar/dumper-2/24__2023.02.13__requestsrisat.py
        all_kodetran_prop = content.xpath(base + '//input[1][@type="hidden"]/@name')
        all_kodetran_val = content.xpath(base + '//input[1][@type="hidden"]/@value')
        all_stat_prop = content.xpath(base + '//input[2][@type="hidden"]/@name')
        all_stat_val = content.xpath(base + '//input[2][@type="hidden"]/@value')
        all_submitbtn = content.xpath(base + '//tr[7]//input[@type="submit"]/@name')

        # The number of rows
        # Assumes the 'all_' array size equals the number of data rows
        number_of_row = len(all_kodetran_prop)

        # The starting row coordinate of the active sheet
        row_start = 3

        # The maximum number of 'DATA ANGGOTA' table data row
        y_max_row = 0

        # Iterating through each entry row element
        # Assumes all the 'all_' arrays in the previous code block
        # are of the same length/size
        # Copy-pasted from: /ssynthesia/ghostcity/ar/dumper-2/24__2023.02.13__requestsrisat.py
        control.append_message_area(f'+ Melakukan iterasi terhadap baris tabel dan menulis spreadsheet luaran ...')
        control.set_progress_bar(35)
        temporary_prompt = data_prompt
        for i in range(number_of_row):
            # Noisy preamble logging
            # Please don't use this -_-
            # ---
            # control.append_message_area(f'ITERASI [{i}]')

            # Updating the progress bar status
            control.set_progress_bar(35 + round(45 * (i + 1) / number_of_row))

            # Preparing the AJAX payload
            detail_prompt = {
                'viewstate': temporary_prompt['viewstate'],
                'viewstategen': temporary_prompt['viewstategen'],
                'eventvalidation': temporary_prompt['eventvalidation'],
                'button_name': all_submitbtn[i],
                'kodetran_prop': all_kodetran_prop[i],
                'kodetran_val': all_kodetran_val[i],
                'stat_prop': all_stat_prop[i],
                'stat_val': all_stat_val[i]
            }

            # Obtaining the response data of each individual entry row detail page
            data = self.get_risat_pengabdian_berdisetujui_pengabdian_detil(detail_prompt)
            content = data['html_content']
            response = data['http_response']

            # Terminal logging for detecting errors
            log_string_1 = str(i + 1)
            log_string_2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti2_kiri1_txjudul1"]/text()')[
                0].replace('\r', '').replace('\n', '').strip()
            control.append_message_area(f'+ Memanen data detil: [{log_string_1}] {log_string_2} ...')
            print(f'+ Harvesting detailed data: [{log_string_1}] {log_string_2}')

            # DEVELOPER'S NOTE (DO NOT REMOVE)
            '''
            Variables and their associated data

            'a1' ==> 'No.'
            'b2' ==> 'Judul'
            'c2' ==> 'Tgl. Usulan'
            'd2' ==> 'TKT Saat Ini'
            'e2' ==> 'Level'
            'f2' ==> 'Kategori'
            'g2' ==> 'Skema'
            'h2' ==> 'Rumpun Ilmu'
            'i2' ==> 'Bidang Fokus'
            'j2' ==> 'Tema'
            'k2' ==> 'Topik'
            'l2' ==> 'Lama Kegiatan'
            'm2' ==> 'Kelompok Makro'
            'n2' ==> 'File Proposal'
            'o2' ==> 'Biaya'
            'p2' ==> 'File RAB'
            'q2' ==> 'Mitra'
            'r2' ==> 'Dukungan Biaya'
            's2' ==> 'Surat Dukungan Mitra'
            't2' ==> 'N.I.P'
            'u2' ==> 'N.I.K'
            'v2' ==> 'N.I.D.N'
            'w2' ==> 'Nama Lengkap'
            'x2' ==> 'Jabatan Fungsional'

            In addition, variables starting with 'y' prefix are pertaining
            to the table data of "Data Anggota"
            '''

            # Getting the static (non-variable as in "Data Anggota") detail information
            # ---
            a1 = str(i+1)
            # Use try-except catching to mitigate empty data
            try:
                b2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti2_kiri1_txjudul1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                b2 = ''
            # Use try-except catching to mitigate empty data
            try:
                c2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti2_kiri1_tglusul1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                c2 = ''
            # 'Pengabdian Masyarakat' ain't have TKT (d2) and level (e2) data!
            # ---
            d2 = '-'
            e2 = '-'
            # Use try-except catching to mitigate empty data
            try:
                f2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti2_kiri1_lkat1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                f2 = ''
            # Use try-except catching to mitigate empty data
            try:
                g2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti2_kiri1_ddlskema1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                g2 = ''
            # 'Rumpun Ilmu' scraped data contains multiple array of strings
            # This needs some extra tweaking
            h2 = ''
            # Use try-except catching to mitigate empty data
            try:
                h2_pre = content.xpath('//div[@id="ContentPlaceHolder1_teliti2_kiri1_updx"]/div[1][@class="panel panel-info"]/div[@class="panel-body f12"]/div[5][@class="row f12"]/div[@class="col-sm-4"]//text()')
                for l in h2_pre:
                    l = l.replace('\r','').replace('\n','').strip()
                    h2 += l +', '
                # Remove trailing ', ' characters
                h2 = h2[:-2]
            except IndexError:
                h2 = ''
            # ---
            # Use try-except catching to mitigate empty data
            try:
                i2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti2_kiri1_ddlfokus1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                i2 = ''
            # Use try-except catching to mitigate empty data
            try:
                j2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti2_kiri1_ddltema1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                j2 = ''
            # Some 'Topik' has no data element
            # Use try-except catching to mitigate
            try:
                k2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti2_kiri1_ddltopik1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                k2 = ''
            # Use try-except catching to mitigate empty data
            try:
                l2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti2_kiri1_ddllama1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                l2 = ''
            # Use try-except catching to mitigate empty data
            try:
                m2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti2_kiri1_ddlmakro1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                m2 = ''
            # Use try-except catching to mitigate empty data
            try:
                n2 = content.xpath('//a[@id="ContentPlaceHolder1_teliti2_kiri1_alblcatatan1"]/@href')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                n2 = ''
            # Use try-except catching to mitigate empty data
            try:
                o2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti2_kiri1_tbiayathn1"]/text()')[0].replace('\r','').replace('\n','').replace('Rp.', '').replace(',', '').strip()
            except IndexError:
                o2 = ''
            # Use try-except catching to mitigate empty data
            try:
                p2 = content.xpath('//a[@id="ContentPlaceHolder1_teliti2_kiri1_alblfilerab1"]/@href')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                p2 = ''
            # Some 'Data Detil' has no 'Mitra' data element
            # Use try-except catching to mitigate
            try:
                q2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti2_kiri1_tnmmitra1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                q2 = ''
            # Use try-except catching to mitigate empty data
            try:
                r2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti2_kiri1_tbiayadukung1"]/text()')[0].replace('\r','').replace('\n','').replace('Rp.', '').replace(',', '').strip()
            except IndexError:
                r2 = ''
            # Use try-except catching to mitigate empty data
            try:
                s2 = content.xpath('//a[@id="ContentPlaceHolder1_teliti2_kiri1_alblfilemitradukung1"]/@href')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                s2 = ''
            # Use try-except catching to mitigate empty data
            try:
                t2 = content.xpath('//div[@id="ContentPlaceHolder1_teliti2_kiri1_updx"]/div[3][@class="panel panel-info"]/div[@class="panel-body f12"]/div[1][@class="row"]/div[@class="col-sm-4"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                t2 = ''
            # Use try-except catching to mitigate empty data
            try:
                u2 = content.xpath('//div[@id="ContentPlaceHolder1_teliti2_kiri1_updx"]/div[3][@class="panel panel-info"]/div[@class="panel-body f12"]/div[2][@class="row"]/div[@class="col-sm-4"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                u2 = ''
            # Use try-except catching to mitigate empty data
            try:
                v2 = content.xpath('//div[@id="ContentPlaceHolder1_teliti2_kiri1_updx"]/div[3][@class="panel panel-info"]/div[@class="panel-body f12"]/div[3][@class="row"]/div[@class="col-sm-4"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                v2 = ''
            # Use try-except catching to mitigate empty data
            try:
                w2 = content.xpath('//div[@id="ContentPlaceHolder1_teliti2_kiri1_updx"]/div[3][@class="panel panel-info"]/div[@class="panel-body f12"]/div[4][@class="row"]/div[@class="col-sm-4"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                w2 = ''
            # Use try-except catching to mitigate empty data
            try:
                x2 = content.xpath('//div[@id="ContentPlaceHolder1_teliti2_kiri1_updx"]/div[3][@class="panel panel-info"]/div[@class="panel-body f12"]/div[5][@class="row"]/div[@class="col-sm-4"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                x2 = ''

            # Painting the static (non-variable as in "Data Anggota") detail information
            sheet[f'A{row_start}'].value = a1
            sheet[f'B{row_start}'].value = b2
            sheet[f'C{row_start}'].value = c2
            sheet[f'D{row_start}'].value = d2
            sheet[f'E{row_start}'].value = e2
            sheet[f'F{row_start}'].value = f2
            sheet[f'G{row_start}'].value = g2
            sheet[f'H{row_start}'].value = h2
            sheet[f'I{row_start}'].value = i2
            sheet[f'J{row_start}'].value = j2
            sheet[f'K{row_start}'].value = k2
            sheet[f'L{row_start}'].value = l2
            sheet[f'M{row_start}'].value = m2
            sheet[f'N{row_start}'].value = n2
            sheet[f'O{row_start}'].value = o2
            sheet[f'P{row_start}'].value = p2
            sheet[f'Q{row_start}'].value = q2
            sheet[f'R{row_start}'].value = r2
            sheet[f'S{row_start}'].value = s2
            sheet[f'T{row_start}'].value = t2
            sheet[f'U{row_start}'].value = u2
            sheet[f'V{row_start}'].value = v2
            sheet[f'W{row_start}'].value = w2
            sheet[f'X{row_start}'].value = x2

            # The table which displays 'Identitas Pengusul - Anggota Peneliti'
            # Treated differently, specially
            # 'y_' is a DOM element representing a HTML table row (<tr>)
            y_ = content.xpath(
                '//div[@id="ContentPlaceHolder1_teliti2_kiri1_updx"]/div[4]//table[@class="table"]//tr[position()>1]')

            # The base path
            y_base = '//div[@id="ContentPlaceHolder1_teliti2_kiri1_updx"]/div[4]//table[@class="table"]//tr[position()>1]'

            # Painting the scraped data to the output spreadsheet row
            # Section: 'DATA ANGGOTA'
            # ---
            # Checking if 'Identitas Pengusul - Anggota Peneliti' table data exists
            if len(y_) == 0:
                pass  # --- nope. the data does not exist
            else:

                # Checking if this table's 'LUARAN' data has the most rows
                y_all_rows = len(y_)
                y_max_row = max(y_max_row, y_all_rows)

                # ROW DATA LABEL CONVENTION
                # y_row_a -> "N.I.P"
                # y_row_b -> "Nama"
                # y_row_c -> "Bidang Keahlian"
                # y_row_d -> "Alamat"
                # y_row_e -> "Instansi"
                # y_row_f -> "Email"
                # y_row_g -> "No. HP"
                # y_row_h -> "Peran"
                # y_row_i -> "Tugas"

                y_row_a = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[2]/text()')]

                y_row_b = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[3]/text()')]

                y_row_c = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[4]/text()')]

                y_row_d = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[5]/text()')]

                y_row_e = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[6]/text()')]

                y_row_f = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[7]/text()')]

                y_row_g = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[8]/text()')]

                y_row_h = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[9]/text()')]

                y_row_i = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[10]/text()')]

                # The starting column coordinate for filling the 'LUARAN' table data
                col_start = 25

                # Iterating through each table row and write to the spreadsheet
                # Assumes the lists y_row_a, y_row_b, ... have the same array size
                for j in range(len(y_row_a)):
                    # Painting table data
                    sheet.cell(row=row_start, column=col_start).value = y_row_a[j]
                    sheet.cell(row=row_start, column=col_start + 1).value = y_row_b[j]
                    sheet.cell(row=row_start, column=col_start + 2).value = y_row_c[j]
                    sheet.cell(row=row_start, column=col_start + 3).value = y_row_d[j]
                    sheet.cell(row=row_start, column=col_start + 4).value = y_row_e[j]
                    sheet.cell(row=row_start, column=col_start + 5).value = y_row_f[j]
                    sheet.cell(row=row_start, column=col_start + 6).value = y_row_g[j]
                    sheet.cell(row=row_start, column=col_start + 7).value = y_row_h[j]
                    sheet.cell(row=row_start, column=col_start + 8).value = y_row_i[j]

                    # Incrementing the 'col_start' iterator before continuing the loop
                    col_start += 9
                    continue

            # Incrementing the value of 'row_start' before continuing
            row_start += 1

            # Reopening the "Dana Pengabdian" list page,
            # then assign the AJAX response to the temporary array 'temporary_prompt'
            # The 'data' array is obtained from opening individual entry row detail page
            #
            # This is done only on the 7-multiple row iteration, because
            # a single set of 'viewstate', 'viewstategen', and 'eventvalidation' values of ASPX
            # can only be used to do at most 8 operations before having to be renewed.
            if ( i % 7 == 0 ) and ( i > 0 ):
                temporary_prompt = self.get_risat_pengabdian_berdisetujui_pengabdian(data)
            continue

        # Post-loop logging: appending the header over the 'DATA ANGGOTA' columns
        control.append_message_area(
            f'+ Melengkapi kepala tabel pada bagian "Identitas Pengusul - Anggota Peneliti" ...')
        # ---
        # The starting column for the 'DATA ANGGOTA' data
        col_start = 25
        # Beginning the loop that detects the maximum number of 'LUARAN' rows
        # according to the variable 'y_max_row'
        for i in range(1, y_max_row + 1):
            # Setting the top header
            sheet.merge_cells(
                start_row=1,
                start_column=col_start,
                end_row=1,
                end_column=col_start + 8
            )
            sheet.cell(row=1, column=col_start).value = f'IDENTITAS PENGUSUL — ANGGOTA #{i}'
            sheet.cell(row=1, column=col_start).alignment = Alignment(horizontal='center')
            # Setting the sub headers
            sheet.cell(row=2, column=col_start).value = 'N.I.P'
            sheet.cell(row=2, column=col_start + 1).value = 'Nama'
            sheet.cell(row=2, column=col_start + 2).value = 'Bidang Keahlian'
            sheet.cell(row=2, column=col_start + 3).value = 'Alamat'
            sheet.cell(row=2, column=col_start + 4).value = 'Instansi'
            sheet.cell(row=2, column=col_start + 5).value = 'Email'
            sheet.cell(row=2, column=col_start + 6).value = 'No. HP'
            sheet.cell(row=2, column=col_start + 7).value = 'Peran'
            sheet.cell(row=2, column=col_start + 8).value = 'Tugas'
            # Incrementing the 'col_start' iterator before continuing the loop
            col_start += 9
            continue

        # Post-loop logging: successfully painted the output spreadsheet file
        control.append_message_area(f'+ Tabel sukses dipanen!')
        control.set_progress_bar(85)

        # Asking for the spreadsheet name to save as
        # ---
        # Logging and setting the progress bar
        control.append_message_area(f'+ Menyimpan spreadsheet luaran ...')
        control.set_progress_bar(90)
        # Dealing with file name prompt and saving
        # Using loop to mitigate the user clicking 'cancel'
        # in the file name dialog prompt
        while True:
            # Opening the dialog prompt
            output_spreadsheet = filedialog.asksaveasfilename(
                filetypes=[('Excel files', '*.xlsx')],
                initialfile='Sipesat - Berkas Disetujui Pengabdian Detil Risat.xlsx',
                title='Simpan sebagai ...'
            )

            # 'cancel' button in the dialog prompt is clicked
            if len(output_spreadsheet) == 0:
                # Showing confirmation
                x = messagebox.askyesno(
                    'Nama File Kosong',
                    'Apakah Anda yakin ingin melanjutkan tanpa menyimpan file spreadsheet hasil pemanenan?'
                )
                # Determining whether to break or to continue the loop
                # based on the inversed value of 'x'
                if x:
                    control.append_message_area(
                        f'+ Finalisasi pemanenan data tanpa menyimpan file spreadsheet luaran ...')
                    workbook.close()  # --- closing the workbook without saving
                    break
                else:
                    continue  # --- continuing the loop
            # File name does not end in spreadsheet extension
            elif output_spreadsheet[-5:] != '.xlsx':
                output_spreadsheet = output_spreadsheet + '.xlsx'

            # Saving the spreadsheet
            control.append_message_area(f'LOKASI_SPREADSHEET_LUARAN: {output_spreadsheet}')
            control.set_progress_bar(95)
            workbook.save(output_spreadsheet)

            # Closing the openpyxl workbook
            control.append_message_area(f'+ Menutup file spreadsheet ...')
            control.set_progress_bar(98)
            workbook.close()

            # Breaking the loop
            break

        # Notify for a successful scraping
        control.append_message_area(f'+ Pemanenan selesai pada: {str(dt.now())}')
        control.set_progress_bar(100)
        control.on_notify_successful_scraping()

    # This function harvests "Risat Berkas Direvisi Pengabdian > Data Detil Lengkap" data
    # and then store the harvested data as an excel file
    #
    # Required arguments:
    # - control             --> for updating the progress bar and
    #                           message area of the screen SipesatScrHarvest
    # - username, password  --> the Risat administrator username and password
    def run_harvest_c_1_berdirevisi(self, control, username, password):

        # SipesatScrHarvest messenger
        control.set_header_desc('Panen Data "Risat Berkas Direvisi Pengabdian > Data Detil Lengkap"')
        control.set_help_label('Data sedang dipanen. Silahkan menunggu.')
        control.set_progress_bar(0)
        control.clear_message_area()

        # Preamble logging
        control.append_message_area(f'+ Memulai pemanenan data ...')
        control.append_message_area(f'+ Pemanenan dimulai pada: {str(dt.now())}')
        control.set_progress_bar(5)

        # Preparing the 'data_prompt' arrays
        control.append_message_area(f'+ Log masuk Risat sebagai [{username}] ...')
        control.set_progress_bar(10)
        data_prompt = self.get_risat_login(username, password)
        data_prompt = self.get_risat_pengabdian(data_prompt)
        data_prompt = self.get_risat_pengabdian_berdirevisi_pengabdian(data_prompt)

        # Parsing XML tree content
        control.append_message_area(f'+ Membaca halaman web ...')
        control.set_progress_bar(15)
        content = data_prompt['html_content']

        # Establishing the export spreadsheet file
        control.append_message_area(f'+ Mempersiapkan file spreadsheet luaran ...')
        control.set_progress_bar(20)
        workbook = xl.Workbook()
        sheet = workbook.active
        sheet.title = 'Berkas Direvisi Pengabdian Detil'

        # DEVELOPER'S NOTE (DO NOT REMOVE)
        '''
        Data yang perlu diambil:
        1. Judul
        2. Substansi usulan
        3. Identitas pengusul - ketua
        4. Identitas pengusul - anggota
        5. RAB
        6. Dokumen pendukung

        Urutan kolom excel:
        1. Judul
        2. Substansi usulan
        3. RAB
        4. Dokumen pendukung
        5. Identitas pengusul - ketua
        6. Identitas pengusul - anggota
        '''

        # Preparing the sheet header
        control.append_message_area(f'+ Mempersiapkan kepala lembar spreadsheet ...')
        control.set_progress_bar(25)
        # ---
        # Preparing the "NOMOR" header
        sheet.merge_cells(
            start_row=1,
            start_column=1,
            end_row=2,
            end_column=1
        )
        sheet['A1'].value = 'No.'
        sheet['A1'].alignment = Alignment(horizontal='center')
        # ---
        # Preparing the "JUDUL" header
        sheet.merge_cells(
            start_row=1,
            start_column=2,
            end_row=1,
            end_column=12
        )
        sheet['B1'].value = 'IDENTITAS'
        sheet['B1'].alignment = Alignment(horizontal='center')
        # Preparing the "JUDUL" sub-headers
        sheet['B2'].value = 'Judul'
        sheet['C2'].value = 'Tgl. Usulan'
        sheet['D2'].value = 'TKT Saat Ini'
        sheet['E2'].value = 'Level'
        sheet['F2'].value = 'Kategori'
        sheet['G2'].value = 'Skema'
        sheet['H2'].value = 'Rumpun Ilmu'
        sheet['I2'].value = 'Bidang Fokus'
        sheet['J2'].value = 'Tema'
        sheet['K2'].value = 'Topik'
        sheet['L2'].value = 'Lama Kegiatan'
        # ---
        # Preparing the "SUBSTANSI USULAN" header
        sheet.merge_cells(
            start_row=1,
            start_column=13,
            end_row=1,
            end_column=14
        )
        sheet['M1'].value = 'SUBSTANSI USULAN'
        sheet['M1'].alignment = Alignment(horizontal='center')
        # Preparing the "SUBSTANSI USULAN" sub-headers
        sheet['M2'].value = 'Kelompok Makro'
        sheet['N2'].value = 'File Proposal'
        # ---
        # Preparing the "RAB" header
        sheet.merge_cells(
            start_row=1,
            start_column=15,
            end_row=1,
            end_column=16
        )
        sheet['O1'].value = 'RENCANA ANGGARAN BIAYA'
        sheet['O1'].alignment = Alignment(horizontal='center')
        # Preparing the "RAB" sub-headers
        sheet['O2'].value = 'Biaya'
        sheet['P2'].value = 'File RAB'
        # ---
        # Preparing the "DOKUMEN PENDUKUNG" header
        sheet.merge_cells(
            start_row=1,
            start_column=17,
            end_row=1,
            end_column=19
        )
        sheet['Q1'].value = 'DOKUMEN PENDUKUNG'
        sheet['Q1'].alignment = Alignment(horizontal='center')
        # Preparing the "DOKUMEN PENDUKUNG" sub-headers
        sheet['Q2'].value = 'Mitra'
        sheet['R2'].value = 'Dukungan Biaya'
        sheet['S2'].value = 'Surat Dukungan Mitra'
        # ---
        # Preparing the "IDENTITAS PENGUSUL — KETUA" header
        sheet.merge_cells(
            start_row=1,
            start_column=20,
            end_row=1,
            end_column=24
        )
        sheet['T1'].value = 'IDENTITAS PENGUSUL — KETUA'
        sheet['T1'].alignment = Alignment(horizontal='center')
        # Preparing the "IDENTITAS PENGUSUL — KETUA" sub-headers
        sheet['T2'].value = 'N.I.P'
        sheet['U2'].value = 'N.I.K'
        sheet['V2'].value = 'N.I.D.N'
        sheet['W2'].value = 'Nama Lengkap'
        sheet['X2'].value = 'Jabatan Fungsional'

        # The base XPath location, pointing to each entry row
        base = '//div[@class="mw-100"]//div[@class="form-group f12"]/table[@width="100%"]//tr[@valign="top"]'

        # ---
        # Obtaining the data row values
        control.append_message_area(f'+ Mendapatkan data pada baris tabel ...')
        control.set_progress_bar(30)

        # Reading the HTML entry row hidden ASPX values
        # Copy-pasted from: /ssynthesia/ghostcity/ar/dumper-2/24__2023.02.13__requestsrisat.py
        all_kodetran_prop = content.xpath(base + '//input[1][@type="hidden"]/@name')
        all_kodetran_val = content.xpath(base + '//input[1][@type="hidden"]/@value')
        all_stat_prop = content.xpath(base + '//input[2][@type="hidden"]/@name')
        all_stat_val = content.xpath(base + '//input[2][@type="hidden"]/@value')
        all_submitbtn = content.xpath(base + '//tr[7]//input[@type="submit"]/@name')

        # The number of rows
        # Assumes the 'all_' array size equals the number of data rows
        number_of_row = len(all_kodetran_prop)

        # The starting row coordinate of the active sheet
        row_start = 3

        # The maximum number of 'DATA ANGGOTA' table data row
        y_max_row = 0

        # Iterating through each entry row element
        # Assumes all the 'all_' arrays in the previous code block
        # are of the same length/size
        # Copy-pasted from: /ssynthesia/ghostcity/ar/dumper-2/24__2023.02.13__requestsrisat.py
        control.append_message_area(f'+ Melakukan iterasi terhadap baris tabel dan menulis spreadsheet luaran ...')
        control.set_progress_bar(35)
        temporary_prompt = data_prompt
        for i in range(number_of_row):
            # Noisy preamble logging
            # Please don't use this -_-
            # ---
            # control.append_message_area(f'ITERASI [{i}]')

            # Updating the progress bar status
            control.set_progress_bar(35 + round(45 * (i + 1) / number_of_row))

            # Preparing the AJAX payload
            detail_prompt = {
                'viewstate': temporary_prompt['viewstate'],
                'viewstategen': temporary_prompt['viewstategen'],
                'eventvalidation': temporary_prompt['eventvalidation'],
                'button_name': all_submitbtn[i],
                'kodetran_prop': all_kodetran_prop[i],
                'kodetran_val': all_kodetran_val[i],
                'stat_prop': all_stat_prop[i],
                'stat_val': all_stat_val[i]
            }

            # Obtaining the response data of each individual entry row detail page
            data = self.get_risat_pengabdian_berdirevisi_pengabdian_detil(detail_prompt)
            content = data['html_content']
            response = data['http_response']

            # Terminal logging for detecting errors
            log_string_1 = str(i + 1)
            log_string_2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti3_kiri1_txjudul1"]/text()')[
                0].replace('\r', '').replace('\n', '').strip()
            control.append_message_area(f'+ Memanen data detil: [{log_string_1}] {log_string_2} ...')
            print(f'+ Harvesting detailed data: [{log_string_1}] {log_string_2}')

            # DEVELOPER'S NOTE (DO NOT REMOVE)
            '''
            Variables and their associated data

            'a1' ==> 'No.'
            'b2' ==> 'Judul'
            'c2' ==> 'Tgl. Usulan'
            'd2' ==> 'TKT Saat Ini'
            'e2' ==> 'Level'
            'f2' ==> 'Kategori'
            'g2' ==> 'Skema'
            'h2' ==> 'Rumpun Ilmu'
            'i2' ==> 'Bidang Fokus'
            'j2' ==> 'Tema'
            'k2' ==> 'Topik'
            'l2' ==> 'Lama Kegiatan'
            'm2' ==> 'Kelompok Makro'
            'n2' ==> 'File Proposal'
            'o2' ==> 'Biaya'
            'p2' ==> 'File RAB'
            'q2' ==> 'Mitra'
            'r2' ==> 'Dukungan Biaya'
            's2' ==> 'Surat Dukungan Mitra'
            't2' ==> 'N.I.P'
            'u2' ==> 'N.I.K'
            'v2' ==> 'N.I.D.N'
            'w2' ==> 'Nama Lengkap'
            'x2' ==> 'Jabatan Fungsional'

            In addition, variables starting with 'y' prefix are pertaining
            to the table data of "Data Anggota"
            '''

            # Getting the static (non-variable as in "Data Anggota") detail information
            # ---
            a1 = str(i+1)
            # Use try-except catching to mitigate empty data
            try:
                b2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti3_kiri1_txjudul1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                b2 = ''
            # Use try-except catching to mitigate empty data
            try:
                c2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti3_kiri1_tglusul1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                c2 = ''
            # 'Pengabdian Masyarakat' ain't have TKT (d2) and level (e2) data!
            # ---
            d2 = '-'
            e2 = '-'
            # Use try-except catching to mitigate empty data
            try:
                f2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti3_kiri1_lkat1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                f2 = ''
            # Use try-except catching to mitigate empty data
            try:
                g2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti3_kiri1_ddlskema1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                g2 = ''
            # 'Rumpun Ilmu' scraped data contains multiple array of strings
            # This needs some extra tweaking
            h2 = ''
            # Use try-except catching to mitigate empty data
            try:
                h2_pre = content.xpath('//div[@id="ContentPlaceHolder1_teliti3_kiri1_updx"]/div[1][@class="panel panel-info"]/div[@class="panel-body f12"]/div[5][@class="row f12"]/div[@class="col-sm-4"]//text()')
                for l in h2_pre:
                    l = l.replace('\r','').replace('\n','').strip()
                    h2 += l +', '
                # Remove trailing ', ' characters
                h2 = h2[:-2]
            except IndexError:
                h2 = ''
            # ---
            # Use try-except catching to mitigate empty data
            try:
                i2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti3_kiri1_ddlfokus1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                i2 = ''
            # Use try-except catching to mitigate empty data
            try:
                j2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti3_kiri1_ddltema1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                j2 = ''
            # Some 'Topik' has no data element
            # Use try-except catching to mitigate
            try:
                k2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti3_kiri1_ddltopik1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                k2 = ''
            # Use try-except catching to mitigate empty data
            try:
                l2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti3_kiri1_ddllama1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                l2 = ''
            # Use try-except catching to mitigate empty data
            try:
                m2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti3_kiri1_ddlmakro1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                m2 = ''
            # Use try-except catching to mitigate empty data
            try:
                n2 = content.xpath('//a[@id="ContentPlaceHolder1_teliti3_kiri1_alblcatatan1"]/@href')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                n2 = ''
            # Use try-except catching to mitigate empty data
            try:
                o2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti3_kiri1_tbiayathn1"]/text()')[0].replace('\r','').replace('\n','').replace('Rp.', '').replace(',', '').strip()
            except IndexError:
                o2 = ''
            # Use try-except catching to mitigate empty data
            try:
                p2 = content.xpath('//a[@id="ContentPlaceHolder1_teliti3_kiri1_alblfilerab1"]/@href')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                p2 = ''
            # Some 'Data Detil' has no 'Mitra' data element
            # Use try-except catching to mitigate
            try:
                q2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti3_kiri1_tnmmitra1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                q2 = ''
            # Use try-except catching to mitigate empty data
            try:
                r2 = content.xpath('//span[@id="ContentPlaceHolder1_teliti3_kiri1_tbiayadukung1"]/text()')[0].replace('\r','').replace('\n','').replace('Rp.', '').replace(',', '').strip()
            except IndexError:
                r2 = ''
            # Use try-except catching to mitigate empty data
            try:
                s2 = content.xpath('//a[@id="ContentPlaceHolder1_teliti3_kiri1_alblfilemitradukung1"]/@href')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                s2 = ''
            # Use try-except catching to mitigate empty data
            try:
                t2 = content.xpath('//div[@id="ContentPlaceHolder1_teliti3_kiri1_updx"]/div[3][@class="panel panel-info"]/div[@class="panel-body f12"]/div[1][@class="row"]/div[@class="col-sm-4"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                t2 = ''
            # Use try-except catching to mitigate empty data
            try:
                u2 = content.xpath('//div[@id="ContentPlaceHolder1_teliti3_kiri1_updx"]/div[3][@class="panel panel-info"]/div[@class="panel-body f12"]/div[2][@class="row"]/div[@class="col-sm-4"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                u2 = ''
            # Use try-except catching to mitigate empty data
            try:
                v2 = content.xpath('//div[@id="ContentPlaceHolder1_teliti3_kiri1_updx"]/div[3][@class="panel panel-info"]/div[@class="panel-body f12"]/div[3][@class="row"]/div[@class="col-sm-4"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                v2 = ''
            # Use try-except catching to mitigate empty data
            try:
                w2 = content.xpath('//div[@id="ContentPlaceHolder1_teliti3_kiri1_updx"]/div[3][@class="panel panel-info"]/div[@class="panel-body f12"]/div[4][@class="row"]/div[@class="col-sm-4"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                w2 = ''
            # Use try-except catching to mitigate empty data
            try:
                x2 = content.xpath('//div[@id="ContentPlaceHolder1_teliti3_kiri1_updx"]/div[3][@class="panel panel-info"]/div[@class="panel-body f12"]/div[5][@class="row"]/div[@class="col-sm-4"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                x2 = ''

            # Painting the static (non-variable as in "Data Anggota") detail information
            sheet[f'A{row_start}'].value = a1
            sheet[f'B{row_start}'].value = b2
            sheet[f'C{row_start}'].value = c2
            sheet[f'D{row_start}'].value = d2
            sheet[f'E{row_start}'].value = e2
            sheet[f'F{row_start}'].value = f2
            sheet[f'G{row_start}'].value = g2
            sheet[f'H{row_start}'].value = h2
            sheet[f'I{row_start}'].value = i2
            sheet[f'J{row_start}'].value = j2
            sheet[f'K{row_start}'].value = k2
            sheet[f'L{row_start}'].value = l2
            sheet[f'M{row_start}'].value = m2
            sheet[f'N{row_start}'].value = n2
            sheet[f'O{row_start}'].value = o2
            sheet[f'P{row_start}'].value = p2
            sheet[f'Q{row_start}'].value = q2
            sheet[f'R{row_start}'].value = r2
            sheet[f'S{row_start}'].value = s2
            sheet[f'T{row_start}'].value = t2
            sheet[f'U{row_start}'].value = u2
            sheet[f'V{row_start}'].value = v2
            sheet[f'W{row_start}'].value = w2
            sheet[f'X{row_start}'].value = x2

            # The table which displays 'Identitas Pengusul - Anggota Peneliti'
            # Treated differently, specially
            # 'y_' is a DOM element representing a HTML table row (<tr>)
            y_ = content.xpath(
                '//div[@id="ContentPlaceHolder1_teliti3_kiri1_updx"]/div[4]//table[@class="table"]//tr[position()>1]')

            # The base path
            y_base = '//div[@id="ContentPlaceHolder1_teliti3_kiri1_updx"]/div[4]//table[@class="table"]//tr[position()>1]'

            # Painting the scraped data to the output spreadsheet row
            # Section: 'DATA ANGGOTA'
            # ---
            # Checking if 'Identitas Pengusul - Anggota Peneliti' table data exists
            if len(y_) == 0:
                pass  # --- nope. the data does not exist
            else:

                # Checking if this table's 'LUARAN' data has the most rows
                y_all_rows = len(y_)
                y_max_row = max(y_max_row, y_all_rows)

                # ROW DATA LABEL CONVENTION
                # y_row_a -> "N.I.P"
                # y_row_b -> "Nama"
                # y_row_c -> "Bidang Keahlian"
                # y_row_d -> "Alamat"
                # y_row_e -> "Instansi"
                # y_row_f -> "Email"
                # y_row_g -> "No. HP"
                # y_row_h -> "Peran"
                # y_row_i -> "Tugas"

                y_row_a = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[2]/text()')]

                y_row_b = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[3]/text()')]

                y_row_c = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[4]/text()')]

                y_row_d = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[5]/text()')]

                y_row_e = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[6]/text()')]

                y_row_f = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[7]/text()')]

                y_row_g = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[8]/text()')]

                y_row_h = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[9]/text()')]

                y_row_i = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[10]/text()')]

                # The starting column coordinate for filling the 'LUARAN' table data
                col_start = 25

                # Iterating through each table row and write to the spreadsheet
                # Assumes the lists y_row_a, y_row_b, ... have the same array size
                for j in range(len(y_row_a)):
                    # Painting table data
                    sheet.cell(row=row_start, column=col_start).value = y_row_a[j]
                    sheet.cell(row=row_start, column=col_start + 1).value = y_row_b[j]
                    sheet.cell(row=row_start, column=col_start + 2).value = y_row_c[j]
                    sheet.cell(row=row_start, column=col_start + 3).value = y_row_d[j]
                    sheet.cell(row=row_start, column=col_start + 4).value = y_row_e[j]
                    sheet.cell(row=row_start, column=col_start + 5).value = y_row_f[j]
                    sheet.cell(row=row_start, column=col_start + 6).value = y_row_g[j]
                    sheet.cell(row=row_start, column=col_start + 7).value = y_row_h[j]
                    sheet.cell(row=row_start, column=col_start + 8).value = y_row_i[j]

                    # Incrementing the 'col_start' iterator before continuing the loop
                    col_start += 9
                    continue

            # Incrementing the value of 'row_start' before continuing
            row_start += 1

            # Reopening the "Dana Pengabdian" list page,
            # then assign the AJAX response to the temporary array 'temporary_prompt'
            # The 'data' array is obtained from opening individual entry row detail page
            #
            # This is done only on the 7-multiple row iteration, because
            # a single set of 'viewstate', 'viewstategen', and 'eventvalidation' values of ASPX
            # can only be used to do at most 8 operations before having to be renewed.
            if ( i % 7 == 0 ) and ( i > 0 ):
                temporary_prompt = self.get_risat_pengabdian_berdirevisi_pengabdian(data)
            continue

        # Post-loop logging: appending the header over the 'DATA ANGGOTA' columns
        control.append_message_area(
            f'+ Melengkapi kepala tabel pada bagian "Identitas Pengusul - Anggota Peneliti" ...')
        # ---
        # The starting column for the 'DATA ANGGOTA' data
        col_start = 25
        # Beginning the loop that detects the maximum number of 'LUARAN' rows
        # according to the variable 'y_max_row'
        for i in range(1, y_max_row + 1):
            # Setting the top header
            sheet.merge_cells(
                start_row=1,
                start_column=col_start,
                end_row=1,
                end_column=col_start + 8
            )
            sheet.cell(row=1, column=col_start).value = f'IDENTITAS PENGUSUL — ANGGOTA #{i}'
            sheet.cell(row=1, column=col_start).alignment = Alignment(horizontal='center')
            # Setting the sub headers
            sheet.cell(row=2, column=col_start).value = 'N.I.P'
            sheet.cell(row=2, column=col_start + 1).value = 'Nama'
            sheet.cell(row=2, column=col_start + 2).value = 'Bidang Keahlian'
            sheet.cell(row=2, column=col_start + 3).value = 'Alamat'
            sheet.cell(row=2, column=col_start + 4).value = 'Instansi'
            sheet.cell(row=2, column=col_start + 5).value = 'Email'
            sheet.cell(row=2, column=col_start + 6).value = 'No. HP'
            sheet.cell(row=2, column=col_start + 7).value = 'Peran'
            sheet.cell(row=2, column=col_start + 8).value = 'Tugas'
            # Incrementing the 'col_start' iterator before continuing the loop
            col_start += 9
            continue

        # Post-loop logging: successfully painted the output spreadsheet file
        control.append_message_area(f'+ Tabel sukses dipanen!')
        control.set_progress_bar(85)

        # Asking for the spreadsheet name to save as
        # ---
        # Logging and setting the progress bar
        control.append_message_area(f'+ Menyimpan spreadsheet luaran ...')
        control.set_progress_bar(90)
        # Dealing with file name prompt and saving
        # Using loop to mitigate the user clicking 'cancel'
        # in the file name dialog prompt
        while True:
            # Opening the dialog prompt
            output_spreadsheet = filedialog.asksaveasfilename(
                filetypes=[('Excel files', '*.xlsx')],
                initialfile='Sipesat - Berkas Direvisi Pengabdian Detil Risat.xlsx',
                title='Simpan sebagai ...'
            )

            # 'cancel' button in the dialog prompt is clicked
            if len(output_spreadsheet) == 0:
                # Showing confirmation
                x = messagebox.askyesno(
                    'Nama File Kosong',
                    'Apakah Anda yakin ingin melanjutkan tanpa menyimpan file spreadsheet hasil pemanenan?'
                )
                # Determining whether to break or to continue the loop
                # based on the inversed value of 'x'
                if x:
                    control.append_message_area(
                        f'+ Finalisasi pemanenan data tanpa menyimpan file spreadsheet luaran ...')
                    workbook.close()  # --- closing the workbook without saving
                    break
                else:
                    continue  # --- continuing the loop
            # File name does not end in spreadsheet extension
            elif output_spreadsheet[-5:] != '.xlsx':
                output_spreadsheet = output_spreadsheet + '.xlsx'

            # Saving the spreadsheet
            control.append_message_area(f'LOKASI_SPREADSHEET_LUARAN: {output_spreadsheet}')
            control.set_progress_bar(95)
            workbook.save(output_spreadsheet)

            # Closing the openpyxl workbook
            control.append_message_area(f'+ Menutup file spreadsheet ...')
            control.set_progress_bar(98)
            workbook.close()

            # Breaking the loop
            break

        # Notify for a successful scraping
        control.append_message_area(f'+ Pemanenan selesai pada: {str(dt.now())}')
        control.set_progress_bar(100)
        control.on_notify_successful_scraping()

    # This function harvests "Risat Dana Pengabdian > Data Detil Lengkap" data
    # and then store the harvested data as an excel file
    #
    # Required arguments:
    # - control             --> for updating the progress bar and
    #                           message area of the screen SipesatScrHarvest
    # - username, password  --> the Risat administrator username and password
    def run_harvest_c_1_dana(self, control, username, password):

        # SipesatScrHarvest messenger
        control.set_header_desc('Panen Data "Risat Dana Pengabdian > Data Detil Lengkap"')
        control.set_help_label('Data sedang dipanen. Silahkan menunggu.')
        control.set_progress_bar(0)
        control.clear_message_area()

        # Preamble logging
        control.append_message_area(f'+ Memulai pemanenan data ...')
        control.append_message_area(f'+ Pemanenan dimulai pada: {str(dt.now())}')
        control.set_progress_bar(5)

        # Preparing the 'data_prompt' arrays
        control.append_message_area(f'+ Log masuk Risat sebagai [{username}] ...')
        control.set_progress_bar(10)
        data_prompt = self.get_risat_login(username, password)
        data_prompt = self.get_risat_pengabdian(data_prompt)
        data_prompt = self.get_risat_pengabdian_dana_pengabdian(data_prompt)

        # Parsing XML tree content
        control.append_message_area(f'+ Membaca halaman web ...')
        control.set_progress_bar(15)
        content = data_prompt['html_content']

        # Establishing the export spreadsheet file
        control.append_message_area(f'+ Mempersiapkan file spreadsheet luaran ...')
        control.set_progress_bar(20)
        workbook = xl.Workbook()
        sheet = workbook.active
        sheet.title = 'Dana Pengabdian Detil'

        # DEVELOPER'S NOTE (DO NOT REMOVE)
        '''
        Data yang perlu diambil:
        1. Judul
        2. Substansi usulan
        3. Identitas pengusul - ketua
        4. Identitas pengusul - anggota
        5. RAB
        6. Dokumen pendukung

        Urutan kolom excel:
        1. Judul
        2. Substansi usulan
        3. RAB
        4. Dokumen pendukung
        5. Identitas pengusul - ketua
        6. Identitas pengusul - anggota
        '''

        # Preparing the sheet header
        control.append_message_area(f'+ Mempersiapkan kepala lembar spreadsheet ...')
        control.set_progress_bar(25)
        # ---
        # Preparing the "NOMOR" header
        sheet.merge_cells(
            start_row=1,
            start_column=1,
            end_row=2,
            end_column=1
        )
        sheet['A1'].value = 'No.'
        sheet['A1'].alignment = Alignment(horizontal='center')
        # ---
        # Preparing the "JUDUL" header
        sheet.merge_cells(
            start_row=1,
            start_column=2,
            end_row=1,
            end_column=12
        )
        sheet['B1'].value = 'IDENTITAS'
        sheet['B1'].alignment = Alignment(horizontal='center')
        # Preparing the "JUDUL" sub-headers
        sheet['B2'].value = 'Judul'
        sheet['C2'].value = 'Tgl. Usulan'
        sheet['D2'].value = 'TKT Saat Ini'
        sheet['E2'].value = 'Level'
        sheet['F2'].value = 'Kategori'
        sheet['G2'].value = 'Skema'
        sheet['H2'].value = 'Rumpun Ilmu'
        sheet['I2'].value = 'Bidang Fokus'
        sheet['J2'].value = 'Tema'
        sheet['K2'].value = 'Topik'
        sheet['L2'].value = 'Lama Kegiatan'
        # ---
        # Preparing the "SUBSTANSI USULAN" header
        sheet.merge_cells(
            start_row=1,
            start_column=13,
            end_row=1,
            end_column=14
        )
        sheet['M1'].value = 'SUBSTANSI USULAN'
        sheet['M1'].alignment = Alignment(horizontal='center')
        # Preparing the "SUBSTANSI USULAN" sub-headers
        sheet['M2'].value = 'Kelompok Makro'
        sheet['N2'].value = 'File Proposal'
        # ---
        # Preparing the "RAB" header
        sheet.merge_cells(
            start_row=1,
            start_column=15,
            end_row=1,
            end_column=16
        )
        sheet['O1'].value = 'RENCANA ANGGARAN BIAYA'
        sheet['O1'].alignment = Alignment(horizontal='center')
        # Preparing the "RAB" sub-headers
        sheet['O2'].value = 'Biaya'
        sheet['P2'].value = 'File RAB'
        # ---
        # Preparing the "DOKUMEN PENDUKUNG" header
        sheet.merge_cells(
            start_row=1,
            start_column=17,
            end_row=1,
            end_column=19
        )
        sheet['Q1'].value = 'DOKUMEN PENDUKUNG'
        sheet['Q1'].alignment = Alignment(horizontal='center')
        # Preparing the "DOKUMEN PENDUKUNG" sub-headers
        sheet['Q2'].value = 'Mitra'
        sheet['R2'].value = 'Dukungan Biaya'
        sheet['S2'].value = 'Surat Dukungan Mitra'
        # ---
        # Preparing the "IDENTITAS PENGUSUL — KETUA" header
        sheet.merge_cells(
            start_row=1,
            start_column=20,
            end_row=1,
            end_column=24
        )
        sheet['T1'].value = 'IDENTITAS PENGUSUL — KETUA'
        sheet['T1'].alignment = Alignment(horizontal='center')
        # Preparing the "IDENTITAS PENGUSUL — KETUA" sub-headers
        sheet['T2'].value = 'N.I.P'
        sheet['U2'].value = 'N.I.K'
        sheet['V2'].value = 'N.I.D.N'
        sheet['W2'].value = 'Nama Lengkap'
        sheet['X2'].value = 'Jabatan Fungsional'

        # The base XPath location, pointing to each entry row
        base = '//div[@class="mw-100"]//div[@class="form-group f12"]/table[@width="100%"]//tr[@valign="top"]'

        # ---
        # Obtaining the data row values
        control.append_message_area(f'+ Mendapatkan data pada baris tabel ...')
        control.set_progress_bar(30)

        # Reading the HTML entry row hidden ASPX values
        # Copy-pasted from: /ssynthesia/ghostcity/ar/dumper-2/24__2023.02.13__requestsrisat.py
        all_kodetran_prop = content.xpath(base + '//input[1][@type="hidden"]/@name')
        all_kodetran_val = content.xpath(base + '//input[1][@type="hidden"]/@value')
        all_stat_prop = content.xpath(base + '//input[2][@type="hidden"]/@name')
        all_stat_val = content.xpath(base + '//input[2][@type="hidden"]/@value')
        all_submitbtn = content.xpath(base + '//input[@type="submit"]/@name')

        # The number of rows
        # Assumes the 'all_' array size equals the number of data rows
        number_of_row = len(all_kodetran_prop)

        # The starting row coordinate of the active sheet
        row_start = 3

        # The maximum number of 'DATA ANGGOTA' table data row
        y_max_row = 0

        # Iterating through each entry row element
        # Assumes all the 'all_' arrays in the previous code block
        # are of the same length/size
        # Copy-pasted from: /ssynthesia/ghostcity/ar/dumper-2/24__2023.02.13__requestsrisat.py
        control.append_message_area(f'+ Melakukan iterasi terhadap baris tabel dan menulis spreadsheet luaran ...')
        control.set_progress_bar(35)
        temporary_prompt = data_prompt
        for i in range(number_of_row):
            # Noisy preamble logging
            # Please don't use this -_-
            # ---
            # control.append_message_area(f'ITERASI [{i}]')

            # Updating the progress bar status
            control.set_progress_bar(35 + round(45 * (i + 1) / number_of_row))

            # Preparing the AJAX payload
            detail_prompt = {
                'viewstate': temporary_prompt['viewstate'],
                'viewstategen': temporary_prompt['viewstategen'],
                'eventvalidation': temporary_prompt['eventvalidation'],
                'button_name': all_submitbtn[i],
                'kodetran_prop': all_kodetran_prop[i],
                'kodetran_val': all_kodetran_val[i],
                'stat_prop': all_stat_prop[i],
                'stat_val': all_stat_val[i]
            }

            # Obtaining the response data of each individual entry row detail page
            data = self.get_risat_pengabdian_dana_pengabdian_detil(detail_prompt)
            content = data['html_content']
            response = data['http_response']

            # Terminal logging for detecting errors
            log_string_1 = str(i + 1)
            log_string_2 = content.xpath('//span[@id="ContentPlaceHolder1_danacair1_kiri1_txjudul1"]/text()')[
                0].replace('\r', '').replace('\n', '').strip()
            control.append_message_area(f'+ Memanen data detil: [{log_string_1}] {log_string_2} ...')
            print(f'+ Harvesting detailed data: [{log_string_1}] {log_string_2}')

            # DEVELOPER'S NOTE (DO NOT REMOVE)
            '''
            Variables and their associated data

            'a1' ==> 'No.'
            'b2' ==> 'Judul'
            'c2' ==> 'Tgl. Usulan'
            'd2' ==> 'TKT Saat Ini'
            'e2' ==> 'Level'
            'f2' ==> 'Kategori'
            'g2' ==> 'Skema'
            'h2' ==> 'Rumpun Ilmu'
            'i2' ==> 'Bidang Fokus'
            'j2' ==> 'Tema'
            'k2' ==> 'Topik'
            'l2' ==> 'Lama Kegiatan'
            'm2' ==> 'Kelompok Makro'
            'n2' ==> 'File Proposal'
            'o2' ==> 'Biaya'
            'p2' ==> 'File RAB'
            'q2' ==> 'Mitra'
            'r2' ==> 'Dukungan Biaya'
            's2' ==> 'Surat Dukungan Mitra'
            't2' ==> 'N.I.P'
            'u2' ==> 'N.I.K'
            'v2' ==> 'N.I.D.N'
            'w2' ==> 'Nama Lengkap'
            'x2' ==> 'Jabatan Fungsional'

            In addition, variables starting with 'y' prefix are pertaining
            to the table data of "Data Anggota"
            '''

            # Getting the static (non-variable as in "Data Anggota") detail information
            # ---
            a1 = str(i+1)
            # Use try-except catching to mitigate empty data
            try:
                b2 = content.xpath('//span[@id="ContentPlaceHolder1_danacair1_kiri1_txjudul1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                b2 = ''
            # Use try-except catching to mitigate empty data
            try:
                c2 = content.xpath('//span[@id="ContentPlaceHolder1_danacair1_kiri1_tglusul1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                c2 = ''
            # 'Pengabdian Masyarakat' ain't have TKT (d2) and level (e2) data!
            # ---
            d2 = '-'
            e2 = '-'
            # Use try-except catching to mitigate empty data
            try:
                f2 = content.xpath('//span[@id="ContentPlaceHolder1_danacair1_kiri1_lkat1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                f2 = ''
            # Use try-except catching to mitigate empty data
            try:
                g2 = content.xpath('//span[@id="ContentPlaceHolder1_danacair1_kiri1_ddlskema1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                g2 = ''
            # 'Rumpun Ilmu' scraped data contains multiple array of strings
            # This needs some extra tweaking
            h2 = ''
            # Use try-except catching to mitigate empty data
            try:
                h2_pre = content.xpath('//div[@id="ContentPlaceHolder1_danacair1_kiri1_updx"]/div[1][@class="panel panel-info"]/div[@class="panel-body f12"]/div[5][@class="row f12"]/div[@class="col-sm-4"]//text()')
                for l in h2_pre:
                    l = l.replace('\r','').replace('\n','').strip()
                    h2 += l +', '
                # Remove trailing ', ' characters
                h2 = h2[:-2]
            except IndexError:
                h2 = ''
            # ---
            # Use try-except catching to mitigate empty data
            try:
                i2 = content.xpath('//span[@id="ContentPlaceHolder1_danacair1_kiri1_ddlfokus1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                i2 = ''
            # Use try-except catching to mitigate empty data
            try:
                j2 = content.xpath('//span[@id="ContentPlaceHolder1_danacair1_kiri1_ddltema1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                j2 = ''
            # Some 'Topik' has no data element
            # Use try-except catching to mitigate
            try:
                k2 = content.xpath('//span[@id="ContentPlaceHolder1_danacair1_kiri1_ddltopik1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                k2 = ''
            # Use try-except catching to mitigate empty data
            try:
                l2 = content.xpath('//span[@id="ContentPlaceHolder1_danacair1_kiri1_ddllama1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                l2 = ''
            # Use try-except catching to mitigate empty data
            try:
                m2 = content.xpath('//span[@id="ContentPlaceHolder1_danacair1_kiri1_ddlmakro1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                m2 = ''
            # Use try-except catching to mitigate empty data
            try:
                n2 = content.xpath('//a[@id="ContentPlaceHolder1_danacair1_kiri1_alblcatatan1"]/@href')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                n2 = ''
            # Use try-except catching to mitigate empty data
            try:
                o2 = content.xpath('//span[@id="ContentPlaceHolder1_danacair1_kiri1_tbiayathn1"]/text()')[0].replace('\r','').replace('\n','').replace('Rp.', '').replace(',', '').strip()
            except IndexError:
                o2 = ''
            # Use try-except catching to mitigate empty data
            try:
                p2 = content.xpath('//a[@id="ContentPlaceHolder1_danacair1_kiri1_alblfilerab1"]/@href')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                p2 = ''
            # Some 'Data Detil' has no 'Mitra' data element
            # Use try-except catching to mitigate
            try:
                q2 = content.xpath('//span[@id="ContentPlaceHolder1_danacair1_kiri1_tnmmitra1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                q2 = ''
            # Use try-except catching to mitigate empty data
            try:
                r2 = content.xpath('//span[@id="ContentPlaceHolder1_danacair1_kiri1_tbiayadukung1"]/text()')[0].replace('\r','').replace('\n','').replace('Rp.', '').replace(',', '').strip()
            except IndexError:
                r2 = ''
            # Use try-except catching to mitigate empty data
            try:
                s2 = content.xpath('//a[@id="ContentPlaceHolder1_danacair1_kiri1_alblfilemitradukung1"]/@href')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                s2 = ''
            # Use try-except catching to mitigate empty data
            try:
                t2 = content.xpath('//div[@id="ContentPlaceHolder1_danacair1_kiri1_updx"]/div[3][@class="panel panel-info"]/div[@class="panel-body f12"]/div[1][@class="row"]/div[@class="col-sm-4"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                t2 = ''
            # Use try-except catching to mitigate empty data
            try:
                u2 = content.xpath('//div[@id="ContentPlaceHolder1_danacair1_kiri1_updx"]/div[3][@class="panel panel-info"]/div[@class="panel-body f12"]/div[2][@class="row"]/div[@class="col-sm-4"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                u2 = ''
            # Use try-except catching to mitigate empty data
            try:
                v2 = content.xpath('//div[@id="ContentPlaceHolder1_danacair1_kiri1_updx"]/div[3][@class="panel panel-info"]/div[@class="panel-body f12"]/div[3][@class="row"]/div[@class="col-sm-4"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                v2 = ''
            # Use try-except catching to mitigate empty data
            try:
                w2 = content.xpath('//div[@id="ContentPlaceHolder1_danacair1_kiri1_updx"]/div[3][@class="panel panel-info"]/div[@class="panel-body f12"]/div[4][@class="row"]/div[@class="col-sm-4"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                w2 = ''
            # Use try-except catching to mitigate empty data
            try:
                x2 = content.xpath('//div[@id="ContentPlaceHolder1_danacair1_kiri1_updx"]/div[3][@class="panel panel-info"]/div[@class="panel-body f12"]/div[5][@class="row"]/div[@class="col-sm-4"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                x2 = ''
            
            # Painting the static (non-variable as in "Data Anggota") detail information
            sheet[f'A{row_start}'].value = a1
            sheet[f'B{row_start}'].value = b2
            sheet[f'C{row_start}'].value = c2
            sheet[f'D{row_start}'].value = d2
            sheet[f'E{row_start}'].value = e2
            sheet[f'F{row_start}'].value = f2
            sheet[f'G{row_start}'].value = g2
            sheet[f'H{row_start}'].value = h2
            sheet[f'I{row_start}'].value = i2
            sheet[f'J{row_start}'].value = j2
            sheet[f'K{row_start}'].value = k2
            sheet[f'L{row_start}'].value = l2
            sheet[f'M{row_start}'].value = m2
            sheet[f'N{row_start}'].value = n2
            sheet[f'O{row_start}'].value = o2
            sheet[f'P{row_start}'].value = p2
            sheet[f'Q{row_start}'].value = q2
            sheet[f'R{row_start}'].value = r2
            sheet[f'S{row_start}'].value = s2
            sheet[f'T{row_start}'].value = t2
            sheet[f'U{row_start}'].value = u2
            sheet[f'V{row_start}'].value = v2
            sheet[f'W{row_start}'].value = w2
            sheet[f'X{row_start}'].value = x2

            # The table which displays 'Identitas Pengusul - Anggota Peneliti'
            # Treated differently, specially
            # 'y_' is a DOM element representing a HTML table row (<tr>)
            y_ = content.xpath(
                '//div[@id="ContentPlaceHolder1_danacair1_kiri1_updx"]/div[4]//table[@class="table"]//tr[position()>1]')

            # The base path
            y_base = '//div[@id="ContentPlaceHolder1_danacair1_kiri1_updx"]/div[4]//table[@class="table"]//tr[position()>1]'

            # Painting the scraped data to the output spreadsheet row
            # Section: 'DATA ANGGOTA'
            # ---
            # Checking if 'Identitas Pengusul - Anggota Peneliti' table data exists
            if len(y_) == 0:
                pass  # --- nope. the data does not exist
            else:

                # Checking if this table's 'LUARAN' data has the most rows
                y_all_rows = len(y_)
                y_max_row = max(y_max_row, y_all_rows)

                # ROW DATA LABEL CONVENTION
                # y_row_a -> "N.I.P"
                # y_row_b -> "Nama"
                # y_row_c -> "Bidang Keahlian"
                # y_row_d -> "Alamat"
                # y_row_e -> "Instansi"
                # y_row_f -> "Email"
                # y_row_g -> "No. HP"
                # y_row_h -> "Peran"
                # y_row_i -> "Tugas"

                y_row_a = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[2]/text()')]

                y_row_b = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[3]/text()')]

                y_row_c = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[4]/text()')]

                y_row_d = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[5]/text()')]

                y_row_e = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[6]/text()')]

                y_row_f = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[7]/text()')]

                y_row_g = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[8]/text()')]

                y_row_h = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[9]/text()')]

                y_row_i = [l.replace('\r', '').replace('\n', '').strip()
                           for l in content.xpath(y_base + '/td[10]/text()')]

                # The starting column coordinate for filling the 'LUARAN' table data
                col_start = 25

                # Iterating through each table row and write to the spreadsheet
                # Assumes the lists y_row_a, y_row_b, ... have the same array size
                for j in range(len(y_row_a)):
                    # Painting table data
                    sheet.cell(row=row_start, column=col_start).value = y_row_a[j]
                    sheet.cell(row=row_start, column=col_start + 1).value = y_row_b[j]
                    sheet.cell(row=row_start, column=col_start + 2).value = y_row_c[j]
                    sheet.cell(row=row_start, column=col_start + 3).value = y_row_d[j]
                    sheet.cell(row=row_start, column=col_start + 4).value = y_row_e[j]
                    sheet.cell(row=row_start, column=col_start + 5).value = y_row_f[j]
                    sheet.cell(row=row_start, column=col_start + 6).value = y_row_g[j]
                    sheet.cell(row=row_start, column=col_start + 7).value = y_row_h[j]
                    sheet.cell(row=row_start, column=col_start + 8).value = y_row_i[j]

                    # Incrementing the 'col_start' iterator before continuing the loop
                    col_start += 9
                    continue

            # Incrementing the value of 'row_start' before continuing
            row_start += 1

            # Reopening the "Dana Pengabdian" list page,
            # then assign the AJAX response to the temporary array 'temporary_prompt'
            # The 'data' array is obtained from opening individual entry row detail page
            #
            # This is done only on the 7-multiple row iteration, because
            # a single set of 'viewstate', 'viewstategen', and 'eventvalidation' values of ASPX
            # can only be used to do at most 8 operations before having to be renewed.
            if ( i % 7 == 0 ) and ( i > 0 ):
                temporary_prompt = self.get_risat_pengabdian_dana_pengabdian(data)
            continue

        # Post-loop logging: appending the header over the 'DATA ANGGOTA' columns
        control.append_message_area(
            f'+ Melengkapi kepala tabel pada bagian "Identitas Pengusul - Anggota Peneliti" ...')
        # ---
        # The starting column for the 'DATA ANGGOTA' data
        col_start = 25
        # Beginning the loop that detects the maximum number of 'LUARAN' rows
        # according to the variable 'y_max_row'
        for i in range(1, y_max_row + 1):
            # Setting the top header
            sheet.merge_cells(
                start_row=1,
                start_column=col_start,
                end_row=1,
                end_column=col_start + 8
            )
            sheet.cell(row=1, column=col_start).value = f'IDENTITAS PENGUSUL — ANGGOTA #{i}'
            sheet.cell(row=1, column=col_start).alignment = Alignment(horizontal='center')
            # Setting the sub headers
            sheet.cell(row=2, column=col_start).value = 'N.I.P'
            sheet.cell(row=2, column=col_start + 1).value = 'Nama'
            sheet.cell(row=2, column=col_start + 2).value = 'Bidang Keahlian'
            sheet.cell(row=2, column=col_start + 3).value = 'Alamat'
            sheet.cell(row=2, column=col_start + 4).value = 'Instansi'
            sheet.cell(row=2, column=col_start + 5).value = 'Email'
            sheet.cell(row=2, column=col_start + 6).value = 'No. HP'
            sheet.cell(row=2, column=col_start + 7).value = 'Peran'
            sheet.cell(row=2, column=col_start + 8).value = 'Tugas'
            # Incrementing the 'col_start' iterator before continuing the loop
            col_start += 9
            continue

        # Post-loop logging: successfully painted the output spreadsheet file
        control.append_message_area(f'+ Tabel sukses dipanen!')
        control.set_progress_bar(85)

        # Asking for the spreadsheet name to save as
        # ---
        # Logging and setting the progress bar
        control.append_message_area(f'+ Menyimpan spreadsheet luaran ...')
        control.set_progress_bar(90)
        # Dealing with file name prompt and saving
        # Using loop to mitigate the user clicking 'cancel'
        # in the file name dialog prompt
        while True:
            # Opening the dialog prompt
            output_spreadsheet = filedialog.asksaveasfilename(
                filetypes=[('Excel files', '*.xlsx')],
                initialfile='Sipesat - Dana Pengabdian Detil Risat.xlsx',
                title='Simpan sebagai ...'
            )

            # 'cancel' button in the dialog prompt is clicked
            if len(output_spreadsheet) == 0:
                # Showing confirmation
                x = messagebox.askyesno(
                    'Nama File Kosong',
                    'Apakah Anda yakin ingin melanjutkan tanpa menyimpan file spreadsheet hasil pemanenan?'
                )
                # Determining whether to break or to continue the loop
                # based on the inversed value of 'x'
                if x:
                    control.append_message_area(
                        f'+ Finalisasi pemanenan data tanpa menyimpan file spreadsheet luaran ...')
                    workbook.close()  # --- closing the workbook without saving
                    break
                else:
                    continue  # --- continuing the loop
            # File name does not end in spreadsheet extension
            elif output_spreadsheet[-5:] != '.xlsx':
                output_spreadsheet = output_spreadsheet + '.xlsx'

            # Saving the spreadsheet
            control.append_message_area(f'LOKASI_SPREADSHEET_LUARAN: {output_spreadsheet}')
            control.set_progress_bar(95)
            workbook.save(output_spreadsheet)

            # Closing the openpyxl workbook
            control.append_message_area(f'+ Menutup file spreadsheet ...')
            control.set_progress_bar(98)
            workbook.close()

            # Breaking the loop
            break

        # Notify for a successful scraping
        control.append_message_area(f'+ Pemanenan selesai pada: {str(dt.now())}')
        control.set_progress_bar(100)
        control.on_notify_successful_scraping()

    # This function harvests "Risat Arsip Pengabdian > Data Detil Lengkap" data
    # and then store the harvested data as an excel file
    #
    # Required arguments:
    # - control             --> for updating the progress bar and
    #                           message area of the screen SipesatScrHarvest
    # - username, password  --> the Risat administrator username and password
    def run_harvest_c_1_arsip(self, control, username, password):

        # Dodge Python function documentation
        '''
        '''

        # Give a warning about bugs in 'Arsip' data harvester
        # This bug is due to Risat's server lag, not because of this program by itself
        # (Just in case. As per 2023-06-12, turns out the 'Arsip' page of Risat
        # does not always load in 30 seconds -- depending on circumstances)
        '''
        messagebox.showerror(
            title='Pesan Pengembang',
            message='' +
                    'Dikarenakan kendala teknis pada server risat.uksw.edu, ' +
                    'data arsip Risat membutuhkan waktu sekitar 30 detik untuk dapat termuat. ' +
                    'Setiap pemuatan dari server, sebanyak 7 data dapat dipanen sekaligus. ' +
                    'Sehingga, secara keseluruhan dibutuhkan waktu sekitar 4 menit (dengan asumsi jumlah data sebanyak 50) ' +
                    'untuk dapat memanen semua data arsip Risat. ' +
                    'Antarmuka program SiPe.Sat akan sesekali membeku dan tombol-tombolnya hilang ' +
                    'selama pemanenan berlangsung karena menunggu pemuatan data arsip dari server risat.uksw.edu. ' +
                    '\n\n' +
                    'Pengembang SiPe.Sat meminta maaf atas ketidaknyamanan ini.'
        )
        '''

        # SipesatScrHarvest messenger
        control.set_header_desc('Panen Data "Risat Arsip Pengabdian > Data Detil Lengkap"')
        control.set_help_label('Data sedang dipanen. Silahkan menunggu.')
        control.set_progress_bar(0)
        control.clear_message_area()

        # Preamble logging
        control.append_message_area(f'+ Memulai pemanenan data ...')
        control.append_message_area(f'+ Pemanenan dimulai pada: {str(dt.now())}')
        control.set_progress_bar(5)

        # Preparing the 'data_prompt' arrays
        control.append_message_area(f'+ Log masuk Risat sebagai [{username}] ...')
        control.set_progress_bar(10)
        data_prompt = self.get_risat_login(username, password)
        data_prompt = self.get_risat_pengabdian(data_prompt)
        data_prompt = self.get_risat_pengabdian_arsip(data_prompt)

        # Parsing XML tree content
        control.append_message_area(f'+ Membaca halaman web ...')
        control.set_progress_bar(15)
        content = data_prompt['html_content']

        # Establishing the export spreadsheet file
        control.append_message_area(f'+ Mempersiapkan file spreadsheet luaran ...')
        control.set_progress_bar(20)
        workbook = xl.Workbook()
        sheet = workbook.active
        sheet.title = 'Arsip Pengabdian Detil'

        # DEVELOPER'S NOTE (DO NOT REMOVE)
        '''
        Data yang perlu diambil:
        1. Judul
        2. Substansi usulan
        3. Identitas pengusul - ketua
        4. Identitas pengusul - anggota
        5. RAB
        6. Dokumen pendukung
        
        Urutan kolom excel:
        1. Judul
        2. Substansi usulan
        3. RAB
        4. Dokumen pendukung
        5. Identitas pengusul - ketua
        6. Identitas pengusul - anggota
        '''

        # Preparing the sheet header
        control.append_message_area(f'+ Mempersiapkan kepala lembar spreadsheet ...')
        control.set_progress_bar(25)
        # ---
        # Preparing the "NOMOR" header
        sheet.merge_cells(
            start_row=1,
            start_column=1,
            end_row=2,
            end_column=1
        )
        sheet['A1'].value = 'No.'
        sheet['A1'].alignment = Alignment(horizontal='center')
        # ---
        # Preparing the "JUDUL" header
        sheet.merge_cells(
            start_row=1,
            start_column=2,
            end_row=1,
            end_column=12
        )
        sheet['B1'].value = 'IDENTITAS'
        sheet['B1'].alignment = Alignment(horizontal='center')
        # Preparing the "JUDUL" sub-headers
        sheet['B2'].value = 'Judul'
        sheet['C2'].value = 'Tgl. Usulan'
        sheet['D2'].value = 'TKT Saat Ini'
        sheet['E2'].value = 'Level'
        sheet['F2'].value = 'Kategori'
        sheet['G2'].value = 'Skema'
        sheet['H2'].value = 'Rumpun Ilmu'
        sheet['I2'].value = 'Bidang Fokus'
        sheet['J2'].value = 'Tema'
        sheet['K2'].value = 'Topik'
        sheet['L2'].value = 'Lama Kegiatan'
        # ---
        # Preparing the "SUBSTANSI USULAN" header
        sheet.merge_cells(
            start_row=1,
            start_column=13,
            end_row=1,
            end_column=14
        )
        sheet['M1'].value = 'SUBSTANSI USULAN'
        sheet['M1'].alignment = Alignment(horizontal='center')
        # Preparing the "SUBSTANSI USULAN" sub-headers
        sheet['M2'].value = 'Kelompok Makro'
        sheet['N2'].value = 'File Proposal'
        # ---
        # Preparing the "RAB" header
        sheet.merge_cells(
            start_row=1,
            start_column=15,
            end_row=1,
            end_column=16
        )
        sheet['O1'].value = 'RENCANA ANGGARAN BIAYA'
        sheet['O1'].alignment = Alignment(horizontal='center')
        # Preparing the "RAB" sub-headers
        sheet['O2'].value = 'Biaya'
        sheet['P2'].value = 'File RAB'
        # ---
        # Preparing the "DOKUMEN PENDUKUNG" header
        sheet.merge_cells(
            start_row=1,
            start_column=17,
            end_row=1,
            end_column=19
        )
        sheet['Q1'].value = 'DOKUMEN PENDUKUNG'
        sheet['Q1'].alignment = Alignment(horizontal='center')
        # Preparing the "DOKUMEN PENDUKUNG" sub-headers
        sheet['Q2'].value = 'Mitra'
        sheet['R2'].value = 'Dukungan Biaya'
        sheet['S2'].value = 'Surat Dukungan Mitra'
        # ---
        # Preparing the "IDENTITAS PENGUSUL — KETUA" header
        sheet.merge_cells(
            start_row=1,
            start_column=20,
            end_row=1,
            end_column=24
        )
        sheet['T1'].value = 'IDENTITAS PENGUSUL — KETUA'
        sheet['T1'].alignment = Alignment(horizontal='center')
        # Preparing the "IDENTITAS PENGUSUL — KETUA" sub-headers
        sheet['T2'].value = 'N.I.P'
        sheet['U2'].value = 'N.I.K'
        sheet['V2'].value = 'N.I.D.N'
        sheet['W2'].value = 'Nama Lengkap'
        sheet['X2'].value = 'Jabatan Fungsional'

        # The base XPath location, pointing to each entry row
        base = '//div[@id="ContentPlaceHolder1_arsip1_updg1"]//div[@class="panel-body f12"]/table[@width="100%"]'

        # ---
        # Obtaining the data row values
        control.append_message_area(f'+ Mendapatkan data pada baris tabel ...')
        control.set_progress_bar(30)

        # Reading the HTML entry row hidden ASPX values
        # Copy-pasted from: /ssynthesia/ghostcity/ar/dumper-2/24__2023.02.13__requestsrisat.py
        all_idat_prop = content.xpath(base + '//input[1][@type="hidden"]/@name')
        all_idat_val = content.xpath(base + '//input[1][@type="hidden"]/@value')
        all_itgl_prop = content.xpath(base + '//input[2][@type="hidden"]/@name')
        all_itgl_val = content.xpath(base + '//input[2][@type="hidden"]/@value')
        all_submitbtn = content.xpath(base + '//input[@type="submit"]/@name')

        # The number of rows
        # Assumes the 'all_' array size equals the number of data rows
        number_of_row = len(all_idat_prop)

        # DEBUG
        # Please comment out after use
        # ---
        # print(f'NUMBER_OF_ROW_LEN: {number_of_row}')
        
        # The starting row coordinate of the active sheet
        row_start = 3
        
        # The maximum number of 'DATA ANGGOTA' table data row
        y_max_row = 0

        # Iterating through each entry row element
        # Assumes all the 'all_' arrays in the previous code block
        # are of the same length/size
        # Copy-pasted from: /ssynthesia/ghostcity/ar/dumper-2/24__2023.02.13__requestsrisat.py
        control.append_message_area(f'+ Melakukan iterasi terhadap baris tabel dan menulis spreadsheet luaran ...')
        control.set_progress_bar(35)
        temporary_prompt = data_prompt
        for i in range(number_of_row):
            # Noisy preamble logging
            # Please don't use this -_-
            # ---
            # control.append_message_area(f'ITERASI [{i}]')

            # Updating the progress bar status
            control.set_progress_bar(35 + round(45 * (i + 1) / number_of_row))

            # Preparing the AJAX payload
            detail_prompt = {
                'viewstate' : temporary_prompt['viewstate'],
                'viewstategen' : temporary_prompt['viewstategen'],
                'eventvalidation' : temporary_prompt['eventvalidation'],
                'button_name' : all_submitbtn[i],
                'idat_prop' : all_idat_prop[i],
                'idat_val' : all_idat_val[i],
                'itgl_prop' : all_itgl_prop[i],
                'itgl_val' : all_itgl_val[i]
            }

            # Obtaining the response data of each individual entry row detail page
            data = self.get_risat_pengabdian_arsip_detil(detail_prompt)
            content = data['html_content']
            response = data['http_response']
            
            # Terminal logging for detecting errors
            log_string_1 = str(i+1)
            log_string_2 = content.xpath('//span[@id="ContentPlaceHolder1_arsip1_kiri1_txjudul1"]/text()')[0].replace('\r','').replace('\n','').strip()
            control.append_message_area(f'+ Memanen data detil: [{log_string_1}] {log_string_2} ...')
            print(f'+ Harvesting detailed data: [{log_string_1}] {log_string_2}')
            
            # DEVELOPER'S NOTE (DO NOT REMOVE)
            '''
            Variables and their associated data
            
            'a1' ==> 'No.'
            'b2' ==> 'Judul'
            'c2' ==> 'Tgl. Usulan'
            'd2' ==> 'TKT Saat Ini'
            'e2' ==> 'Level'
            'f2' ==> 'Kategori'
            'g2' ==> 'Skema'
            'h2' ==> 'Rumpun Ilmu'
            'i2' ==> 'Bidang Fokus'
            'j2' ==> 'Tema'
            'k2' ==> 'Topik'
            'l2' ==> 'Lama Kegiatan'
            'm2' ==> 'Kelompok Makro'
            'n2' ==> 'File Proposal'
            'o2' ==> 'Biaya'
            'p2' ==> 'File RAB'
            'q2' ==> 'Mitra'
            'r2' ==> 'Dukungan Biaya'
            's2' ==> 'Surat Dukungan Mitra'
            't2' ==> 'N.I.P'
            'u2' ==> 'N.I.K'
            'v2' ==> 'N.I.D.N'
            'w2' ==> 'Nama Lengkap'
            'x2' ==> 'Jabatan Fungsional'
            
            In addition, variables starting with 'y' prefix are pertaining
            to the table data of "Data Anggota"
            '''

            # Getting the static (non-variable as in "Data Anggota") detail information
            # ---
            a1 = str(i+1)
            # Use try-except catching to mitigate empty data
            try:
                b2 = content.xpath('//span[@id="ContentPlaceHolder1_arsip1_kiri1_txjudul1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                b2 = ''
            # Use try-except catching to mitigate empty data
            try:
                c2 = content.xpath('//span[@id="ContentPlaceHolder1_arsip1_kiri1_tglusul1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                c2 = ''
            d2 = '-'
            e2 = '-'
            # Use try-except catching to mitigate empty data
            try:
                f2 = content.xpath('//span[@id="ContentPlaceHolder1_arsip1_kiri1_lkat1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                f2 = ''
            # Use try-except catching to mitigate empty data
            try:
                g2 = content.xpath('//span[@id="ContentPlaceHolder1_arsip1_kiri1_ddlskema1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                g2 = ''
            # 'Rumpun Ilmu' scraped data contains multiple array of strings
            # This needs some extra tweaking
            h2 = ''
            # Use try-except catching to mitigate empty data
            try:
                h2_pre = content.xpath('//div[@id="ContentPlaceHolder1_arsip1_kiri1_updx"]/div[1][@class="panel panel-info"]/div[@class="panel-body f12"]/div[7][@class="row f12"]/div[@class="col-sm-4"]//text()')
                for l in h2_pre:
                    l = l.replace('\r','').replace('\n','').strip()
                    h2 += l +', '
                    # Remove trailing ', ' characters
                    h2 = h2[:-2]
            except IndexError:
                h2 = ''
            # ---
            # Use try-except catching to mitigate empty data
            try:
                i2 = content.xpath('//span[@id="ContentPlaceHolder1_arsip1_kiri1_ddlfokus1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                i2 = ''
            # Use try-except catching to mitigate empty data
            try:
                j2 = content.xpath('//span[@id="ContentPlaceHolder1_arsip1_kiri1_ddltema1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                j2 = ''
            # Use try-except catching to mitigate data
            try:
                k2 = content.xpath('//span[@id="ContentPlaceHolder1_arsip1_kiri1_ddltopik1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                k2 = ''
            # Use try-except catching to mitigate empty data
            try:
                l2 = content.xpath('//span[@id="ContentPlaceHolder1_arsip1_kiri1_ddllama1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                l2 = ''
            # Use try-except catching to mitigate empty data
            try:
                m2 = content.xpath('//span[@id="ContentPlaceHolder1_arsip1_kiri1_ddlmakro1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                m2 = ''
            # Use try-except catching to mitigate empty data
            try:
                n2 = content.xpath('//a[@id="ContentPlaceHolder1_arsip1_kiri1_alblcatatan1"]/@href')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                n2 = ''
            # Use try-except catching to mitigate empty data
            try:
                o2 = content.xpath('//span[@id="ContentPlaceHolder1_arsip1_kiri1_tbiayathn1"]/text()')[0].replace('\r','').replace('\n','').replace('Rp.', '').replace(',', '').strip()
            except IndexError:
                o2 = ''
            # Use try-except catching to mitigate empty data
            try:
                p2 = content.xpath('//a[@id="ContentPlaceHolder1_arsip1_kiri1_alblfilerab1"]/@href')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                p2 = ''
            # Some 'Data Detil' has no 'Mitra' data element
            # Use try-except catching to mitigate
            try:
                q2 = content.xpath('//span[@id="ContentPlaceHolder1_arsip1_kiri1_tnmmitra1"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                q2 = ''
            # Use try-except catching to mitigate empty data
            try:
                r2 = content.xpath('//span[@id="ContentPlaceHolder1_arsip1_kiri1_tbiayadukung1"]/text()')[0].replace('\r','').replace('\n','').replace('Rp.', '').replace(',', '').strip()
            except IndexError:
                r2 = ''
            # Use try-except catching to mitigate empty data
            try:
                s2 = content.xpath('//a[@id="ContentPlaceHolder1_arsip1_kiri1_alblfilemitradukung1"]/@href')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                s2 = ''
            # Use try-except catching to mitigate empty data
            try:
                t2 = content.xpath('//div[@id="ContentPlaceHolder1_arsip1_kiri1_updx"]/div[3][@class="panel panel-info"]/div[@class="panel-body f12"]/div[1][@class="row"]/div[@class="col-sm-4"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                t2 = ''
            # Use try-except catching to mitigate empty data
            try:
                u2 = content.xpath('//div[@id="ContentPlaceHolder1_arsip1_kiri1_updx"]/div[3][@class="panel panel-info"]/div[@class="panel-body f12"]/div[2][@class="row"]/div[@class="col-sm-4"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                u2 = ''
            # Use try-except catching to mitigate empty data
            try:
                v2 = content.xpath('//div[@id="ContentPlaceHolder1_arsip1_kiri1_updx"]/div[3][@class="panel panel-info"]/div[@class="panel-body f12"]/div[3][@class="row"]/div[@class="col-sm-4"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                v2 = ''
            # Use try-except catching to mitigate empty data
            try:
                w2 = content.xpath('//div[@id="ContentPlaceHolder1_arsip1_kiri1_updx"]/div[3][@class="panel panel-info"]/div[@class="panel-body f12"]/div[4][@class="row"]/div[@class="col-sm-4"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                w2 = ''
            # Use try-except catching to mitigate empty data
            try:
                x2 = content.xpath('//div[@id="ContentPlaceHolder1_arsip1_kiri1_updx"]/div[3][@class="panel panel-info"]/div[@class="panel-body f12"]/div[5][@class="row"]/div[@class="col-sm-4"]/text()')[0].replace('\r','').replace('\n','').strip()
            except IndexError:
                x2 = ''

            # Painting the static (non-variable as in "Data Anggota") detail information
            sheet[f'A{row_start}'].value = a1
            sheet[f'B{row_start}'].value = b2
            sheet[f'C{row_start}'].value = c2
            sheet[f'D{row_start}'].value = d2
            sheet[f'E{row_start}'].value = e2
            sheet[f'F{row_start}'].value = f2
            sheet[f'G{row_start}'].value = g2
            sheet[f'H{row_start}'].value = h2
            sheet[f'I{row_start}'].value = i2
            sheet[f'J{row_start}'].value = j2
            sheet[f'K{row_start}'].value = k2
            sheet[f'L{row_start}'].value = l2
            sheet[f'M{row_start}'].value = m2
            sheet[f'N{row_start}'].value = n2
            sheet[f'O{row_start}'].value = o2
            sheet[f'P{row_start}'].value = p2
            sheet[f'Q{row_start}'].value = q2
            sheet[f'R{row_start}'].value = r2
            sheet[f'S{row_start}'].value = s2
            sheet[f'T{row_start}'].value = t2
            sheet[f'U{row_start}'].value = u2
            sheet[f'V{row_start}'].value = v2
            sheet[f'W{row_start}'].value = w2
            sheet[f'X{row_start}'].value = x2

            # DEBUG
            # Please comment out after use
            # ---
            '''
            print()
            print('-------------------------------------------------------------------------------------------------------')
            print(f' > DEBUG LOGGING FOR DATA ITERATION NO. {i}: NON-ANGGOTA DATA')
            print('-------------------------------------------------------------------------------------------------------')
            print(f'A1 = {a1}')
            print(f'B2 = {b2}')
            print(f'C2 = {c2}')
            print(f'D2 = {d2}')
            print(f'E2 = {e2}')
            print(f'F2 = {f2}')
            print(f'G2 = {g2}')
            print(f'H2 = {h2}')
            print(f'I2 = {i2}')
            print(f'J2 = {j2}')
            print(f'K2 = {k2}')
            print(f'L2 = {l2}')
            print(f'M2 = {m2}')
            print(f'N2 = {n2}')
            print(f'O2 = {o2}')
            print(f'P2 = {p2}')
            print(f'Q2 = {q2}')
            print(f'R2 = {r2}')
            print(f'S2 = {s2}')
            print(f'T2 = {t2}')
            print(f'U2 = {u2}')
            print(f'V2 = {v2}')
            print(f'W2 = {w2}')
            print(f'X2 = {x2}')
            print('-------------------------------------------------------------------------------------------------------', end='')
            print()
            '''
            
            # The table which displays 'Identitas Pengusul - Anggota Peneliti'
            # Treated differently, specially
            # 'y_' is a DOM element representing a HTML table row (<tr>)
            y_ = content.xpath('//div[@id="ContentPlaceHolder1_arsip1_kiri1_updx"]/div[4]//table[@class="table"]//tr[position()>1]')
            
            # The base path
            y_base = '//div[@id="ContentPlaceHolder1_arsip1_kiri1_updx"]/div[4]//table[@class="table"]//tr[position()>1]'
            
            # Painting the scraped data to the output spreadsheet row
            # Section: 'DATA ANGGOTA'
            # ---
            # Checking if 'Identitas Pengusul - Anggota Peneliti' table data exists
            if len(y_) == 0:
                pass  # --- nope. the data does not exist
            else:
                
                # Checking if this table's 'LUARAN' data has the most rows
                y_all_rows = len(y_)
                y_max_row = max(y_max_row, y_all_rows)
                
                # ROW DATA LABEL CONVENTION
                # y_row_a -> "N.I.P"
                # y_row_b -> "Nama"
                # y_row_c -> "Bidang Keahlian"
                # y_row_d -> "Alamat"
                # y_row_e -> "Instansi"
                # y_row_f -> "Email"
                # y_row_g -> "No. HP"
                # y_row_h -> "Peran"
                # y_row_i -> "Tugas"
                
                y_row_a = [l.replace('\r','').replace('\n','').strip()
                      for l in content.xpath(y_base + '/td[2]/text()')]

                y_row_b = [l.replace('\r','').replace('\n','').strip()
                      for l in content.xpath(y_base + '/td[3]/text()')]

                y_row_c = [l.replace('\r','').replace('\n','').strip()
                      for l in content.xpath(y_base + '/td[4]/text()')]

                y_row_d = [l.replace('\r','').replace('\n','').strip()
                      for l in content.xpath(y_base + '/td[5]/text()')]

                y_row_e = [l.replace('\r','').replace('\n','').strip()
                      for l in content.xpath(y_base + '/td[6]/text()')]

                y_row_f = [l.replace('\r','').replace('\n','').strip()
                      for l in content.xpath(y_base + '/td[7]/text()')]

                y_row_g = [l.replace('\r','').replace('\n','').strip()
                      for l in content.xpath(y_base + '/td[8]/text()')]

                y_row_h = [l.replace('\r','').replace('\n','').strip()
                      for l in content.xpath(y_base + '/td[9]/text()')]

                y_row_i = [l.replace('\r','').replace('\n','').strip()
                      for l in content.xpath(y_base + '/td[10]/text()')]

                # The starting column coordinate for filling the 'LUARAN' table data
                col_start = 25
                
                # Iterating through each table row and write to the spreadsheet
                # Assumes the lists y_row_a, y_row_b, ... have the same array size
                for j in range(len(y_row_a)):

                    # Painting table data
                    sheet.cell(row=row_start, column=col_start).value = y_row_a[j]
                    sheet.cell(row=row_start, column=col_start+1).value = y_row_b[j]
                    sheet.cell(row=row_start, column=col_start+2).value = y_row_c[j]
                    sheet.cell(row=row_start, column=col_start+3).value = y_row_d[j]
                    sheet.cell(row=row_start, column=col_start+4).value = y_row_e[j]
                    sheet.cell(row=row_start, column=col_start+5).value = y_row_f[j]
                    sheet.cell(row=row_start, column=col_start+6).value = y_row_g[j]
                    sheet.cell(row=row_start, column=col_start+7).value = y_row_h[j]
                    sheet.cell(row=row_start, column=col_start+8).value = y_row_i[j]

                    # DEBUG
                    # Please comment out after use
                    # ---
                    '''
                    print()
                    print('-------------------------------------------------------------------------------------------------------')
                    print(f' > DEBUG LOGGING FOR DATA ITERATION NO. {i}: ANGGOTA DATA {j}')
                    print('-------------------------------------------------------------------------------------------------------')
                    print(f'y_row_a[{j}] = {y_row_a[j]}')
                    print(f'y_row_b[{j}] = {y_row_b[j]}')
                    print(f'y_row_c[{j}] = {y_row_c[j]}')
                    print(f'y_row_d[{j}] = {y_row_d[j]}')
                    print(f'y_row_e[{j}] = {y_row_e[j]}')
                    print(f'y_row_f[{j}] = {y_row_f[j]}')
                    print(f'y_row_g[{j}] = {y_row_g[j]}')
                    print(f'y_row_h[{j}] = {y_row_h[j]}')
                    print(f'y_row_i[{j}] = {y_row_i[j]}')
                    print('-------------------------------------------------------------------------------------------------------', end='')
                    print()
                    '''

                    # Incrementing the 'col_start' iterator before continuing the loop
                    col_start += 9
                    continue
                
            # Incrementing the value of 'row_start' before continuing
            row_start += 1

            # Reopening the "Arsip Pengabdian" list page,
            # then assign the AJAX response to the temporary array 'temporary_prompt'
            # The 'data' array is obtained from opening individual entry row detail page
            #
            # This is done only on the 7-multiple row iteration, because
            # a single set of 'viewstate', 'viewstategen', and 'eventvalidation' values of ASPX
            # can only be used to do at most 8 operations before having to be renewed.
            if ( i % 7 == 0 ) and ( i > 0 ):
                temporary_prompt = self.get_risat_pengabdian_arsip(data)
            continue
        
        # Post-loop logging: appending the header over the 'DATA ANGGOTA' columns
        control.append_message_area(f'+ Melengkapi kepala tabel pada bagian "Identitas Pengusul - Anggota Pengabdian" ...')
        # ---
        # The starting column for the 'DATA ANGGOTA' data
        col_start = 25
        # Beginning the loop that detects the maximum number of 'LUARAN' rows
        # according to the variable 'y_max_row'
        for i in range(1, y_max_row+1):
            # Setting the top header
            sheet.merge_cells(
                start_row=1,
                start_column=col_start,
                end_row=1,
                end_column=col_start+8
            )
            sheet.cell(row=1, column=col_start).value = f'IDENTITAS PENGUSUL — ANGGOTA #{i}'
            sheet.cell(row=1, column=col_start).alignment = Alignment(horizontal='center')
            # Setting the sub headers
            sheet.cell(row=2, column=col_start).value = 'N.I.P'
            sheet.cell(row=2, column=col_start+1).value = 'Nama'
            sheet.cell(row=2, column=col_start+2).value = 'Bidang Keahlian'
            sheet.cell(row=2, column=col_start+3).value = 'Alamat'
            sheet.cell(row=2, column=col_start+4).value = 'Instansi'
            sheet.cell(row=2, column=col_start+5).value = 'Email'
            sheet.cell(row=2, column=col_start+6).value = 'No. HP'
            sheet.cell(row=2, column=col_start+7).value = 'Peran'
            sheet.cell(row=2, column=col_start+8).value = 'Tugas'
            # Incrementing the 'col_start' iterator before continuing the loop
            col_start += 9
            continue

        # Post-loop logging: successfully painted the output spreadsheet file
        control.append_message_area(f'+ Tabel sukses dipanen!')
        control.set_progress_bar(85)

        # Asking for the spreadsheet name to save as
        # ---
        # Logging and setting the progress bar
        control.append_message_area(f'+ Menyimpan spreadsheet luaran ...')
        control.set_progress_bar(90)
        # Dealing with file name prompt and saving
        # Using loop to mitigate the user clicking 'cancel'
        # in the file name dialog prompt
        while True:
            # Opening the dialog prompt
            output_spreadsheet = filedialog.asksaveasfilename(
                filetypes=[('Excel files', '*.xlsx')],
                initialfile='Sipesat - Arsip Pengabdian Detil Risat.xlsx',
                title='Simpan sebagai ...'
            )

            # 'cancel' button in the dialog prompt is clicked
            if len(output_spreadsheet) == 0:
                # Showing confirmation
                x = messagebox.askyesno(
                    'Nama File Kosong',
                    'Apakah Anda yakin ingin melanjutkan tanpa menyimpan file spreadsheet hasil pemanenan?'
                )
                # Determining whether to break or to continue the loop
                # based on the inversed value of 'x'
                if x:
                    control.append_message_area(
                        f'+ Finalisasi pemanenan data tanpa menyimpan file spreadsheet luaran ...')
                    workbook.close()  # --- closing the workbook without saving
                    break
                else:
                    continue  # --- continuing the loop
            # File name does not end in spreadsheet extension
            elif output_spreadsheet[-5:] != '.xlsx':
                output_spreadsheet = output_spreadsheet + '.xlsx'

            # Saving the spreadsheet
            control.append_message_area(f'LOKASI_SPREADSHEET_LUARAN: {output_spreadsheet}')
            control.set_progress_bar(95)
            workbook.save(output_spreadsheet)

            # Closing the openpyxl workbook
            control.append_message_area(f'+ Menutup file spreadsheet ...')
            control.set_progress_bar(98)
            workbook.close()

            # Breaking the loop
            break

        # Notify for a successful scraping
        control.append_message_area(f'+ Pemanenan selesai pada: {str(dt.now())}')
        control.set_progress_bar(100)
        control.on_notify_successful_scraping()

    # This function harvests "Risat Laporan Akhir Pengabdian > Data Detil Lengkap" data
    # and then store the harvested data as an excel file
    #
    # Required arguments:
    # - control             --> for updating the progress bar and
    #                           message area of the screen SipesatScrHarvest
    # - username, password  --> the Risat administrator username and password
    def run_harvest_c_1_lapakhir(self, control, username, password):

        # SipesatScrHarvest messenger
        control.set_header_desc('Panen Data "Risat Laporan Akhir Pengabdian > Data Detil Lengkap"')
        control.set_help_label('Data sedang dipanen. Silahkan menunggu.')
        control.set_progress_bar(0)
        control.clear_message_area()

        # Preamble logging
        control.append_message_area(f'+ Memulai pemanenan data ...')
        control.append_message_area(f'+ Pemanenan dimulai pada: {str(dt.now())}')
        control.set_progress_bar(5)

        # Preparing the 'data_prompt' arrays
        control.append_message_area(f'+ Log masuk Risat sebagai [{username}] ...')
        control.set_progress_bar(10)
        data_prompt = self.get_risat_login(username, password)
        data_prompt = self.get_risat_pengabdian(data_prompt)
        data_prompt = self.get_risat_pengabdian_pelak_kegi(data_prompt)
        data_prompt = self.get_risat_pengabdian_pelak_kegi_lapakhir(data_prompt)

        # Parsing XML tree content
        control.append_message_area(f'+ Membaca halaman web ...')
        control.set_progress_bar(15)
        content = data_prompt['html_content']

        # Establishing the export spreadsheet file
        control.append_message_area(f'+ Mempersiapkan file spreadsheet luaran ...')
        control.set_progress_bar(20)
        workbook = xl.Workbook()
        sheet = workbook.active
        sheet.title = 'Lap. Akhir Detil Pengabdian'

        # Preparing the sheet header
        control.append_message_area(f'+ Mempersiapkan kepala lembar spreadsheet ...')
        control.set_progress_bar(25)
        # ---
        # Preparing the "IDENTITY" header
        sheet.merge_cells('A1:I1')
        sheet['A1'].value = 'IDENTITAS'
        sheet['A1'].alignment = Alignment(horizontal='center')
        # Preparing the "IDENTITY" sub-headers
        sheet['A2'].value = 'No.'
        sheet['B2'].value = 'Judul'
        sheet['C2'].value = 'Ketua'
        sheet['D2'].value = 'Bidang Fokus'
        sheet['E2'].value = 'Tgl. Usulan'
        sheet['F2'].value = 'Jumlah Anggota'
        sheet['G2'].value = 'Biaya'
        sheet['H2'].value = 'Dana Disetujui'
        sheet['I2'].value = 'Lama Kegiatan'

        # The base XPath location, pointing to each entry row
        base = '//div[@id="ContentPlaceHolder1_lapakhirbp3m1_updg1"]//div[@class="panel-body f12"]/table[@width="100%"]'

        # ---
        # Obtaining the data row values
        control.append_message_area(f'+ Mendapatkan data pada baris tabel ...')
        control.set_progress_bar(30)

        # HYPOTHESIS:
        # Xpath cannot detect 'tbody' element.
        # So instead of using 'table/tbody/tr', use 'table//tr' instead
        #
        # RESULT:
        # The hypothesis is correct.
        # Therefore, don't mention 'tbody' in any of the following Xpath paths

        a2 = [str(i)
              for i in range(1, len(content.xpath(base)) + 1)]

        b2 = [l.strip()
              for l in content.xpath(base + '//tr[@valign="top"]/td[@colspan="3"]/b/text()')]

        c2 = [l.replace('Ketua:', '').strip()
              for l in content.xpath(base + '//tr[2][not(descendant::table) and not(ancestor::table[@class="table"])]/td[2]/text()')]

        d2 = [l.replace('Bidang Fokus:', '').strip()
              for l in content.xpath(base + '//tr[3][not(descendant::table) and not(ancestor::table[@class="table"])]/td[@colspan="3"]/text()')]

        e2 = [l.replace('Tgl Usulan:', '').strip()
              for l in content.xpath(base + '//tr[4][not(descendant::table) and not(ancestor::table[@class="table"])]/td[@colspan="3"]/text()')]

        f2 = [l.replace('Jml Anggota:', '').strip()
              for l in content.xpath(base + '//tr[5][not(descendant::table) and not(ancestor::table[@class="table"])]/td[@colspan="3"]/text()')]

        g2 = [l.replace('Biaya:', '').replace('Rp.', '').replace(',', '').split('disetujui')[0].strip()
              for l in content.xpath(base + '//tr[6][not(descendant::table) and not(ancestor::table[@class="table"])]/td[@colspan="3"]/text()')]

        h2 = [l.replace('Biaya:', '').replace('Rp.', '').replace(',', '').split('disetujui')[1].strip()
              for l in content.xpath(base + '//tr[6][not(descendant::table) and not(ancestor::table[@class="table"])]/td[@colspan="3"]/text()')]

        i2 = [l.replace('Lama Kegiatan:', '').strip()
              for l in content.xpath(base + '//tr[7][not(descendant::table) and not(ancestor::table[@class="table"])]/td[@colspan="3"]/text()')]

        # The table which displays 'Luaran Laporan Akhir'
        # Treated differently, specially
        # 'j_' is a DOM element representing a HTML table
        j_ = content.xpath(base + '//tr[9]/td[@colspan="3"]/table[@class="table"]')

        # The starting row coordinate of the active sheet
        row_start = 3

        # DEBUG
        # Please comment out after use
        # ---
        '''
        print()
        print('-------------------------------------------------------------------------------------------------------')
        for a in (a2, b2, c2, d2, e2, f2, g2, h2, i2, j_):
            print('LEN', len(a))
            print('CONTENT', a)
            print()
        print('-------------------------------------------------------------------------------------------------------', end='')
        print()
        '''

        # The maximum number of 'LUARAN' table data row
        j_max_row = 0

        # Iterating through each table row and write to the table
        # Assumes the lists a1, b1, c1, ... have the same array size
        control.append_message_area(f'+ Melakukan iterasi terhadap baris tabel dan menulis spreadsheet luaran ...')
        control.set_progress_bar(35)
        for i in range(len(a2)):
            # Noisy preamble logging
            # Please don't use this -_-
            # ---
            # control.append_message_area(f'ITERASI [{i}]')

            # Updating the progress bar status
            control.set_progress_bar(35 + round(45 * (i + 1) / (len(a2))))

            # Painting the scraped data to the output spreadsheet row
            # Section 1: 'IDENTITY'
            sheet[f'A{row_start}'] = a2[i]
            sheet[f'B{row_start}'] = b2[i]
            sheet[f'C{row_start}'] = c2[i]
            sheet[f'D{row_start}'] = d2[i]
            sheet[f'E{row_start}'] = e2[i]
            sheet[f'F{row_start}'] = f2[i]
            sheet[f'G{row_start}'] = g2[i]
            sheet[f'H{row_start}'] = h2[i]
            sheet[f'I{row_start}'] = i2[i]

            # Painting the scraped data to the output spreadsheet row
            # Section 2: 'LUARAN'
            # ---
            # Checking if 'luaran' table data exists
            data_luaran = j_[i].xpath('.//td')
            if len(data_luaran) == 0:
                pass  # --- nope. the data does not exist
            else:

                # Checking if this table's 'LUARAN' data has the most rows
                j_all_rows = j_[i].xpath('.//tr[not(descendant::th)]')
                j_max_row = max(j_max_row, len(j_all_rows))

                # ROW DATA LABEL CONVENTION
                # j_row_a -> "No. Luaran"
                # j_row_b -> "Jenis Luaran"
                # j_row_c -> "Status Luaran"
                # j_row_d -> "File Laporan Akhir"
                # j_row_e -> "Alamat URL"
                # j_row_f -> "Status Akhir Luaran"
                # j_row_g -> "File PTJ Keuangan"

                j_row_a = [l.strip()
                      for l in j_[i].xpath('.//tr[not(descendant::th)]/td[1]/text()')]

                j_row_b = [l.strip()
                      for l in j_[i].xpath('.//tr[not(descendant::th)]/td[2]/text()')]

                j_row_c = [l.strip()
                      for l in j_[i].xpath('.//tr[not(descendant::th)]/td[3]/text()')]

                j_row_d = [l.strip()
                      for l in j_[i].xpath('.//tr[not(descendant::th)]/td[4]/a/@href')]

                j_row_e = [l.strip()
                      for l in j_[i].xpath('.//tr[not(descendant::th)]/td[5]/a/@href')]

                j_row_f = [l.strip()
                      for l in j_[i].xpath('.//tr[not(descendant::th)]/td[6]/text()')]

                j_row_g = [l.strip()
                      for l in j_[i].xpath('.//tr[not(descendant::th)]/td[7]/a/@href')]

                # The starting column coordinate for filling the 'LUARAN' table data
                col_start = 10

                # DEBUG
                # Please comment out after use
                # ---
                '''
                print()
                print(
                    '-------------------------------------------------------------------------------------------------------')
                for b in (j_row_a, j_row_b, j_row_c, j_row_d, j_row_e, j_row_f, j_row_g):
                    print('LEN', len(b))
                    print('CONTENT', b)
                    print()
                print(
                    '-------------------------------------------------------------------------------------------------------',
                    end='')
                print()
                '''

                # Iterating through each table row and write to the spreadsheet
                # Assumes the lists j_row_a, j_row_b, ... have the same array size
                for j in range(len(j_row_a)):

                    # Painting table data
                    sheet.cell(row=row_start, column=col_start).value = j_row_b[j]
                    sheet.cell(row=row_start, column=col_start+1).value = j_row_c[j]
                    sheet.cell(row=row_start, column=col_start+2).value = j_row_d[j]
                    sheet.cell(row=row_start, column=col_start+3).value = j_row_e[j]
                    sheet.cell(row=row_start, column=col_start+4).value = j_row_f[j]
                    sheet.cell(row=row_start, column=col_start+5).value = j_row_g[j]

                    # Incrementing the 'col_start' iterator before continuing the loop
                    col_start += 6
                    continue

            # Incrementing the 'row_start' iterator
            # Then continue the loop
            row_start += 1
            continue

        # Post-loop logging: appending the header over the 'LUARAN' columns
        control.append_message_area(f'+ Melengkapi kepala tabel pada bagian "LUARAN" ...')
        # ---
        # The starting column for the 'LUARAN' data
        col_start = 10
        # Beginning the loop that detects the maximum number of 'LUARAN' rows
        # according to the variable 'j_max_row'
        for i in range(1, j_max_row+1):
            # Setting the top header
            sheet.merge_cells(
                start_row=1,
                start_column=col_start,
                end_row=1,
                end_column=col_start+5
            )
            sheet.cell(row=1, column=col_start).value = f'LUARAN #{i}'
            sheet.cell(row=1, column=col_start).alignment = Alignment(horizontal='center')
            # Setting the sub headers
            sheet.cell(row=2, column=col_start).value = 'Jenis Luaran'
            sheet.cell(row=2, column=col_start+1).value = 'Status Luaran'
            sheet.cell(row=2, column=col_start+2).value = 'File Lap. Akhir'
            sheet.cell(row=2, column=col_start+3).value = 'Alamat URL'
            sheet.cell(row=2, column=col_start+4).value = 'Status Akhir Luaran'
            sheet.cell(row=2, column=col_start+5).value = 'File PTJ Keuangan'
            # Incrementing the 'col_start' iterator before continuing the loop
            col_start += 6
            continue

        # Post-loop logging: successfully painted the output spreadsheet file
        control.append_message_area(f'+ Tabel sukses dipanen!')
        control.set_progress_bar(85)

        # Asking for the spreadsheet name to save as
        # ---
        # Logging and setting the progress bar
        control.append_message_area(f'+ Menyimpan spreadsheet luaran ...')
        control.set_progress_bar(90)
        # Dealing with file name prompt and saving
        # Using loop to mitigate the user clicking 'cancel'
        # in the file name dialog prompt
        while True:
            # Opening the dialog prompt
            output_spreadsheet = filedialog.asksaveasfilename(
                filetypes=[('Excel files', '*.xlsx')],
                initialfile='Sipesat - Lap Akhir Pengabdian Detil Risat.xlsx',
                title='Simpan sebagai ...'
            )

            # 'cancel' button in the dialog prompt is clicked
            if len(output_spreadsheet) == 0:
                # Showing confirmation
                x = messagebox.askyesno(
                    'Nama File Kosong',
                    'Apakah Anda yakin ingin melanjutkan tanpa menyimpan file spreadsheet hasil pemanenan?'
                )
                # Determining whether to break or to continue the loop
                # based on the inversed value of 'x'
                if x:
                    control.append_message_area(
                        f'+ Finalisasi pemanenan data tanpa menyimpan file spreadsheet luaran ...')
                    workbook.close()  # --- closing the workbook without saving
                    break
                else:
                    continue  # --- continuing the loop
            # File name does not end in spreadsheet extension
            elif output_spreadsheet[-5:] != '.xlsx':
                output_spreadsheet = output_spreadsheet + '.xlsx'

            # Saving the spreadsheet
            control.append_message_area(f'LOKASI_SPREADSHEET_LUARAN: {output_spreadsheet}')
            control.set_progress_bar(95)
            workbook.save(output_spreadsheet)

            # Closing the openpyxl workbook
            control.append_message_area(f'+ Menutup file spreadsheet ...')
            control.set_progress_bar(98)
            workbook.close()

            # Breaking the loop
            break

        # Notify for a successful scraping
        control.append_message_area(f'+ Pemanenan selesai pada: {str(dt.now())}')
        control.set_progress_bar(100)
        control.on_notify_successful_scraping()

# -------------------------- CONSTANT PRESETS -------------------------- #

# 'FRAME_CLASSES' is a tuple that defines all the frame classes of the file
# - The first class specified in this tuple will be the frame displayed
#   at the very beginning of the application after launch
# - Need to have at least two classes, otherwise the following error will be casted:
#   TypeError: 'type' object is not iterable
FRAME_CLASSES = (SipesatScrLogin, SipesatScrMainMenu, SipesatScrComService, SipesatScrResearch, SipesatScrHarvest,)

# The following constants define font presets used in styling Tkinter widgets
FONT_HEADER_TITLE = ('Segoe UI', 30, 'bold')
FONT_HEADER_DESC = ('Segoe UI', 20, 'italic')
FONT_REGULAR = ('Segoe UI', 12)
FONT_FORM_INPUT = ('Courier New', 10, 'bold')
FONT_RADIO = ('Courier New', 10)
FONT_PROGRESS_VALUE = ('Courier New', 14)

# The following constants define the string presets used as template and localization
STRING_HEADER_TITLE = 'SiPe.Sat'
STRING_HEADER_DESC = 'Sistem Pemanen Risat'

# Constants that define application identity
APP_NAME = 'Sistem Pemanen Risat - UKSW 2023'

# Back-End constants that specify the temporary folder name prefix
# in which the recursively scraped 'detail' pages are stored into
SCRAPE_TEMP_DIR_PREFIX = 'sipesat-123-'

# -------------------------- DEVELOPMENT TEST -------------------------- #

# Modules import (for development purposes only)
from getpass import getpass as input_pass

# The development test class
class DevelopmentTest():

    # On 2023-03-28
    # Testing out opening the detail page of 'penelitian arsip' entry row
    def experiment_1(self):
        # Preamble
        print('+ Running experiment_1: Opening "Penelitian Arsip Detil" page ...')

        # Initializing the harvester back-end
        backend = BackEndHarvester()

        # Opening the 'Penelitian Arsip' page
        data_prompt = backend.get_risat_login(self.user_, self.pass_)
        data_prompt = backend.get_risat_penelitian(data_prompt)
        data_prompt = backend.get_risat_penelitian_arsip(data_prompt)

        # Setting of sample detail page
        data_prompt['button_name'] = 'ctl00$ContentPlaceHolder1$arsip1$repusul1$ctl01$btdetil1'
        data_prompt['idat_prop'] = 'ctl00$ContentPlaceHolder1$arsip1$repusul1$ctl01$idat1'
        data_prompt['idat_val'] = '12AAFE6D-7C03-4998-9CB6-1BF4ECA37831'
        data_prompt['itgl_prop'] = 'ctl00$ContentPlaceHolder1$arsip1$repusul1$ctl01$itgl1'
        data_prompt['itgl_val'] = '2021/01/21'

        # Opening the detail page of 'Penelitian Arsip'
        data_prompt = backend.get_risat_penelitian_arsip_detil(data_prompt)

        # Writing the HTTP response to an external file
        print('+ Dumping the HTTP response ...')
        fo = open(HTTP_RESPONSE_DUMPER, 'w')
        fo.write(str(data_prompt['http_response']))

    # On 2023-03-28
    # Testing scraping the 'Penelitian > Pelaksanaan Kegiatan > Lap. Akhir' page
    def experiment_2(self):
        # Preamble
        print('+ Running experiment_2: Opening "Penelitian Lap. Akhir" page ...')

        # Initializing the harvester back-end
        backend = BackEndHarvester()

        # Opening the 'Penelitian Lap. Akhir' page
        data_prompt = backend.get_risat_login(self.user_, self.pass_)
        data_prompt = backend.get_risat_penelitian(data_prompt)
        data_prompt = backend.get_risat_penelitian_pelak_kegi(data_prompt)
        data_prompt = backend.get_risat_penelitian_pelak_kegi_lapakhir(data_prompt)

        # Writing the the HTTP response to an external file
        print('+ Dumping the HTTP response ...')
        fo = open(HTTP_RESPONSE_DUMPER, 'w')
        fo.write(str(data_prompt['http_response']))

    # On 2023-03-28
    # Testing scraping the 'Pengabdian > Pelaksanaan Kegiatan > Lap. Akhir' page
    def experiment_3(self):
        # Preamble
        print('+ Running experiment_3: Opening "Pengabdian Lap. Akhir" page ...')

        # Initializing the harvester back-end
        backend = BackEndHarvester()

        # Opening the 'Pengabdian Lap. Akhir' page
        data_prompt = backend.get_risat_login(self.user_, self.pass_)
        data_prompt = backend.get_risat_pengabdian(data_prompt)
        data_prompt = backend.get_risat_pengabdian_pelak_kegi(data_prompt)
        data_prompt = backend.get_risat_pengabdian_pelak_kegi_lapakhir(data_prompt)

        # Writing the the HTTP response to an external file
        print('+ Dumping the HTTP response ...')
        fo = open(HTTP_RESPONSE_DUMPER, 'w')
        fo.write(str(data_prompt['http_response']))

    # On 2023-03-28
    # Testing out opening the detail page of 'pengabdian dana pengabdian' entry row
    def experiment_4(self):
        # Preamble
        print('+ Running experiment_4: Opening "Pengabdian Dana Pengabdian Detil" page ...')

        # Initializing the harvester back-end
        backend = BackEndHarvester()

        # Opening the 'Pengabdian Dana Pengabdian' page
        data_prompt = backend.get_risat_login(self.user_, self.pass_)
        data_prompt = backend.get_risat_pengabdian(data_prompt)
        data_prompt = backend.get_risat_pengabdian_dana_pengabdian(data_prompt)

        # Setting of sample detail page
        data_prompt['button_name'] = 'ctl00$ContentPlaceHolder1$danacair1$repusulan1$ctl01$btdetil1'
        data_prompt['kodetran_prop'] = 'ctl00$ContentPlaceHolder1$danacair1$repusulan1$ctl01$kodetran1'
        data_prompt['kodetran_val'] = 'D480D002-80B3-42A4-919F-08EDD11F62D5'
        data_prompt['stat_prop'] = 'ctl00$ContentPlaceHolder1$danacair1$repusulan1$ctl01$stat1'
        data_prompt['stat_val'] = 'C'

        # Opening the detail page of 'Pengabdian Dana Pengabdian'
        data_prompt = backend.get_risat_pengabdian_dana_pengabdian_detil(data_prompt)

        # Writing the HTTP response to an external file
        print('+ Dumping the HTTP response ...')
        fo = open(HTTP_RESPONSE_DUMPER, 'w')
        fo.write(str(data_prompt['http_response']))

    # On 2023-03-28
    # Testing out opening the detail page of 'pengabdian arsip' entry row
    def experiment_5(self):
        # Preamble
        print('+ Running experiment_5: Opening "Pengabdian Arsip Detil" page ...')

        # Initializing the harvester back-end
        backend = BackEndHarvester()

        # Opening the 'Pengabdian Arsip' page
        data_prompt = backend.get_risat_login('', '')
        data_prompt = backend.get_risat_pengabdian(data_prompt)
        data_prompt = backend.get_risat_pengabdian_arsip(data_prompt)

        # Setting of sample detail page
        data_prompt['button_name'] = 'ctl00$ContentPlaceHolder1$arsip1$repusul1$ctl01$btdetil1'
        data_prompt['idat_prop'] = 'ctl00$ContentPlaceHolder1$arsip1$repusul1$ctl01$idat1'
        data_prompt['idat_val'] = 'EB969B69-8F86-4790-995F-3833400C42D8'
        data_prompt['itgl_prop'] = 'ctl00$ContentPlaceHolder1$arsip1$repusul1$ctl01$itgl1'
        data_prompt['itgl_val'] = '2021/02/09'

        # Opening the detail page of 'Penelitian Arsip'
        data_prompt = backend.get_risat_pengabdian_arsip_detil(data_prompt)

        # Writing the HTTP response to an external file
        print('+ Dumping the HTTP response ...')
        fo = open(HTTP_RESPONSE_DUMPER, 'w')
        fo.write(str(data_prompt['http_response']))

    # On 2023-03-28
    # Testing out running the recursive autoscraper of 'Dana Penelitian'
    def experiment_6(self):
        # Preamble
        print('+ Running experiment_6: Running the recursive autoscraper of "Dana Penelitian" ...')

        # Initializing the harvester back-end
        backend = BackEndHarvester()

        # Getting the array of files of the scraped detail pages
        array_of_files = backend.get_auto_risat_detil('dana_penelitian', self.user_, self.pass_)

        # Printing the array
        print('+ Printing the array ...')
        print(array_of_files)

    # On 2023-05-22
    # Testing out the scraper of 'Dana Penelitian Detil'
    # Running the function 'test_harvest_r_1_dana'
    def experiment_7(self):
        # Preamble
        print('+ Running experiment_7: Testing out the scraper of "Dana Penelitian Detil" ...')

        # Initializing the harvester back-end
        backend = BackEndHarvester()

        # Getting the array of files of the scraped detail pages
        backend.test_harvest_r_1_dana(self.user_, self.pass_)

        # Done!
        print('+ Done!')

    # Wrapper of the developmental tester
    def launch_experiment_wrapper(self):

        # Processing the request
        # This switching-cases require Python version >= v3.10
        match self.nmbr_:
            case 1:
                self.experiment_1()
            case 2:
                self.experiment_2()
            case 3:
                self.experiment_3()
            case 4:
                self.experiment_4()
            case 5:
                self.experiment_5()
            case 6:
                self.experiment_6()
            case 7:
                self.experiment_7()
            case _:
                print('+ Error! Command not available!')

        # At the end of the experiment test, exit the app
        # so that the main GUI won't load
        print('+ Exitting the development test ...')
        exit()

    # The '__init__' function
    def __init__(self):
        # Preamble
        print('+ Beginning the development test ...')

        # Prompting username and password
        self.user_ = input('Please enter your username\n >>> ')
        self.pass_ = input_pass(f'Enter the password for [{self.user_}]\n >>> ')
        self.nmbr_ = int(input('Please enter the experiment number (integer)\n >>> '))

# Constants used in development tests only
HTTP_RESPONSE_DUMPER = '/tmp/http_dumper.html'

# ------------------------- APPLICATION LAUNCH ------------------------- #

# Development testing
# Not required for end-user usages
# Should be commented out by final release to the public
# dev = DevelopmentTest()
# dev.launch_experiment_wrapper()

# Initializing the GUI
app_gui = MainGUI()
app_gui.mainloop()
